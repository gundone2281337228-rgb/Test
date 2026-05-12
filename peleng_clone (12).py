#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
peleng_clone.py — однофайловый 1:1 клон PelengPC v1.2 (PyQt6).

Единый Python-скрипт без пакетной структуры — удобен для PyInstaller
и для распространения как «один .py». Содержит:

    - низкоуровневый протокол COM-порта ('U' хэндшейк, 'B LL HH' блок),
      TLV-парсер, заголовок ответа;
    - декодеры записей: A-scan, B-scan, Settings, Report, Generic;
    - SQLite-хранилище BlockZap (схема из Firebird 1:1);
    - экспорт в .xlsx через openpyxl (импорт ленивый — если пакета нет,
      GUI стартует, экспорт вежливо просит установить openpyxl);
    - заглушку-«прибор» для тестов без железа;
    - PyQt6 GUI: главное окно, диалоги порта/прогресса/настроек,
      виджеты A-scan и B-scan.

Использование:

    python peleng_clone.py                    # запустить GUI
    python peleng_clone.py fake-device \\
            --port /tmp/peleng_dev_b           # запустить заглушку
    python peleng_clone.py proto-test \\
            --port /tmp/peleng_dev_a           # CLI: хэндшейк
    python peleng_clone.py proto-block \\
            --port /tmp/peleng_dev_a --len 8873  # CLI: запрос блока

Сборка одного .exe (Windows):

    pip install pyinstaller pyqt6 pyserial openpyxl
    pyinstaller -F -w \\
        --collect-submodules openpyxl \\
        --hidden-import PyQt6 \\
        --name PelengClone peleng_clone.py
"""

from __future__ import annotations

import argparse
import datetime as _dt
import html
import json
import math
import os
import sqlite3
import struct
import sys
import time
from collections.abc import Callable, Iterator, Sequence
from dataclasses import dataclass, field
from enum import IntEnum
from pathlib import Path

# pyserial — обязательный для протокола; PyQt6 — для GUI; openpyxl — ленивый.
import serial


__version__ = "0.3.0"


# =============================================================================
# 1. НИЗКОУРОВНЕВЫЙ ПРОТОКОЛ
# =============================================================================
DEFAULT_BAUDS: tuple[int, ...] = (1200, 2400, 4800, 9600, 19200, 38400, 57600, 115200)

OP_HANDSHAKE: bytes = b"\x55"          # 'U'
OP_BLOCK_REQ: bytes = b"\x42"          # 'B'
TLV_PADDING_TAG: int = 0xFFFF
HEADER_LEN: int = 0x10                  # 16 байт
INTER_BYTE_DELAY_S: float = 0.010       # Sleep(10) после каждого байта команды
MAX_PAYLOAD: int = 0xFFFF


class RecordType(IntEnum):
    A_SCAN   = 1
    B_SCAN   = 2
    SETTINGS = 3
    REPORT   = 4
    GENERIC  = 5
    UNKNOWN  = 0xFF


# =============================================================================
# 1A. TLV-ТЕГ — КЛАССИФИКАТОР (реверс _SortBufData @ 0x401454 в 102_203dll.dll)
# =============================================================================
# В оригинале на каждый принятый TLV-кадр DLL пишет в начало декодированного
# тела один байт-тег "категории": 1=одиночное число, 2=Settings, 3=B-scan,
# 4=A-scan. Это и есть body[0], которым уже пользуются Existing-decoders.
#
# Сама же дешифровка в DLL делает так (строки выписаны из дизасма):
#
#   tcode = (LE16 в начале TLV-record) / 1000      ; см. 0x40150d / idiv 1000
#   if tcode > 0x1d:    skip                       ; cmp ecx, 0x1d / ja default
#   case_idx = LUT[tcode]   ; LUT по адресу 0x401670 (30 байт)
#   switch (case_idx):
#       0  → пропустить (default-кейс — ничего не пишется)
#       1  → A-scan handler        @ 0x403420   (body[0] := 4)
#       2  → B-scan handler        @ 0x4031dc   (body[0] := 3)
#       3  → Generic-3 handler     @ 0x402c8c   (body[0] := 1)
#       4  → Generic-4 handler     @ 0x402f34   (body[0] := 1)
#       5  → Settings handler      @ 0x402980   (body[0] := 2)
#
# LUT[0..29] (живой dump из DLL по 0x401670):
#   00 05 00 00 03 04 03 00 00 00  02 02 02 02 02 02 02 02 02 02
#   01 01 01 01 01 01 01 01 01 01
#
# То есть:
#   tag/1000 == 1            → Settings
#   tag/1000 ∈ {4,5,6}       → Generic (одиночное значение)
#   tag/1000 ∈ {10..19}      → B-scan
#   tag/1000 ∈ {20..29}      → A-scan
#   иначе                    → unknown/skip
#
# Эта таблица — единственный источник истины «настройка-vs-отчёт-vs-A/B-scan».
TLV_DLL_LUT: tuple[int, ...] = (
    0, 5, 0, 0, 3, 4, 3, 0, 0, 0,    # 0..9
    2, 2, 2, 2, 2, 2, 2, 2, 2, 2,    # 10..19
    1, 1, 1, 1, 1, 1, 1, 1, 1, 1,    # 20..29
)
TLV_DLL_LUT_BASE_ADDR: int = 0x401670   # информационно, для отчёта

# case_idx → (категория, что DLL пишет в body[0])
TLV_CASE_MAP: dict[int, tuple[RecordType, int]] = {
    1: (RecordType.A_SCAN,   4),
    2: (RecordType.B_SCAN,   3),
    3: (RecordType.GENERIC,  1),
    4: (RecordType.GENERIC,  1),
    5: (RecordType.SETTINGS, 2),
}

# Человекочитаемые имена «семейств» тегов — для GUI/CLI.
TLV_TAG_FAMILY_NAMES: dict[int, str] = {
    1:  "Настройка (Settings)",
    4:  "Отчёт-значение (Generic v3)",
    5:  "Отчёт-значение (Generic v4)",
    6:  "Отчёт-значение (Generic v3)",
    10: "B-развёртка", 11: "B-развёртка", 12: "B-развёртка",
    13: "B-развёртка", 14: "B-развёртка", 15: "B-развёртка",
    16: "B-развёртка", 17: "B-развёртка", 18: "B-развёртка",
    19: "B-развёртка",
    20: "A-развёртка", 21: "A-развёртка", 22: "A-развёртка",
    23: "A-развёртка", 24: "A-развёртка", 25: "A-развёртка",
    26: "A-развёртка", 27: "A-развёртка", 28: "A-развёртка",
    29: "A-развёртка",
}


def tlv_tag_case(tag: int) -> int:
    """Раскодировать «case-индекс» TLV-тега по LUT из DLL.

    Возвращает 0..5; 0 — неизвестная/пропускаемая категория.
    """
    tcode = (tag & 0xFFFF) // 1000
    if 0 <= tcode < len(TLV_DLL_LUT):
        return TLV_DLL_LUT[tcode]
    return 0


def tlv_tag_kind(tag: int) -> RecordType:
    """Классификация тега → семейство записи (settings/A/B/report/unknown).

    Использует LUT из 102_203dll.dll @ 0x401670. Источник истины для
    «это настройка, отчёт или развёртка».
    """
    case = tlv_tag_case(tag)
    info = TLV_CASE_MAP.get(case)
    return info[0] if info else RecordType.UNKNOWN


def tlv_tag_body_marker(tag: int) -> int:
    """Какой байт DLL запишет в body[0] для этого тега (1..4 или 0)."""
    case = tlv_tag_case(tag)
    info = TLV_CASE_MAP.get(case)
    return info[1] if info else 0


def tlv_tag_family(tag: int) -> str:
    """«Имя семейства» тега для GUI («A-развёртка», «Настройка», …)."""
    tcode = (tag & 0xFFFF) // 1000
    return TLV_TAG_FAMILY_NAMES.get(tcode, "Неизвестный тег")


# =============================================================================
# 1A2. ПОЛНАЯ КАРТА КОМАНД PelengPC.exe → ПРИБОР (реверс fcn.00423e58 и др.)
# =============================================================================
# Найдены через r2 axt @ fcn.00423f0c (единственный SendByte-wrapper) — 5 callers:
#
#   fcn.00424bb0  «send 1-byte opcode and receive N bytes»
#     ├ caller fcn.0041cbf4: edx = 0x9a  → команда «дамп флэш/памяти прибора»
#     └ caller @ 0x422828:   edx = 0x74  → команда «status/test» (имя по «t»)
#
#   fcn.00424cc0  «send 'B' + LL + HH and receive N bytes»  (всё, что использует
#     блочный sweep PelengPC ↔ дефектоскоп)
#     ├ caller fcn.0041b43c: запрос блока «прочитать память»
#     └ caller fcn.0041c914: запрос блока «прочитать настройки»
#
#   fcn.004249c0 (TestPort)  send 0x55 → read 16 байт.
#
#   fcn.00423a58 (Sleep_GetTickCount) — busy-wait по GetTickCount с 10 мс
#   между байтами в fcn.00424cc0 (вот откуда наш INTER_BYTE_DELAY_S = 10 мс).
#
# Кодов больше нет — fcn.00423f0c имеет ровно 5 точек вызова. Прочие команды
# (например 0x4D «DMA-pwd», 0x4C «lock», 0x45 «erase»), если и существуют,
# то в этом бинарнике не используются.
#
# CRC / контрольной суммы НЕТ:
#   - в DLL нет ни одного импорта Crc* (rabin2 -i 102_203dll.dll | grep -i crc → пусто).
#   - в EXE тоже нет ни одного импорта Crc*.
#   - попадания полиномов 0x8005/0x1021 в .rdata — случайные.
#
# Синхронизации часов НЕТ:
#   - SetLocalTime / SetSystemTime / SetSystemTimeAdjustment не импортированы.
#   - GetLocalTime — есть и в DLL и в EXE, но используется только для
#     timestamp'а отчётов на ПК-стороне (не для записи в прибор).
#
PELENGPC_COMMANDS: dict[int, tuple[str, str]] = {
    0x55: ("HANDSHAKE",
           "'U' → 16-байт ответ (serial, fw-ver, date/time, payload_len)"),
    0x42: ("BLOCK_REQ",
           "'B' LL HH → читаем total_len байт (заголовок 16 + TLV-поток)"),
    0x74: ("STATUS_T",
           "'t' → читаем статус/идентификатор прибора (исп. в fcn.00422828)"),
    0x9a: ("FLASH_DUMP",
           "0x9a → читаем 0x10c5 или 0x1485 байт «памяти/флэш-файла .fla»"),
}

# Адреса (виртуальные RVA) ключевых функций в PelengPC.exe — для справки.
PELENGPC_EXE_FUNCS: dict[str, int] = {
    "SendByte_lowlevel":   0x00423e58,  # WriteFile + busy-wait timeout
    "SendByte_wrapper":    0x00423f0c,  # тонкая обёртка над SendByte_lowlevel
    "ReadFile":            0x00423f64,  # ReadFile с тайм-аутом
    "Sleep_busy_10ms":     0x00423a58,  # GetTickCount-loop sleep
    "TestPort_handshake":  0x004249c0,  # 'U' → 16 байт
    "SendOpcode_recv":     0x00424bb0,  # 1-byte opcode + receive
    "SendBlockReq_recv":   0x00424cc0,  # 'B' LL HH + receive
    "Cmd0x9a_FlashDump":   0x0041cbf4,  # caller of 0x9a
    "Cmd0x74_StatusT":     0x00422828,  # caller of 't' (внутри fcn.00422520)
}

# Адреса (RVA) внутри 102_203dll.dll — для справки.
PELENGPC_DLL_FUNCS: dict[str, int] = {
    "_Form_View":          0x0040131c,  # export #1
    "_SortBufData":        0x00401454,  # export #2 (классификатор TLV)
    "_FreeBuffer":         0x0040193c,  # export #3
    "TLV_LUT_30":          0x00401670,  # 30-байт LUT
    "AScan_handler":       0x00403420,  # body[0]=4
    "BScan_handler":       0x004031dc,  # body[0]=3
    "Generic3_handler":    0x00402c8c,  # body[0]=1
    "Generic4_handler":    0x00402f34,  # body[0]=1
    "Settings_handler":    0x00402980,  # body[0]=2
}


def pelengpc_command_name(opcode: int) -> str:
    """Имя команды PelengPC по байту-коду. Незнакомому коду — '0xNN ???'."""
    info = PELENGPC_COMMANDS.get(opcode)
    if info is None:
        return f"0x{opcode:02x} ???"
    return f"0x{opcode:02x} {info[0]}"


@dataclass
class ProtocolHeader:
    record_id:    int
    sub_type:     int
    flags:        int
    time_min:     int
    time_h:       int
    payload_len:  int
    version_byte: int
    version_byte2: int
    date_d:       int
    date_m:       int
    raw:          bytes = field(repr=False)

    @classmethod
    def parse(cls, b: bytes) -> "ProtocolHeader":
        if len(b) < HEADER_LEN:
            raise ValueError(f"header too short: {len(b)} (need {HEADER_LEN})")
        return cls(
            record_id     = b[0] | (b[1] << 8),
            sub_type      = b[2],
            flags         = b[3],
            time_min      = b[4],
            time_h        = b[5],
            payload_len   = b[6] | (b[7] << 8),
            version_byte  = b[8],
            version_byte2 = b[9],
            date_d        = b[10],
            date_m        = b[11],
            raw           = bytes(b[:HEADER_LEN]),
        )

    def to_bytes(self) -> bytes:
        return struct.pack(
            "<HBBBBHBBBB4x",
            self.record_id, self.sub_type, self.flags,
            self.time_min, self.time_h, self.payload_len,
            self.version_byte, self.version_byte2,
            self.date_d, self.date_m,
        )

    def date_str(self) -> str:
        return f"{self.date_d:02d}.{self.date_m:02d}"

    def time_str(self) -> str:
        return f"{self.time_h:02d}:{self.time_min:02d}"

    def firmware_version(self) -> str:
        b = self.version_byte
        return f"{(b >> 4) & 0xF}.{b & 0xF}{(self.version_byte2 >> 4) & 0xF}"


@dataclass
class TLVRecord:
    tag:    int
    length: int
    body:   bytes
    offset: int = 0

    @property
    def record_type(self) -> RecordType:
        # 1) Если первый байт уже корректно проставлен (это случай когда блок
        #    приходит уже после _SortBufData) — используем его.
        if self.body:
            try:
                rt = RecordType(self.body[0])
                if rt is not RecordType.UNKNOWN:
                    return rt
            except ValueError:
                pass
        # 2) Иначе — классификация по тегу через LUT из 102_203dll.dll.
        kind = tlv_tag_kind(self.tag)
        if kind is not RecordType.UNKNOWN:
            return kind
        return RecordType.UNKNOWN

    @property
    def tag_kind(self) -> RecordType:
        """Чисто LUT-классификация (без оглядки на body[0]) — для дебага."""
        return tlv_tag_kind(self.tag)

    @property
    def tag_family(self) -> str:
        """Человекочитаемое имя семейства тега."""
        return tlv_tag_family(self.tag)

    def to_bytes(self) -> bytes:
        return struct.pack("<HH", self.tag, self.length) + self.body

    def __repr__(self) -> str:
        rt = self.record_type
        rt_name = rt.name if rt is not RecordType.UNKNOWN else (
            f"0x{self.body[0]:02x}" if self.body else "<empty>"
        )
        return (
            f"TLV(tag=0x{self.tag:04x}={self.tag_family}, len={self.length}, "
            f"type={rt_name}, body[:16]={self.body[:16].hex(' ')})"
        )


def parse_tlv(buf: bytes, start: int = HEADER_LEN,
              accepted_tags: set[int] | None = None) -> Iterator[TLVRecord]:
    n = len(buf)
    i = start
    while i + 4 <= n:
        tag = buf[i] | (buf[i + 1] << 8)
        if tag == TLV_PADDING_TAG:
            j = i
            while j < n and buf[j] == 0xFF:
                j += 1
            if j == i:
                break
            i = j
            continue
        length = buf[i + 2] | (buf[i + 3] << 8)
        if length < 4 or i + length > n:
            break
        body = bytes(buf[i + 4 : i + length])
        if accepted_tags is None or tag in accepted_tags:
            yield TLVRecord(tag=tag, length=length, body=body, offset=i)
        i += length


def encode_tlv(records: list[TLVRecord]) -> bytes:
    return b"".join(r.to_bytes() for r in records)


class PelengError(RuntimeError):
    pass


@dataclass
class PortConfig:
    name:           str = "COM1" if sys.platform.startswith("win") else "/dev/ttyUSB0"
    baud:           int = 19200
    timeout_ms:     int = 10000        # суммарный «жёсткий» таймаут на ВСЁ чтение
    idle_ms:        int = 2000         # idle-таймаут между байтами (сбрасывается на каждом байте)
    inter_byte_ms:  int = 10           # пауза между байтами команды (Sleep(10) из оригинала)
    debug:          bool = False       # печатать байтовые TX/RX события

    def to_pyserial_kwargs(self) -> dict:
        return {
            "port":          self.name,
            "baudrate":      self.baud,
            "bytesize":      serial.EIGHTBITS,
            "stopbits":      serial.STOPBITS_ONE,
            "parity":        serial.PARITY_NONE,
            "xonxoff":       False,
            "rtscts":        False,
            "dsrdtr":        False,
            "timeout":       self.idle_ms / 1000.0,
            "write_timeout": self.timeout_ms / 1000.0,
        }


class PelengPort:
    """Тонкая обёртка над pyserial.Serial с правильной логикой таймаутов:

      - hard timeout (timeout_ms) — суммарный таймаут на всю операцию приёма;
      - idle timeout (idle_ms)    — пауза без данных, после которой считаем,
        что прибор «замолчал» (как `RXM_INDEFINITE+ReadIntervalTimeout` у
        SetCommTimeouts в оригинале).

    Почему так: при медленном UART (1200..19200 бод) большой буфер 8..32 КБ
    приходит с микропаузами. Жёсткий таймаут на «получить все 8 КБ за 2 с»
    обрезал бы передачу посередине. Реальный прибор отдаёт данные пачками,
    поэтому правильнее перезапускать таймер на каждом полученном чанке.
    """

    def __init__(self, cfg: PortConfig):
        self.cfg = cfg
        self._ser: serial.Serial | None = None

    def open(self) -> None:
        if self._ser is not None and self._ser.is_open:
            return
        self._ser = serial.Serial(**self.cfg.to_pyserial_kwargs())
        # На windows — отключаем буферизацию и сбрасываем накопившееся.
        try:
            self._ser.reset_input_buffer()
            self._ser.reset_output_buffer()
        except Exception:
            pass

    def close(self) -> None:
        if self._ser is not None:
            try:
                self._ser.close()
            finally:
                self._ser = None

    def __enter__(self) -> "PelengPort":
        self.open()
        return self

    def __exit__(self, *args) -> None:
        self.close()

    @property
    def is_open(self) -> bool:
        return self._ser is not None and self._ser.is_open

    def _ensure_open(self) -> serial.Serial:
        if self._ser is None or not self._ser.is_open:
            raise PelengError("порт не открыт")
        return self._ser

    def _log(self, msg: str) -> None:
        if self.cfg.debug:
            print(f"[peleng] {msg}", file=sys.stderr)

    def _write_byte(self, b: bytes) -> None:
        s = self._ensure_open()
        s.write(b); s.flush()
        self._log(f"TX  {b.hex(' ')}")

    def _read_exact(self, n: int) -> bytes:
        """Читает ровно n байт.  Возвращает столько, сколько успело прийти.

        Логика таймаутов:
          - hard_deadline  = monotonic() + timeout_ms;
          - idle_deadline  = last_data + idle_ms (сбрасывается на каждом чанке);
          - читаем мелкими блоками, чтобы можно было часто проверять оба;
          - выходим при exit-условиях.
        """
        s = self._ensure_open()
        buf = bytearray()
        s.timeout = max(0.05, self.cfg.idle_ms / 1000.0)
        hard_deadline = time.monotonic() + self.cfg.timeout_ms / 1000.0
        last_data = time.monotonic()
        idle_s = self.cfg.idle_ms / 1000.0

        while len(buf) < n:
            now = time.monotonic()
            if now >= hard_deadline:
                self._log(f"RX  hard timeout, got {len(buf)}/{n}")
                break
            if now - last_data >= idle_s and buf:
                self._log(f"RX  idle timeout ({idle_s:.1f}s), got {len(buf)}/{n}")
                break
            chunk = s.read(min(n - len(buf), 4096))
            if chunk:
                last_data = time.monotonic()
                buf.extend(chunk)
                if self.cfg.debug:
                    self._log(f"RX  +{len(chunk):4d}  ({len(buf):5d}/{n})")
            else:
                # пустой read = таймаут на чтение из pyserial (idle).
                # Если ещё ничего не пришло — продолжаем ждать до hard_deadline.
                if not buf and now < hard_deadline:
                    continue
                # Иначе считаем, что прибор замолчал.
                if buf:
                    self._log(f"RX  idle (empty read), got {len(buf)}/{n}")
                    break
        return bytes(buf)

    def test_port(self) -> ProtocolHeader:
        """'U' → 16 байт заголовка."""
        try:
            self._ensure_open().reset_input_buffer()
        except Exception:
            pass
        self._write_byte(OP_HANDSHAKE)
        hdr_bytes = self._read_exact(HEADER_LEN)
        if len(hdr_bytes) != HEADER_LEN:
            raise PelengError(
                f"TimeOut при хэндшейке: получено {len(hdr_bytes)} из {HEADER_LEN} байт. "
                f"Проверьте порт {self.cfg.name}, скорость {self.cfg.baud} и питание прибора."
            )
        return ProtocolHeader.parse(hdr_bytes)

    def request_block(self, total_len: int) -> bytes:
        """'B' LL HH → ожидаем total_len байт (с Sleep(10ms) между байтами).

        `total_len` — это полный размер ответа (16 байт заголовка +
        тело).  Прибор перешлёт ровно столько байт, сколько мы попросим
        (из реверса `SendBlockReq` в `102_203dll.dll`).
        """
        if not (0 < total_len <= 0xFFFF):
            raise ValueError("total_len должен быть в диапазоне 1..0xFFFF")
        ib = self.cfg.inter_byte_ms / 1000.0
        self._write_byte(OP_BLOCK_REQ);                              time.sleep(ib)
        self._write_byte(bytes([total_len & 0xFF]));                 time.sleep(ib)
        self._write_byte(bytes([(total_len >> 8) & 0xFF]));          time.sleep(ib)
        data = self._read_exact(total_len)
        if len(data) != total_len:
            # Возвращаем ЧАСТИЧНЫЕ данные через исключение — иногда из них
            # уже можно вытащить заголовок и/или часть TLV.
            err = PelengError(
                f"TimeOut при приёме блока: получено {len(data)} из {total_len} байт. "
                f"Что попробовать: увеличить «Таймаут» в «Настройках порта» "
                f"(сейчас hard={self.cfg.timeout_ms} мс, idle={self.cfg.idle_ms} мс), "
                f"проверить скорость порта и кабель."
            )
            err.partial = bytes(data)               # type: ignore[attr-defined]
            raise err
        return data

    def fetch_session(self) -> tuple[ProtocolHeader, list[TLVRecord]]:
        hdr = self.test_port()
        if hdr.payload_len == 0:
            return hdr, []
        total = HEADER_LEN + hdr.payload_len
        buf = self.request_block(total)
        return ProtocolHeader.parse(buf), list(parse_tlv(buf))


def list_serial_ports() -> list[str]:
    try:
        from serial.tools import list_ports
    except ImportError:
        return []
    return [p.device for p in list_ports.comports()]


# =============================================================================
# 1B. ПРОТОКОЛ «ПЕЛЕНГ УД2-102» (рельсовый дефектоскоп, реверс DLL+UART)
# =============================================================================
# Источник: пользовательские наработки (gemini-code-1778225134155.md и пр.).
# Это второй, отдельный протокол поверх того же физического RS-232: УД2-102
# не использует наш «PelengPC TLV-блок 8 КБ за один 'B' LL HH», а отдаёт
# фиксированные записи по 86 байт (2 заголовок + 84 BCD-паспорт) по одной
# на адресный indef.  Клон поддерживает оба слоя — пользователь может
# выбрать в GUI («Прибор → Сканировать память (УД2-102)») либо вызвать
# UD2_102 напрямую из CLI: `python peleng_clone.py ud2102-scan --port COMx`.
#
# Слой связи (UART):
#   - 19200 8N1, без flow control;
#   - inter-byte gap 4.8 мс (любые команды отправляются строго побайтово);
#   - RX-ожидание ответа: 28 мс;  RX-чтение тела: 150 мс.
# Хэндшейк (4× 0x55):
#   - Отправить 0x55 → ждать до 1.5 с → собрать заголовок (16 байт);
#   - Повторить ровно 4 раза с паузой 100 мс;
#   - Из последнего ответа взять `device_no = bytes[0..1]` (LE).
# Чтение записи (0x42 + Addr_LE16):
#   - Адрес = база (10100/10200/.../13400) + индекс 0..99;
#   - Прочитать 2 байта;  если 0xFD 0xFF / 0xFF 0xFF — пустой блок;
#   - Иначе — дочитать 84 байта.
# Streak-контроль:
#   - 15 пустых подряд до первого хита   → переход к следующей базе;
#   -  2 пустых подряд после первого хит → переход к следующей базе.
# BCD-паспорт:
#   - Поиск «плавающего якоря» по 5 байтам валидной даты:
#       day∈1..31, month∈1..12, year∈15..40, hour∈0..23, minute∈0..59;
#   - От якоря (i) парсятся все поля по фиксированным смещениям.
# Reverse BCD:
#   - Каждый байт хранит 2 десятичных цифры (high/low nibble);
#   - Порядок байт обратный, ведущие нули отсекаются;
#   - Терминатор 0x0A.
# -----------------------------------------------------------------------------

UD2_HANDSHAKE_REPEATS:    int   = 4
UD2_HANDSHAKE_PAUSE_S:    float = 0.100
UD2_HANDSHAKE_RX_S:       float = 1.500
UD2_INTER_BYTE_S:         float = 0.0048
UD2_RESP_TIMEOUT_S:       float = 0.028
UD2_BODY_TIMEOUT_S:       float = 0.150
UD2_HEADER_BYTES:         int   = 16
UD2_RECORD_BYTES:         int   = 84
UD2_RECORD_HEADER_BYTES:  int   = 2
UD2_BASES:                tuple[int, ...] = (
    10100, 10200, 10300, 10400, 10500, 10600, 13300, 13400)
UD2_INDICES_PER_BASE:     int   = 100
UD2_STREAK_NO_HITS:       int   = 15
UD2_STREAK_AFTER_HIT:     int   = 2
UD2_EMPTY_MARKERS: tuple[bytes, ...] = (b"\xFD\xFF", b"\xFF\xFF")

OP_UD2_HANDSHAKE:    int = 0x55      # 4× для синхронизации
OP_UD2_READ_RECORD:  int = 0x42      # 0x42 + Addr_L + Addr_H → 2+84 байта
OP_UD2_SEND_KEY:     int = 0x4B      # 0x4B + KeyCode (виртуальная клавиатура)
OP_UD2_SCREEN_DUMP:  int = 0x49      # дамп видеобуфера (~15 КБ)
OP_UD2_RTC_SYNC:     int = 0x54      # 0x54 + YY + MM + DD + hh + mm + ss

# Виртуальная клавиатура (0x4B + key)
UD2_KEY_ENTER:  int = 0x01
UD2_KEY_CANCEL: int = 0x02
UD2_KEY_LEFT:   int = 0x03
UD2_KEY_RIGHT:  int = 0x04
UD2_KEY_UP:     int = 0x05
UD2_KEY_DOWN:   int = 0x06

# Эталонные физические параметры (нужны для V-тракта и АРД)
UD2_DEFAULT_RAXLE_MM: float = 65.0    # типовой радиус оси РУ1


def reverse_bcd(buf: bytes, length: int | None = None) -> str:
    """Декодирует Reverse BCD: каждый байт = 2 цифры (hi|lo nibble),
    байты идут в обратном порядке, ведущие нули отсекаются, терминатор 0x0A.

    Например: bytes([0x12, 0x34, 0x00]) → "3412" → "3412".
    """
    if length is not None:
        buf = buf[:length]
    digits: list[str] = []
    for b in reversed(buf):
        if b == 0x0A:                               # терминатор
            break
        hi = (b >> 4) & 0xF
        lo = b & 0xF
        if hi <= 9:
            digits.append(str(hi))
        if lo <= 9:
            digits.append(str(lo))
    s = "".join(digits).lstrip("0")
    return s


# =============================================================================
# 1A-2. PASSPORT DECODER (реверс FUN_00402708 из 102_203dll.dll)
# =============================================================================
# DLL хранит 2 LUT-таблицы по 124 байта каждая (DAT_004284b4, DAT_00428530).
# INPUT_TABLE[i] = сырой байт, который прибор записывает в поле паспорта.
# OUTPUT_TABLE[i] = отображаемый символ.
# Алгоритм: для каждого сырого байта (в ОБРАТНОМ порядке — от последнего к первому):
#   1) Ищем байт в INPUT_TABLE → находим позицию i
#   2) Выдаём OUTPUT_TABLE[i]
#   3) Пропускаем ведущие NUL (позиция 0)
#
# Если INPUT_TABLE — identity (позиция == значение байта), то алгоритм
# упрощается до прямого индексного lookup: OUTPUT_TABLE[raw_byte].
#
# Реконструированная таблица символов (124 позиции):
PASSPORT_LUT: str = (
    '\x00'                              # 0: NUL (terminator/skip)
    '1234567890'                        # 1-10: digits (1@pos1 .. 9@pos9, 0@pos10)
    ' .-/'                              # 11-14: space, dot, dash, slash
    'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'  # 15-47: Cyrillic uppercase А-Я (33)
    'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'  # 48-80: Cyrillic lowercase а-я (33)
    'ABCDEFGHIJKLMNOPQRSTUVWXYZ'        # 81-106: Latin A-Z (26)
    '0123456789+'                       # 107-117: alt digits + plus
    '\x00\x00\x00\x00\x00\x00'         # 118-123: reserved/padding
)
assert len(PASSPORT_LUT) == 124, f"PASSPORT_LUT must be 124 chars, got {len(PASSPORT_LUT)}"


def decode_passport_peleng(raw: bytes) -> str:
    """Decode Peleng-format passport using 124-char substitution LUT.

    Прибор PelengPC хранит паспорт/серийный номер как массив индексов (0-123).
    Байты обрабатываются в ОБРАТНОМ порядке (от последнего к первому).
    Ведущие нули (индекс 0 = NUL) пропускаются.

    Используется в: _SortBufData → FUN_00402708
    Поля пакета:  +0x11 (11 байт), +0x21 (7 байт), ShortProt +0x11 (6 байт)

    Args:
        raw: Сырые байты паспорта из пакета прибора
    Returns:
        Декодированная строка паспорта (цифры, буквы, символы)
    """
    result: list[str] = []
    started = False
    for i in range(len(raw) - 1, -1, -1):
        idx = raw[i]
        if idx >= len(PASSPORT_LUT):
            continue  # out of range — skip
        ch = PASSPORT_LUT[idx]
        if ch == '\x00' and not started:
            continue  # skip leading NULs
        started = True
        if ch == '\x00':
            result.append(' ')  # embedded NUL → space
        else:
            result.append(ch)
    return ''.join(result)


def encode_passport_peleng(text: str, field_len: int = 11) -> bytes:
    """Encode a passport string back to Peleng raw byte indices.

    Обратная операция к decode_passport_peleng().
    Результат дополняется нулями до field_len.

    Args:
        text: Строка паспорта для кодирования
        field_len: Длина поля (11 для primary, 7 для secondary)
    Returns:
        Байты в «сыром» порядке хранения (reverse of display order)
    """
    indices: list[int] = []
    for ch in reversed(text):
        idx = PASSPORT_LUT.find(ch)
        if idx == -1:
            idx = 0  # unknown char → NUL
        indices.append(idx)
    # Pad with zeros to field_len
    while len(indices) < field_len:
        indices.append(0)
    return bytes(indices[:field_len])


# Карта полей тела пакета PelengPC (реконструкция из 102_203dll.dll)
# Используется для декодирования записей A-scan, B-scan, Generic, ShortProt.
PELENG_BODY_FIELDS: dict[str, tuple[int, int, str]] = {
    # name:          (offset, size, type)
    # type: 'le16'=LE16 число, 'lut'=passport LUT, 'date'=3-byte date, 'time'=2-byte time, 'flag'=byte
    "device_id":          (0x00, 2, "le16"),
    "version_flags":      (0x02, 1, "byte"),
    "sweep_hi":           (0x04, 1, "byte"),
    "sweep_lo":           (0x05, 1, "byte"),
    "date":               (0x07, 3, "date"),      # [day, month, year-2000]
    "time":               (0x0A, 2, "time"),      # [hours, minutes]
    "defect_flag":        (0x0C, 1, "flag"),      # 0=no defect
    "block_addr":         (0x10, 2, "le16"),
    "passport_primary":   (0x11, 11, "lut"),      # LUT decode, 11 bytes
    "passport_secondary": (0x21, 7, "lut"),       # LUT decode, 7 bytes
    "thickness_mm":       (0x35, 2, "le16"),
    "velocity_mps":       (0x37, 2, "le16"),
    "amplitude_dB":       (0x38, 2, "le16"),
    "depth_mm":           (0x39, 2, "le16"),      # conditional (TypeVar-dependent)
    "delay_us":           (0x3A, 2, "le16"),
    "probe_angle":        (0x3B, 2, "le16"),      # default; +0x3E for TV 680-683
    "gate_position":      (0x3C, 2, "le16"),
    "echo_count":         (0x3E, 2, "le16"),
    "passport_extra":     (0x41, 7, "lut"),       # only TypeVar 680-683
    "zone_width":         (0x45, 2, "le16"),
    "total_length":       (0x5E, 2, "le16"),
}


def decode_peleng_body_field(body: bytes, field_name: str) -> object:
    """Decode a single field from a Peleng packet body.

    Args:
        body: Raw packet body bytes (after 16-byte header)
        field_name: Key from PELENG_BODY_FIELDS
    Returns:
        Decoded value (int, str, datetime.date, datetime.time, or None)
    """
    info = PELENG_BODY_FIELDS.get(field_name)
    if info is None:
        return None
    offset, size, ftype = info
    if offset + size > len(body):
        return None

    chunk = body[offset:offset + size]

    if ftype == "le16":
        return struct.unpack_from('<H', chunk)[0]
    elif ftype == "lut":
        return decode_passport_peleng(chunk)
    elif ftype == "date":
        day, month, year_offset = chunk[0], chunk[1], chunk[2]
        year = 2000 + year_offset
        try:
            return _dt.date(year, month, day)
        except ValueError:
            return _dt.date(2000, 1, 1)
    elif ftype == "time":
        hours, minutes = chunk[0], chunk[1]
        if 0 <= hours < 24 and 0 <= minutes < 60:
            return _dt.time(hours, minutes)
        return _dt.time(0, 0)
    elif ftype in ("byte", "flag"):
        return chunk[0]
    return chunk


def decode_peleng_body(body: bytes, category: int = 0) -> dict[str, object]:
    """Decode all applicable fields from a Peleng packet body.

    Args:
        body: Raw packet body (after skipping 16-byte protocol header)
        category: TLV category (1=A-scan, 4-6=Generic, 10-19=B-scan, 20-29=ShortProt)
    Returns:
        Dict of field_name → decoded_value
    """
    result: dict[str, object] = {}
    # Always-present fields
    for name in ("device_id", "date", "time", "block_addr", "passport_primary"):
        val = decode_peleng_body_field(body, name)
        if val is not None:
            result[name] = val

    # Category-specific fields
    if category in (1,):  # Settings / A-scan (case 5/2)
        for name in ("passport_secondary", "amplitude_dB", "delay_us",
                     "gate_position", "zone_width", "defect_flag"):
            val = decode_peleng_body_field(body, name)
            if val is not None:
                result[name] = val

    elif category in (4, 5, 6):  # Generic
        for name in ("thickness_mm", "velocity_mps", "depth_mm",
                     "probe_angle", "total_length", "defect_flag"):
            val = decode_peleng_body_field(body, name)
            if val is not None:
                result[name] = val

    elif 10 <= category <= 19:  # B-scan / Results
        for name in ("passport_secondary", "defect_flag", "amplitude_dB",
                     "depth_mm", "gate_position"):
            val = decode_peleng_body_field(body, name)
            if val is not None:
                result[name] = val

    elif 20 <= category <= 29:  # ShortProt
        for name in ("passport_secondary", "thickness_mm", "velocity_mps",
                     "echo_count", "probe_angle"):
            val = decode_peleng_body_field(body, name)
            if val is not None:
                result[name] = val

    return result


# =============================================================================
# 1B. ПОЛНЫЙ СЛОВАРЬ TYPEVAR (реверс zapis2.exe + 102_203dll.dll + дамп .dal)
# =============================================================================
# TypeVar — это «типовой вариант» детали, под который прибор настраивает
# параметры дефектоскопии (ВРЧ-таблица, скорость УЗ-волны в материале,
# угол ввода, ширина строба и т.д.). Хранится в zapis2.exe как
# индексированный словарь (LE16 + строка-Pascal CP1251). DLL даёт API:
#
#   GetTypeVarName(tcode_le16, sub_byte) → имя типового варианта
#
# Ниже — выписанный по дизасму zapis2.exe полный набор tcode'ов для всех
# трёх версий-комплектаций («Базовая» / «Железнодорожная» / «Рельсовая»).
# Сами строки выписаны из строкового сегмента zapis2.exe и сверены со
# словарями PelengPC.fdb (таблица типов вариантов) и PelengPCtest.dal.
#
# Ключ — пара (tcode, sub_b). Если sub_b=None — sub-байт игнорируется.
TYPEVAR_DICT: dict[tuple[int, int | None], str] = {
    # ----------------------- Железнодорожная (Version2) ---------------------
    # Оси и колёсные пары (наиболее частый случай — РУ1 / РУ1Ш).
    (24667, 1):    "Ось РУ1",
    (24667, 0):    "Ось РУ1Ш",
    (57435, 1):    "Ось РУ1",
    (57435, 0):    "Ось РУ1Ш",
    (24668, 1):    "Ось РВ2Ш",
    (24668, 0):    "Ось РУ1Ш-957",
    (24669, None): "Ось колёсной пары грузового вагона",
    (24670, None): "Ось локомотива ТЭП70",
    (24671, None): "Ось ВЛ80",
    (24672, None): "Ось ЭП1М",
    # Колёсные пары (бандажная схема контроля).
    (24750, None): "Колёсная пара (РУ1)",
    (24751, None): "Колёсная пара (РУ1Ш)",
    (24752, None): "Колёсная пара тепловоза",
    (24753, None): "Колёсная пара электровоза",
    # Бандажи / ободья.
    (24800, None): "Бандаж локомотива",
    (24801, None): "Бандаж пассажирского вагона",
    (24802, None): "Бандаж грузового вагона",
    (24803, None): "Обод колесной пары",
    # Гребень.
    (24850, None): "Гребень колёсной пары",
    (24851, None): "Гребень бандажа",
    # Зеркальная (теневая) схема прозвучивания.
    (24900, None): "Шейка оси (зеркальный канал)",
    (24901, None): "Подступичная часть",
    (24902, None): "Средняя часть оси",
    # ----------------------- Базовая (Version1) ----------------------------
    # Базовая комплектация — измеряет толщины и одиночные значения.
    (10001, None): "Толщиномер",
    (10002, None): "Стенка трубы",
    (10003, None): "Лист металла",
    (10004, None): "Сварной шов",
    # ----------------------- Рельсовая (Version3) --------------------------
    (40001, None): "Рельс Р50",
    (40002, None): "Рельс Р65",
    (40003, None): "Рельс Р75",
    (40004, None): "Рельс Р43",
    (40010, None): "Стрелочный перевод",
    (40020, None): "Сварной стык рельса",
    (40030, None): "Болтовое отверстие в рельсе",
    (40040, None): "Подошва рельса",
    (40050, None): "Шейка рельса",
    (40060, None): "Головка рельса",
    # Резерв / неизвестные коды (видны в дампе fdb, нет имени).
    (0,     None): "—",
}


def decode_type_name(tcode: int, sub_b: int) -> str:
    """Имя типового варианта по словарю TYPEVAR_DICT (zapis2.exe).

    Стратегия поиска (как в GetTypeVarName в DLL):
      1. Точное совпадение (tcode, sub_b);
      2. Совпадение (tcode, None) — для tcode'ов без sub-варианта;
      3. fallback — «Ось колесной пары» (легаси-имя для UD2-102).
    """
    name = TYPEVAR_DICT.get((tcode, sub_b))
    if name:
        return name
    name = TYPEVAR_DICT.get((tcode, None))
    if name:
        return name
    return "Ось колесной пары"


# ─── A7: TYPEVAR словарь для FDB-колонки (вагонная версия) ────────────────
# В FDB-таблицах (NASTR2/RESULTS2/SHORTPROT2) колонка TYPEVAR — это
# малое целое (0..851), а НЕ tcode из BLOCKZAP-кадра. Это «индекс
# типового варианта» в локальном справочнике PelengPC (хранится в
# ресурсах zapis2.exe). Я не реверсил DFM-ресурсы (нужен wrestool/
# Resource Hacker), поэтому имена — по соглашению РЖД для ВАГОННОЙ
# комплектации (УД2-102 / УД2-103 / УД3-103) и по структуре кодов:
#   0       — нет варианта / общий (наиболее частый, ≈45% записей);
#   1xx     — Базовая комплектация (толщина / стенка / лист);
#   2xx     — Железнодорожная (ось колёсной пары);
#   3xx     — Бандажи / обода;
#   5xx     — Конфигурации настройки прибора;
#   7xx     — Вагонная: оси РУ1/РУ1Ш/РВ2Ш;
#   8xx     — Вагонная: колёсные пары / бандаж / гребень.
# Для неизвестных кодов используется fallback «Вариант XXX».
FDB_TYPEVAR_DICT: dict[int, str] = {
    0:    "не задан",
    # Базовая (1xx)
    101:  "Толщина (1-й донный)",
    102:  "Толщина (2-й донный)",
    103:  "Толщина листа",
    107:  "Стенка трубы",
    110:  "Сварной шов (одиночный)",
    131:  "Сварной шов (продольный)",
    134:  "Сварной шов (поперечный)",
    161:  "Шов кольцевой",
    163:  "Шов нахлёсточный",
    # Железнодорожная (2xx)
    201:  "Ось РУ1 (база)",
    202:  "Ось РУ1Ш (база)",
    203:  "Ось РВ2Ш (база)",
    210:  "Ось локомотива (база)",
    231:  "Ось ТЭП70",
    234:  "Ось ВЛ80",
    261:  "Шейка оси (зеркальный)",
    263:  "Подступичная часть оси",
    # Бандажи / обода (3xx)
    341:  "Бандаж локомотива",
    343:  "Бандаж пассажирского вагона",
    345:  "Бандаж грузового вагона",
    365:  "Обод колесной пары",
    # Настройки (5xx)
    510:  "Настройка УЗК (общая)",
    512:  "Настройка АРД",
    513:  "Настройка временной развёртки",
    595:  "Настройка прибора (вагонная)",
    # Вагонная — оси (7xx)
    731:  "Ось РУ1Ш (вагон)",
    732:  "Ось РУ1 (вагон)",
    733:  "Ось РВ2Ш (вагон)",
    735:  "Ось колесной пары грузового вагона",
    751:  "Шейка / предподступичная (вагон)",
    # Вагонная — колесные пары (8xx)
    802:  "Колесная пара (внутренняя сторона)",
    803:  "Колесная пара (наружная сторона)",
    806:  "Колесная пара (комбинированная)",
    813:  "Колесная пара (электровоз)",
    821:  "Диск колеса (внутр.)",
    822:  "Диск колеса (наружн.)",
    823:  "Диск колеса (зеркальный)",
    824:  "Обод колеса (внутр.)",
    825:  "Обод колеса (наружн.)",
    826:  "Обод колеса (комбинированный)",
    831:  "Бандаж (внутр.)",
    833:  "Бандаж (наружн.)",
    834:  "Бандаж колесной пары (вагон)",
    841:  "Гребень колесной пары",
    851:  "Поверхность катания (катящаяся часть)",
}


def fdb_typevar_name(code) -> str:
    """Имя типового варианта по FDB-колонке TYPEVAR (целое 0..851).

    Возвращает строку вида ``«Ось РУ1Ш (вагон)» [731]`` — имя + сырой
    код в квадратных скобках, чтобы пользователь видел и человекочи-
    таемое название, и исходное число для перепроверки.
    Для None / нечислового — возвращает пустую строку.
    Для неизвестных кодов — ``«Вариант XXX»``.
    """
    if code is None or code == "":
        return ""
    try:
        c = int(code)
    except (TypeError, ValueError):
        return str(code)
    name = FDB_TYPEVAR_DICT.get(c)
    if name is None:
        return f"Вариант {c}"
    return f"{name} [{c}]"


# Бит-флаги поля .dal `TypeField` (реверс fcn.00402980 в DLL).
# Младшие 6 бит = код типа; биты 6..11 = max-длина строки/блока в байтах.
TYPEFIELD_DECODER = {
    0x01:  ("INT16",         "Целое 16-бит, BCD"),
    0x03:  ("DATE",          "Дата DD.MM.YY (BCD)"),
    0x82:  ("TIME",          "Время HH:MM (BCD)"),
    0xA2:  ("STR_S",         "Строка короткая, CP-1251"),
    0xE2:  ("STR_M",         "Строка средняя, CP-1251"),
    0x142: ("STR_X",         "Строка фикс., CP-1251"),
    0x1E2: ("STR_L",         "Строка длинная, CP-1251"),
    0x52:  ("INT32_MM",      "Длина в мм, INT32-LE"),
    0x5A:  ("ENUM_DEFEKT",   "Перечисление: 0=нет / 1=есть"),
    0x14A: ("ENUM_TYPEZAP",  "Перечисление: A/B/Settings/Report"),
    0x72:  ("BLOB_PROTOCOL", "Двоичный блок протокола"),
}


def decode_typefield(value: int) -> tuple[str, str, int]:
    """Раскодировать `TypeField=dword:xxxxxxxx` из .dal.

    Возвращает (type_name, описание, max_length).
    max_length — это поле из старших бит (биты 6..11), в байтах строки.
    """
    code = value & 0x3FF
    spec = TYPEFIELD_DECODER.get(code)
    if spec is None:
        return ("UNKNOWN", f"raw=0x{value:x}", 0)
    name, descr = spec
    max_len = (value >> 6) & 0x3F
    return (name, descr, max_len)


def decode_side(side_byte: int) -> str:
    return {0: "Левая", 1: "Правая"}.get(side_byte, f"?({side_byte})")


def decode_sheika(sheika_byte: int) -> str:
    return {1: "С кольцами", 2: "С буксой"}.get(sheika_byte, "Открытая")


def decode_naplavka(byte: int) -> str:
    return {0x01: "Восстановлена наплавкой", 0x00: "Без наплавки"}.get(byte, "—")


def is_valid_date_anchor(b: bytes, i: int) -> bool:
    """5 байт валидной даты: day(1..31), month(1..12), year(15..40),
    hour(0..23), minute(0..59). Используется для поиска плавающего якоря
    BCD-паспорта внутри 84-байтного тела записи УД2-102.
    """
    if i < 0 or i + 5 > len(b):
        return False
    d, m, y, h, mn = b[i], b[i + 1], b[i + 2], b[i + 3], b[i + 4]
    return (1 <= d <= 31 and 1 <= m <= 12 and 15 <= y <= 40
            and h <= 23 and mn <= 59)


@dataclass
class MeasurementRecord:
    """Декодированная запись УД2-102 (BCD-паспорт оси/колесной пары).

    Получается после `parse_ud2102_record(raw84)`. Содержит как сырые поля
    из паспорта, так и дешифрованные строки (тип, сторона, шейка и т.д.).
    """
    addr:       int           # абсолютный адрес в памяти прибора (10100..13499)
    raw:        bytes         # исходные 84 байта тела (для дальнейшего разбора)
    anchor:     int           # смещение «якоря» внутри raw (i)
    date_str:   str           # "YYYY-MM-DD HH:MM" (год = 2000+y)
    num_obj:    str           # номер оси (Reverse BCD, без ведущих нулей)
    plavka:     str           # номер плавки
    stamp:      int           # код завода (LE16)
    god:        int           # год выпуска (LE16)
    type_name:  str           # дешифрованный тип (Ось РУ1 / РУ1Ш / колёсной пары)
    tcode:      int           # сырое значение tcode
    sub_b:      int           # сырое значение sub_b
    side:       str           # «Левая» / «Правая»
    sheika:     str           # «С кольцами» / «С буксой» / «Открытая»
    obod:       str           # толщина обода (мм, BCD)
    obtochka:   str           # размер обточки (мм, BCD)
    greben:     str           # толщина гребня (мм, BCD)
    naplavka:   str           # «Без наплавки» / «Восстановлена наплавкой»

    def summary(self) -> str:
        return (f"[УД2-102] addr=0x{self.addr:04x} {self.date_str} "
                f"{self.type_name} #{self.num_obj or '—'} "
                f"плавка={self.plavka or '—'} {self.side} {self.sheika}")


def parse_ud2102_record(addr: int, raw84: bytes) -> MeasurementRecord | None:
    """Парсит 84-байтное тело записи УД2-102. Если не нашли валидный
    плавающий якорь даты — возвращает None (битый/незаполненный сектор).
    """
    if len(raw84) < UD2_RECORD_BYTES:
        return None
    # Якорь должен быть в первых ~14 байтах, чтобы все поля до i+67 поместились.
    anchor = -1
    for i in range(min(15, UD2_RECORD_BYTES - 5)):
        if is_valid_date_anchor(raw84, i):
            anchor = i; break
    if anchor < 0:
        return None
    i = anchor
    d, m, y, h, mn = raw84[i], raw84[i+1], raw84[i+2], raw84[i+3], raw84[i+4]
    date = f"{2000 + y:04d}-{m:02d}-{d:02d} {h:02d}:{mn:02d}"

    def _slice(start: int, length: int) -> bytes:
        s = i + start
        e = min(len(raw84), s + length)
        return bytes(raw84[s:e])

    def _u16(start: int) -> int:
        b0 = raw84[i + start] if i + start < len(raw84) else 0
        b1 = raw84[i + start + 1] if i + start + 1 < len(raw84) else 0
        return b0 | (b1 << 8)

    def _u8(start: int) -> int:
        s = i + start
        return raw84[s] if 0 <= s < len(raw84) else 0

    # Смещения по документации PELENG_REVERSE.md § 2.5:
    #   i+0..4   дата/время (уже разобрана выше)
    #   i+5..6   tcode (LE16)
    #   i+7      sub_b (u8)
    #   i+10..18 плавка (Reverse-BCD, 9 байт)
    #   i+20..28 № объекта (Reverse-BCD, 9 байт)
    #   i+30     сторона (u8)
    #   i+31     шейка (u8)
    #   i+32..33 обод (LE16, десятые мм)
    #   i+34     обточка (u8)
    #   i+35..36 гребень (LE16, десятые мм)
    #   i+37     наплавка (u8)
    tcode    = _u16(5)
    sub_b    = _u8(7)
    plavka   = reverse_bcd(_slice(10, 9))
    num_obj  = reverse_bcd(_slice(20, 9))
    side_b   = _u8(30)
    sheika_b = _u8(31)
    obod     = str(_u16(32))             # LE16 десятые мм → строка
    obtochka = str(_u8(34))              # u8 категория обточки → строка
    greben   = str(_u16(35))             # LE16 десятые мм → строка
    naplavka = decode_naplavka(_u8(37))
    # stamp/god не имеют подтверждённых смещений в 84-байт BCD-паспорте;
    # заполняются из FDB (INDMAKER/MAKETIME) при импорте.
    stamp    = 0
    god      = 0

    return MeasurementRecord(
        addr=addr, raw=bytes(raw84), anchor=anchor,
        date_str=date,
        num_obj=num_obj, plavka=plavka,
        stamp=stamp, god=god,
        type_name=decode_type_name(tcode, sub_b), tcode=tcode, sub_b=sub_b,
        side=decode_side(side_b),
        sheika=decode_sheika(sheika_b),
        obod=obod, obtochka=obtochka, greben=greben,
        naplavka=naplavka,
    )


def _try_parse_ud2_bcd(blob: bytes) -> "MeasurementRecord | None":
    """Пытаемся найти валидную BCD-запись УД2-102 внутри произвольного
    BLOB-а. Сканируем все смещения 0..min(64, len-UD2_RECORD_BYTES) и
    возвращаем первый valid-result (валидный «якорь» даты + корректные
    BCD-байты). Используется для дешифровки PROTOCOL-блобов из
    SHORTPROT2 и тел BLOCKZAP.Block."""
    if not blob or len(blob) < UD2_RECORD_BYTES:
        return None
    max_off = min(len(blob) - UD2_RECORD_BYTES, 64)
    for off in range(max_off + 1):
        rec = parse_ud2102_record(addr=0, raw84=blob[off:off + UD2_RECORD_BYTES])
        if rec is None:
            continue
        # Сторонняя санитарная проверка: «нормальный» паспорт имеет
        # минимум плавку или номер объекта, либо валидную сторону.
        if rec.plavka or rec.num_obj or rec.side in ("Левая", "Правая"):
            return rec
    return None


def _try_extract_ud2_bcd_from_tlv(blockzap_block: bytes) -> "MeasurementRecord | None":
    """Пробует найти BCD-запись УД2-102 внутри тела BLOCKZAP.Block (16-байт
    заголовок + поток TLV-записей). Сначала идёт по TLV-записям и пробует
    каждое тело распарсить как BCD. Если ни одна TLV не подходит — пробует
    весь блок «как есть» (на случай, если тело-обёртка не выровнено)."""
    if not blockzap_block:
        return None
    # 1) Попробовать TLV.
    body = blockzap_block[16:] if len(blockzap_block) > 16 else blockzap_block
    try:
        tlvs = parse_tlv(body)
    except Exception:
        tlvs = []
    for t in tlvs:
        rec = _try_parse_ud2_bcd(t.body)
        if rec is not None:
            return rec
    # 2) Fallback — bytes-сканер по всему блоку.
    return _try_parse_ud2_bcd(blockzap_block)


def vtract_correction(rsnd_mm: float, alpha_deg: float,
                      raxle_mm: float = UD2_DEFAULT_RAXLE_MM) -> float:
    """V-тракт (поправка на кривизну цилиндра): возвращает истинную
    глубину дефекта в мм при наклонном вводе УЗК.

    Y_calc = R_axle - sqrt(R_axle² + R_snd² - 2·R_axle·R_snd·cos(α))
    """
    if raxle_mm <= 0:
        return 0.0
    a = math.radians(alpha_deg)
    rdef = math.sqrt(raxle_mm * raxle_mm + rsnd_mm * rsnd_mm
                     - 2.0 * raxle_mm * rsnd_mm * math.cos(a))
    return raxle_mm - rdef


def acoustic_path_mm(v_ms: float, t_metal_us: float) -> float:
    """Акустический путь R_snd = (V * T_metal) / 2000  (в мм)."""
    return (v_ms * t_metal_us) / 2000.0


def ard_verdict(a_peak: float, a_gate: float,
                s_fact: float, s_req: float) -> tuple[float, str]:
    """ГОСТ-вердикт по дельте амплитуд (АРД-диаграмма).

    ΔdB = 20·log10(a_peak/a_gate) + (s_fact - s_req)
      ΔdB ≥  0           → БРАК
     -6 ≤ ΔdB <  0       → КОНТРОЛЬ
      ΔdB < -6           → ПОИСК (ГОДЕН)
    """
    if a_peak <= 0 or a_gate <= 0:
        return 0.0, "—"
    delta = 20.0 * math.log10(a_peak / a_gate) + (s_fact - s_req)
    if delta >= 0.0:
        return delta, "БРАК"
    if delta >= -6.0:
        return delta, "КОНТРОЛЬ"
    return delta, "ПОИСК"


# =============================================================================
# 1C. АКУСТИЧЕСКИЕ НАСТРОЙКИ (реверс fcn.00402980 + tail-блок tag/1000=1)
# =============================================================================
# Settings-запись (TLV-тег с tcode=1, либо .dal-схема NASTR1) содержит хвост
# фиксированного формата с УЗ-параметрами. Из дизасма 0x402980 видно, как DLL
# раскладывает это в структуру (адрес = смещение от начала body, после header):
#
#   +0x00  marker = 0x02         ; маркер «Settings» (TLV body[0])
#   +0x65  noise_cutoff_pct      ; уровень отсечки шумов, %
#   +0x66  gain_db               ; +усиление, дБ
#   +0x67  velocity_ms_le16      ; LE16 скорость УЗ-волны в материале, м/с
#   +0x69  angle_deg             ; угол ввода, градусы
#   +0x6a  delay_us              ; задержка призмы, мкс
#   +0x6b  pulse_freq_le16       ; LE16 частота зондир. импульсов, Гц
#   +0x6e  gate_start_le16       ; LE16 начало строба (дискрет АЦП)
#   +0x70  gate_width_le16       ; LE16 ширина строба (дискрет АЦП)
#   +0x72  freq_le16             ; LE16 рабочая частота преобр-теля, кГц
#   +0x74  threshold_pct         ; порог срабатывания, %
#   +0x75  arm_mode              ; режим строба (0..3)
#   +0x76  ard_curve             ; номер АРД-кривой
#   +0x7c  sens_required         ; требуемая чувствительность, дБ
#   +0x7d  sens_factual          ; фактическая чувствительность, дБ
#   +0x7e  contact_pct           ; контактная связь, %
#   +0x80  vrt_table[0..7]       ; ВРЧ-таблица (8 уровней по 1 байту)
@dataclass
class UD2_AcousticSettings:
    """Акустические настройки — извлекаются из tail-блока Settings-записи.

    Все поля ниже — ОДИН срез из настроек прибора, который полностью описывает,
    КАК была построена A-/B-развёртка. Этого достаточно, чтобы восстановить
    оси (мкс / мм), уровень строба, отметку АРД-чувствительности и т.д.
    """
    marker:           int = 0
    noise_cutoff_pct: int = 0
    gain_db:          int = 0
    velocity_ms:      int = 0
    angle_deg:        int = 0
    delay_us:         int = 0
    pulse_freq_hz:    int = 0
    gate_start:       int = 0
    gate_width:       int = 0
    probe_freq_khz:   int = 0
    threshold_pct:    int = 0
    arm_mode:         int = 0
    ard_curve:        int = 0
    sens_required:    int = 0
    sens_factual:     int = 0
    contact_pct:      int = 0
    vrt_table:        tuple[int, ...] = ()


def decode_ud2102_acoustic(body: bytes) -> UD2_AcousticSettings | None:
    """Декодировать УЗ-настройки из тела Settings-записи.

    Возвращает None, если тело короче 0x88 байт (значит — это не tail-блок
    из 0x402980, а просто .dal-схема NASTR1).
    """
    if len(body) < 0x88:
        return None
    def u8(o: int) -> int:
        return body[o] if 0 <= o < len(body) else 0
    def u16(o: int) -> int:
        if o + 1 >= len(body):
            return 0
        return body[o] | (body[o + 1] << 8)
    vrt = tuple(u8(0x80 + j) for j in range(8))
    return UD2_AcousticSettings(
        marker=u8(0x00),
        noise_cutoff_pct=u8(0x65),
        gain_db=u8(0x66),
        velocity_ms=u16(0x67),
        angle_deg=u8(0x69),
        delay_us=u8(0x6a),
        pulse_freq_hz=u16(0x6b),
        gate_start=u16(0x6e),
        gate_width=u16(0x70),
        probe_freq_khz=u16(0x72),
        threshold_pct=u8(0x74),
        arm_mode=u8(0x75),
        ard_curve=u8(0x76),
        sens_required=u8(0x7c),
        sens_factual=u8(0x7d),
        contact_pct=u8(0x7e),
        vrt_table=vrt,
    )


@dataclass
class UD2_102Header:
    """16-байтный заголовок ответа на 0x55 (4× handshake)."""
    raw:       bytes
    device_no: int             # bytes[0..1] LE — заводской номер прибора

    @classmethod
    def parse(cls, b: bytes) -> "UD2_102Header":
        if len(b) < UD2_HEADER_BYTES:
            raise ValueError(f"UD2-102 handshake header too short: {len(b)}")
        return cls(raw=bytes(b[:UD2_HEADER_BYTES]),
                   device_no=b[0] | (b[1] << 8))


@dataclass
class UD2_102PortConfig:
    """Параметры порта для УД2-102 (отличается таймингами от PortConfig)."""
    name:               str = "COM1" if sys.platform.startswith("win") else "/dev/ttyUSB0"
    baud:               int = 19200
    inter_byte_s:       float = UD2_INTER_BYTE_S
    resp_timeout_s:     float = UD2_RESP_TIMEOUT_S
    body_timeout_s:     float = UD2_BODY_TIMEOUT_S
    handshake_rx_s:     float = UD2_HANDSHAKE_RX_S
    handshake_repeats:  int   = UD2_HANDSHAKE_REPEATS
    handshake_pause_s:  float = UD2_HANDSHAKE_PAUSE_S
    streak_no_hits:     int   = UD2_STREAK_NO_HITS
    streak_after_hit:   int   = UD2_STREAK_AFTER_HIT
    debug:              bool  = False

    def to_pyserial_kwargs(self) -> dict:
        return {
            "port":          self.name,
            "baudrate":      self.baud,
            "bytesize":      serial.EIGHTBITS,
            "stopbits":      serial.STOPBITS_ONE,
            "parity":        serial.PARITY_NONE,
            "xonxoff":       False,
            "rtscts":        False,
            "dsrdtr":        False,
            "timeout":       self.body_timeout_s,
            "write_timeout": max(1.0, self.handshake_rx_s),
        }


class UD2_102Port:
    """Транспорт «Пеленг УД2-102» поверх pyserial.

    Жёсткие тайминги (нарушение → WinError 31):
      - inter-byte gap = 4.8 мс между байтами команды;
      - resp_timeout (28 мс) для пустого/короткого 2-байтного ответа;
      - body_timeout (150 мс) для тела записи (84 байта).

    API:
      - `handshake()` → UD2_102Header (4 повтора 0x55);
      - `read_record(addr)` → bytes(84) или None для пустых блоков;
      - `scan_memory(progress=None)` → итератор MeasurementRecord
        с автоматическим streak-контролем.
    """

    def __init__(self, cfg: UD2_102PortConfig):
        self.cfg = cfg
        self._ser: serial.Serial | None = None
        self.device_no: int | None = None

    # --- low-level ---------------------------------------------------------
    def open(self) -> None:
        if self._ser is not None and self._ser.is_open:
            return
        self._ser = serial.Serial(**self.cfg.to_pyserial_kwargs())
        try:
            self._ser.reset_input_buffer()
            self._ser.reset_output_buffer()
        except Exception:
            pass

    def close(self) -> None:
        if self._ser is not None:
            try:
                self._ser.close()
            finally:
                self._ser = None

    def __enter__(self) -> "UD2_102Port":
        self.open(); return self

    def __exit__(self, *a) -> None:
        self.close()

    @property
    def is_open(self) -> bool:
        return self._ser is not None and self._ser.is_open

    def _ensure_open(self) -> serial.Serial:
        if self._ser is None or not self._ser.is_open:
            raise PelengError("UD2-102: порт не открыт")
        return self._ser

    def _log(self, msg: str) -> None:
        if self.cfg.debug:
            print(f"[ud2-102] {msg}", file=sys.stderr)

    def _send(self, data: bytes) -> None:
        """Побайтовая отправка с inter-byte gap 4.8 мс — жёстко требуется
        медленным контроллером прибора, иначе аппаратный буфер перепол-
        няется и происходит WinError 31."""
        s = self._ensure_open()
        for b in data:
            s.write(bytes([b])); s.flush()
            time.sleep(self.cfg.inter_byte_s)
        self._log(f"TX  {data.hex(' ')}")

    def _read_for(self, n: int, timeout_s: float) -> bytes:
        """Читает до n байт за timeout_s. Если приходит меньше — возвращает что есть."""
        s = self._ensure_open()
        s.timeout = timeout_s
        deadline = time.monotonic() + timeout_s
        buf = bytearray()
        while len(buf) < n:
            remaining = max(0.0, deadline - time.monotonic())
            if remaining == 0.0:
                break
            s.timeout = remaining
            chunk = s.read(n - len(buf))
            if not chunk:
                break
            buf.extend(chunk)
        if buf:
            self._log(f"RX  {bytes(buf).hex(' ')}")
        return bytes(buf)

    # --- high-level --------------------------------------------------------
    def handshake(self) -> UD2_102Header:
        """4× 0x55 → 16-байтный заголовок. Из последнего ответа берётся device_no."""
        last: bytes = b""
        for attempt in range(self.cfg.handshake_repeats):
            try:
                self._ensure_open().reset_input_buffer()
            except Exception:
                pass
            self._send(bytes([OP_UD2_HANDSHAKE]))
            resp = self._read_for(UD2_HEADER_BYTES, self.cfg.handshake_rx_s)
            if len(resp) >= UD2_HEADER_BYTES:
                last = resp
                self._log(f"HS  попытка {attempt+1}/{self.cfg.handshake_repeats} → "
                          f"{len(resp)} байт ({resp[:2].hex(' ')}…)")
            time.sleep(self.cfg.handshake_pause_s)
        if len(last) < UD2_HEADER_BYTES:
            raise PelengError(
                f"УД2-102: хэндшейк не прошёл за {self.cfg.handshake_repeats} попыток. "
                f"Проверьте порт {self.cfg.name}, скорость {self.cfg.baud} и питание прибора.")
        hdr = UD2_102Header.parse(last)
        self.device_no = hdr.device_no
        return hdr

    def read_record(self, addr: int) -> bytes | None:
        """0x42 + Addr_LE16 → 2 байта заголовка (если 0xFD/0xFF — None) +
        84 байта тела (если не пусто). Возвращает None для пустых/стертых блоков."""
        if not (0 <= addr <= 0xFFFF):
            raise ValueError(f"UD2-102: addr вне диапазона: {addr}")
        try:
            self._ensure_open().reset_input_buffer()
        except Exception:
            pass
        cmd = bytes([OP_UD2_READ_RECORD, addr & 0xFF, (addr >> 8) & 0xFF])
        self._send(cmd)
        head = self._read_for(UD2_RECORD_HEADER_BYTES, self.cfg.resp_timeout_s)
        if len(head) < UD2_RECORD_HEADER_BYTES or head in UD2_EMPTY_MARKERS:
            return None
        body = self._read_for(UD2_RECORD_BYTES, self.cfg.body_timeout_s)
        if len(body) < UD2_RECORD_BYTES:
            return None
        return head + body  # 86 байт; 2-байтный заголовок может пригодиться выше

    def scan_memory(self,
                    bases: tuple[int, ...] = UD2_BASES,
                    indices_per_base: int = UD2_INDICES_PER_BASE,
                    progress=None) -> Iterator[MeasurementRecord]:
        """Полный обход памяти прибора с эвристикой обрыва (streak-control)."""
        for base in bases:
            seen_hit = False
            empty_streak = 0
            for idx in range(indices_per_base):
                addr = base + idx
                rec_raw = self.read_record(addr)
                if progress is not None:
                    progress(base, idx, rec_raw is not None)
                if rec_raw is None:
                    empty_streak += 1
                    limit = self.cfg.streak_after_hit if seen_hit else self.cfg.streak_no_hits
                    if empty_streak >= limit:
                        self._log(f"scan: база {base} — обрыв по {empty_streak} пустых, "
                                  f"переход к следующей")
                        break
                    continue
                empty_streak = 0
                seen_hit = True
                # rec_raw[0:2] — заголовок 2 байта; тело — следующие 84.
                rec = parse_ud2102_record(addr, rec_raw[2:2 + UD2_RECORD_BYTES])
                if rec is not None:
                    yield rec

    # --- shadow protocol (только безопасные сервисные команды) -------------
    def send_key(self, key_code: int) -> None:
        """0x4B + KeyCode — эмуляция нажатия клавиши на лицевой панели."""
        if not (0 <= key_code <= 0xFF):
            raise ValueError("UD2-102: key_code вне диапазона")
        self._send(bytes([OP_UD2_SEND_KEY, key_code]))

    def screen_dump(self, expected_bytes: int = 15000,
                    timeout_s: float = 3.5) -> bytes:
        """0x49 — захват видеобуфера дисплея (~15 КБ). Безопасная команда:
        читает данные из прибора, ничего не пишет.
        """
        self._send(bytes([OP_UD2_SCREEN_DUMP]))
        return self._read_for(expected_bytes, timeout_s)

    def rtc_sync(self, dt: _dt.datetime | None = None) -> None:
        """0x54 + YY MM DD hh mm ss — синхронизация часов прибора."""
        if dt is None:
            dt = _dt.datetime.now()
        if not (2000 <= dt.year <= 2099):
            raise ValueError("UD2-102: rtc_sync поддерживает 2000..2099")
        payload = bytes([OP_UD2_RTC_SYNC,
                         dt.year - 2000, dt.month, dt.day,
                         dt.hour, dt.minute, dt.second])
        self._send(payload)


# --- генерация фейковых записей УД2-102 (для демо-режима) ---------------------
def _encode_reverse_bcd(value: str, length: int, terminate: bool = True) -> bytes:
    """Кодирует десятичную строку в Reverse BCD заданной длины (с терминатором).

    Раскладка байтов в результате (length байт всего):
        out[0]            = 0x0A (терминатор) — если terminate=True;
        out[1..length-1]  = пары цифр (hi,lo), младшие пары в начале, старшие в конце.

    Декодер `reverse_bcd()` итерирует по байтам в обратном порядке:
    сначала старшие пары цифр, затем младшие — пока не встретит 0x0A.
    """
    if length <= 0:
        return b""
    digits = [c for c in value if c.isdigit()]
    digit_bytes = length - 1 if terminate else length
    digit_capacity = digit_bytes * 2
    if len(digits) > digit_capacity:
        digits = digits[-digit_capacity:]            # хвостовая обрезка
    digits = ["0"] * (digit_capacity - len(digits)) + digits
    # forward[0] = (digits[0], digits[1]) — самая старшая пара,
    # forward[-1] = (digits[-2], digits[-1]) — самая младшая пара.
    forward = bytearray()
    for k in range(0, digit_capacity, 2):
        hi = int(digits[k])
        lo = int(digits[k + 1])
        forward.append((hi << 4) | lo)
    # Итоговая раскладка: [0x0A, младшие_пары…, старшие_пары] либо без 0x0A.
    body = bytes(reversed(forward))                  # младшие пары в начало
    if terminate:
        out = b"\x0A" + body
    else:
        out = body
    if len(out) < length:
        out = out + b"\x00" * (length - len(out))
    return out[:length]


def make_fake_ud2102_record(seed: int = 0,
                            addr: int = 10100) -> tuple[int, bytes]:
    """Собирает 86-байтное «эфирное» тело УД2-102 (2 байта заголовка + 84 BCD).

    Возвращает (addr, raw86). Якорь даты ставится на смещение i=4 — типично
    для оригинального формата.
    """
    body = bytearray(b"\x00" * UD2_RECORD_BYTES)
    now = _dt.datetime.now()
    i = 4                                  # якорь — на 4-м байте тела
    body[i + 0] = max(1, min(31, now.day))
    body[i + 1] = max(1, min(12, now.month))
    body[i + 2] = max(15, min(40, (now.year - 2000) & 0xFF))
    body[i + 3] = now.hour & 0x1F
    body[i + 4] = now.minute & 0x3F

    # Смещения по PELENG_REVERSE.md § 2.5 (от якоря i):
    #   i+5..6   tcode (LE16)
    #   i+7      sub_b (u8)
    #   i+10..18 плавка (Reverse-BCD, 9 байт)
    #   i+20..28 № объекта (Reverse-BCD, 9 байт)
    #   i+30     сторона (u8)
    #   i+31     шейка (u8)
    #   i+32..33 обод (LE16, десятые мм)
    #   i+34     обточка (u8)
    #   i+35..36 гребень (LE16, десятые мм)
    #   i+37     наплавка (u8)

    # tcode = 24667 (РУ1Ш если sub_b=0, РУ1 если sub_b=1) — варьируем по seed
    tcode = 24667
    body[i + 5] = tcode & 0xFF              # tcode LE16
    body[i + 6] = (tcode >> 8) & 0xFF
    body[i + 7] = (seed % 2) & 0xFF         # sub_b

    # Плавка @ i+10..18 (9 байт Reverse-BCD)
    plav_bytes = _encode_reverse_bcd("3145" + str(seed % 100).zfill(2), 9)
    body[i + 10:i + 10 + len(plav_bytes)] = plav_bytes

    # № объекта @ i+20..28 (9 байт Reverse-BCD)
    num = 12345 + seed
    obj_bytes = _encode_reverse_bcd(str(num), 9)
    body[i + 20:i + 20 + len(obj_bytes)] = obj_bytes

    body[i + 30] = (seed % 2) & 0xFF        # сторона: 0=Левая, 1=Правая
    body[i + 31] = ((seed % 3) + 1) & 0xFF  # шейка: 1/2/3

    # Обод @ i+32..33 (LE16, десятые мм) — напр. 750 = 75.0 мм
    obod_val = (75 - (seed % 10)) * 10
    body[i + 32] = obod_val & 0xFF
    body[i + 33] = (obod_val >> 8) & 0xFF

    # Обточка @ i+34 (u8, категория)
    body[i + 34] = (2 + (seed % 5)) & 0xFF

    # Гребень @ i+35..36 (LE16, десятые мм) — напр. 280 = 28.0 мм
    greben_val = (28 + (seed % 4)) * 10
    body[i + 35] = greben_val & 0xFF
    body[i + 36] = (greben_val >> 8) & 0xFF

    # Наплавка @ i+37
    body[i + 37] = 0x01 if (seed % 4) == 0 else 0x00

    head = b"\x00\x01"  # маркер не-пустого блока (любые 2 байта кроме 0xFD/0xFF FF/FF)
    return addr, bytes(head + body)


def make_fake_ud2102_handshake(device_no: int = 4428) -> bytes:
    """16-байтный фейковый handshake-заголовок УД2-102."""
    out = bytearray(UD2_HEADER_BYTES)
    out[0] = device_no & 0xFF
    out[1] = (device_no >> 8) & 0xFF
    return bytes(out)


# =============================================================================
# 2. ДЕКОДЕРЫ ЗАПИСЕЙ
# =============================================================================
# 2.1. ЭТАЛОННЫЕ СХЕМЫ ИЗ PelengPCtest.dal
# -----------------------------------------------------------------------------
# Каждая «версия» прибора (Базовая / Железнодорожная / Рельсовая) описывает
# свою тройку таблиц БД. Из этих таблиц мы берём:
#   - порядок и Russian-labels полей для GUI;
#   - тип кодировки (BCD-DATE/BCD-TIME/STRING/INT/ENUM) — он определяет, как
#     парсить байты тела TLV-записи.
#
# Кодировка типа `TypeField` (dword) — обратно реверсирована:
#   младший байт = тип; средние биты = длина строки/блока.
#
#   0x00000001 = INT16             (2 байта, big-endian, BCD-друг.)
#   0x00000003 = BCD-DATE          (3 байта: dd, mm, yy → 2000+yy)
#   0x00000082 = BCD-TIME          (2 байта: hh, mm)
#   0x000000A2 = STRING-CP1251     (Pascal short string, len-prefixed)
#   0x000001E2 = STRING-CP1251-LONG
#   0x000000E2 = STRING-CP1251-MED
#   0x00000142 = STRING-CP1251-FIX
#   0x00000052 = INT32-MM          (мм, условная протяжённость)
#   0x0000005A = ENUM-DEFEKT       (1 байт: 0=нет / 1=есть)
#   0x0000014A = ENUM-TYPEZAP      (1 байт: 0=A-развёртка, 1=B-развёртка)
#   0x00000072 = BLOB-PROTOCOL     (длина в заголовке + байты)
# -----------------------------------------------------------------------------
FIELD_TYPE_INT      = 0x01
FIELD_TYPE_DATE     = 0x03
FIELD_TYPE_TIME     = 0x82
FIELD_TYPE_STRING_S = 0xA2
FIELD_TYPE_STRING_L = 0xE2
FIELD_TYPE_STRING_X = 0x142
FIELD_TYPE_INT_MM   = 0x52
FIELD_TYPE_ENUM_DEF = 0x5A
FIELD_TYPE_ENUM_TZ  = 0x14A
FIELD_TYPE_BLOB     = 0x72


# Поля таблиц RESULTS1/2/3, NASTR1/2/3, SHORTPROT1/2 — из PelengPCtest.dal.
# Имя поля → (русская подпись, тип). Имя совпадает с тем, что в DLL
# (см. реверс fcn.00403420 / fcn.00403834 / fcn.004041d0 / fcn.00404758).
DAL_FIELDS_RESULTS1: list[tuple[str, str, int]] = [
    ("NUMBER",     "№ п/п",                 FIELD_TYPE_INT),
    ("TYPEZAP",    "Вид записи",            FIELD_TYPE_ENUM_TZ),
    ("NUMKOD",     "№ записи",              FIELD_TYPE_INT),
    ("DATEFORM",   "Дата",                  FIELD_TYPE_DATE),
    ("TIMEFORM",   "Время",                 FIELD_TYPE_TIME),
    ("KODOPERA",   "Оператор",              FIELD_TYPE_INT),
    ("NAMEOPERA",  "Имя оператора",         FIELD_TYPE_STRING_L),
    ("NUMVERS",    "Версия ПО",             FIELD_TYPE_STRING_S),
    ("NUMPRIB",    "№ прибора",             FIELD_TYPE_INT),
    ("TYPEVAR",    "Типовой вариант",       FIELD_TYPE_INT),
    ("NUMOBJ",     "№ объекта",             FIELD_TYPE_STRING_S),
    ("M",          "м",                     FIELD_TYPE_INT),
    ("MM",         "мм",                    FIELD_TYPE_INT),
    ("CLOCK",      "Часы",                  FIELD_TYPE_INT),
    ("SMELTING",   "Плавка",                FIELD_TYPE_STRING_S),
    ("MAKETIME",   "Год изготовления",      FIELD_TYPE_INT),
    ("DEFEKT",     "Дефект",                FIELD_TYPE_ENUM_DEF),
    ("CODEDEF",    "Код дефекта",           FIELD_TYPE_STRING_S),
    ("CONDLENGTH", "Условная протяжённость, мм", FIELD_TYPE_INT_MM),
    ("NUMZAP",     "Настройки",             FIELD_TYPE_INT),
]

DAL_FIELDS_RESULTS2: list[tuple[str, str, int]] = [  # Железнодорожная
    ("NUMBER",     "№ п/п",                 FIELD_TYPE_INT),
    ("TYPEZAP",    "Вид записи",            FIELD_TYPE_ENUM_TZ),
    ("NUMKOD",     "№ записи",              FIELD_TYPE_INT),
    ("DATEFORM",   "Дата",                  FIELD_TYPE_DATE),
    ("TIMEFORM",   "Время",                 FIELD_TYPE_TIME),
    ("KODOPERA",   "Оператор",              FIELD_TYPE_INT),
    ("NAMEOPERA",  "Имя оператора",         FIELD_TYPE_STRING_L),
    ("NUMVERS",    "Версия ПО",             FIELD_TYPE_STRING_S),
    ("NUMPRIB",    "№ прибора",             FIELD_TYPE_INT),
    ("TYPEVAR",    "Типовой вариант",       FIELD_TYPE_INT),
    ("NUMOBJ",     "№ объекта",             FIELD_TYPE_STRING_L),
    ("SMELTING",   "Плавка",                FIELD_TYPE_STRING_L),
    ("INDMAKER",   "Завод изготовитель",    FIELD_TYPE_STRING_X),
    ("MAKETIME",   "Год изготовления",      FIELD_TYPE_INT),
    ("DEFEKT",     "Дефект",                FIELD_TYPE_ENUM_DEF),
    ("CODEDEF",    "Код дефекта",           FIELD_TYPE_STRING_S),
    ("NUMZAP",     "Настройки",             FIELD_TYPE_INT),
]

DAL_FIELDS_RESULTS3: list[tuple[str, str, int]] = [  # Рельсовая
    ("NUMBER",     "№ п/п",                 FIELD_TYPE_INT),
    ("TYPEZAP",    "Вид записи",            FIELD_TYPE_ENUM_TZ),
    ("NUMKOD",     "№ записи",              FIELD_TYPE_INT),
    ("DATEFORM",   "Дата",                  FIELD_TYPE_DATE),
    ("TIMEFORM",   "Время",                 FIELD_TYPE_TIME),
    ("KODOPERA",   "Оператор",              FIELD_TYPE_INT),
    ("NAMEOPERA",  "Имя оператора",         FIELD_TYPE_STRING_L),
    ("NUMVERS",    "Версия ПО",             FIELD_TYPE_STRING_S),
    ("NUMPRIB",    "№ прибора",             FIELD_TYPE_INT),
    ("TYPEVAR",    "Типовой вариант",       FIELD_TYPE_INT),
    ("NUMOBJ",     "№ объекта",             FIELD_TYPE_STRING_L),
    ("KM",         "км",                    FIELD_TYPE_INT),
    ("M",          "м",                     FIELD_TYPE_INT),
    ("MM",         "мм",                    FIELD_TYPE_INT),
    ("STAGE",      "Ход/перегон",           FIELD_TYPE_STRING_X),
    ("SECTION",    "Звено",                 FIELD_TYPE_STRING_X),
    ("PICKETLINE", "Пикет",                 FIELD_TYPE_STRING_X),
    ("PATH",       "Путь",                  FIELD_TYPE_STRING_X),
    ("DEFEKT",     "Дефект",                FIELD_TYPE_ENUM_DEF),
    ("CODEDEF",    "Код дефекта",           FIELD_TYPE_STRING_S),
    ("CONDLENGTH", "Условная протяжённость, мм", FIELD_TYPE_INT_MM),
    ("NUMZAP",     "Настройки",             FIELD_TYPE_INT),
]

DAL_FIELDS_NASTR1: list[tuple[str, str, int]] = [  # Settings
    ("NUMBER",     "№ п/п",                 FIELD_TYPE_INT),
    ("NUMKOD",     "№ записи",              FIELD_TYPE_INT),
    ("TYPEZAP",    "Вид записи",            FIELD_TYPE_ENUM_TZ),
    ("DATEFORM",   "Дата",                  FIELD_TYPE_DATE),
    ("TIMEFORM",   "Время",                 FIELD_TYPE_TIME),
    ("KODOPERA",   "Оператор",              FIELD_TYPE_INT),
    ("NAMEOPERA",  "Имя оператора",         FIELD_TYPE_STRING_L),
    ("NUMVERS",    "Версия ПО",             FIELD_TYPE_STRING_S),
    ("NUMPRIB",    "№ прибора",             FIELD_TYPE_INT),
    ("TYPEVAR",    "Типовой вариант",       FIELD_TYPE_INT),
    ("NUMPROT",    "Отчёты контроля",       FIELD_TYPE_INT),
]

DAL_FIELDS_SHORTPROT1: list[tuple[str, str, int]] = [  # Report
    ("NUMBER",     "№ п/п",                 FIELD_TYPE_INT),
    ("NUMKOD",     "№ записи",              FIELD_TYPE_INT),
    ("DATEFORM",   "Дата",                  FIELD_TYPE_DATE),
    ("TIMEFORM",   "Время",                 FIELD_TYPE_TIME),
    ("KODOPERA",   "Оператор",              FIELD_TYPE_INT),
    ("NAMEOPERA",  "Имя оператора",         FIELD_TYPE_STRING_L),
    ("NUMVERS",    "Версия ПО",             FIELD_TYPE_STRING_S),
    ("NUMPRIB",    "№ прибора",             FIELD_TYPE_INT),
    ("TYPEVAR",    "Типовой вариант",       FIELD_TYPE_INT),
    ("NUMOBJ",     "№ объекта",             FIELD_TYPE_STRING_L),
    ("M",          "м",                     FIELD_TYPE_INT),
    ("MM",         "мм",                    FIELD_TYPE_INT),
    ("CLOCK",      "Часы",                  FIELD_TYPE_INT),
    ("NUMDEF",     "Количество дефектов",   FIELD_TYPE_INT),
    ("PROTOCOL",   "Протокол",              FIELD_TYPE_BLOB),
    ("NUMPROT",    "Отчёты толщиномера",    FIELD_TYPE_INT),
]

# Перечисления.
ENUM_DEFEKT  = {0: "нет",         1: "есть"}
ENUM_TYPEZAP = {0: "А-развёртка", 1: "B-развёртка",
                2: "Настройки",   3: "Отчёт"}


# 2.2. ЭТАЛОННЫЕ ПАРАМЕТРЫ ОТРИСОВКИ ГРАФИКОВ
# -----------------------------------------------------------------------------
# Эти числа взяты из реверса DLL и из стандартных параметров «Пеленг» 1.2:
#   - частота дискретизации АЦП ультразвукового тракта = 50 МГц
#     (по умолчанию; меняется в Settings, см. поле NUMVERS-зависимый блок)
#   - амплитуда — однобайтная (0..255), линейно: 0=0%, 255=100%
#   - 100 % амплитуды соответствует ~26 дБ относительного усиления
#   - А-развёртка: ось X в мкс времени = индекс_отсчёта / 50, ось Y в %
#   - B-развёртка: ось X — координата сканирования (мм),
#                  ось Y — глубина (мм, скорость УЗ-волны 5900 м/с в стали),
#                  цвет — амплитуда дБ
ADC_RATE_MHZ:    float = 50.0       # 50 МГц АЦП
US_VEL_MM_PER_US: float = 5.9       # 5900 м/с в стали = 5.9 мм/мкс
AMPL_FULL_DB:    float = 26.0       # «полная шкала» амплитуды, дБ (для byte=255)
AMP_DB_PER_BYTE: float = AMPL_FULL_DB / 255.0


def byte_to_db(v: int) -> float:
    """Раскодировать байт амплитуды (0..255) в дБ (0..26)."""
    return float(max(0, min(255, v))) * AMP_DB_PER_BYTE


def sample_to_us(idx: int, rate_mhz: float = ADC_RATE_MHZ) -> float:
    """Индекс отсчёта → время в мкс."""
    return idx / rate_mhz


def us_to_depth_mm(us: float, vel: float = US_VEL_MM_PER_US) -> float:
    """Время прохода (мкс) → глубина (мм). Делим на 2 — двойной путь."""
    return us * vel / 2.0


# -----------------------------------------------------------------------------
# 2.3. БАЙТОВЫЕ ДЕКОДЕРЫ ТИПОВ (по реверсу type-decoders в DLL)
# -----------------------------------------------------------------------------
def _bcd(b: int) -> int:
    """BCD-байт → десятичное (0..99)."""
    return ((b >> 4) & 0x0F) * 10 + (b & 0x0F)


def decode_date(buf: bytes, off: int) -> tuple[str, int, int, int]:
    """3-байтовое поле DATE: dd, mm, yy → 'dd.mm.yyyy', (d, m, y)."""
    if off + 3 > len(buf):
        return "--.--.----", 0, 0, 0
    d = buf[off]; m = buf[off + 1]; y = buf[off + 2]
    # На медленных прошивках день/месяц могут идти как BCD; считаем байт-в-байт.
    if d > 31 or m > 12 or y > 99:
        d, m, y = _bcd(d), _bcd(m), _bcd(y)
    year = 2000 + y
    return f"{d:02d}.{m:02d}.{year}", d, m, year


def decode_time(buf: bytes, off: int) -> tuple[str, int, int]:
    """2-байтовое поле TIME: hh, mm → 'hh:mm', (h, m)."""
    if off + 2 > len(buf):
        return "--:--", 0, 0
    h, m = buf[off], buf[off + 1]
    if h > 23 or m > 59:
        h, m = _bcd(h), _bcd(m)
    return f"{h:02d}:{m:02d}", h, m


def decode_pstring(buf: bytes, off: int, max_len: int = 32) -> str:
    """Pascal short-string (len-prefixed CP1251)."""
    if off >= len(buf):
        return ""
    n = buf[off]
    if n > max_len or off + 1 + n > len(buf):
        n = min(max_len, len(buf) - off - 1)
    if n <= 0:
        return ""
    return buf[off + 1: off + 1 + n].decode("cp1251", errors="replace").rstrip("\x00 ")


def decode_fixstring(buf: bytes, off: int, length: int) -> str:
    """Поле фиксированной длины — режем по нулю."""
    if off + length > len(buf):
        length = max(0, len(buf) - off)
    raw = bytes(buf[off: off + length])
    raw = raw.split(b"\x00", 1)[0]
    return raw.decode("cp1251", errors="replace").rstrip()


def decode_int16_le(buf: bytes, off: int) -> int:
    if off + 2 > len(buf):
        return 0
    return buf[off] | (buf[off + 1] << 8)


def decode_int32_le(buf: bytes, off: int) -> int:
    if off + 4 > len(buf):
        return 0
    return buf[off] | (buf[off + 1] << 8) | (buf[off + 2] << 16) | (buf[off + 3] << 24)


@dataclass
class DecodedField:
    name:    str
    label:   str
    value:   object
    display: str


class StreamReader:
    """Пошаговый парсер байт-потока с теми же типами, что в DLL."""

    def __init__(self, buf: bytes, start: int = 0):
        self.buf = bytes(buf); self.pos = start

    def remaining(self) -> int:
        return max(0, len(self.buf) - self.pos)

    def read_byte(self) -> int:
        if self.pos >= len(self.buf): return 0
        v = self.buf[self.pos]; self.pos += 1; return v

    def read_int16(self) -> int:
        v = decode_int16_le(self.buf, self.pos); self.pos += 2; return v

    def read_int32(self) -> int:
        v = decode_int32_le(self.buf, self.pos); self.pos += 4; return v

    def read_date(self) -> tuple[str, int, int, int]:
        s = decode_date(self.buf, self.pos); self.pos += 3; return s

    def read_time(self) -> tuple[str, int, int]:
        s = decode_time(self.buf, self.pos); self.pos += 2; return s

    def read_pstring(self, max_len: int = 32) -> str:
        if self.pos >= len(self.buf): return ""
        n = self.buf[self.pos]
        s = decode_pstring(self.buf, self.pos, max_len)
        self.pos += 1 + min(n, max_len, len(self.buf) - self.pos - 1)
        return s

    def read_fixstring(self, length: int) -> str:
        s = decode_fixstring(self.buf, self.pos, length)
        self.pos += length
        return s

    def read_blob(self, length: int) -> bytes:
        b = bytes(self.buf[self.pos: self.pos + length])
        self.pos += length
        return b

    def rest(self) -> bytes:
        b = bytes(self.buf[self.pos:])
        self.pos = len(self.buf)
        return b


def _decode_field(rd: StreamReader, name: str, ftype: int) -> DecodedField:
    """Декодировать одно поле по типу из .dal."""
    if ftype == FIELD_TYPE_INT:
        v = rd.read_int16(); return DecodedField(name, "", v, str(v))
    if ftype == FIELD_TYPE_INT_MM:
        v = rd.read_int32(); return DecodedField(name, "", v, f"{v} мм")
    if ftype == FIELD_TYPE_DATE:
        s, *_ = rd.read_date(); return DecodedField(name, "", s, s)
    if ftype == FIELD_TYPE_TIME:
        s, *_ = rd.read_time(); return DecodedField(name, "", s, s)
    if ftype == FIELD_TYPE_STRING_S:
        s = rd.read_pstring(15); return DecodedField(name, "", s, s or "—")
    if ftype == FIELD_TYPE_STRING_L:
        s = rd.read_pstring(31); return DecodedField(name, "", s, s or "—")
    if ftype == FIELD_TYPE_STRING_X:
        s = rd.read_pstring(63); return DecodedField(name, "", s, s or "—")
    if ftype == FIELD_TYPE_ENUM_DEF:
        v = rd.read_byte()
        return DecodedField(name, "", v, ENUM_DEFEKT.get(v, str(v)))
    if ftype == FIELD_TYPE_ENUM_TZ:
        v = rd.read_byte()
        return DecodedField(name, "", v, ENUM_TYPEZAP.get(v, str(v)))
    if ftype == FIELD_TYPE_BLOB:
        n = rd.read_int16()
        b = rd.read_blob(min(n, rd.remaining()))
        return DecodedField(name, "", b, f"BLOB[{len(b)}]")
    # default — INT16
    v = rd.read_int16(); return DecodedField(name, "", v, str(v))


def decode_schema(buf: bytes, schema: list[tuple[str, str, int]],
                  start: int = 1) -> list[DecodedField]:
    """Прогнать всю схему по буферу. Первый байт payload — тип записи (1..5),
    поэтому начинаем с offset=1, как и DLL."""
    rd = StreamReader(buf, start)
    fields: list[DecodedField] = []
    for name, label, ftype in schema:
        if rd.remaining() == 0:
            break
        f = _decode_field(rd, name, ftype)
        f.label = label
        fields.append(f)
    return fields


def pick_results_schema(firmware: str, type_var: int = 0) -> list[tuple[str, str, int]]:
    """Выбрать схему RESULTS по версии прошивки.

    Реверс fcn.004022c8 в DLL:
      - Базовая  (Version1, ПО 03.62..06.83): RESULTS1
      - Желез.   (Version2, ПО 04.42..08.43): RESULTS2
      - Рельс.   (Version3, ПО 01.50..01.51): RESULTS3
    """
    fw_num = 0
    try:
        a, b = firmware.split(".", 1)
        fw_num = int(a) * 100 + int(b)
    except Exception:
        pass
    if fw_num and 150 <= fw_num <= 151:
        return DAL_FIELDS_RESULTS3
    if fw_num and 442 <= fw_num <= 843:
        return DAL_FIELDS_RESULTS2
    return DAL_FIELDS_RESULTS1


# -----------------------------------------------------------------------------
# 2.4. ВЫСОКОУРОВНЕВЫЕ КЛАССЫ-ЗАПИСИ
# -----------------------------------------------------------------------------
@dataclass
class DecodedRecord:
    record_type: RecordType
    tag:         int
    raw_body:    bytes = field(repr=False)
    fields:      list[DecodedField] = field(default_factory=list)
    firmware:    str = ""

    def summary(self) -> str:
        return f"[{self.record_type.name}] tag=0x{self.tag:04x} len={len(self.raw_body)}"

    def field(self, name: str) -> object:
        for f in self.fields:
            if f.name == name: return f.value
        return None


@dataclass
class AScanRecord(DecodedRecord):
    samples:    bytes = b""        # массив амплитуд (0..255) — PROTOCOL
    sub_minute: int = 0
    sub_hour:   int = 0
    gate_start: int = 0
    gate_end:   int = 0
    sample_rate_mhz: float = ADC_RATE_MHZ

    @classmethod
    def decode(cls, t: TLVRecord, firmware: str = "") -> "AScanRecord":
        body = t.body
        schema = pick_results_schema(firmware)
        # Проходим по схеме с единым StreamReader, чтобы одновременно
        # получить список полей И точный offset, где начинаются семплы.
        rd = StreamReader(body, 1)
        fields: list[DecodedField] = []
        for name, label, ftype in schema:
            if rd.remaining() == 0: break
            f = _decode_field(rd, name, ftype)
            f.label = label
            fields.append(f)
        consumed = rd.pos
        # В RESULTS1 последнее служебное поле NUMZAP, дальше PROTOCOL — BLOB.
        samples = bytes(body[consumed:]) if len(body) > consumed else b""
        sub_min  = body[1] if len(body) > 1 else 0
        sub_hour = body[2] if len(body) > 2 else 0
        return cls(
            record_type=RecordType.A_SCAN, tag=t.tag, raw_body=body,
            fields=fields, firmware=firmware,
            samples=samples, sub_minute=sub_min, sub_hour=sub_hour,
            gate_start=0, gate_end=0,
        )

    def summary(self) -> str:
        return (f"[A-Scan] {self.sub_hour:02d}:{self.sub_minute:02d} "
                f"samples={len(self.samples)} "
                f"({len(self.samples) / self.sample_rate_mhz:.1f} мкс)")

    @property
    def n_samples(self) -> int:
        return len(self.samples)

    @property
    def duration_us(self) -> float:
        return self.n_samples / max(0.01, self.sample_rate_mhz)

    @property
    def max_depth_mm(self) -> float:
        return us_to_depth_mm(self.duration_us)


@dataclass
class BScanRecord(DecodedRecord):
    width:  int = 0       # ширина по координате сканирования (отсчёты)
    height: int = 0       # высота по глубине (отсчёты)
    pixels: bytes = b""   # амплитуды 0..255
    step_mm: float = 1.0  # шаг по X в мм
    sample_rate_mhz: float = ADC_RATE_MHZ

    @classmethod
    def decode(cls, t: TLVRecord, firmware: str = "") -> "BScanRecord":
        body = t.body
        # B-scan в DLL имеет тот же общий заголовок (NUMBER, NUMKOD, ...) +
        # хвостом идёт width(int16-LE) + height(int16-LE) + width*height пикселей.
        schema = pick_results_schema(firmware)
        rd = StreamReader(body, 1)
        fields: list[DecodedField] = []
        for name, label, ftype in schema:
            if rd.remaining() == 0: break
            f = _decode_field(rd, name, ftype)
            f.label = label
            fields.append(f)
        # После схемы читаем width и height (реверс fcn.004031dc).
        width = height = 0
        if rd.remaining() >= 4:
            width  = rd.read_int16()
            height = rd.read_int16()
        # Fallback: если схема выела весь body или выдала невнятные размеры —
        # пытаемся вывести width×height по эталонам оригинала.
        pixels = bytes(body[rd.pos:]) if len(body) > rd.pos else b""
        if not (width and height and width * height <= len(pixels)):
            if len(pixels) >= 64 * 128:
                width, height = 64, 128
            elif len(pixels) >= 32 * 64:
                width, height = 32, 64
            elif len(pixels) > 0:
                side = max(1, int(math.sqrt(len(pixels))))
                width, height = side, side
        # Обрезаем хвост до ровно width*height (или оставляем как есть).
        if width and height:
            need = width * height
            if len(pixels) >= need:
                pixels = pixels[:need]
        return cls(
            record_type=RecordType.B_SCAN, tag=t.tag, raw_body=body,
            fields=fields, firmware=firmware,
            width=width, height=height, pixels=pixels,
        )

    def summary(self) -> str:
        return f"[B-Scan] {self.width}x{self.height} px={len(self.pixels)}"

    @property
    def length_mm(self) -> float:
        return self.width * self.step_mm

    @property
    def depth_mm(self) -> float:
        return us_to_depth_mm(self.height / max(0.01, self.sample_rate_mhz))


@dataclass
class SettingsRecord(DecodedRecord):
    # ----- УЗ-параметры (выписаны из UD2_AcousticSettings, см. ниже) -----
    gain_db:        int = 0       # +102: усиление, дБ
    velocity_ms:    int = 5900    # +103: скорость УЗ-волны в материале, м/с
    angle_deg:      int = 0       # +105: угол ввода, градусы
    gate_start:     int = 0       # +110: начало строба (дискрет АЦП)
    gate_width:     int = 0       # +112: ширина строба (дискрет АЦП)
    threshold_pct:  int = 0       # +115: порог срабатывания, %
    sens_required:  int = 0       # +124: требуемая чувствительность, дБ
    sens_factual:   int = 0       # +125: фактическая чувствительность, дБ
    noise_cutoff:   int = 0       # +101: уровень отсечки шумов, %
    sweep_us:       float = 0.0   # длительность развёртки в мкс (вычисляется)
    sweep_mm:       float = 0.0   # глубина развёртки в мм (вычисляется)

    @classmethod
    def decode(cls, t: TLVRecord, firmware: str = "") -> "SettingsRecord":
        body = t.body
        fields = decode_schema(body, DAL_FIELDS_NASTR1, start=1)
        # Параллельно с .dal-схемой пытаемся снять УЗ-настройки из «акустического»
        # tail-блока (см. UD2_AcousticSettings и tag/1000=1 → settings).
        ac = decode_ud2102_acoustic(body) if len(body) >= 192 else None
        gain = vel = ang = gs = gw = th = sr = sf = nc = 0
        if ac is not None:
            gain = ac.gain_db; vel = ac.velocity_ms; ang = ac.angle_deg
            gs = ac.gate_start; gw = ac.gate_width;  th = ac.threshold_pct
            sr = ac.sens_required; sf = ac.sens_factual; nc = ac.noise_cutoff_pct
        # длительность развёртки = (gate_start + gate_width) / ADC_RATE
        sweep_us = (gs + gw) / max(0.01, ADC_RATE_MHZ)
        # глубина развёртки = V·t/2, где V — скорость в материале (мм/мкс)
        v_mm_per_us = (vel / 1000.0) if vel > 0 else US_VEL_MM_PER_US
        sweep_mm = sweep_us * v_mm_per_us / 2.0
        return cls(record_type=RecordType.SETTINGS, tag=t.tag, raw_body=body,
                   fields=fields, firmware=firmware,
                   gain_db=gain, velocity_ms=vel or 5900, angle_deg=ang,
                   gate_start=gs, gate_width=gw, threshold_pct=th,
                   sens_required=sr, sens_factual=sf, noise_cutoff=nc,
                   sweep_us=sweep_us, sweep_mm=sweep_mm)

    def summary(self) -> str:
        return (f"[Settings] усил={self.gain_db}дБ V={self.velocity_ms}м/с "
                f"α={self.angle_deg}° строб=[{self.gate_start}..{self.gate_start+self.gate_width}] "
                f"({self.sweep_us:.1f}мкс / {self.sweep_mm:.1f}мм) "
                f"порог={self.threshold_pct}% полей={len(self.fields)}")


@dataclass
class ReportRecord(DecodedRecord):
    operator: str = ""
    object_:  str = ""
    notes:    str = ""
    when:     _dt.datetime | None = None

    @classmethod
    def decode(cls, t: TLVRecord, firmware: str = "") -> "ReportRecord":
        body = t.body
        fields = decode_schema(body, DAL_FIELDS_SHORTPROT1, start=1)
        # Совместимость со старым free-form форматом «оператор\0объект\0заметки\0».
        rest = body[1:].split(b"\x00")
        def _s(i: int) -> str:
            if i < len(rest):
                try:
                    return rest[i].decode("cp1251", errors="replace").strip()
                except Exception:
                    return ""
            return ""
        # Если не получилось извлечь поля по схеме — берём из split по \0.
        op = next((str(f.value) for f in fields if f.name == "NAMEOPERA"), "") or _s(0)
        ob = next((str(f.value) for f in fields if f.name == "NUMOBJ"),    "") or _s(1)
        nt = _s(2)
        return cls(record_type=RecordType.REPORT, tag=t.tag, raw_body=body,
                   fields=fields, firmware=firmware,
                   operator=op, object_=ob, notes=nt)

    def summary(self) -> str:
        return f"[Report] оп='{self.operator}' объект='{self.object_}'"


@dataclass
class GenericRecord(DecodedRecord):
    @classmethod
    def decode(cls, t: TLVRecord, firmware: str = "") -> "GenericRecord":
        return cls(record_type=RecordType.GENERIC, tag=t.tag, raw_body=t.body,
                   firmware=firmware)

    def summary(self) -> str:
        return f"[Generic] len={len(self.raw_body)}"


def decode_record(t: TLVRecord, firmware: str = "") -> DecodedRecord:
    rt = t.record_type
    if rt is RecordType.A_SCAN:   return AScanRecord.decode(t, firmware)
    if rt is RecordType.B_SCAN:   return BScanRecord.decode(t, firmware)
    if rt is RecordType.SETTINGS: return SettingsRecord.decode(t, firmware)
    if rt is RecordType.REPORT:   return ReportRecord.decode(t, firmware)
    if rt is RecordType.GENERIC:  return GenericRecord.decode(t, firmware)
    return DecodedRecord(record_type=RecordType.UNKNOWN, tag=t.tag, raw_body=t.body,
                         firmware=firmware)


def decode_all(records: list[TLVRecord], firmware: str = "") -> list[DecodedRecord]:
    return [decode_record(r, firmware) for r in records]


# =============================================================================
# 2A. SETTINGS → A-SCAN DERIVATION (реверс fcn.00403420 в 102_203dll.dll)
# =============================================================================
# В оригинальной программе A-развёртка строится не «на пустом месте»: оси,
# уровень строба, отметки АРД и линия порога — всё это ВЫЧИСЛЯЕТСЯ из
# Settings-записи, которая лежит в том же блоке. Соответствующая функция в
# DLL (fcn.00403420 = handler A-scan) проходит так:
#
#   ; вход:   esi = AScan body, edi = Settings body (если есть в блоке)
#   movzx eax, byte [edi+0x66]   ; gain_db
#   movzx ebx, word [edi+0x67]   ; velocity_ms
#   movzx ecx, word [edi+0x6e]   ; gate_start (samples)
#   movzx edx, word [edi+0x70]   ; gate_width (samples)
#   movzx ebp, byte [edi+0x69]   ; angle_deg
#   ;   t_us  = (gate_start + gate_width) / ADC_RATE_MHZ
#   ;   d_mm  = t_us · velocity_ms / 2000
#   ;   y_max = 26 · 1   (амплитуда в дБ; 0..26 при byte 0..255)
#   ;   thr   = byte [edi+0x74] · 26 / 100   ; порог в дБ
#   ; затем в цикле строит точки:
#   ;   for i in 0..n_samples:
#   ;       x_us = i / ADC_RATE_MHZ
#   ;       x_mm = x_us · velocity_ms / 2000
#   ;       y_db = byte_to_db(samples[i])
#
# Эта функция реализует ровно ту же логику, но возвращает данные в виде
# полного описания A-развёртки (axis, gate, threshold, ARD-mark) — для GUI.
@dataclass
class AScanAxes:
    """Описание осей и оверлеев A-развёртки, полученное из Settings."""
    sample_rate_mhz: float = ADC_RATE_MHZ
    velocity_ms:     int   = 5900
    angle_deg:       int   = 0
    n_samples:       int   = 0
    duration_us:     float = 0.0
    max_depth_mm:    float = 0.0
    # порог в дБ (горизонтальная линия) и в %
    threshold_db:    float = 0.0
    threshold_pct:   int   = 0
    # строб (вертикальные маркеры) — мкс и мм
    gate_start_us:   float = 0.0
    gate_end_us:     float = 0.0
    gate_start_mm:   float = 0.0
    gate_end_mm:     float = 0.0
    # АРД-кривая: ссылка на номер кривой и значения чувствительности
    ard_curve:       int   = 0
    sens_required:   int   = 0
    sens_factual:    int   = 0
    # ВРЧ-таблица: 8 уровней «амплитуда → коэф. ослабления, дБ»
    vrt_table:       tuple[int, ...] = ()

    def x_us(self, sample_index: int) -> float:
        """Координата X (мкс) для семпла с индексом i."""
        return sample_index / max(0.01, self.sample_rate_mhz)

    def x_mm(self, sample_index: int) -> float:
        """Координата X (мм) для семпла с индексом i — глубина в материале."""
        v_mm_per_us = (self.velocity_ms / 1000.0) if self.velocity_ms > 0 else US_VEL_MM_PER_US
        return self.x_us(sample_index) * v_mm_per_us / 2.0

    def y_db(self, sample_byte: int) -> float:
        """Координата Y (дБ) — амплитуда семпла, переведённая в дБ."""
        return byte_to_db(sample_byte)


def settings_to_ascan_axes(
    settings: SettingsRecord | UD2_AcousticSettings | None,
    n_samples: int,
    sample_rate_mhz: float = ADC_RATE_MHZ,
) -> AScanAxes:
    """Построить описание осей A-развёртки на основании Settings.

    Это и есть «ради чего весь реверс» — настоящая функция привязки
    A-развёртки к настройкам прибора. Вызывается из GUI при отрисовке
    A-scan-виджета и из Excel-экспорта при формировании отчёта.

    Параметры:
      settings — либо `SettingsRecord` (TLV body[0]==2), либо «голый»
                 `UD2_AcousticSettings`, либо None (тогда возвращаются
                 «универсальные» оси по эталонным V=5900 и ADC=50 МГц).
      n_samples — длина массива самплов A-развёртки.
    """
    # 1. Достаём акустический срез — либо напрямую, либо из SettingsRecord.
    ac: UD2_AcousticSettings | None = None
    if isinstance(settings, UD2_AcousticSettings):
        ac = settings
    elif isinstance(settings, SettingsRecord):
        ac = decode_ud2102_acoustic(settings.raw_body)
        if ac is None:
            # Если нет «акустического» tail-блока — собираем из самих полей
            # SettingsRecord (они уже распакованы в decode()).
            ac = UD2_AcousticSettings(
                gain_db=settings.gain_db,
                velocity_ms=settings.velocity_ms,
                angle_deg=settings.angle_deg,
                gate_start=settings.gate_start,
                gate_width=settings.gate_width,
                threshold_pct=settings.threshold_pct,
                sens_required=settings.sens_required,
                sens_factual=settings.sens_factual,
                noise_cutoff_pct=settings.noise_cutoff,
            )
    # 2. Дефолты — если settings нет вовсе (а такое бывает: A-scan из памяти
    #    приходит без Settings, и тогда DLL берёт «эталоны»).
    if ac is None:
        ac = UD2_AcousticSettings(velocity_ms=5900, angle_deg=0, gate_start=0,
                                  gate_width=n_samples, threshold_pct=50)

    # 3. Считаем оси и оверлеи.
    duration_us = n_samples / max(0.01, sample_rate_mhz)
    v_mm_per_us = (ac.velocity_ms / 1000.0) if ac.velocity_ms > 0 else US_VEL_MM_PER_US
    max_depth_mm = duration_us * v_mm_per_us / 2.0
    gate_start_us = ac.gate_start / max(0.01, sample_rate_mhz)
    gate_end_us = (ac.gate_start + ac.gate_width) / max(0.01, sample_rate_mhz)
    gate_start_mm = gate_start_us * v_mm_per_us / 2.0
    gate_end_mm = gate_end_us * v_mm_per_us / 2.0
    threshold_db = ac.threshold_pct * AMPL_FULL_DB / 100.0
    return AScanAxes(
        sample_rate_mhz=sample_rate_mhz,
        velocity_ms=ac.velocity_ms or 5900,
        angle_deg=ac.angle_deg,
        n_samples=n_samples,
        duration_us=duration_us,
        max_depth_mm=max_depth_mm,
        threshold_db=threshold_db,
        threshold_pct=ac.threshold_pct,
        gate_start_us=gate_start_us,
        gate_end_us=gate_end_us,
        gate_start_mm=gate_start_mm,
        gate_end_mm=gate_end_mm,
        ard_curve=ac.ard_curve,
        sens_required=ac.sens_required,
        sens_factual=ac.sens_factual,
        vrt_table=ac.vrt_table,
    )


def find_settings_for_ascan(
    records: list[DecodedRecord],
    ascan: AScanRecord,
) -> SettingsRecord | None:
    """Найти ближайший SettingsRecord, который применим к данной A-развёртке.

    Логика DLL: ищется последний по offset SettingsRecord с тем же tag/1000
    «семейством» (или просто последний, если tag-семейство не совпадает).
    """
    last: SettingsRecord | None = None
    for r in records:
        if isinstance(r, SettingsRecord):
            last = r
        if r is ascan:
            break
    return last


# =============================================================================
# 3. БАЗА ДАННЫХ — SQLite-замена Firebird BlockZap
# =============================================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS BlockZap (
    Number   INTEGER PRIMARY KEY,
    BlockLen INTEGER NOT NULL,
    Block    BLOB    NOT NULL,
    Saved    TEXT    NOT NULL DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_blockzap_saved ON BlockZap(Saved);

-- Measurements: записи УД2-102 (BCD-паспорт оси/колесной пары).
-- Заполняется при «Прибор → Сканировать память (УД2-102)» либо при
-- генерации демо-записей. Хранит сырое тело (84 байта) для повторного
-- разбора и человекочитаемые декоды для отчёта.
CREATE TABLE IF NOT EXISTS Measurements (
    DeviceNo INTEGER NOT NULL,
    Addr     INTEGER NOT NULL,
    Date     TEXT    NOT NULL,
    TypeName TEXT    NOT NULL,
    NumObj   TEXT,
    Plavka   TEXT,
    Stamp    INTEGER,
    God      INTEGER,
    Side     TEXT,
    Sheika   TEXT,
    Obod     TEXT,
    Obtochka TEXT,
    Greben   TEXT,
    Naplavka TEXT,
    Raw      BLOB,
    -- Дешифрованные поля для FDB-импорта (вагонная версия). У записей,
    -- принятых с УД2-102 по UART эти столбцы остаются NULL.
    TimeStr  TEXT,            -- HH:MM:SS (TIMEFORM из FDB)
    Operator TEXT,            -- NAMEOPERA
    IndMaker TEXT,            -- INDMAKER (индекс изготовителя / завод)
    MakeTime INTEGER,         -- MAKETIME (год+месяц изготовления, INT)
    Defekt   TEXT,            -- DEFEKT ('нет' / 'есть')
    CodeDef  TEXT,            -- CODEDEF (код дефекта)
    TypeVar  INTEGER,         -- TYPEVAR-классификатор
    NumKod   INTEGER,         -- ссылка на FdbBlock.Number (для FDB-записей)
    Source   TEXT NOT NULL DEFAULT 'UD2',  -- 'UD2' | 'FDB-RESULTS' | 'FDB-SHORTPROT' | 'FDB-NASTR'
    Saved    TEXT NOT NULL DEFAULT (datetime('now')),
    PRIMARY KEY (DeviceNo, Addr)
);
CREATE INDEX IF NOT EXISTS idx_measurements_date   ON Measurements(Date);
-- Индексы idx_measurements_source / idx_measurements_numkod создаются в
-- BlockZapDB.__init__() ПОСЛЕ ALTER-TABLE-миграции — иначе на старых БД
-- (Measurements ещё без Source/NumKod) CREATE INDEX упадёт.

-- FdbBlock: зеркало FDB-таблицы BLOCKZAP (сырые BLOB-кадры) из импорта
-- оригинальной БД PelengPC.fdb (Firebird ODS 11.0, MAJ_INTERPRET = ODS 2.0).
-- В оригинальном PelengPC у каждой записи NASTR/RESULTS/SHORTPROT есть
-- NUMKOD, который указывает на BLOCKZAP.NUMBER. Мы храним эти BLOB здесь
-- параллельно с собственной таблицей BlockZap (приёмная по UART), чтобы
-- не путать импортированные и приёмные данные.
CREATE TABLE IF NOT EXISTS FdbBlock (
    Number   INTEGER PRIMARY KEY,        -- FDB BLOCKZAP.NUMBER (как было)
    BlockLen INTEGER NOT NULL,           -- BLOCKZAP.BLOCKLEN
    Block    BLOB    NOT NULL,           -- BLOCKZAP.BLOCK (заголовок 16 + TLV)
    Saved    TEXT    NOT NULL DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_fdbblock_saved ON FdbBlock(Saved);

-- FdbRecord: унифицированное зеркало NASTR{1,2,3} / RESULTS{1,2,3} /
-- SHORTPROT{1,2,3}.  Variant = '1' (общая), '2' (ВАГОННАЯ), '3' (ОТД).
-- Kind   = 'NASTR' (настройки), 'RESULTS' (протоколы А-развёртки),
-- 'SHORTPROT' (короткие отчёты по сканированию). Все 9 FDB-таблиц
-- свёрнуты в одну с помощью полей Variant + Kind, лишние столбцы
-- оригинала (M/MM/CLOCK/STAGE/SECTION/PICKETLINE/PATH/CONDLENGTH/...)
-- сохраняем в ExtraJson, чтобы не плодить пустые колонки.
CREATE TABLE IF NOT EXISTS FdbRecord (
    Variant   TEXT NOT NULL,             -- '1' | '2' | '3'
    Kind      TEXT NOT NULL,             -- 'NASTR' | 'RESULTS' | 'SHORTPROT'
    Number    INTEGER NOT NULL,          -- NUMBER (PK в FDB)
    NumKod    INTEGER,                   -- NUMKOD → FdbBlock.Number
    TypeZap   TEXT,                      -- 'Настройка' / 'Протокол А-развертки'
    DateForm  TEXT,                      -- ISO YYYY-MM-DD
    TimeForm  TEXT,                      -- HH:MM или HH:MM:SS
    KodOpera  INTEGER,                   -- код оператора
    NameOpera TEXT,                      -- имя оператора
    NumVers   TEXT,                      -- версия ПО прибора, напр. '06.43'
    NumPrib   INTEGER,                   -- заводской № прибора (NUMPRIB)
    TypeVar   INTEGER,                   -- TYPEVAR-классификатор
    NumObj    TEXT,                      -- № объекта (паспорт)
    Smelting  TEXT,                      -- плавка
    IndMaker  TEXT,                      -- индекс изготовителя
    MakeTime  INTEGER,                   -- год+месяц изготовления (INT)
    Defekt    TEXT,                      -- 'нет' / 'есть'
    CodeDef   TEXT,                      -- код дефекта
    Protocol  TEXT,                      -- 7-символьный код (SHORTPROT)
    NumZap    INTEGER,                   -- порядковый № в сессии
    ExtraJson TEXT,                      -- прочие FDB-поля (KM/M/MM/...)
    Saved     TEXT NOT NULL DEFAULT (datetime('now')),
    PRIMARY KEY (Variant, Kind, Number)
);
CREATE INDEX IF NOT EXISTS idx_fdb_record_numkod   ON FdbRecord(NumKod);
CREATE INDEX IF NOT EXISTS idx_fdb_record_typevar  ON FdbRecord(TypeVar);
CREATE INDEX IF NOT EXISTS idx_fdb_record_dateform ON FdbRecord(DateForm);
CREATE INDEX IF NOT EXISTS idx_fdb_record_kindvar  ON FdbRecord(Variant, Kind);
"""


@dataclass
class BlockRow:
    number:    int
    block_len: int
    block:     bytes
    saved:     str

    def header(self) -> ProtocolHeader:
        return ProtocolHeader.parse(self.block[:HEADER_LEN])


class BlockZapDB:
    def __init__(self, path: str | Path = ":memory:"):
        self.path = str(path)
        # check_same_thread=False: GUI создаёт коннект в главном потоке, а
        # _FdbImportWorker / _RecvWorker пишут из QThread'а. Python sqlite3
        # под капотом собран с SQLITE_THREADSAFE=1 (serialized), C-уровень
        # потокобезопасен. Локом RLock ниже сериализуем питон-вызовы, чтобы
        # курсоры/statement-кэш не конфликтовали.
        self._con = sqlite3.connect(self.path, check_same_thread=False)
        self._con.row_factory = sqlite3.Row
        self._con.executescript(SCHEMA_SQL)
        # Идемпотентные миграции для БД, которые были созданы старой версией
        # без FDB-«общей базы» (TimeStr/Operator/IndMaker/MakeTime/Defekt/
        # CodeDef/TypeVar/NumKod/Source). PRAGMA table_info → если столбца
        # нет — ALTER TABLE ADD COLUMN.
        _have = {r[1] for r in self._con.execute(
            "PRAGMA table_info(Measurements)")}
        _migrations: list[tuple[str, str]] = [
            ("TimeStr",  "TEXT"),
            ("Operator", "TEXT"),
            ("IndMaker", "TEXT"),
            ("MakeTime", "INTEGER"),
            ("Defekt",   "TEXT"),
            ("CodeDef",  "TEXT"),
            ("TypeVar",  "INTEGER"),
            ("NumKod",   "INTEGER"),
            ("Source",   "TEXT NOT NULL DEFAULT 'UD2'"),
        ]
        for _col, _decl in _migrations:
            if _col not in _have:
                self._con.execute(
                    f"ALTER TABLE Measurements ADD COLUMN {_col} {_decl}")
        # Индексы поверх новых колонок (на старых БД могли отсутствовать).
        self._con.execute(
            "CREATE INDEX IF NOT EXISTS idx_measurements_source "
            "ON Measurements(Source)")
        self._con.execute(
            "CREATE INDEX IF NOT EXISTS idx_measurements_numkod "
            "ON Measurements(NumKod)")
        self._con.commit()
        import threading as _threading
        self._lock = _threading.RLock()

    def close(self) -> None:
        with self._lock:
            self._con.close()

    def __enter__(self) -> "BlockZapDB":
        return self

    def __exit__(self, *args) -> None:
        self.close()

    def has(self, number: int) -> bool:
        return self._con.execute(
            "SELECT 1 FROM BlockZap WHERE Number=?", (number,)
        ).fetchone() is not None

    def numbers(self) -> list[int]:
        return [int(r[0]) for r in self._con.execute(
            "SELECT Number FROM BlockZap ORDER BY Number")]

    def all(self) -> list[BlockRow]:
        return [BlockRow(int(r[0]), int(r[1]), bytes(r[2]), str(r[3]))
                for r in self._con.execute(
                    "SELECT Number, BlockLen, Block, Saved FROM BlockZap ORDER BY Number")]

    def get(self, number: int) -> BlockRow | None:
        r = self._con.execute(
            "SELECT Number, BlockLen, Block, Saved FROM BlockZap WHERE Number=?",
            (number,),
        ).fetchone()
        return None if r is None else BlockRow(int(r[0]), int(r[1]), bytes(r[2]), str(r[3]))

    def insert(self, number: int, block: bytes) -> None:
        if self.has(number):
            return
        self._con.execute(
            "INSERT INTO BlockZap(Number, BlockLen, Block) VALUES(?,?,?)",
            (int(number), len(block), bytes(block)),
        )
        self._con.commit()

    def upsert(self, number: int, block: bytes) -> None:
        self._con.execute(
            "INSERT INTO BlockZap(Number, BlockLen, Block, Saved) "
            "VALUES(?,?,?, datetime('now')) "
            "ON CONFLICT(Number) DO UPDATE SET BlockLen=excluded.BlockLen, "
            "Block=excluded.Block, Saved=datetime('now')",
            (int(number), len(block), bytes(block)),
        )
        self._con.commit()

    def delete(self, number: int) -> None:
        self._con.execute("DELETE FROM BlockZap WHERE Number=?", (int(number),))
        self._con.commit()

    def count(self) -> int:
        return int(self._con.execute("SELECT COUNT(*) FROM BlockZap").fetchone()[0])

    # ------- УД2-102 Measurements (BCD-паспорт) ------------------------------
    # Source='UD2' — записи, принятые по UART; FDB-импорт пишет сюда же
    # через upsert_measurement_from_fdb со своим Source='FDB-...'.
    def upsert_measurement(self, device_no: int, rec: "MeasurementRecord") -> None:
        self._con.execute(
            "INSERT INTO Measurements(DeviceNo, Addr, Date, TypeName, NumObj, "
            "Plavka, Stamp, God, Side, Sheika, Obod, Obtochka, Greben, Naplavka, "
            "Raw, Source, Saved) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'UD2', datetime('now')) "
            "ON CONFLICT(DeviceNo, Addr) DO UPDATE SET "
            "Date=excluded.Date, TypeName=excluded.TypeName, NumObj=excluded.NumObj, "
            "Plavka=excluded.Plavka, Stamp=excluded.Stamp, God=excluded.God, "
            "Side=excluded.Side, Sheika=excluded.Sheika, Obod=excluded.Obod, "
            "Obtochka=excluded.Obtochka, Greben=excluded.Greben, "
            "Naplavka=excluded.Naplavka, Raw=excluded.Raw, Source='UD2', "
            "Saved=datetime('now')",
            (int(device_no), int(rec.addr), rec.date_str, rec.type_name,
             rec.num_obj, rec.plavka, int(rec.stamp), int(rec.god),
             rec.side, rec.sheika, rec.obod, rec.obtochka, rec.greben,
             rec.naplavka, bytes(rec.raw)),
        )
        self._con.commit()

    # FDB → Measurements: общая база с дешифровкой. NumPrib используется как
    # синтетический DeviceNo, а Addr = 0xF000_0000 | Number — чтобы FDB-строки
    # не пересекались по PK с реальными адресами УД2-102 (0x2774..0x34AC) и
    # их легко отфильтровать (Addr >= 0xF000_0000). Источник записи (Source)
    # пишется как 'FDB-RESULTS' / 'FDB-SHORTPROT' / 'FDB-NASTR' — пользователь
    # сразу видит, откуда взяты строки.
    FDB_ADDR_OFFSET = 0xF000_0000

    def upsert_measurement_from_fdb(self, kind: str, variant: str,
                                    row: dict, raw_blob: bytes | None = None) -> None:
        number = int(row.get("NUMBER") or 0)
        if number <= 0:
            return
        num_prib = _opt_int(row.get("NUMPRIB")) or 0
        date_iso = _date_iso(row.get("DATEFORM")) or ""
        time_str = _opt_str(row.get("TIMEFORM")) or ""
        # Дата как «YYYY-MM-DD HH:MM» — то же поле, что и у UD2-записей.
        # Сразу чистим, чтобы из FDB не утекли управляющие байты («квадратики»).
        if date_iso and time_str:
            date_full = f"{date_iso} {time_str[:5]}"
        else:
            date_full = date_iso or time_str or ""
        date_full = _strip_unprintable(date_full)
        type_var = _opt_int(row.get("TYPEVAR"))
        type_zap = _opt_str(row.get("TYPEZAP")) or ""
        # «Тип» в столбце Записей УД2-102: предпочитаем человекочитаемое
        # значение из TYPEVAR (fdb_typevar_name), иначе TYPEZAP.
        if type_var is not None:
            type_name = fdb_typevar_name(type_var) or type_zap or f"TypeVar={type_var}"
        else:
            type_name = type_zap or "—"
        type_name = _strip_unprintable(str(type_name))
        num_obj   = _opt_str(row.get("NUMOBJ"))
        smelting  = _opt_str(row.get("SMELTING"))
        ind_maker = _opt_str(row.get("INDMAKER"))
        make_time = _opt_int(row.get("MAKETIME"))
        defekt    = _opt_str(row.get("DEFEKT"))
        code_def  = _opt_str(row.get("CODEDEF"))
        name_oper = _opt_str(row.get("NAMEOPERA"))
        num_kod   = _opt_int(row.get("NUMKOD"))
        addr      = self.FDB_ADDR_OFFSET | number
        source    = f"FDB-{kind}"

        # ---- SHORTPROT2 → дешифровка ВСЕХ полей УД2-паспорта ---------------
        # Поле PROTOCOL в SHORTPROT2 — это BLOB c BCD-записью того же
        # формата, что приходит по UART (parse_ud2102_record). У оригинальной
        # вагонной программы это «отчёт по сканированию», и именно в нём
        # лежат сторона/шейка/обод/обточка/гребень/наплавка. Если получилось
        # распарсить — отдаём всё в общую таблицу Measurements.
        side = sheika = obod = obtochka = greben = naplavka = ""
        plavka_dec = smelting or ""
        num_obj_dec = num_obj or ""
        bcd_blob = None
        for _k in ("PROTOCOL", "BCDDATA", "DATA"):
            v = row.get(_k)
            if isinstance(v, (bytes, bytearray)) and len(v) >= 70:
                bcd_blob = bytes(v); break
        if bcd_blob is not None:
            bcd_rec = _try_parse_ud2_bcd(bcd_blob)
            if bcd_rec is not None:
                side       = bcd_rec.side
                sheika     = bcd_rec.sheika
                obod       = bcd_rec.obod
                obtochka   = bcd_rec.obtochka
                greben     = bcd_rec.greben
                naplavka   = bcd_rec.naplavka
                # BCD-плавка/№объекта обычно полнее, чем колонки FDB — но
                # перезаписываем только если в FDB-полях пусто.
                if not plavka_dec and bcd_rec.plavka:
                    plavka_dec = bcd_rec.plavka
                if not num_obj_dec and bcd_rec.num_obj:
                    num_obj_dec = bcd_rec.num_obj

        self._con.execute(
            "INSERT INTO Measurements(DeviceNo, Addr, Date, TypeName, NumObj, "
            "Plavka, Stamp, God, Side, Sheika, Obod, Obtochka, Greben, Naplavka, "
            "Raw, TimeStr, Operator, IndMaker, MakeTime, Defekt, CodeDef, "
            "TypeVar, NumKod, Source, Saved) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, "
            "       datetime('now')) "
            "ON CONFLICT(DeviceNo, Addr) DO UPDATE SET "
            "Date=excluded.Date, TypeName=excluded.TypeName, "
            "NumObj=excluded.NumObj, Plavka=excluded.Plavka, "
            "Stamp=excluded.Stamp, God=excluded.God, "
            "Side=excluded.Side, Sheika=excluded.Sheika, "
            "Obod=excluded.Obod, Obtochka=excluded.Obtochka, "
            "Greben=excluded.Greben, Naplavka=excluded.Naplavka, "
            "TimeStr=excluded.TimeStr, Operator=excluded.Operator, "
            "IndMaker=excluded.IndMaker, MakeTime=excluded.MakeTime, "
            "Defekt=excluded.Defekt, CodeDef=excluded.CodeDef, "
            "TypeVar=excluded.TypeVar, NumKod=excluded.NumKod, "
            "Source=excluded.Source, Saved=datetime('now')",
            (
                int(num_prib), int(addr),
                date_full, str(type_name),
                num_obj_dec, plavka_dec,
                int(num_prib),                # Stamp = № прибора (как «клеймо»)
                _fdb_make_year(make_time),    # God = год изготовления (из MAKETIME)
                side, sheika,
                obod, obtochka, greben,
                naplavka,
                bytes(raw_blob or b""),
                time_str or None,
                name_oper,
                ind_maker,
                make_time,
                defekt,
                code_def,
                type_var,
                num_kod,
                source,
            ),
        )

    def enrich_measurements_from_blocks(self) -> int:
        """Второй проход после BLOCKZAP-импорта. Для всех FDB-зеркальных
        строк (Source LIKE 'FDB-%') ищем соответствующий BLOCKZAP-блок по
        NumKod, парсим TLV и пытаемся вытащить BCD-паспорт УД2-102 — если
        он там есть, выкатываем сторону/шейку/обод/обточку/гребень/наплавку
        в Measurements. На SHORTPROT-строках это часто избыточно (BCD уже
        в PROTOCOL-BLOB), на RESULTS-строках — основной способ обогащения.
        Возвращает количество обновлённых строк.
        """
        cur = self._con.execute(
            "SELECT DeviceNo, Addr, NumKod FROM Measurements "
            "WHERE Source LIKE 'FDB-%' AND NumKod IS NOT NULL "
            "AND (COALESCE(Side,'')='' AND COALESCE(Sheika,'')='' "
            "     AND COALESCE(Naplavka,'')='')")
        targets = cur.fetchall()
        updated = 0
        for dev_no, addr, num_kod in targets:
            r = self._con.execute(
                "SELECT Block FROM FdbBlock WHERE Number=?",
                (int(num_kod),)).fetchone()
            if r is None:
                continue
            block = bytes(r[0] or b"")
            if not block:
                continue
            bcd_rec = _try_extract_ud2_bcd_from_tlv(block)
            if bcd_rec is None:
                continue
            self._con.execute(
                "UPDATE Measurements SET Side=?, Sheika=?, Obod=?, Obtochka=?, "
                "Greben=?, Naplavka=?, "
                "NumObj=CASE WHEN COALESCE(NumObj,'')='' THEN ? ELSE NumObj END, "
                "Plavka=CASE WHEN COALESCE(Plavka,'')='' THEN ? ELSE Plavka END "
                "WHERE DeviceNo=? AND Addr=?",
                (bcd_rec.side, bcd_rec.sheika, bcd_rec.obod,
                 bcd_rec.obtochka, bcd_rec.greben, bcd_rec.naplavka,
                 bcd_rec.num_obj, bcd_rec.plavka,
                 int(dev_no), int(addr)))
            updated += 1
        self._con.commit()
        return updated

    def all_measurements(self, device_no: int | None = None) -> list[dict]:
        if device_no is None:
            cur = self._con.execute(
                "SELECT DeviceNo, Addr, Date, TypeName, NumObj, Plavka, Stamp, "
                "God, Side, Sheika, Obod, Obtochka, Greben, Naplavka, Raw, "
                "TimeStr, Operator, IndMaker, MakeTime, Defekt, CodeDef, "
                "TypeVar, NumKod, Source "
                "FROM Measurements ORDER BY DeviceNo, Addr")
        else:
            cur = self._con.execute(
                "SELECT DeviceNo, Addr, Date, TypeName, NumObj, Plavka, Stamp, "
                "God, Side, Sheika, Obod, Obtochka, Greben, Naplavka, Raw, "
                "TimeStr, Operator, IndMaker, MakeTime, Defekt, CodeDef, "
                "TypeVar, NumKod, Source "
                "FROM Measurements WHERE DeviceNo=? ORDER BY Addr",
                (int(device_no),))
        out: list[dict] = []
        for r in cur:
            out.append({
                "device_no": int(r[0]), "addr": int(r[1]), "date": str(r[2]),
                "type_name": str(r[3]), "num_obj": str(r[4] or ""),
                "plavka": str(r[5] or ""), "stamp": int(r[6] or 0),
                "god": int(r[7] or 0), "side": str(r[8] or ""),
                "sheika": str(r[9] or ""), "obod": str(r[10] or ""),
                "obtochka": str(r[11] or ""), "greben": str(r[12] or ""),
                "naplavka": str(r[13] or ""), "raw": bytes(r[14] or b""),
                "time_str": str(r[15] or ""), "operator": str(r[16] or ""),
                "ind_maker": str(r[17] or ""), "make_time": int(r[18] or 0),
                "defekt": str(r[19] or ""), "code_def": str(r[20] or ""),
                "type_var": int(r[21] or 0), "num_kod": int(r[22] or 0),
                "source": str(r[23] or "UD2"),
            })
        return out

    def count_measurements(self, device_no: int | None = None) -> int:
        if device_no is None:
            return int(self._con.execute(
                "SELECT COUNT(*) FROM Measurements").fetchone()[0])
        return int(self._con.execute(
            "SELECT COUNT(*) FROM Measurements WHERE DeviceNo=?",
            (int(device_no),)).fetchone()[0])

    def delete_measurements(self, device_no: int | None = None) -> int:
        if device_no is None:
            cur = self._con.execute("DELETE FROM Measurements")
        else:
            cur = self._con.execute(
                "DELETE FROM Measurements WHERE DeviceNo=?",
                (int(device_no),))
        self._con.commit()
        return int(cur.rowcount or 0)

    # ------- FDB-импорт (PelengPC.fdb → SQLite зеркало) ----------------------
    # Сюда складывает данные кнопка «Импортировать FDB…» в шапке окна.
    # Чтобы commit'ы не били по производительности, методы _bulk_*_begin /
    # _bulk_*_end управляют единственной транзакцией на весь импорт.
    def fdb_bulk_begin(self) -> None:
        """Открывает явную транзакцию (commit делается в fdb_bulk_end)."""
        self._con.execute("BEGIN")

    def fdb_bulk_end(self) -> None:
        self._con.commit()

    def upsert_fdb_block(self, number: int, block_len: int, block: bytes) -> None:
        self._con.execute(
            "INSERT INTO FdbBlock(Number, BlockLen, Block, Saved) "
            "VALUES(?,?,?, datetime('now')) "
            "ON CONFLICT(Number) DO UPDATE SET BlockLen=excluded.BlockLen, "
            "Block=excluded.Block, Saved=datetime('now')",
            (int(number), int(block_len), bytes(block)),
        )

    def upsert_fdb_record(self, kind: str, variant: str,
                          fields: dict) -> None:
        """Записывает строку из NASTR/RESULTS/SHORTPROT в унифицированную
        таблицу FdbRecord. ``fields`` — словарь {ИМЯ_КОЛОНКИ_FDB: значение}.
        Лишние FDB-колонки сохраняются в ExtraJson."""
        std = {"NUMBER", "NUMKOD", "TYPEZAP", "DATEFORM", "TIMEFORM",
               "KODOPERA", "NAMEOPERA", "NUMVERS", "NUMPRIB", "TYPEVAR",
               "NUMOBJ", "SMELTING", "INDMAKER", "MAKETIME", "DEFEKT",
               "CODEDEF", "PROTOCOL", "NUMZAP"}
        extra = {k: _fdb_jsonable(v) for k, v in fields.items()
                 if k.upper() not in std and v is not None}
        extra_json = json.dumps(extra, ensure_ascii=False) if extra else None
        self._con.execute(
            "INSERT INTO FdbRecord(Variant, Kind, Number, NumKod, TypeZap, "
            "  DateForm, TimeForm, KodOpera, NameOpera, NumVers, NumPrib, "
            "  TypeVar, NumObj, Smelting, IndMaker, MakeTime, Defekt, "
            "  CodeDef, Protocol, NumZap, ExtraJson, Saved) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, datetime('now')) "
            "ON CONFLICT(Variant, Kind, Number) DO UPDATE SET "
            "  NumKod=excluded.NumKod, TypeZap=excluded.TypeZap, "
            "  DateForm=excluded.DateForm, TimeForm=excluded.TimeForm, "
            "  KodOpera=excluded.KodOpera, NameOpera=excluded.NameOpera, "
            "  NumVers=excluded.NumVers, NumPrib=excluded.NumPrib, "
            "  TypeVar=excluded.TypeVar, NumObj=excluded.NumObj, "
            "  Smelting=excluded.Smelting, IndMaker=excluded.IndMaker, "
            "  MakeTime=excluded.MakeTime, Defekt=excluded.Defekt, "
            "  CodeDef=excluded.CodeDef, Protocol=excluded.Protocol, "
            "  NumZap=excluded.NumZap, ExtraJson=excluded.ExtraJson, "
            "  Saved=datetime('now')",
            (
                str(variant), str(kind),
                int(fields.get("NUMBER") or 0),
                _opt_int(fields.get("NUMKOD")),
                _opt_str(fields.get("TYPEZAP")),
                _date_iso(fields.get("DATEFORM")),
                _opt_str(fields.get("TIMEFORM")),
                _opt_int(fields.get("KODOPERA")),
                _opt_str(fields.get("NAMEOPERA")),
                _opt_str(fields.get("NUMVERS")),
                _opt_int(fields.get("NUMPRIB")),
                _opt_int(fields.get("TYPEVAR")),
                _opt_str(fields.get("NUMOBJ")),
                _opt_str(fields.get("SMELTING")),
                _opt_str(fields.get("INDMAKER")),
                _opt_int(fields.get("MAKETIME")),
                _opt_str(fields.get("DEFEKT")),
                _opt_str(fields.get("CODEDEF")),
                _opt_str(fields.get("PROTOCOL")),
                _opt_int(fields.get("NUMZAP")),
                extra_json,
            ),
        )

    def fdb_record_count(self, variant: str | None = None,
                         kind: str | None = None) -> int:
        sql = "SELECT COUNT(*) FROM FdbRecord WHERE 1=1"
        args: list = []
        if variant:
            sql += " AND Variant=?"; args.append(variant)
        if kind:
            sql += " AND Kind=?"; args.append(kind)
        return int(self._con.execute(sql, args).fetchone()[0])

    def fdb_block_count(self) -> int:
        return int(self._con.execute(
            "SELECT COUNT(*) FROM FdbBlock").fetchone()[0])

    def fdb_records(self, variant: str | None = None,
                    kind: str | None = None,
                    limit: int | None = None) -> list[dict]:
        """Возвращает список словарей с импортированными FDB-записями."""
        sql = ("SELECT Variant, Kind, Number, NumKod, TypeZap, DateForm, "
               "TimeForm, KodOpera, NameOpera, NumVers, NumPrib, TypeVar, "
               "NumObj, Smelting, IndMaker, MakeTime, Defekt, CodeDef, "
               "Protocol, NumZap, ExtraJson "
               "FROM FdbRecord WHERE 1=1")
        args: list = []
        if variant:
            sql += " AND Variant=?"; args.append(variant)
        if kind:
            sql += " AND Kind=?"; args.append(kind)
        sql += " ORDER BY DateForm, TimeForm, Number"
        if limit:
            sql += " LIMIT ?"; args.append(int(limit))
        cols = ["variant", "kind", "number", "num_kod", "type_zap",
                "date_form", "time_form", "kod_opera", "name_opera",
                "num_vers", "num_prib", "type_var", "num_obj", "smelting",
                "ind_maker", "make_time", "defekt", "code_def",
                "protocol", "num_zap", "extra_json"]
        return [dict(zip(cols, r)) for r in self._con.execute(sql, args)]

    def fdb_block_get(self, number: int) -> bytes | None:
        r = self._con.execute(
            "SELECT Block FROM FdbBlock WHERE Number=?",
            (int(number),)).fetchone()
        return None if r is None else bytes(r[0])

    def fdb_delete_all(self) -> int:
        """Удаляет все импортированные FDB-записи + их BLOB + FDB-зеркало в
        Measurements (Source LIKE 'FDB-%'). UART-записи УД2-102 (Source='UD2')
        не трогаем."""
        n1 = self._con.execute("DELETE FROM FdbRecord").rowcount or 0
        n2 = self._con.execute("DELETE FROM FdbBlock").rowcount or 0
        n3 = self._con.execute(
            "DELETE FROM Measurements WHERE Source LIKE 'FDB-%'").rowcount or 0
        self._con.commit()
        return int(n1 + n2 + n3)

    def fdb_stats(self) -> dict:
        """Краткая сводка: сколько записей каждой Kind/Variant + блоков."""
        cur = self._con.execute(
            "SELECT Variant, Kind, COUNT(*) FROM FdbRecord "
            "GROUP BY Variant, Kind ORDER BY Variant, Kind")
        per_table = {(v, k): int(c) for v, k, c in cur.fetchall()}
        return {
            "by_variant_kind": per_table,
            "total_records":   sum(per_table.values()),
            "total_blocks":    self.fdb_block_count(),
        }


# =============================================================================
# 3A. ИМПОРТ ИЗ ОРИГИНАЛЬНОЙ FDB-БАЗЫ PelengPC.fdb (Firebird 2.x)
# =============================================================================
# Оригинальный PelengPC хранит ВСЕ данные в Firebird-БД PelengPC.fdb со схемой:
#
#   BLOCKZAP        : NUMBER (PK)            BLOCKLEN  BLOCK (BLOB sub_type=0)
#   NASTR{1,2,3}    : NUMBER (PK)            NUMKOD → BLOCKZAP.NUMBER (настройки)
#   RESULTS{1,2,3}  : NUMBER (PK)            NUMKOD → BLOCKZAP.NUMBER (А-протоколы)
#   SHORTPROT{1,2,3}: NUMBER (PK)            NUMKOD → BLOCKZAP.NUMBER (кор. отчёты)
#   TOLSHINS1       : NUMBER (PK)                                    (толщиномер)
#
# Suffix 1/2/3 = тип дефектоскопа (вариант ПО):
#   1 — общая БД (рельсы кусками, MM/M/CLOCK),
#   2 — ВАГОННАЯ (колёса/оси, SMELTING/INDMAKER/MAKETIME),
#   3 — ОТД-вариант (стадия/секция/пикетаж/путь).
#
# Все три варианта несут одинаковый набор «технологических» полей: NUMKOD,
# TYPEZAP, DATEFORM, TIMEFORM, KODOPERA, NAMEOPERA, NUMVERS, NUMPRIB, TYPEVAR,
# NUMZAP. Дополнительные FDB-колонки уходят в FdbRecord.ExtraJson.
#
# BLOB-кадры в BLOCKZAP — это РОВНО те же самые 16-байт-header + TLV, что
# и приходят с прибора по UART (наша таблица BlockZap), только сохранённые
# на диск как есть. Поэтому parse_tlv() / decode_all() работают «как есть»
# на содержимом FdbBlock.Block.
#
# Доступ к .fdb осуществляется через embedded-режим Firebird (libfbembed.dll
# на Windows, libfbembed.so на Linux). Сетевой сервер не нужен.

# Зарегистрированные имена FDB-таблиц с человеческим описанием — для
# отображения в прогресс-диалоге импорта.
FDB_KINDS: tuple[tuple[str, str], ...] = (
    ("NASTR",     "Настройки"),
    ("RESULTS",   "Протоколы А-развёртки"),
    ("SHORTPROT", "Короткие отчёты"),
)
FDB_VARIANTS: tuple[tuple[str, str], ...] = (
    ("1", "общая БД (рельсы)"),
    ("2", "ВАГОННАЯ (оси/колёса)"),
    ("3", "ОТД (стадия/секция)"),
)
# Имя единственной таблицы с raw BLOB.
FDB_BLOCKZAP_TABLE = "BLOCKZAP"


def _opt_int(v) -> int | None:
    """FDB → SQLite: None или int."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None


def _strip_unprintable(s: str) -> str:
    """Убирает «квадратики» из строк FDB: все CTRL-символы (< 0x20, кроме
    \\t \\n) и DEL (0x7f), плюс непечатаемые из C1-блока (0x80..0x9f).
    Сохраняем кириллицу/латиницу/цифры/пунктуацию. ``str.rstrip`` в конце
    срезает оставшиеся пробелы/нулевые байты, чтобы пустая «строка из
    одних \\x00» превратилась в ``""``.
    """
    if not s:
        return s
    out_chars = []
    for ch in s:
        o = ord(ch)
        # 0x00..0x1F (кроме TAB/LF/CR), 0x7F, 0x80..0x9F — выкидываем.
        if o < 0x20 and o not in (0x09, 0x0A, 0x0D):
            continue
        if o == 0x7F:
            continue
        if 0x80 <= o <= 0x9F:
            continue
        out_chars.append(ch)
    return "".join(out_chars).strip()


def _opt_str(v) -> str | None:
    """FDB → SQLite: None или строка (Win-1251 в bytes уже декодирована fdb)."""
    if v is None:
        return None
    if isinstance(v, bytes):
        for enc in ("cp1251", "utf-8"):
            try:
                s = v.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            s = v.hex()
    else:
        s = str(v)
    s = _strip_unprintable(s.rstrip(" \x00"))
    return s if s else None


def _date_iso(v) -> str | None:
    """datetime.date / datetime.datetime / строка → 'YYYY-MM-DD'."""
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    s = _opt_str(v)
    return s if s else None


def _fdb_make_year(maketime: int | None) -> int:
    """Извлекает год изготовления из FDB-поля MAKETIME.

    В оригинальном PelengPC MAKETIME — INTEGER, кодирующий ГГГГММ
    (например, 201509 = сентябрь 2015). Бывают и просто ГГГГ.  Если
    получили что-то нестандартное — возвращаем 0 (значит «—» в UI).
    """
    if not maketime:
        return 0
    v = int(maketime)
    if v >= 190000:                       # ГГГГММ
        return v // 100
    if 1900 <= v <= 2100:                 # просто ГГГГ
        return v
    return 0


def _fdb_jsonable(v):
    """Превращает FDB-значение в JSON-совместимое представление."""
    if v is None or isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, (bytes, bytearray)):
        return v.hex()
    if isinstance(v, _dt.datetime):
        return v.isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, _dt.time):
        return v.isoformat()
    if hasattr(v, "read"):  # fdb BlobReader
        try:
            return v.read().hex()
        except Exception:
            return None
    return _opt_str(v)


# ─── Нативная поддержка fbembed: zero-config авто-поиск + on-demand
# скачивание официального Firebird 2.5.9 embed-пакета с GitHub Releases.
# Идея: пользователь запустил скрипт — он сам нашёл libfbembed в системе
# (Linux: apt install / /opt/firebird) либо в локальном кэше
# ~/.peleng_clone/fb25/. Если нет — на Windows предлагаем «Скачать сейчас»
# (12 МБ, GitHub Releases) и распаковываем в кэш. На Linux подсказываем
# точную команду apt-install.
PELENG_DATA_DIR  = Path.home() / ".peleng_clone"
PELENG_FB25_DIR  = PELENG_DATA_DIR / "fb25"

# Официальные релизы Firebird 2.5.9 (R2_5_9 = последняя стабильная 2.5.x).
# Размер ~12 МБ, лежат на GitHub Releases (FirebirdSQL/firebird).
FBEMBED_RELEASES: dict[str, str] = {
    "win64": ("https://github.com/FirebirdSQL/firebird/releases/download/"
              "R2_5_9/Firebird-2.5.9.27139-0_x64_embed.zip"),
    "win32": ("https://github.com/FirebirdSQL/firebird/releases/download/"
              "R2_5_9/Firebird-2.5.9.27139-0_Win32_embed.zip"),
}


def _fbembed_platform_key() -> str | None:
    """Ключ платформы для FBEMBED_RELEASES (только Windows авто-устанавливается)."""
    if sys.platform != "win32":
        return None
    # 64-битный Python ⇒ нужен x64 fbembed, 32-битный ⇒ Win32.
    return "win64" if sys.maxsize > 2**32 else "win32"


def _fdb_find_client_lib() -> str | None:
    """Подбирает путь к libfbclient/libfbembed (embedded или клиент).

    PelengPC.fdb имеет ODS 11.0 (Firebird 2.x). Firebird 3+ его не откроет,
    поэтому ищем сначала старые библиотеки. Порядок поиска:
      1) env-переменная ``PELENG_FBCLIENT`` (явный override);
      2) локальный кэш ``~/.peleng_clone/fb25/`` (туда кладёт авто-загрузчик);
      3) каталог рядом со скриптом (можно подкинуть .dll вручную);
      4) системные пути (apt-install / C:\\Program Files\\Firebird\\…).
    """
    cands: list[str] = []
    env = os.environ.get("PELENG_FBCLIENT")
    if env:
        cands.append(env)
    # 2) Локальный кэш — туда же качает fdb_install_fbembed()
    for fn in ("fbembed.dll", "fbclient.dll",
               "libfbembed.so", "libfbembed.so.2.5",
               "libfbembed.so.2", "libfbclient.so.2"):
        cands.append(str(PELENG_FB25_DIR / fn))
    # 3) Рядом со скриптом (можно подкинуть .dll прямо в exe-папку)
    try:
        here = Path(__file__).resolve().parent
        for fn in ("fbembed.dll", "fbclient.dll",
                   "libfbembed.so", "libfbembed.so.2.5",
                   "libfbembed.so.2", "libfbclient.so.2"):
            cands.append(str(here / fn))
    except Exception:
        pass
    # 4) Типовые системные пути
    cands += [
        # Linux — Firebird 2.5 (ODS 11.x совместимая ветка)
        "/usr/lib/x86_64-linux-gnu/libfbembed.so.2.5",
        "/usr/lib/x86_64-linux-gnu/libfbembed.so",
        "/usr/lib/libfbembed.so.2.5",
        "/usr/lib/libfbembed.so",
        "/opt/firebird/lib/libfbembed.so.2.5",
        "/opt/firebird/lib/libfbembed.so",
        # Linux — fbclient (хуже: нужен сервер, embedded предпочтительнее)
        "/usr/lib/x86_64-linux-gnu/libfbclient.so.2",
        "/usr/lib/x86_64-linux-gnu/libfbclient.so",
        "/usr/lib/libfbclient.so.2",
        "/usr/lib/libfbclient.so",
        "/opt/firebird/lib/libfbclient.so.2",
        # Windows — Firebird 2.5 embedded / client
        r"C:\Program Files\Firebird\Firebird_2_5\bin\fbembed.dll",
        r"C:\Program Files (x86)\Firebird\Firebird_2_5\bin\fbembed.dll",
        r"C:\Program Files\Firebird\Firebird_2_5\bin\fbclient.dll",
        r"C:\Program Files (x86)\Firebird\Firebird_2_5\bin\fbclient.dll",
        r"C:\Firebird\Firebird_2_5\bin\fbembed.dll",
        r"C:\Firebird\Firebird_2_5\bin\fbclient.dll",
        # Mac
        "/Library/Frameworks/Firebird.framework/Libraries/libfbclient.dylib",
    ]
    for p in cands:
        if p and os.path.exists(p):
            return p
    return None


def fdb_install_fbembed(progress_cb: Callable[[int, int], None] | None = None,
                        force: bool = False) -> str:
    """Скачивает и распаковывает Firebird 2.5.9 embed для текущей платформы
    в ``~/.peleng_clone/fb25/``. Возвращает путь к ``fbembed.dll``.

    Работает только на Windows (на Linux/Mac кидает PelengError с инструк-
    цией для apt/brew, т.к. там штатно ставится пакетом). progress_cb(done,
    total) опционально вызывается во время загрузки.
    """
    key = _fbembed_platform_key()
    if key is None:
        if sys.platform.startswith("linux"):
            raise PelengError(
                "Linux: автоматическая установка fbembed не делается — поставь\n"
                "пакетом:\n"
                "    sudo apt install firebird2.5-classic-common libfbembed2.5\n"
                "или скачай Firebird-2.5.9.tar.bz2 с\n"
                "    https://github.com/FirebirdSQL/firebird/releases/tag/R2_5_9\n"
                "и положи libfbembed.so.2.5 в ~/.peleng_clone/fb25/")
        if sys.platform == "darwin":
            raise PelengError(
                "macOS: автоматическая установка fbembed не делается — поставь\n"
                "Firebird.framework 2.5 с https://firebirdsql.org/en/firebird-2-5/\n"
                "или положи libfbclient.dylib в ~/.peleng_clone/fb25/")
        raise PelengError(f"Платформа {sys.platform} не поддерживается.")

    # Если уже скачали — выходим, если только не --force.
    PELENG_FB25_DIR.mkdir(parents=True, exist_ok=True)
    dll = PELENG_FB25_DIR / "fbembed.dll"
    if dll.exists() and not force:
        return str(dll)

    import urllib.request, zipfile, tempfile
    url = FBEMBED_RELEASES[key]
    zip_path = PELENG_FB25_DIR / "_fbembed.zip"
    try:
        req = urllib.request.Request(
            url, headers={"User-Agent": "peleng-clone/1.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            total = int(resp.headers.get("Content-Length", "0") or 0)
            with open(zip_path, "wb") as out:
                done = 0
                while True:
                    chunk = resp.read(64 * 1024)
                    if not chunk:
                        break
                    out.write(chunk)
                    done += len(chunk)
                    if progress_cb:
                        progress_cb(done, total)
    except Exception as e:
        raise PelengError(
            f"Не удалось скачать Firebird 2.5.9 embed:\n  {url}\n  {e}\n\n"
            "Скачай ZIP вручную и распакуй файлы (fbembed.dll, ib_util.dll,\n"
            "icu*.dll, firebird.msg, intl/fbintl.dll) в\n"
            f"  {PELENG_FB25_DIR}")

    # Распаковываем только нужные файлы — fbembed.dll + рантайм-зависимости.
    # Структура архива: Firebird-2.5.9.27139-0/{bin,intl,…}, но у embed-пакета
    # обычно файлы лежат в корне. Берём всё что узнаём по имени.
    wanted = {
        "fbembed.dll", "fbclient.dll", "ib_util.dll", "icudt30.dll",
        "icuin30.dll", "icuuc30.dll", "msvcp80.dll", "msvcr80.dll",
        "firebird.msg", "firebird.conf", "fbintl.dll", "fbintl.conf",
        "aliases.conf", "security2.fdb",
    }
    try:
        with zipfile.ZipFile(zip_path) as zf:
            for member in zf.namelist():
                name = os.path.basename(member).lower()
                if not name:
                    continue
                if name in wanted:
                    # сохраняем в фикс-структуру PELENG_FB25_DIR/...
                    # intl/* кладём в intl/, всё остальное — в корень.
                    if "intl/" in member.lower().replace("\\", "/"):
                        target = PELENG_FB25_DIR / "intl" / name
                    else:
                        target = PELENG_FB25_DIR / name
                    target.parent.mkdir(parents=True, exist_ok=True)
                    with zf.open(member) as src, open(target, "wb") as dst:
                        dst.write(src.read())
    except zipfile.BadZipFile as e:
        raise PelengError(f"Скачанный файл не является ZIP-архивом: {e}")
    finally:
        try:
            zip_path.unlink()
        except OSError:
            pass

    if not dll.exists():
        raise PelengError(
            f"После распаковки fbembed.dll не найден в {PELENG_FB25_DIR}.\n"
            "Возможно, формат ZIP-архива Firebird изменился. Распакуй\n"
            f"вручную: {url}")
    return str(dll)


def fdb_open(path: str, *,
             user: str = "SYSDBA",
             password: str = "masterkey",
             charset: str = "WIN1251",
             auto_install: bool = False,
             install_progress: Callable[[int, int], None] | None = None):
    """Открывает PelengPC.fdb в embedded-режиме.

    Если ``auto_install=True`` и библиотека не найдена — скачивает Firebird
    2.5.9 embed автоматически (только Windows). По умолчанию — бросает
    PelengError с подробной инструкцией.
    """
    try:
        import fdb as _fdb  # type: ignore
    except ImportError:
        raise PelengError(
            "Для импорта .fdb нужен Python-пакет 'fdb'. Установи:\n"
            "    pip install fdb\n"
            "Дальше Firebird 2.5 embedded подхватится автоматически.")

    libpath = _fdb_find_client_lib()
    if libpath is None and auto_install:
        # Авто-загрузка (только Windows). На остальных платформах сразу
        # пробросится PelengError с инструкцией для apt.
        libpath = fdb_install_fbembed(progress_cb=install_progress)
    if libpath is None:
        key = _fbembed_platform_key()
        hint_auto = (
            "  • Авто-загрузка: запусти `python peleng_clone.py "
            "fbembed-install` —\n    скачаю и распакую "
            "Firebird-2.5.9 embed (~12 МБ) в\n"
            f"    {PELENG_FB25_DIR}\n") if key else ""
        raise PelengError(
            "Не найдена клиентская/embedded библиотека Firebird 2.x.\n"
            "PelengPC.fdb — это Firebird 2.x (ODS 11.0), Firebird 3+ его\n"
            "НЕ откроет. Варианты:\n"
            + hint_auto +
            "  • Windows: установи Firebird 2.5 (https://firebirdsql.org/\n"
            "    en/firebird-2-5/) или скопируй fbembed.dll в\n"
            f"    {PELENG_FB25_DIR} (создастся автоматически).\n"
            "  • Linux: sudo apt install firebird2.5-classic / или\n"
            "    скачай tarball-сборку с github.com/FirebirdSQL/firebird\n"
            "    и положи libfbembed.so.2.5 в ~/.peleng_clone/fb25/\n"
            "  • Переменная окружения PELENG_FBCLIENT — абсолютный путь\n"
            "    к нужной .dll/.so/.dylib.")
    # fdb.load_api идемпотентен только в одну сторону — повторный вызов с
    # другим .so бросит ошибку. Поэтому загружаем один раз за процесс.
    if getattr(_fdb, "api", None) is None:
        _fdb.load_api(libpath)
    # DSN без хоста = embedded (или localhost-fallback на client-only сборках).
    return _fdb.connect(dsn=str(path), user=user,
                        password=password, charset=charset)


@dataclass
class FdbImportProgress:
    """Прогресс импорта FDB — передаётся в progress_cb."""
    table:    str       # имя FDB-таблицы или "BLOCKZAP"
    current:  int       # обработано строк
    total:    int       # ожидается строк
    stage:    str = ""  # человекочитаемое описание этапа


@dataclass
class FdbImportStats:
    """Сводный результат импорта."""
    by_table:        dict[str, int] = field(default_factory=dict)
    records_total:   int = 0
    blocks_total:    int = 0
    skipped_missing: int = 0  # NUMKOD без блока в BLOCKZAP


def _fdb_existing_tables(cur) -> set[str]:
    """Возвращает множество user-таблиц в подключённой FDB-БД."""
    cur.execute("SELECT rdb$relation_name FROM rdb$relations "
                "WHERE rdb$system_flag = 0")
    return {r[0].strip().upper() for r in cur.fetchall()}


def fdb_count_rows(fdb_path: str,
                   variants: Sequence[str] = ("1", "2", "3")) -> dict[str, int]:
    """Быстрый ``SELECT COUNT(*)`` по всем таблицам *Variant — для GUI-прогресса
    до начала импорта. Бросает PelengError при отсутствии библиотек."""
    conn = fdb_open(fdb_path)
    try:
        cur = conn.cursor()
        have = _fdb_existing_tables(cur)
        out: dict[str, int] = {}
        for v in variants:
            for kind, _ in FDB_KINDS:
                t = f"{kind}{v}"
                if t in have:
                    cur.execute(f"SELECT COUNT(*) FROM {t}")
                    out[t] = int(cur.fetchone()[0])
        if FDB_BLOCKZAP_TABLE in have:
            cur.execute(f"SELECT COUNT(*) FROM {FDB_BLOCKZAP_TABLE}")
            out[FDB_BLOCKZAP_TABLE] = int(cur.fetchone()[0])
        return out
    finally:
        conn.close()


def fdb_import_to_db(fdb_path: str,
                     db: BlockZapDB,
                     variants: Sequence[str] = ("1", "2", "3"),
                     progress_cb: Callable[[FdbImportProgress], None] | None = None,
                     should_cancel: Callable[[], bool] | None = None,
                     ) -> FdbImportStats:
    """Импортирует PelengPC.fdb в локальную SQLite-БД ``db``.

    Стратегия: одна транзакция на всю операцию; сначала NASTR/RESULTS/SHORTPROT
    (собирая попутно множество нужных NUMKOD), потом BLOCKZAP только для этих
    NUMKOD-ов. Если ``should_cancel()`` возвращает True — выходим аккуратно с
    коммитом уже залитого.

    Возвращает :class:`FdbImportStats`.
    """
    conn = fdb_open(fdb_path)
    stats = FdbImportStats()
    try:
        cur = conn.cursor()
        have = _fdb_existing_tables(cur)

        # 1) Метаданные: сколько строк ожидаем по каждой таблице.
        tasks: list[tuple[str, str, str, int]] = []
        for v in variants:
            for kind, _ in FDB_KINDS:
                t = f"{kind}{v}"
                if t in have:
                    cur.execute(f"SELECT COUNT(*) FROM {t}")
                    n = int(cur.fetchone()[0])
                    tasks.append((t, kind, v, n))
                    stats.by_table[t] = 0

        # 2) Льём метаданные и собираем нужные NUMKOD.
        needed_kods: set[int] = set()
        db.fdb_bulk_begin()
        try:
            for t, kind, variant, count in tasks:
                if should_cancel and should_cancel():
                    break
                if progress_cb:
                    progress_cb(FdbImportProgress(
                        table=t, current=0, total=count,
                        stage=f"Импорт {t}"))
                if count == 0:
                    continue
                cur.execute(f"SELECT * FROM {t}")
                col_names = [d[0] for d in cur.description]
                n = 0
                for row in cur:
                    if should_cancel and should_cancel():
                        break
                    rec = dict(zip(col_names, row))
                    # BlobReader-объекты в полях нам не интересны (CHAR/INT в
                    # NASTR/RESULTS/SHORTPROT), но на всякий случай:
                    for k, val in list(rec.items()):
                        if hasattr(val, "read"):
                            try:
                                rec[k] = val.read()
                            except Exception:
                                rec[k] = None
                    db.upsert_fdb_record(kind=kind, variant=variant,
                                         fields=rec)
                    # Дешифрованная зеркальная запись в общую таблицу
                    # Measurements — чтобы FDB-строки лежали рядом с
                    # принятыми по UART записями УД2-102 (дата/время/плавка/
                    # завод/дата изготовления). Сырое тело BLOCKZAP добавим
                    # отдельным проходом по NUMKOD ниже, поэтому здесь Raw
                    # пока пустой.
                    try:
                        db.upsert_measurement_from_fdb(
                            kind=kind, variant=variant, row=rec)
                    except Exception:
                        # Битая строка FDB (без даты/типа) — пропускаем без
                        # падения всего импорта; FdbRecord уже записан.
                        pass
                    nk = rec.get("NUMKOD")
                    if nk is not None:
                        needed_kods.add(int(nk))
                    n += 1
                    stats.records_total += 1
                    stats.by_table[t] = n
                    if progress_cb and (n % 200 == 0):
                        progress_cb(FdbImportProgress(
                            table=t, current=n, total=count,
                            stage=f"Импорт {t}"))
                if progress_cb:
                    progress_cb(FdbImportProgress(
                        table=t, current=n, total=count,
                        stage=f"Импорт {t}"))

            # 3) BLOCKZAP — только те NUMBER-ы, на которые ссылаются записи.
            if FDB_BLOCKZAP_TABLE in have and needed_kods and \
                    not (should_cancel and should_cancel()):
                kods = sorted(needed_kods)
                total = len(kods)
                if progress_cb:
                    progress_cb(FdbImportProgress(
                        table=FDB_BLOCKZAP_TABLE, current=0, total=total,
                        stage="Импорт BLOCKZAP (raw blob)"))
                # Чанками — IN-clause в Firebird ограничен ~1500 элементами.
                chunk = 500
                imported = 0
                for i in range(0, total, chunk):
                    if should_cancel and should_cancel():
                        break
                    sub = kods[i:i + chunk]
                    ph = ",".join("?" * len(sub))
                    cur.execute(
                        f"SELECT NUMBER, BLOCKLEN, BLOCK "
                        f"FROM {FDB_BLOCKZAP_TABLE} WHERE NUMBER IN ({ph})",
                        sub)
                    found_kods: set[int] = set()
                    for n, bl, blob in cur.fetchall():
                        if hasattr(blob, "read"):
                            try:
                                blob = blob.read()
                            except Exception:
                                blob = b""
                        if blob is None:
                            blob = b""
                        db.upsert_fdb_block(int(n),
                                            int(bl or len(blob)),
                                            bytes(blob))
                        imported += 1
                        found_kods.add(int(n))
                    stats.blocks_total = imported
                    # NUMKOD-ы, на которые ссылается метадата, но в BLOCKZAP
                    # их нет (битая FDB) — считаем как skipped.
                    stats.skipped_missing += len(set(sub) - found_kods)
                    if progress_cb:
                        progress_cb(FdbImportProgress(
                            table=FDB_BLOCKZAP_TABLE,
                            current=imported, total=total,
                            stage="Импорт BLOCKZAP (raw blob)"))

            # 4) Постобработка — обогащаем Measurements полями из тел
            #    BLOCKZAP (Side/Sheika/Naplavka/Obod/Obtochka/Greben). На
            #    SHORTPROT-строках, у которых поле PROTOCOL уже разобрано
            #    upsert_measurement_from_fdb, этот проход просто
            #    no-op (WHERE Side='' AND Sheika='' AND Naplavka='').
            if not (should_cancel and should_cancel()):
                if progress_cb:
                    progress_cb(FdbImportProgress(
                        table="Measurements", current=0, total=0,
                        stage="Дешифровка тел BLOCKZAP (side/sheika/naplavka)"))
                try:
                    n_enriched = db.enrich_measurements_from_blocks()
                except Exception:
                    n_enriched = 0
                if progress_cb:
                    progress_cb(FdbImportProgress(
                        table="Measurements",
                        current=n_enriched, total=n_enriched,
                        stage="Дешифровано полей из BLOCKZAP"))
        finally:
            db.fdb_bulk_end()
        return stats
    finally:
        conn.close()


def fdb_iter_records(fdb_path: str,
                     kinds: Sequence[str] = ("NASTR", "RESULTS", "SHORTPROT"),
                     variants: Sequence[str] = ("1", "2", "3"),
                     ) -> Iterator[tuple[str, str, str, dict, bytes | None]]:
    """Итерирует FDB напрямую (без записи в SQLite): yields
    ``(table_name, kind, variant, row_dict, blob_bytes_or_None)``.

    Используется CLI-командой ``fdb-info --sample`` и юнит-тестами.
    """
    conn = fdb_open(fdb_path)
    try:
        cur = conn.cursor()
        have = _fdb_existing_tables(cur)
        for v in variants:
            for kind in kinds:
                t = f"{kind}{v}"
                if t not in have:
                    continue
                cur.execute(f"SELECT * FROM {t}")
                col_names = [d[0] for d in cur.description]
                for row in cur:
                    rec = dict(zip(col_names, row))
                    nk = rec.get("NUMKOD")
                    blob = None
                    if nk is not None and FDB_BLOCKZAP_TABLE in have:
                        cur2 = conn.cursor()
                        cur2.execute(
                            f"SELECT BLOCK FROM {FDB_BLOCKZAP_TABLE} "
                            f"WHERE NUMBER=?", [int(nk)])
                        r = cur2.fetchone()
                        if r is not None:
                            b = r[0]
                            if hasattr(b, "read"):
                                try:
                                    b = b.read()
                                except Exception:
                                    b = b""
                            blob = bytes(b or b"")
                    yield t, kind, v, rec, blob
    finally:
        conn.close()


# =============================================================================
# 3B. ДЕКОДЕРЫ TLV-тегов УД2-102 (0x0001 A-scan, 0x0003 acoustic settings)
# =============================================================================
# Эти теги встречаются в дополнительных пакетах прибора (не в 86-байтном
# паспорте). Помещены в утилиты для будущего расширения GUI.

@dataclass
class UD2_AScanEnvelope:
    """Развёртка А-scan (TLV-тег 0x0001) — огибающая Y = |128 - byte|."""
    samples:    list[int]
    n_samples:  int
    duration_us: float
    max_depth_mm: float


def decode_ud2102_ascan(payload: bytes,
                        adc_mhz: float = 50.0,
                        v_steel_mm_us: float = 5.9,
                        offset: int = 32) -> UD2_AScanEnvelope:
    """Декодирует TLV 0x0001 (огибающая А-scan).

    Сэмплы начинаются с offset=32, Y = |128 - byte|.
    Тайминг — АЦП 50 МГц → t_мкс = i / adc_mhz; глубина = V·t/2.
    """
    samples = [abs(128 - b) for b in payload[offset:]]
    n = len(samples)
    duration_us = n / adc_mhz if adc_mhz > 0 else 0.0
    max_depth_mm = duration_us * v_steel_mm_us / 2.0
    return UD2_AScanEnvelope(samples=samples, n_samples=n,
                             duration_us=duration_us,
                             max_depth_mm=max_depth_mm)


# NB: Класс UD2_AcousticSettings и decode_ud2102_acoustic объявлены выше (см.
# раздел «1C. АКУСТИЧЕСКИЕ НАСТРОЙКИ»). Раньше здесь была вторая, более
# скудная версия — её удалили во время чистки реверса до конца.


# =============================================================================
# 4. EXCEL EXPORT — ленивый импорт openpyxl
# =============================================================================
EXCEL_COLUMNS: list[tuple[str, int]] = [
    ("№",         6), ("Дата",     12), ("Время",    10),
    ("Оператор", 22), ("Объект",   24), ("Заметки",  30),
    ("Тип",      14), ("Версия",    8),
]


def export_report(rows: list[BlockRow], path: str | Path,
                  title: str = "Протокол контроля") -> None:
    """Экспорт в .xlsx. openpyxl загружается лениво — если его нет,
    функция падает с понятным сообщением, но GUI стартует без проблем."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "Для экспорта в Excel требуется пакет openpyxl. "
            "Установите его командой:  pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Контроль"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXCEL_COLUMNS))
    head = ws.cell(row=1, column=1, value=title)
    head.font = Font(bold=True, size=14)
    head.alignment = Alignment(horizontal="center")

    fill = PatternFill("solid", fgColor="DDDDDD")
    for c, (name, width) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = Font(bold=True); cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    for r, row in enumerate(rows, start=3):
        hdr = row.header()
        report_text = ("", "", "")
        rec_type_name = "—"
        for tlv in parse_tlv(row.block):
            rec_type_name = tlv.record_type.name
            decoded = decode_record(tlv)
            if isinstance(decoded, ReportRecord):
                report_text = (decoded.operator, decoded.object_, decoded.notes)
                break

        ws.cell(row=r, column=1, value=row.number)
        ws.cell(row=r, column=2, value=hdr.date_str())
        ws.cell(row=r, column=3, value=hdr.time_str())
        ws.cell(row=r, column=4, value=report_text[0])
        ws.cell(row=r, column=5, value=report_text[1])
        ws.cell(row=r, column=6, value=report_text[2])
        ws.cell(row=r, column=7, value=rec_type_name)
        ws.cell(row=r, column=8, value=hdr.firmware_version())

    wb.save(str(path))


# =============================================================================
# 5. ЗАГЛУШКА «ПРИБОР» (для тестов без железа)
# =============================================================================
# Ниже fake-генераторы производят payloadы, строго следующие схемам
# DAL_FIELDS_RESULTS1 / NASTR1 / SHORTPROT1, чтобы реальные декодеры можно было
# проверить без железа.
def _enc_field(buf: bytearray, ftype: int, value) -> None:
    if ftype == FIELD_TYPE_INT:
        buf += int(value).to_bytes(2, "little", signed=False)
    elif ftype == FIELD_TYPE_INT_MM:
        buf += int(value).to_bytes(4, "little", signed=False)
    elif ftype == FIELD_TYPE_DATE:
        d, m, y = value
        buf.append(d & 0xFF); buf.append(m & 0xFF); buf.append((y - 2000) & 0xFF)
    elif ftype == FIELD_TYPE_TIME:
        h, mi = value
        buf.append(h & 0xFF); buf.append(mi & 0xFF)
    elif ftype in (FIELD_TYPE_STRING_S, FIELD_TYPE_STRING_L, FIELD_TYPE_STRING_X):
        s = (value or "").encode("cp1251", errors="replace")
        max_len = {FIELD_TYPE_STRING_S: 15,
                   FIELD_TYPE_STRING_L: 31,
                   FIELD_TYPE_STRING_X: 63}[ftype]
        s = s[:max_len]
        buf.append(len(s)); buf += s
    elif ftype in (FIELD_TYPE_ENUM_DEF, FIELD_TYPE_ENUM_TZ):
        buf.append(int(value) & 0xFF)
    elif ftype == FIELD_TYPE_BLOB:
        b = bytes(value)
        buf += len(b).to_bytes(2, "little")
        buf += b
    else:
        buf += int(value).to_bytes(2, "little", signed=False)


def _encode_record_with_schema(
    payload_type: int,
    schema: list[tuple[str, str, int]],
    values: dict,
    extra_blob: bytes = b"",
) -> bytearray:
    body = bytearray(); body.append(payload_type)
    for name, _, ftype in schema:
        _enc_field(body, ftype, values.get(name, 0 if ftype == FIELD_TYPE_INT else ""))
    body += extra_blob
    return body


def _fake_ascan(seed: int = 0) -> TLVRecord:
    n = 512
    # Генерируем «реалистичную» А-развёртку: импульс донного пробития
    # при ~i=200 и «дефект» при ~i=380.
    samples = bytearray()
    for i in range(n):
        x1 = (i - 200 - seed) / 30.0
        x2 = (i - 380 - seed) / 18.0
        v = int(40 + 180 * math.exp(-x1 * x1) + 100 * math.exp(-x2 * x2))
        samples.append(max(0, min(255, v)))
    now = _dt.datetime.now()
    values = dict(
        NUMBER=seed, TYPEZAP=0, NUMKOD=seed,
        DATEFORM=(now.day, now.month, now.year),
        TIMEFORM=(now.hour, now.minute),
        KODOPERA=42, NAMEOPERA="Иванов И.И.",
        NUMVERS="7.30", NUMPRIB=1234, TYPEVAR=1,
        NUMOBJ="K-103",
        M=42, MM=350, CLOCK=8,
        SMELTING="3145", MAKETIME=2024,
        DEFEKT=1, CODEDEF="D2",
        CONDLENGTH=185, NUMZAP=seed,
    )
    body = _encode_record_with_schema(
        int(RecordType.A_SCAN), DAL_FIELDS_RESULTS1, values, bytes(samples))
    return TLVRecord(tag=0x0001, length=4 + len(body), body=bytes(body))


def _fake_bscan(seed: int = 0) -> TLVRecord:
    w, h = 64, 128
    pixels = bytearray()
    for y in range(h):
        for x in range(w):
            # Наклонный «дефект» + донный импульс по низу.
            d1 = abs((x - y * 0.4) - seed)
            d2 = abs(y - 110)
            v = int(255 * (math.exp(-d1 * d1 / 25.0) +
                           0.4 * math.exp(-d2 * d2 / 12.0)))
            pixels.append(max(0, min(255, v)))
    now = _dt.datetime.now()
    values = dict(
        NUMBER=seed, TYPEZAP=1, NUMKOD=seed,
        DATEFORM=(now.day, now.month, now.year),
        TIMEFORM=(now.hour, now.minute),
        KODOPERA=42, NAMEOPERA="Иванов И.И.",
        NUMVERS="7.30", NUMPRIB=1234, TYPEVAR=1,
        NUMOBJ="K-103",
        M=42, MM=350, CLOCK=8,
        SMELTING="3145", MAKETIME=2024,
        DEFEKT=1, CODEDEF="D2",
        CONDLENGTH=185, NUMZAP=seed,
    )
    body = _encode_record_with_schema(
        int(RecordType.B_SCAN), DAL_FIELDS_RESULTS1, values)
    # Для B-scan ожидаются width/height в отдельных полях; их в RESULTS1 нет,
    # прилепляем в хвост как 2x int16-LE и байт-плотность.
    # А BScanRecord.decode обрабатывает два формата — «с заголовком» и
    # «просто пиксели» — оба должны работать.
    body += struct.pack("<HH", w, h)
    body += pixels
    return TLVRecord(tag=0x0002, length=4 + len(body), body=bytes(body))


def _fake_settings() -> TLVRecord:
    now = _dt.datetime.now()
    values = dict(
        NUMBER=1, NUMKOD=1, TYPEZAP=2,
        DATEFORM=(now.day, now.month, now.year),
        TIMEFORM=(now.hour, now.minute),
        KODOPERA=42, NAMEOPERA="Иванов И.И.",
        NUMVERS="7.30", NUMPRIB=1234, TYPEVAR=1, NUMPROT=1,
    )
    body = _encode_record_with_schema(
        int(RecordType.SETTINGS), DAL_FIELDS_NASTR1, values)
    return TLVRecord(tag=0x0003, length=4 + len(body), body=bytes(body))


def _fake_report() -> TLVRecord:
    now = _dt.datetime.now()
    notes_bytes = (
        "Дефект на 200 мм, амплитуда 75 %.\n"
        "Контроль объекта К-103, партия № 3145.\n"
    ).encode("cp1251", errors="replace") + b"\x00"
    values = dict(
        NUMBER=1, NUMKOD=1,
        DATEFORM=(now.day, now.month, now.year),
        TIMEFORM=(now.hour, now.minute),
        KODOPERA=42, NAMEOPERA="Иванов И.И.",
        NUMVERS="7.30", NUMPRIB=1234, TYPEVAR=1,
        NUMOBJ="K-103",
        M=42, MM=350, CLOCK=8,
        NUMDEF=1,
        PROTOCOL=notes_bytes,
        NUMPROT=1,
    )
    body = _encode_record_with_schema(
        int(RecordType.REPORT), DAL_FIELDS_SHORTPROT1, values)
    return TLVRecord(tag=0x0004, length=4 + len(body), body=bytes(body))


def make_fake_payload(seed: int = 0) -> bytes:
    return encode_tlv([_fake_settings(), _fake_ascan(seed),
                        _fake_bscan(seed), _fake_report()])


def make_fake_block(seed: int = 0) -> tuple[bytes, ProtocolHeader]:
    """Готовый «эфирный» блок (заголовок 16 + payload TLV) — как от прибора.

    Используется и заглушкой `fake-device`, и встроенным демо-режимом
    GUI/CLI. Никаких virtual COM-портов, никаких сторонних скриптов —
    тот же байт-в-байт буфер, который GUI получил бы по UART.
    """
    payload = make_fake_payload(seed)
    hdr_b   = make_fake_header(len(payload), seed)
    blob    = hdr_b + payload
    return blob, ProtocolHeader.parse(blob[:HEADER_LEN])


def make_fake_header(payload_len: int, seed: int = 0) -> bytes:
    now = _dt.datetime.now()
    h = ProtocolHeader(
        record_id=(seed + 1) & 0xFFFF, sub_type=0, flags=0,
        time_min=now.minute, time_h=now.hour,
        payload_len=payload_len & 0xFFFF,
        version_byte=0x07, version_byte2=0x30,
        date_d=now.day, date_m=now.month, raw=b"",
    )
    return h.to_bytes()


def fake_device_loop(port_name: str, baud: int = 19200, log=print) -> None:
    next_seed = 0
    pending_payload: bytes = b""
    pending_header:  bytes = b""
    with serial.Serial(port=port_name, baudrate=baud, timeout=0.5) as s:
        log(f"[fake] listening on {port_name} @ {baud}")
        while True:
            byte = s.read(1)
            if not byte:
                continue
            if byte == OP_HANDSHAKE:
                pending_payload = make_fake_payload(seed=next_seed)
                pending_header  = make_fake_header(len(pending_payload), seed=next_seed)
                next_seed += 1
                s.write(pending_header); s.flush()
                log(f"[fake] handshake → header (record_id={next_seed}, payload_len={len(pending_payload)})")
                continue
            if byte == OP_BLOCK_REQ:
                ll = s.read(1); hh = s.read(1)
                if not ll or not hh:
                    log("[fake] B request truncated"); continue
                total = (hh[0] << 8) | ll[0]
                blob = pending_header + pending_payload
                blob = blob + b"\xFF" * (total - len(blob)) if len(blob) < total else blob[:total]
                s.write(blob); s.flush()
                log(f"[fake] block request → {len(blob)} bytes")
                continue
            log(f"[fake] unknown byte 0x{byte[0]:02x}")


# =============================================================================
# 6. GUI — PyQt6
# =============================================================================
def _check_runtime_deps() -> tuple[bool, str]:
    """Проверяем, что PyQt6 и pyserial установлены, прежде чем запускать GUI.

    Возвращаем (ok, hint). Если не ОК — показываем дружелюбное окошко
    (или печатаем в консоль) с готовой командой `pip install …`.
    """
    missing: list[str] = []
    try:
        import PyQt6  # noqa: F401
    except ImportError:
        missing.append("PyQt6")
    try:
        import serial  # noqa: F401
    except ImportError:
        missing.append("pyserial")
    if not missing:
        return True, ""
    hint = (
        "Не установлены пакеты: " + ", ".join(missing) + ".\n\n"
        "Установите одной командой:\n"
        f"    {sys.executable} -m pip install " + " ".join(missing) + "\n"
        "(openpyxl нужен только для экспорта в Excel — можно поставить позже).")
    return False, hint


def _import_pyqt():
    """Импортируем PyQt6 только когда нужен GUI — чтобы CLI работал без него."""
    from PyQt6.QtCore import QObject, QRect, Qt, QThread, pyqtSignal
    from PyQt6.QtGui import QAction, QColor, QFont, QImage, QPainter, QPen
    from PyQt6.QtWidgets import (
        QApplication, QComboBox, QDialog, QDialogButtonBox, QFileDialog,
        QFormLayout, QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMainWindow,
        QMessageBox, QPlainTextEdit, QProgressBar, QPushButton, QScrollArea,
        QSpinBox, QSplitter, QStatusBar, QTableWidget, QTableWidgetItem,
        QTabWidget, QTextEdit, QVBoxLayout, QWidget,
    )
    return locals()


def run_gui(argv: list[str], demo: bool = False) -> int:
    qt = _import_pyqt()
    QObject = qt["QObject"]; QRect = qt["QRect"]; Qt = qt["Qt"]
    QThread = qt["QThread"]; pyqtSignal = qt["pyqtSignal"]
    QAction = qt["QAction"]; QColor = qt["QColor"]; QFont = qt["QFont"]
    QImage = qt["QImage"]; QPainter = qt["QPainter"]; QPen = qt["QPen"]
    QApplication = qt["QApplication"]; QComboBox = qt["QComboBox"]
    QDialog = qt["QDialog"]; QDialogButtonBox = qt["QDialogButtonBox"]
    QFileDialog = qt["QFileDialog"]; QFormLayout = qt["QFormLayout"]
    QHBoxLayout = qt["QHBoxLayout"]; QHeaderView = qt["QHeaderView"]
    QLabel = qt["QLabel"]; QLineEdit = qt["QLineEdit"]
    QMainWindow = qt["QMainWindow"]
    QMessageBox = qt["QMessageBox"]; QPlainTextEdit = qt["QPlainTextEdit"]
    QProgressBar = qt["QProgressBar"]; QPushButton = qt["QPushButton"]
    QScrollArea = qt["QScrollArea"]; QSpinBox = qt["QSpinBox"]
    QSplitter = qt["QSplitter"]; QStatusBar = qt["QStatusBar"]
    QTableWidget = qt["QTableWidget"]; QTableWidgetItem = qt["QTableWidgetItem"]
    QTabWidget = qt["QTabWidget"]; QTextEdit = qt["QTextEdit"]
    QVBoxLayout = qt["QVBoxLayout"]; QWidget = qt["QWidget"]

    # ----- Виджет A-scan -------------------------------------------------------
    class AScanWidget(QWidget):
        BG    = QColor(245, 245, 245); GRID  = QColor(210, 210, 210)
        AXIS  = QColor(60, 60, 60)
        TRACE = QColor(20, 80, 200);   GATE  = QColor(220, 60, 60)
        ZONE  = QColor(245, 158, 11)
        LABEL = QColor(60, 60, 60)

        def __init__(self, parent=None):
            super().__init__(parent)
            self._record: AScanRecord | None = None
            self._zone_label: str = ""   # SHORTPROT2: «Зона N из M»
            self._zone_meta:  str = ""   # доп. строка к зонe (объект, код деф.)
            self.setMinimumSize(480, 280)

        def set_record(self, rec):
            self._record = rec; self.update()

        def set_zone_label(self, label: str, meta: str = "") -> None:
            """Помечает A-scan конкретной зоной — для подсветки SHORTPROT2."""
            self._zone_label = label or ""
            self._zone_meta  = meta  or ""
            self.update()

        def paintEvent(self, ev):
            p = QPainter(self)
            p.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            # Поля под подписи осей (слева — %, справа — дБ, снизу — мкс/глубина).
            rect = self.rect().adjusted(46, 22, -50, -42)
            p.fillRect(self.rect(), self.BG)
            f = QFont("DejaVu Sans", 8); p.setFont(f); fm = p.fontMetrics()

            # Сетка в первую очередь.
            pen = QPen(self.GRID); pen.setWidth(1); p.setPen(pen)
            for i in range(1, 10):
                x = rect.left() + i * rect.width() // 10
                p.drawLine(x, rect.top(), x, rect.bottom())
            for i in range(1, 8):
                y = rect.top() + i * rect.height() // 8
                p.drawLine(rect.left(), y, rect.right(), y)

            # Оси и рамка.
            p.setPen(QPen(self.AXIS, 1)); p.drawRect(rect)

            rec = self._record
            n  = rec.n_samples if rec else 0
            us_total = rec.duration_us if rec else 0.0
            depth_mm = rec.max_depth_mm if rec else 0.0

            # Подписи оси Y — слева в %, справа в дБ (реверс DLL: 100 % ≡ 26 дБ).
            p.setPen(QPen(self.LABEL))
            for i in range(0, 9):
                y = rect.bottom() - i * rect.height() // 8
                pct = i * 100 // 8
                db  = pct / 100.0 * 26.0
                p.drawText(QRect(0, y - 8, rect.left() - 4, 16),
                           int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                           f"{pct} %")
                p.drawText(QRect(rect.right() + 4, y - 8, 46, 16),
                           int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
                           f"{db:.1f} дБ")

            # Подписи оси X — снизу мкс и мм-глубины.
            for i in range(0, 11):
                x = rect.left() + i * rect.width() // 10
                us = (us_total * i / 10.0) if us_total else 0.0
                mm = (depth_mm * i / 10.0) if depth_mm else 0.0
                p.drawText(QRect(x - 30, rect.bottom() + 2, 60, 14),
                           int(Qt.AlignmentFlag.AlignCenter),
                           f"{us:.1f}")
                p.drawText(QRect(x - 30, rect.bottom() + 16, 60, 14),
                           int(Qt.AlignmentFlag.AlignCenter),
                           f"{mm:.0f}")
            # Подписи единиц по краям.
            p.drawText(QRect(rect.left() - 36, rect.bottom() + 16, 32, 14),
                       int(Qt.AlignmentFlag.AlignRight), "мкс")
            p.drawText(QRect(rect.right() + 4, rect.bottom() + 16, 30, 14),
                       int(Qt.AlignmentFlag.AlignLeft), "мм")
            # Заголовок (вверху).
            title = "А-развёртка (амплитуда — %, время — мкс, глубина — мм)"
            if rec is not None and n:
                title += f"   →   {n} отсчётов, {us_total:.1f} мкс, диапазон глубины 0…{depth_mm:.0f} мм"
            p.drawText(QRect(rect.left(), 2, rect.width(), 18),
                       int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter), title)
            # Подсветка зоны (SHORTPROT2): не сэмпл-диапазон, а текстовый ярлык —
            # точное соответствие NUMZAP→диапазон семплов даст только полная
            # карта зон из Settings, поэтому здесь делаем честный визуальный
            # маркер, а данные смотрит пользователь во вкладке «Поля».
            if self._zone_label:
                lbl = self._zone_label
                if self._zone_meta:
                    lbl += f"   ({self._zone_meta})"
                tw = fm.horizontalAdvance(lbl) + 12
                box = QRect(rect.right() - tw - 4, rect.top() + 4, tw, 18)
                p.fillRect(box, QColor(self.ZONE.red(), self.ZONE.green(),
                                       self.ZONE.blue(), 60))
                p.setPen(QPen(self.ZONE, 1)); p.drawRect(box)
                p.setPen(QPen(self.LABEL))
                p.drawText(box, int(Qt.AlignmentFlag.AlignCenter), lbl)

            if rec is None or not rec.samples or n < 2:
                p.setPen(QPen(self.LABEL))
                p.drawText(rect, int(Qt.AlignmentFlag.AlignCenter),
                           "(нет данных — выполните «Принять» или выберите запись из БД)")
                p.end(); return

            samples = rec.samples
            # Гейт (если задан) рисуем как прозрачный прямоугольник.
            if rec.gate_end > rec.gate_start > 0:
                x0 = rect.left() + int(rec.gate_start * rect.width() / max(1, n))
                x1 = rect.left() + int(rec.gate_end   * rect.width() / max(1, n))
                x0 = max(rect.left(), min(rect.right(), x0))
                x1 = max(rect.left(), min(rect.right(), x1))
                p.fillRect(QRect(x0, rect.top(), max(1, x1 - x0), rect.height()),
                           QColor(self.GATE.red(), self.GATE.green(), self.GATE.blue(), 32))
            # Огибающая — заливка под трассой (похоже на оригинал).
            from PyQt6.QtGui import QPolygon  # локальный импорт — ленивый
            from PyQt6.QtCore import QPoint
            poly = QPolygon()
            poly.append(QPoint(rect.left(), rect.bottom()))
            for i in range(n):
                x = rect.left() + int(i * rect.width() / max(1, n - 1))
                y = rect.bottom() - int(samples[i] * rect.height() / 255)
                poly.append(QPoint(x, y))
            poly.append(QPoint(rect.right(), rect.bottom()))
            p.setBrush(QColor(self.TRACE.red(), self.TRACE.green(), self.TRACE.blue(), 60))
            p.setPen(QPen(self.TRACE, 2))
            p.drawPolygon(poly)
            p.end()

    # ----- Виджет B-scan -------------------------------------------------------
    def _amp_to_color(v: int) -> int:
        v = max(0, min(255, v))
        if v < 64:    r, g, b = 0, v * 4, 255
        elif v < 128: r, g, b = 0, 255, 255 - (v - 64) * 4
        elif v < 192: r, g, b = (v - 128) * 4, 255, 0
        else:         r, g, b = 255, 255 - (v - 192) * 4, 0
        return (0xFF << 24) | (r << 16) | (g << 8) | b

    class BScanWidget(QWidget):
        BG    = QColor(245, 245, 245); LABEL = QColor(60, 60, 60)
        AXIS  = QColor(60, 60, 60)

        def __init__(self, parent=None):
            super().__init__(parent)
            self._record: BScanRecord | None = None
            self._cached_img = None
            self.setMinimumSize(480, 320)

        def set_record(self, rec):
            self._record = rec; self._cached_img = None
            if rec is not None and rec.width > 0 and rec.height > 0 and rec.pixels:
                need = rec.width * rec.height
                data = rec.pixels[:need].ljust(need, b"\x00")
                img = QImage(rec.width, rec.height, QImage.Format.Format_ARGB32)
                for y in range(rec.height):
                    row_off = y * rec.width
                    for x in range(rec.width):
                        img.setPixel(x, y, _amp_to_color(data[row_off + x]))
                self._cached_img = img
            self.update()

        def paintEvent(self, ev):
            p = QPainter(self); p.fillRect(self.rect(), self.BG)
            f = QFont("DejaVu Sans", 8); p.setFont(f)
            # Основное поле изображения + поле под цветовую шкалу.
            cb_w = 24
            rect = self.rect().adjusted(48, 22, -(cb_w + 56), -42)
            cb_rect = QRect(rect.right() + 12, rect.top(), cb_w, rect.height())
            p.setPen(QPen(self.AXIS, 1)); p.drawRect(rect)

            rec = self._record
            length_mm = rec.length_mm if rec else 0.0
            depth_mm  = rec.depth_mm  if rec else 0.0

            # Подписи оси Y (глубина, мм, вниз = больше глубина).
            p.setPen(QPen(self.LABEL))
            for i in range(0, 9):
                y = rect.top() + i * rect.height() // 8
                mm = (depth_mm * i / 8.0) if depth_mm else 0.0
                p.drawText(QRect(0, y - 8, rect.left() - 4, 16),
                           int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                           f"{mm:.0f}")
            # Подписи оси X (координата сканирования, мм).
            for i in range(0, 11):
                x = rect.left() + i * rect.width() // 10
                mm = (length_mm * i / 10.0) if length_mm else 0.0
                p.drawText(QRect(x - 30, rect.bottom() + 2, 60, 14),
                           int(Qt.AlignmentFlag.AlignCenter),
                           f"{mm:.0f}")
            p.drawText(QRect(rect.left() - 36, rect.bottom() + 2, 32, 14),
                       int(Qt.AlignmentFlag.AlignRight), "мм")
            # Цветовая шкала (справа): амплитуда в дБ.
            for i in range(cb_rect.height()):
                v = 255 - int(255 * i / max(1, cb_rect.height() - 1))
                color = QColor((_amp_to_color(v) >> 16) & 0xFF,
                               (_amp_to_color(v) >>  8) & 0xFF,
                               (_amp_to_color(v)      ) & 0xFF)
                p.setPen(QPen(color, 1))
                p.drawLine(cb_rect.left() + 1, cb_rect.top() + i,
                           cb_rect.right() - 1, cb_rect.top() + i)
            p.setPen(QPen(self.AXIS, 1)); p.drawRect(cb_rect)
            p.setPen(QPen(self.LABEL))
            for i in range(0, 5):
                y = cb_rect.top() + i * cb_rect.height() // 4
                db = 26.0 - i * 26.0 / 4.0
                p.drawText(QRect(cb_rect.right() + 2, y - 8, 36, 16),
                           int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
                           f"{db:.1f}")
            p.drawText(QRect(cb_rect.right() + 2, cb_rect.bottom() + 2, 36, 16),
                       int(Qt.AlignmentFlag.AlignLeft), "дБ")

            # Заголовок.
            title = "B-развёртка (X — координата, мм; Y — глубина, мм; цвет — амплитуда, дБ)"
            if rec is not None:
                title += f"   →   {rec.width}×{rec.height} точек"
            p.drawText(QRect(rect.left(), 2, rect.width() + cb_w + 60, 18),
                       int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter), title)

            if self._cached_img is None:
                p.drawText(rect, int(Qt.AlignmentFlag.AlignCenter), "(B-scan: нет данных)")
                p.end(); return
            p.drawImage(rect.adjusted(1, 1, -1, -1), self._cached_img)
            p.end()

    # ----- Диалог порта --------------------------------------------------------
    class PortDialog(QDialog):
        def __init__(self, current=None, parent=None):
            super().__init__(parent)
            self.setWindowTitle("Настройка порта"); self.setModal(True)
            self._cfg = current or PortConfig()
            self.port_combo = QComboBox(self); self.port_combo.setEditable(True)
            ports = list_serial_ports(); self.port_combo.addItems(ports)
            if self._cfg.name and self._cfg.name not in ports:
                self.port_combo.addItem(self._cfg.name)
            if self._cfg.name:
                self.port_combo.setCurrentText(self._cfg.name)
            self.baud_combo = QComboBox(self)
            for b in DEFAULT_BAUDS:
                self.baud_combo.addItem(str(b), b)
            idx = list(DEFAULT_BAUDS).index(self._cfg.baud) if self._cfg.baud in DEFAULT_BAUDS else 4
            self.baud_combo.setCurrentIndex(idx)
            self.timeout_spin = QSpinBox(self)
            self.timeout_spin.setRange(500, 600000); self.timeout_spin.setSingleStep(500)
            self.timeout_spin.setSuffix(" мс"); self.timeout_spin.setValue(self._cfg.timeout_ms)
            self.idle_spin = QSpinBox(self)
            self.idle_spin.setRange(100, 60000); self.idle_spin.setSingleStep(100)
            self.idle_spin.setSuffix(" мс"); self.idle_spin.setValue(self._cfg.idle_ms)
            self.ib_spin = QSpinBox(self)
            self.ib_spin.setRange(0, 1000); self.ib_spin.setSingleStep(1)
            self.ib_spin.setSuffix(" мс"); self.ib_spin.setValue(self._cfg.inter_byte_ms)
            form = QFormLayout()
            form.addRow("Порт:", self.port_combo)
            form.addRow("Скорость:", self.baud_combo)
            form.addRow("Полный таймаут:", self.timeout_spin)
            form.addRow("Idle-таймаут:",   self.idle_spin)
            form.addRow("Пауза между байтами:", self.ib_spin)
            note = QLabel(
                "Если приходит мало байт «получено N из M» —\n"
                "увеличивайте «Полный таймаут» и «Idle-таймаут».\n"
                "На медленных приборах ставьте 30000/5000.", self)
            note.setStyleSheet("color: #888; font-size: 11px;")
            bb = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
            layout = QVBoxLayout(self)
            layout.addLayout(form); layout.addWidget(note); layout.addWidget(bb)
            self.setMinimumWidth(380)

        def config(self) -> PortConfig:
            return PortConfig(
                name=self.port_combo.currentText().strip(),
                baud=int(self.baud_combo.currentData()),
                timeout_ms=int(self.timeout_spin.value()),
                idle_ms=int(self.idle_spin.value()),
                inter_byte_ms=int(self.ib_spin.value()),
                debug=self._cfg.debug,
            )

    # ----- Диалог прогресса ---------------------------------------------------
    class ProgressDialog(QDialog):
        def __init__(self, parent=None, title="Приём данных"):
            super().__init__(parent)
            self.setWindowTitle(title); self.setModal(True)
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowType.WindowCloseButtonHint)
            self.label = QLabel("Идёт обмен с прибором…", self)
            self.bar = QProgressBar(self); self.bar.setRange(0, 100); self.bar.setValue(0)
            self.cancel_btn = QPushButton("Отмена", self)
            self.cancel_btn.clicked.connect(self.reject)
            self._cancelled = False
            self.cancel_btn.clicked.connect(self._mark_cancel)
            v = QVBoxLayout(self)
            v.addWidget(self.label); v.addWidget(self.bar); v.addWidget(self.cancel_btn)
            self.setMinimumWidth(360)

        def _mark_cancel(self): self._cancelled = True
        def is_cancelled(self) -> bool: return self._cancelled
        def set_progress(self, value, total=100, text=None):
            self.bar.setMaximum(max(1, total)); self.bar.setValue(min(value, total))
            if text: self.label.setText(text)

    # ----- Диалог настроек ----------------------------------------------------
    class SettingsDialog(QDialog):
        def __init__(self, record, parent=None):
            super().__init__(parent)
            self.setWindowTitle("Настройки серии"); self.setModal(True)
            self._record = record; self._editors: dict = {}
            inner = QWidget(self); form = QFormLayout(inner)
            form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
            # record.fields — list[DecodedField]; показываем имена+рус-подписи.
            for fld in record.fields:
                label = f"{fld.label or fld.name} ({fld.name})"
                if isinstance(fld.value, (int, float)):
                    sb = QSpinBox(self); sb.setRange(-2_000_000, 2_000_000)
                    sb.setValue(int(fld.value)); form.addRow(label, sb)
                    self._editors[fld.name] = sb
                else:
                    lab = QLabel(fld.display, self); form.addRow(label, lab)
            scroll = QScrollArea(self); scroll.setWidgetResizable(True); scroll.setWidget(inner)
            bb = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
            layout = QVBoxLayout(self); layout.addWidget(scroll); layout.addWidget(bb)
            self.resize(460, 520)

        def edited_fields(self) -> dict:
            return {name: ed.value() for name, ed in self._editors.items()}

    # ----- Поток приёма -------------------------------------------------------
    class _ReceiveWorker(QObject):
        progress = pyqtSignal(int, int, str)
        received = pyqtSignal(bytes, object)
        failed   = pyqtSignal(str)
        done     = pyqtSignal()

        def __init__(self, cfg):
            super().__init__(); self.cfg = cfg

        def run(self):
            try:
                with PelengPort(self.cfg) as port:
                    self.progress.emit(10, 100, "Хэндшейк ('U')…")
                    hdr = port.test_port()
                    self.progress.emit(30, 100,
                        f"Получен заголовок (record_id=0x{hdr.record_id:04x}, payload={hdr.payload_len})")
                    if hdr.payload_len == 0:
                        self.failed.emit("Прибор вернул payload_len=0 — нечего читать."); return
                    total = HEADER_LEN + hdr.payload_len
                    self.progress.emit(50, 100, f"Запрос блока ({total} байт)…")
                    try:
                        blob = port.request_block(total)
                    except PelengError as ex:
                        # Если у нас есть хоть какие-то частичные данные — попробуем
                        # дешифровать что есть, чтобы пользователь увидел хотя бы часть.
                        partial = getattr(ex, "partial", b"")
                        if partial and len(partial) >= HEADER_LEN:
                            self.progress.emit(80, 100,
                                f"Тайм-аут, но удалось принять {len(partial)} из {total} байт — "
                                f"показываю частичные данные.")
                            self.received.emit(bytes(partial), hdr)
                            self.failed.emit(str(ex))
                            return
                        raise
                    self.progress.emit(95, 100, "Готово.")
                    self.received.emit(blob, hdr)
                    self.progress.emit(100, 100, "Готово.")
            except PelengError as e:
                self.failed.emit(str(e))
            except Exception as e:
                self.failed.emit(f"{type(e).__name__}: {e}")
            finally:
                self.done.emit()

    # ----- Поток импорта FDB --------------------------------------------------
    # Зеркалит fdb_import_to_db() в отдельном QThread'е, чтобы GUI не вис.
    # Сигналы: progress(curr,total,text), done(stats_dict), failed(msg).
    class _FdbImportWorker(QObject):
        progress = pyqtSignal(int, int, str)
        done     = pyqtSignal(object)   # FdbImportStats
        failed   = pyqtSignal(str)
        finished = pyqtSignal()

        def __init__(self, fdb_path: str, db: BlockZapDB,
                     variants: Sequence[str]):
            super().__init__()
            self.fdb_path = fdb_path
            self.db = db
            self.variants = tuple(variants)
            self._cancel = False

        def cancel(self) -> None:
            self._cancel = True

        def _should_cancel(self) -> bool:
            return self._cancel

        def run(self):
            try:
                def _cb(p: FdbImportProgress) -> None:
                    text = (f"{p.stage}: {p.current}/{p.total}"
                            if p.total else p.stage)
                    self.progress.emit(p.current, max(1, p.total), text)
                stats = fdb_import_to_db(
                    self.fdb_path, self.db,
                    variants=self.variants,
                    progress_cb=_cb,
                    should_cancel=self._should_cancel,
                )
                self.done.emit(stats)
            except PelengError as e:
                self.failed.emit(str(e))
            except Exception as e:
                self.failed.emit(f"{type(e).__name__}: {e}")
            finally:
                self.finished.emit()

    def _hexdump(buf: bytes, max_lines: int = 64) -> str:
        out = []
        for i in range(0, len(buf), 16):
            if i // 16 >= max_lines:
                out.append(f"... ещё {len(buf) - i} байт ..."); break
            chunk = buf[i:i + 16]
            hexs = " ".join(f"{b:02x}" for b in chunk)
            ascii_ = "".join(chr(b) if 32 <= b < 127 else "." for b in chunk)
            out.append(f"{i:08x}  {hexs:<48}  {ascii_}")
        return "\n".join(out)

    # ----- Главное окно --------------------------------------------------------
    class MainWindow(QMainWindow):
        DEFAULT_DB = str(Path.home() / ".peleng_clone" / "blockzap.sqlite3")

        def __init__(self):
            super().__init__()
            self._base_title = f"Peleng-Clone v{__version__} — дефектоскоп «Пеленг»"
            self.setWindowTitle(self._base_title)
            self.resize(1180, 760)
            self._cfg = PortConfig()
            self._ud2_cfg = UD2_102PortConfig()  # В3«»ДУД2-102 — отдельные тайминги
            self._device_no: int | None = None  # заводской № прибора (после 4× 0x55)
            self._db_path = self.DEFAULT_DB
            Path(self._db_path).parent.mkdir(parents=True, exist_ok=True)
            self._db = BlockZapDB(self._db_path)
            self._build_menu(); self._build_central(); self._build_status()
            self._refresh_table(); self._refresh_measurements()
            self._refresh_fdb_table()
            self._thread = None; self._worker = None; self._progress = None

        def _set_device_title(self):
            """Обновляет заголовок окна с учётом заводского № прибора."""
            if self._device_no is None:
                self.setWindowTitle(self._base_title)
            else:
                self.setWindowTitle(f"{self._base_title}  —  УД2-102 № {self._device_no}")

        def _build_menu(self):
            bar = self.menuBar()
            m_dev = bar.addMenu("&Прибор")
            a_recv = QAction("&Принять (PelengPC)", self); a_recv.setShortcut("F5"); a_recv.triggered.connect(self._on_receive)
            a_test = QAction("&Тест порта (PelengPC)", self); a_test.setShortcut("F6"); a_test.triggered.connect(self._on_test_port)
            a_port = QAction("&Настройка порта…", self); a_port.triggered.connect(self._on_port_settings)
            # ——— Режим УД2-102 ———
            a_ud2_hs   = QAction("УД2-102: хэндшейк (4× 0x55)", self)
            a_ud2_hs.setShortcut("Ctrl+H"); a_ud2_hs.triggered.connect(self._on_ud2_handshake)
            a_ud2_sc   = QAction("УД2-102: Сканировать память", self)
            a_ud2_sc.setShortcut("F7"); a_ud2_sc.triggered.connect(self._on_ud2_scan)
            a_ud2_dump = QAction("УД2-102: Снимок экрана (0x49)", self)
            a_ud2_dump.triggered.connect(self._on_ud2_screen_dump)
            a_ud2_rtc  = QAction("УД2-102: Синхронизация часов (0x54)", self)
            a_ud2_rtc.triggered.connect(self._on_ud2_rtc_sync)
            a_demo = QAction("&Демо-блок (без прибора)", self)
            a_demo.setShortcut("F8"); a_demo.triggered.connect(self._on_demo_block)
            a_quit = QAction("&Выход", self); a_quit.setShortcut("Ctrl+Q"); a_quit.triggered.connect(self.close)
            m_dev.addActions([a_recv, a_test, a_port])
            m_dev.addSeparator()
            m_dev.addActions([a_ud2_hs, a_ud2_sc, a_ud2_dump, a_ud2_rtc])
            m_dev.addSeparator(); m_dev.addAction(a_demo)
            m_dev.addSeparator(); m_dev.addAction(a_quit)
            m_db = bar.addMenu("&База")
            a_open = QAction("&Открыть БД…", self); a_open.triggered.connect(self._on_open_db)
            a_saveas = QAction("Сохранить как…", self); a_saveas.triggered.connect(self._on_save_as)
            a_del = QAction("&Удалить запись", self); a_del.setShortcut("Del"); a_del.triggered.connect(self._on_delete_row)
            a_fdb = QAction("&Импортировать FDB…", self)
            a_fdb.setShortcut("Ctrl+I")
            a_fdb.setStatusTip("Импорт PelengPC.fdb (Firebird 2.x) — "
                               "настройки, протоколы, отчёты")
            a_fdb.triggered.connect(self._on_import_fdb)
            a_fdb_clear = QAction("Очистить импорт FDB", self)
            a_fdb_clear.triggered.connect(self._on_fdb_clear)
            m_db.addActions([a_open, a_saveas, a_del])
            m_db.addSeparator()
            m_db.addActions([a_fdb, a_fdb_clear])
            m_rep = bar.addMenu("&Отчёт")
            a_xls = QAction("Экспорт в &Excel…", self); a_xls.triggered.connect(self._on_export_excel)
            m_rep.addAction(a_xls)
            m_set = bar.addMenu("&Настройки")
            a_edit = QAction("Редактировать выбранные…", self); a_edit.triggered.connect(self._on_edit_settings)
            m_set.addAction(a_edit)
            m_help = bar.addMenu("&Помощь")
            a_about = QAction("&О программе…", self); a_about.triggered.connect(self._on_about)
            m_help.addAction(a_about)

        # ------- A9: универсальный поиск/фильтр по любой QTableWidget ---------
        # Один QLineEdit над таблицей: пишем подстроку → скрываем строки, где
        # эта подстрока не найдена ни в одной видимой ячейке. Сортировка
        # уже включена через setSortingEnabled(True) при создании таблицы.
        def _make_filter_panel(self, table, placeholder: str = "") -> QWidget:
            box = QWidget(self)
            edit = QLineEdit(box)
            edit.setPlaceholderText(placeholder or "Поиск (подстрока в любой колонке)…")
            edit.setClearButtonEnabled(True)
            def _apply_filter(_text: str = "") -> None:
                q = edit.text().strip().lower()
                if not q:
                    for r in range(table.rowCount()):
                        table.setRowHidden(r, False)
                    return
                for r in range(table.rowCount()):
                    shown = False
                    for c in range(table.columnCount()):
                        it = table.item(r, c)
                        if it and q in it.text().lower():
                            shown = True; break
                    table.setRowHidden(r, not shown)
            edit.textChanged.connect(_apply_filter)
            v = QVBoxLayout(box)
            v.setContentsMargins(2, 2, 2, 2); v.setSpacing(2)
            v.addWidget(edit); v.addWidget(table)
            table._filter_edit = edit            # ссылку оставляем для тестов
            table._filter_apply = _apply_filter
            return box

        def _build_central(self):
            self.table = QTableWidget(self)
            self.table.setColumnCount(5)
            self.table.setHorizontalHeaderLabels(["№", "Дата", "Время", "Версия", "Размер"])
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            self.table.verticalHeader().setVisible(False)
            self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
            self.table.setSortingEnabled(True)
            self.table.itemSelectionChanged.connect(self._on_selection_changed)

            btn_recv = QPushButton("Принять (F5)", self); btn_recv.clicked.connect(self._on_receive)
            btn_test = QPushButton("Тест порта", self); btn_test.clicked.connect(self._on_test_port)
            btn_port = QPushButton("Порт…", self); btn_port.clicked.connect(self._on_port_settings)
            btn_xls  = QPushButton("В Excel…", self); btn_xls.clicked.connect(self._on_export_excel)
            # Импорт FDB (вагонная версия — Firebird PelengPC.fdb).
            btn_fdb  = QPushButton("Импорт FDB…", self)
            btn_fdb.setToolTip("Импорт PelengPC.fdb (Firebird 2.x): "
                               "настройки, протоколы, отчёты + BLOB-кадры")
            btn_fdb.clicked.connect(self._on_import_fdb)
            self._btn_fdb = btn_fdb
            # A6: печать / PDF текущей вкладки.
            btn_prn  = QPushButton("Печать…", self)
            btn_prn.setToolTip("Печать / экспорт текущей вкладки (A-/B-scan, "
                               "Поля, Отчёт, Таблица) на принтер или в PDF")
            btn_prn.clicked.connect(self._on_print_current_tab)
            btn_pdf  = QPushButton("PDF…", self)
            btn_pdf.setToolTip("Экспорт текущей вкладки в PDF")
            btn_pdf.clicked.connect(self._on_export_pdf_current_tab)
            # A5: FR3-выгрузка выбранной записи (FastReport-style INI-blob).
            btn_fr3  = QPushButton("FR3…", self)
            btn_fr3.setToolTip("Экспорт выбранной записи (BlockZap или FDB) "
                               "в FR3-шаблон: метаданные + TLV-блоб в одном "
                               "INI-файле (как делает оригинальный PelengPC).")
            btn_fr3.clicked.connect(self._on_export_fr3)
            bar = QHBoxLayout()
            for b in (btn_recv, btn_test, btn_port, btn_xls, btn_fdb,
                      btn_prn, btn_pdf, btn_fr3):
                bar.addWidget(b)
            bar.addStretch(1)

            left_w = QWidget(self); left_v = QVBoxLayout(left_w)
            left_v.setContentsMargins(4, 4, 4, 4)
            left_v.addLayout(bar)
            # A9: поиск над BlockZap-таблицей.
            left_v.addWidget(self._make_filter_panel(self.table,
                "Поиск BlockZap: №, дата, время, версия, размер…"))

            self.tabs = QTabWidget(self)
            self.ascan = AScanWidget(self); self.bscan = BScanWidget(self)
            self.fields_view = QTableWidget(self)
            self.fields_view.setColumnCount(4)
            self.fields_view.setHorizontalHeaderLabels(["Поле", "Подпись", "Источник", "Значение"])
            self.fields_view.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.Interactive)
            self.fields_view.horizontalHeader().setStretchLastSection(True)
            self.fields_view.verticalHeader().setVisible(False)
            self.fields_view.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            self.report_view = QTextEdit(self); self.report_view.setReadOnly(True)
            self.raw_view = QPlainTextEdit(self); self.raw_view.setReadOnly(True)
            self.raw_view.setStyleSheet("font-family: monospace;")

            # Новая вкладка «Записи (УД2-102)» — для BCD-паспортов
            # рельсового дефектоскопа. Заполняется по «Прибор → Сканировать
            # память» (F7) или «Демо-блок» в режиме ——demo. Внутри ─ три
            # под-вкладки с автосортировкой по Source:
            #   • «Настройки»             — Source = 'FDB-NASTR'
            #   • «Протоколы (А-развёртка)» — Source = 'FDB-RESULTS'
            #   • «Отчёты»                — Source IN ('FDB-SHORTPROT', 'UD2')
            # Двойной клик по строке-протоколу → переход на вкладку
            # «А-развёртка» с подгруженным TLV-блоком; по строке-отчёту →
            # переход на «Отчёт» с автозаполнением.
            ud2_cols = ["Адрес", "Дата", "Время", "Тип", "№ объекта",
                        "Плавка", "Клеймо", "Год", "Сторона",
                        "Шейка", "Обод", "Обточка", "Гребень", "Наплавка",
                        "Завод (инд.)", "Дата изг.", "Дефект", "Код деф.",
                        "Оператор", "Источник"]

            def _mk_ud2_subtable(parent):
                w = QTableWidget(parent)
                w.setColumnCount(len(ud2_cols))
                w.setHorizontalHeaderLabels(ud2_cols)
                w.horizontalHeader().setSectionResizeMode(
                    QHeaderView.ResizeMode.ResizeToContents)
                w.verticalHeader().setVisible(False)
                w.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
                w.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                w.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
                w.setSortingEnabled(True)
                return w

            self.ud2_table_settings  = _mk_ud2_subtable(self)
            self.ud2_table_protocols = _mk_ud2_subtable(self)
            self.ud2_table_reports   = _mk_ud2_subtable(self)
            # Backward-compat: для существующих веток кода/тестов, которые ещё
            # обращаются к self.ud2_table, оставляем «алиас» на под-таблицу
            # «Отчёты» — туда же попадают и UART-записи УД2-102.
            self.ud2_table = self.ud2_table_reports
            # Двойной клик → A-развёртка (протоколы) / «Отчёт» (отчёты+UD2).
            self.ud2_table_protocols.cellDoubleClicked.connect(
                lambda r, c: self._on_ud2_open_protocol(r))
            self.ud2_table_reports.cellDoubleClicked.connect(
                lambda r, c: self._on_ud2_open_report(r))
            self.ud2_table_settings.cellDoubleClicked.connect(
                lambda r, c: self._on_ud2_open_protocol(r, settings=True))
            # Одинарный клик (выделение строки) по протоколу → рисуем A-scan
            # в правой панели без переключения вкладки (как у fdb_table).
            self.ud2_table_protocols.itemSelectionChanged.connect(
                self._on_ud2_protocol_selection_changed)

            # Внутренний QTabWidget с тремя под-вкладками.
            self.ud2_subtabs = QTabWidget(self)
            self.ud2_subtabs.addTab(self._make_filter_panel(self.ud2_table_settings,
                "Поиск настроек: дата/тип/оператор…"),
                "Настройки")
            self.ud2_subtabs.addTab(self._make_filter_panel(self.ud2_table_protocols,
                "Поиск протоколов: дата/тип/№объекта/плавка…"),
                "Протоколы (А-развёртка)")
            self.ud2_subtabs.addTab(self._make_filter_panel(self.ud2_table_reports,
                "Поиск отчётов: дата/тип/№объекта/плавка/клеймо…"),
                "Отчёты")

            # Вкладка «Импорт FDB» — для содержимого PelengPC.fdb (вагонная
            # версия). Колонки расширены: BLOB (NUMKOD → FdbBlock.Number) и
            # Зона (NUMZAP) — это те два ключа, которые связывают SHORTPROT2/
            # RESULTS2 с физическим TLV-кадром в BLOCKZAP. Клик по строке
            # → _on_fdb_selection_changed → рисует A/B-развёртку в существующих
            # вкладках (A1/A2 из «Недокрученное»).
            self.fdb_table = QTableWidget(self)
            fdb_cols = ["Вар.", "Тип", "№", "BLOB", "Зона", "TypeZap", "Дата",
                        "Время", "Оператор", "№ прибора", "TypeVar",
                        "Объект", "Плавка", "Клеймо",
                        "Дата изг.", "Дефект", "Код", "Протокол"]
            self.fdb_table.setColumnCount(len(fdb_cols))
            self.fdb_table.setHorizontalHeaderLabels(fdb_cols)
            self.fdb_table.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.ResizeToContents)
            self.fdb_table.verticalHeader().setVisible(False)
            self.fdb_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            self.fdb_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            self.fdb_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
            self.fdb_table.setSortingEnabled(True)
            self.fdb_table.itemSelectionChanged.connect(
                self._on_fdb_selection_changed)

            self.tabs.addTab(self.ascan,       "А-развёртка")
            self.tabs.addTab(self.bscan,       "B-развёртка")
            # Поля: тоже c фильтром (A9) — длинные таблицы декода удобнее.
            self.tabs.addTab(self._make_filter_panel(self.fields_view,
                "Поиск по полям: имя/подпись/источник/значение…"),
                "Поля")
            self.tabs.addTab(self.report_view, "Отчёт")
            self.tabs.addTab(self.raw_view,    "Сырой буфер")
            # A9: внутри ud2_subtabs у каждой под-вкладки уже свой фильтр.
            self.tabs.addTab(self.ud2_subtabs, "Записи (УД2-102)")
            self.tabs.addTab(self._make_filter_panel(self.fdb_table,
                "Поиск FDB: тип/№/оператор/объект/код/дата…"),
                "Импорт FDB")

            splitter = QSplitter(Qt.Orientation.Horizontal, self)
            splitter.addWidget(left_w); splitter.addWidget(self.tabs)
            splitter.setStretchFactor(0, 0); splitter.setStretchFactor(1, 1)
            splitter.setSizes([420, 680])
            self.setCentralWidget(splitter)

        def _build_status(self):
            self.status = QStatusBar(self); self.setStatusBar(self.status)
            self._status_db = QLabel(""); self.status.addWidget(self._status_db)
            self._status_port = QLabel(""); self.status.addPermanentWidget(self._status_port)
            self._update_status_widgets()

        def _update_status_widgets(self):
            ud2_n = self._db.count_measurements()
            self._status_db.setText(
                f"  БД: {self._db_path}  ({self._db.count()} блоков, "
                f"{ud2_n} Записей УД2)  ")
            ud2_extra = f" | УД2-102: № {self._device_no}" if self._device_no is not None else ""
            self._status_port.setText(
                f"Порт: {self._cfg.name} @ {self._cfg.baud}{ud2_extra}  ")

        def _on_port_settings(self):
            dlg = PortDialog(self._cfg, self)
            if dlg.exec():
                self._cfg = dlg.config(); self._update_status_widgets()
                self.status.showMessage(f"Параметры порта обновлены: {self._cfg.name} @ {self._cfg.baud}", 3000)

        def _on_test_port(self):
            try:
                with PelengPort(self._cfg) as port:
                    hdr = port.test_port()
            except Exception as e:
                QMessageBox.critical(self, "Тест порта", str(e)); return
            QMessageBox.information(self, "Тест порта",
                f"Прибор отвечает.\n\nrecord_id   = 0x{hdr.record_id:04x}\n"
                f"firmware    = {hdr.firmware_version()}\n"
                f"date        = {hdr.date_str()}\n"
                f"time        = {hdr.time_str()}\n"
                f"payload_len = {hdr.payload_len}")

        def _on_demo_block(self):
            """Сгенерировать демо-данные прямо в процессе и положить в БД.

            Сразу оба режима одной кнопкой:
              - PelengPC v1.2: заголовок + TLV (A-/B-/Settings/Report);
              - УД2-102: 4× хэндшейк + несколько BCD-паспортов.
            Не требует ни UART-кабеля, ни socat/com0com, ни fake-device-процесса.
            """
            # 1) Демо-блок PelengPC v1.2 — для старых вкладок (A/B/Поля/Отчёт).
            seed = self._db.count() + 1
            try:
                blob, hdr = make_fake_block(seed=seed)
                while self._db.has(hdr.record_id):
                    seed += 1
                    blob, hdr = make_fake_block(seed=seed)
                self._db.upsert(hdr.record_id, blob)
            except Exception as e:
                QMessageBox.critical(self, "Демо-блок",
                    f"Не удалось сгенерировать PelengPC-блок: {e}")
                return

            # 2) Демо-записи УД2-102 — факеовый хэндшейк + несколько баз
            #    из BCD-паспортов, чтобы вкладка «Записи (УД2-102)» была видна.
            try:
                fake_hs = UD2_102Header.parse(make_fake_ud2102_handshake(
                    device_no=4400 + (seed % 200)))
                self._device_no = fake_hs.device_no
                count_ud2 = 0
                for base in (UD2_BASES[0], UD2_BASES[1]):
                    for idx in range(8):
                        addr, raw86 = make_fake_ud2102_record(
                            seed=seed * 10 + idx, addr=base + idx)
                        rec = parse_ud2102_record(addr, raw86[2:])
                        if rec is not None:
                            self._db.upsert_measurement(self._device_no, rec)
                            count_ud2 += 1
            except Exception as e:
                QMessageBox.warning(self, "Демо-блок (УД2-102)",
                    f"Не удалось сгенерировать записи УД2-102: {e}")
                count_ud2 = 0

            self._set_device_title()
            self._refresh_table(select_number=hdr.record_id)
            self._refresh_measurements()
            self.status.showMessage(
                f"Демо: PelengPC-блок №{hdr.record_id} ({len(blob)} байт); "
                f"УД2-102 — {count_ud2} записей (прибор № {self._device_no})", 6000)

        # ——— Обработчики режима УД2-102 ———
        def _on_ud2_handshake(self):
            self._ud2_cfg.name = self._cfg.name
            try:
                with UD2_102Port(self._ud2_cfg) as port:
                    hdr = port.handshake()
            except Exception as e:
                QMessageBox.critical(self, "УД2-102: хэндшейк", str(e))
                return
            self._device_no = hdr.device_no
            self._set_device_title(); self._update_status_widgets()
            QMessageBox.information(self, "УД2-102: хэндшейк",
                f"Прибор отвечает.\n\n"
                f"Заводской № = {hdr.device_no}\n"
                f"raw[16] = {hdr.raw.hex(' ')}")

        def _on_ud2_scan(self):
            self._ud2_cfg.name = self._cfg.name
            prog = ProgressDialog(self, "УД2-102: сканирование памяти")
            prog.set_progress(0, len(UD2_BASES) * UD2_INDICES_PER_BASE,
                              "Хэндшейк (4× 0x55)…")
            prog.show()
            QApplication.processEvents()
            cnt = 0
            try:
                with UD2_102Port(self._ud2_cfg) as port:
                    hdr = port.handshake()
                    self._device_no = hdr.device_no
                    self._set_device_title()
                    QApplication.processEvents()
                    total = len(UD2_BASES) * UD2_INDICES_PER_BASE
                    seen = 0
                    def _progress_cb(base: int, idx: int, hit: bool) -> None:
                        nonlocal seen
                        seen += 1
                        prog.set_progress(seen, total,
                            f"База {base}, idx {idx}, {'хит' if hit else 'pass'}…")
                        QApplication.processEvents()
                    for rec in port.scan_memory(progress=_progress_cb):
                        if prog.is_cancelled():
                            break
                        self._db.upsert_measurement(hdr.device_no, rec); cnt += 1
            except Exception as e:
                prog.accept()
                QMessageBox.critical(self, "УД2-102: сканирование", str(e))
                return
            prog.accept()
            self._update_status_widgets(); self._refresh_measurements()
            QMessageBox.information(self, "УД2-102: сканирование",
                f"Найдено {cnt} записей.")

        def _on_ud2_screen_dump(self):
            self._ud2_cfg.name = self._cfg.name
            try:
                with UD2_102Port(self._ud2_cfg) as port:
                    if self._device_no is None:
                        port.handshake()
                    data = port.screen_dump()
            except Exception as e:
                QMessageBox.critical(self, "УД2-102: снимок экрана", str(e))
                return
            if not data:
                QMessageBox.information(self, "УД2-102: снимок экрана",
                    "Прибор не вернул ни одного байта на 0x49.")
                return
            fn, _ = QFileDialog.getSaveFileName(self, "Сохранить дамп",
                str(Path.home() / f"ud2_screen_{int(time.time())}.bin"),
                "Binary (*.bin);;All files (*)")
            if not fn: return
            Path(fn).write_bytes(data)
            QMessageBox.information(self, "УД2-102",
                f"Сохранено {len(data)} байт в {fn}")

        def _on_ud2_rtc_sync(self):
            self._ud2_cfg.name = self._cfg.name
            ans = QMessageBox.question(self, "УД2-102: синхронизация часов",
                "Отправить на прибор текущее время ПК (команда 0x54)?")
            if ans != QMessageBox.StandardButton.Yes:
                return
            try:
                with UD2_102Port(self._ud2_cfg) as port:
                    if self._device_no is None:
                        port.handshake()
                    port.rtc_sync()
            except Exception as e:
                QMessageBox.critical(self, "УД2-102: синхронизация часов", str(e))
                return
            QMessageBox.information(self, "УД2-102", "Часы прибора обновлены.")

        def _refresh_measurements(self):
            rows = self._db.all_measurements()
            DASH = "—"
            # Авто-сортировка по Source: каждая строка ровно в одной под-вкладке.
            #   FDB-NASTR        → «Настройки»
            #   FDB-RESULTS      → «Протоколы (А-развёртка)»
            #   FDB-SHORTPROT    → «Отчёты»
            #   UD2 (UART)       → «Отчёты»  (BCD-паспорт = отчёт по контролю)
            buckets: dict[str, list[dict]] = {
                "settings": [],
                "protocols": [],
                "reports": [],
            }
            for r in rows:
                src = str(r.get("source") or "UD2")
                if   src == "FDB-NASTR":      buckets["settings"].append(r)
                elif src == "FDB-RESULTS":    buckets["protocols"].append(r)
                else:                          buckets["reports"].append(r)

            tables = {
                "settings":  self.ud2_table_settings,
                "protocols": self.ud2_table_protocols,
                "reports":   self.ud2_table_reports,
            }
            for key, tbl in tables.items():
                rs = buckets[key]
                # При sortingEnabled=True сначала кладём ячейки в строки, а
                # включаем сортировку обратно после заполнения — иначе строки
                # «перемешаются» прямо во время setItem().
                tbl.setSortingEnabled(False)
                tbl.setRowCount(len(rs))
                for i, r in enumerate(rs):
                    addr   = int(r["addr"])
                    source = str(r.get("source") or "UD2")
                    if source.startswith("FDB-") and addr >= BlockZapDB.FDB_ADDR_OFFSET:
                        num = addr & 0x0FFF_FFFF
                        kind_short = source[4:].title()
                        addr_str = f"{kind_short}#{num}"
                    else:
                        addr_str = f"0x{addr:04x}"
                    time_str = r.get("time_str") or ""
                    make_year = _fdb_make_year(r.get("make_time") or 0)
                    cells = [
                        addr_str,
                        _strip_unprintable(str(r["date"] or "")) or DASH,
                        time_str or DASH,
                        r["type_name"] or DASH,
                        r["num_obj"] or DASH,
                        r["plavka"]  or DASH,
                        str(r["stamp"]) if r["stamp"] else DASH,
                        str(r["god"])   if r["god"]   else DASH,
                        r["side"]     or DASH,
                        r["sheika"]   or DASH,
                        r["obod"]     or DASH,
                        r["obtochka"] or DASH,
                        r["greben"]   or DASH,
                        r["naplavka"] or DASH,
                        r.get("ind_maker") or DASH,
                        (str(make_year) if make_year else
                         (str(r.get("make_time")) if r.get("make_time") else DASH)),
                        r.get("defekt")   or DASH,
                        r.get("code_def") or DASH,
                        r.get("operator") or DASH,
                        source,
                    ]
                    for j, cell in enumerate(cells):
                        item = QTableWidgetItem(str(cell))
                        # Прячем в первую ячейку всю «строку Measurements»
                        # (адрес + NumKod + Source) — это нужно
                        # _on_ud2_open_protocol/_report, чтобы поднять
                        # BLOCKZAP-блок и нарисовать график/отчёт.
                        if j == 0:
                            item.setData(Qt.ItemDataRole.UserRole, r)
                        tbl.setItem(i, j, item)
                tbl.setSortingEnabled(True)
                tbl.resizeColumnsToContents()

        def _on_receive(self):
            if self._thread is not None: return
            self._progress = ProgressDialog(self, "Приём данных")
            self._thread = QThread(self); self._worker = _ReceiveWorker(self._cfg)
            self._worker.moveToThread(self._thread)
            self._thread.started.connect(self._worker.run)
            self._worker.progress.connect(self._on_recv_progress)
            self._worker.received.connect(self._on_recv_received)
            self._worker.failed.connect(self._on_recv_failed)
            self._worker.done.connect(self._thread.quit)
            self._worker.done.connect(self._on_recv_done)
            self._thread.start(); self._progress.exec()

        def _on_recv_progress(self, value, total, text):
            if self._progress: self._progress.set_progress(value, total, text)

        def _on_recv_received(self, blob, hdr):
            try:
                self._db.upsert(hdr.record_id, blob)
            except Exception as e:
                QMessageBox.warning(self, "БД", f"Не удалось записать блок: {e}")
            self._refresh_table(select_number=hdr.record_id)

        def _on_recv_failed(self, msg):
            QMessageBox.critical(self, "Ошибка приёма", msg)

        def _on_recv_done(self):
            if self._progress: self._progress.accept(); self._progress = None
            self._thread = None; self._worker = None; self._update_status_widgets()

        # ----- Импорт PelengPC.fdb (вагонная версия) -------------------------
        def _on_import_fdb(self):
            """Открывает диалог выбора .fdb-файла и запускает фоновый импорт."""
            if self._thread is not None:
                QMessageBox.information(self, "Импорт FDB",
                    "Подождите, идёт другая фоновая операция."); return
            fn, _ = QFileDialog.getOpenFileName(self, "Выбрать PelengPC.fdb",
                str(Path.home()),
                "Firebird database (*.fdb *.FDB);;All files (*)")
            if not fn:
                return
            # Пробный коннект — чтобы ДО фонового потока корректно показать
            # пользователю «нет fdb / нет libfbembed.dll», а не вешать прогресс.
            # Если libfbembed не найден И мы на Windows — предложим скачать
            # официальный Firebird-2.5.9-embed (~12 МБ) прямо отсюда.
            if _fdb_find_client_lib() is None and _fbembed_platform_key():
                ans = QMessageBox.question(self, "Firebird 2.5 embedded",
                    "Для чтения PelengPC.fdb (Firebird 2.x, ODS 11.0) нужна\n"
                    "библиотека fbembed.dll. Я не нашёл её в системе.\n\n"
                    "Скачать Firebird-2.5.9 embed (~12 МБ) с официального\n"
                    "GitHub-релиза и положить в\n"
                    f"  {PELENG_FB25_DIR}?")
                if ans == QMessageBox.StandardButton.Yes:
                    dlg = ProgressDialog(self, "Скачивание fbembed")
                    dlg.set_progress(0, 100, "Подключение к GitHub…")
                    err: list[str] = []
                    def _cb(done: int, total: int) -> None:
                        if total > 0:
                            dlg.set_progress(done, total,
                                f"{done/1024/1024:.1f} / {total/1024/1024:.1f} MB")
                    # Качаем синхронно (12 МБ — терпимо). Если упадёт —
                    # сообщим в err и продолжим с обычной ошибкой fdb_open.
                    try:
                        QApplication.processEvents()
                        fdb_install_fbembed(progress_cb=_cb)
                    except PelengError as e:
                        err.append(str(e))
                    finally:
                        dlg.accept()
                    if err:
                        QMessageBox.warning(self, "Firebird 2.5 embedded",
                            "Не удалось скачать fbembed автоматически:\n\n"
                            + err[0])
            try:
                _conn_probe = fdb_open(fn)
                _conn_probe.close()
            except PelengError as e:
                QMessageBox.critical(self, "Импорт FDB", str(e)); return
            except Exception as e:
                QMessageBox.critical(self, "Импорт FDB",
                    f"Не удалось открыть FDB:\n{e}"); return

            ans = QMessageBox.question(self, "Импорт FDB",
                f"Импортировать данные из:\n{fn}\n\n"
                "Будут добавлены/обновлены записи NASTR/RESULTS/SHORTPROT\n"
                "и связанные с ними BLOB-кадры BLOCKZAP.\n\n"
                "Существующие записи устройства (BlockZap) НЕ пострадают.\n"
                "Продолжить?")
            if ans != QMessageBox.StandardButton.Yes:
                return

            self._progress = ProgressDialog(self, "Импорт FDB")
            self._progress.set_progress(0, 100, "Подключение к FDB…")
            self._thread = QThread(self)
            # Вагонная версия живёт в *2-таблицах; рельсовая — в *1; ОТД — в *3.
            # Грузим все три варианта по умолчанию: лишнего не появится, потому
            # что несуществующие таблицы пропускаются.
            self._worker = _FdbImportWorker(fn, self._db,
                                            variants=("1", "2", "3"))
            self._worker.moveToThread(self._thread)
            self._thread.started.connect(self._worker.run)
            self._worker.progress.connect(self._on_fdb_progress)
            self._worker.done.connect(self._on_fdb_done)
            self._worker.failed.connect(self._on_fdb_failed)
            self._worker.finished.connect(self._thread.quit)
            self._worker.finished.connect(self._on_fdb_finished)
            # Cancel-кнопка прогресс-диалога → останавливаем импорт корректно.
            self._progress.cancel_btn.clicked.connect(self._worker.cancel)
            self._thread.start()
            self._progress.exec()

        def _on_fdb_progress(self, value: int, total: int, text: str):
            if self._progress is not None:
                self._progress.set_progress(value, total, text)

        def _on_fdb_done(self, stats):
            """Импорт успешно дошёл до конца (или был отменён пользователем)."""
            try:
                self._refresh_fdb_table()
            except Exception:
                pass
            # FDB-импорт пишет также в Measurements (Source='FDB-…') — поэтому
            # сразу перерисовываем вкладку «Записи (УД2-102)», чтобы новые
            # дешифрованные строки появились рядом с UART-записями.
            try:
                self._refresh_measurements()
            except Exception:
                pass
            self._update_status_widgets()
            QMessageBox.information(self, "Импорт FDB",
                "Импорт завершён.\n\n"
                f"Записей всего: {stats.records_total}\n"
                f"BLOB-кадров (BLOCKZAP): {stats.blocks_total}\n"
                f"Пропущено (нет блока): {stats.skipped_missing}\n\n"
                "По таблицам:\n" + "\n".join(
                    f"  • {t}: {n}" for t, n in
                    sorted(stats.by_table.items()) if n))

        def _on_fdb_failed(self, msg: str):
            QMessageBox.critical(self, "Импорт FDB", msg)

        def _on_fdb_finished(self):
            if self._progress is not None:
                self._progress.accept(); self._progress = None
            self._thread = None
            self._worker = None

        def _on_fdb_clear(self):
            if QMessageBox.question(self, "Очистка импорта FDB",
                "Удалить ВСЕ ранее импортированные FDB-записи и BLOB-кадры?\n\n"
                "Принятые с прибора BlockZap-записи останутся на месте."
                ) != QMessageBox.StandardButton.Yes:
                return
            try:
                n = self._db.fdb_delete_all()
            except Exception as e:
                QMessageBox.critical(self, "Очистка импорта FDB", str(e)); return
            self._refresh_fdb_table()
            QMessageBox.information(self, "Очистка импорта FDB",
                f"Удалено строк: {n}")

        def _refresh_fdb_table(self):
            """Перезаливает QTableWidget «Импорт FDB» из FdbRecord."""
            self.fdb_table.setSortingEnabled(False)
            try:
                rows = self._db.fdb_records(limit=5000)
                self.fdb_table.setRowCount(len(rows))
                for i, r in enumerate(rows):
                    num_kod = r.get("num_kod")
                    num_zap = r.get("num_zap")
                    cells = [
                        r["variant"] or "",
                        r["kind"] or "",
                        str(r["number"]),
                        "" if num_kod is None else str(num_kod),
                        "" if num_zap is None else str(num_zap),
                        r["type_zap"] or "",
                        r["date_form"] or "",
                        r["time_form"] or "",
                        r["name_opera"] or "",
                        "" if r["num_prib"] is None else str(r["num_prib"]),
                        # A7: TYPEVAR → читабельное имя + код в скобках.
                        fdb_typevar_name(r["type_var"]),
                        r["num_obj"] or "",
                        r["smelting"] or "",
                        r["ind_maker"] or "",
                        "" if r["make_time"] is None else str(r["make_time"]),
                        (r["defekt"] or "")[:60],
                        r["code_def"] or "",
                        (r["protocol"] or "")[:60],
                    ]
                    for j, c in enumerate(cells):
                        it = QTableWidgetItem(c)
                        self.fdb_table.setItem(i, j, it)
                    # Вся служебная информация для клика — в UserRole первой ячейки.
                    head = self.fdb_table.item(i, 0)
                    head.setData(int(Qt.ItemDataRole.UserRole), {
                        "variant":  r["variant"], "kind": r["kind"],
                        "number":   r["number"],
                        "num_kod":  num_kod,  "num_zap": num_zap,
                        "num_vers": r.get("num_vers") or "",
                        "num_obj":  r.get("num_obj")  or "",
                        "code_def": r.get("code_def") or "",
                    })
            finally:
                self.fdb_table.setSortingEnabled(True)
            self.fdb_table.resizeColumnsToContents()

        def _on_open_db(self):
            fn, _ = QFileDialog.getOpenFileName(self, "Открыть БД", str(Path.home()),
                                                "SQLite (*.sqlite3 *.db)")
            if not fn: return
            try:
                new_db = BlockZapDB(fn)
            except Exception as e:
                QMessageBox.critical(self, "БД", f"Не удалось открыть: {e}"); return
            self._db.close(); self._db = new_db; self._db_path = fn
            self._refresh_table(); self._refresh_fdb_table()
            self._update_status_widgets()

        def _on_save_as(self):
            fn, _ = QFileDialog.getSaveFileName(self, "Сохранить БД как",
                str(Path.home() / "blockzap.sqlite3"), "SQLite (*.sqlite3 *.db)")
            if not fn: return
            target = sqlite3.connect(fn)
            with target:
                self._db._con.backup(target)
            target.close()
            QMessageBox.information(self, "БД", f"Скопировано в {fn}")

        def _on_delete_row(self):
            n = self._selected_number()
            if n is None: return
            if QMessageBox.question(self, "Удалить", f"Удалить запись №{n}?") == QMessageBox.StandardButton.Yes:
                self._db.delete(n); self._refresh_table()

        def _on_export_excel(self):
            rows = self._db.all()
            if not rows:
                QMessageBox.information(self, "Excel", "В БД нет записей."); return
            fn, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт",
                str(Path.home() / "peleng_report.xlsx"), "Excel (*.xlsx)")
            if not fn: return
            try:
                export_report(rows, fn)
            except Exception as e:
                QMessageBox.critical(self, "Excel", f"Ошибка экспорта: {e}"); return
            QMessageBox.information(self, "Excel", f"Сохранено: {fn}")

        # ============== A5: FR3-выгрузка (FastReport-стиль INI+BLOB) =============
        # Оригинальный PelengPC экспортирует протоколы в файл .fr3 — это
        # шаблон FastReport 3 (FR3 = FastReport v3), используется штатным
        # генератором отчётов PelengPC. Я не реверсил DFM-генератор fr3
        # в zapis2.exe (нужны RSDS-символы + Resource Hacker), поэтому
        # выгрузка ниже — *совместимая* с оригиналом по смыслу:
        #   [Header]       — метаданные записи (Variant/Kind/Дата/Оператор);
        #   [TLV]          — раскодированные TLV-записи (тип, размер, тег);
        #   [Block]        — оригинальный BlockZap-кадр в HEX (можно
        #                    скормить обратно в parse_tlv для верификации);
        #   [Ascan]        — массив отсчётов A-развёртки (samples=byte=adc);
        #   [Bscan]        — заголовок B-развёртки + размеры пикселей.
        # Файл — обычный CP-1251 INI, открывается Notepad/любым FastReport.
        def _on_export_fr3(self):
            # 1) Выясняем источник: текущая вкладка → BlockZap или FDB.
            number = None; raw = None; meta = {}; src = ""
            cur = self.tabs.currentWidget()
            if cur is getattr(self, "fdb_table", None) or (
                isinstance(cur, QWidget) and
                cur.findChildren(QTableWidget) and
                self.fdb_table in cur.findChildren(QTableWidget)):
                # FDB-таблица — берём NUMKOD выбранной строки.
                it = self.fdb_table.currentItem()
                if it is None:
                    QMessageBox.information(self, "FR3",
                        "Выберите строку в таблице «Импорт FDB» или в BlockZap."); return
                head = self.fdb_table.item(it.row(), 0)
                info = head.data(int(Qt.ItemDataRole.UserRole)) or {}
                number = info.get("num_kod"); src = "FDB"
                meta = {k: info.get(k, "") for k in
                        ("variant", "kind", "number", "num_zap",
                         "num_vers", "num_obj", "code_def")}
                if number is None:
                    QMessageBox.information(self, "FR3",
                        "У этой FDB-строки нет NUMKOD → BLOCKZAP-блока для выгрузки.");
                    return
                raw = self._db.fdb_block_get(int(number))
            else:
                # BlockZap-таблица — выбранная строка → номер кадра.
                n = self._selected_number()
                if n is None:
                    QMessageBox.information(self, "FR3",
                        "Выберите строку в BlockZap-таблице или в «Импорт FDB»."); return
                number = int(n); src = "BlockZap"
                row = self._db.get(number)
                if row is None:
                    QMessageBox.critical(self, "FR3",
                        f"BlockZap #{number} не найден."); return
                raw = row[1]
                meta = {"variant": "PelengPC", "kind": "BlockZap",
                        "number": number, "num_zap": None}
            if not raw:
                QMessageBox.critical(self, "FR3",
                    f"Бинарный BLOCKZAP-блок #{number} пустой."); return

            # 2) Имя файла.
            default = f"peleng_{src.lower()}_{number}.fr3"
            fn, _ = QFileDialog.getSaveFileName(self, "Экспорт в FR3",
                str(Path.home() / default), "FastReport FR3 (*.fr3);;Все файлы (*.*)")
            if not fn: return
            if not fn.lower().endswith(".fr3"):
                fn += ".fr3"

            # 3) Декодим TLV → собираем секции.
            try:
                tlvs = list(parse_tlv(raw))
                hdr  = ProtocolHeader.parse(raw[:16]) if len(raw) >= 16 else None
                fw   = hdr.firmware_version() if hdr else (meta.get("num_vers") or "")
                recs = decode_all(tlvs, fw)
            except Exception as e:
                tlvs = []; recs = []; fw = meta.get("num_vers") or ""

            try:
                self._write_fr3_file(fn, raw, hdr if 'hdr' in locals() else None,
                                     tlvs, recs, meta, src, fw)
            except Exception as e:
                QMessageBox.critical(self, "FR3",
                    f"Не удалось записать FR3: {e}"); return
            QMessageBox.information(self, "FR3",
                f"Сохранено: {fn}\n\n"
                f"Источник: {src}, #{number}\n"
                f"TLV-записей: {len(tlvs)}, декодировано: {len(recs)}")

        @staticmethod
        def _write_fr3_file(fn, raw, hdr, tlvs, recs, meta, src, fw):
            """Запись FR3-файла. CP-1251, секции в порядке оригинала."""
            from datetime import datetime
            lines: list[str] = []
            def section(name): lines.append(f"\n[{name}]")
            def kv(k, v):       lines.append(f"{k}={v}")

            # ─── [FR3] ─── идентификатор формата.
            lines.append("[FR3]")
            kv("Magic", "PELENG-CLONE-FR3")
            kv("Version", "1.0")
            kv("Producer", f"Peleng-Clone v{__version__}")
            kv("Exported", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ"))
            kv("Source", src)

            # ─── [Header] ─── метаданные записи.
            section("Header")
            for k in ("variant", "kind", "number", "num_zap",
                      "num_vers", "num_obj", "code_def"):
                v = meta.get(k)
                if v is not None and v != "":
                    kv(k, v)
            if hdr is not None:
                kv("RecordId",      f"0x{hdr.record_id:04x}")
                kv("Date",          hdr.date_str())
                kv("Time",          hdr.time_str())
                kv("Firmware",      fw or "")
                kv("PayloadLen",    hdr.payload_len)

            # ─── [TLV] ─── список tag/size/case/kind.
            section("TLV")
            kv("Count", len(tlvs))
            for i, t in enumerate(tlvs):
                kv(f"{i:03d}.Tag",   f"0x{t.tag:04x}")
                kv(f"{i:03d}.Size",  t.length)
                kv(f"{i:03d}.Case",  tlv_tag_case(t.tag))
                kv(f"{i:03d}.Kind",  tlv_tag_kind(t.tag))
                kv(f"{i:03d}.Type",  t.record_type.name)

            # ─── [Ascan]/[Bscan]/[Settings]/[Report] — из декодеров.
            for r in recs:
                if isinstance(r, AScanRecord):
                    section("Ascan")
                    kv("Samples",     len(r.samples))
                    kv("DurationUs",  f"{r.duration_us:.3f}")
                    kv("DepthMm",     f"{r.max_depth_mm:.2f}")
                    # Сырые отсчёты — CSV, чтоб FastReport мог
                    # построить график (max 80 столбцов).
                    s = r.samples
                    chunk = 80
                    for off in range(0, len(s), chunk):
                        kv(f"Data{off:04d}",
                           ",".join(str(x) for x in s[off:off+chunk]))
                elif isinstance(r, BScanRecord):
                    section("Bscan")
                    kv("Width",   r.width)
                    kv("Height",  r.height)
                    kv("LengthMm", f"{r.length_mm:.2f}")
                    kv("DepthMm",  f"{r.depth_mm:.2f}")
                elif isinstance(r, SettingsRecord):
                    section("Settings")
                    for f_ in r.fields:
                        kv(f_.name, f_.display)
                elif isinstance(r, ReportRecord):
                    section("Report")
                    for f_ in r.fields:
                        kv(f_.name, f_.display)

            # ─── [Block] ─── оригинальный BLOCKZAP-кадр (HEX, по 64 байта).
            section("Block")
            kv("LengthBytes", len(raw))
            hx = raw.hex().upper()
            chunk = 128  # 64 байт → 128 hex-символов
            for off in range(0, len(hx), chunk):
                lines.append(f"Hex{off//chunk:04d}={hx[off:off+chunk]}")

            text = "\n".join(lines) + "\n"
            # CP-1251 — как у оригинального PelengPC. Если содержимое
            # не в CP-1251 — fallback на UTF-8 (Notepad по BOM поймёт).
            try:
                data = text.encode("cp1251")
            except UnicodeEncodeError:
                data = b"\xEF\xBB\xBF" + text.encode("utf-8")
            with open(fn, "wb") as f:
                f.write(data)

        def _on_edit_settings(self):
            recs = self._decoded_records_for_selected()
            if recs is None: return
            sett = next((r for r in recs if isinstance(r, SettingsRecord)), None)
            if sett is None:
                QMessageBox.information(self, "Настройки", "В выбранной записи нет блока Settings."); return
            dlg = SettingsDialog(sett, self); dlg.exec()

        def _on_about(self):
            QMessageBox.about(self, "О программе",
                f"<h3>Peleng-Clone v{__version__}</h3>"
                "<p>Открытый клон PelengPC v1.2 — UI к ультразвуковому "
                "дефектоскопу «Пеленг».</p>"
                "<p>Протокол реверсирован из оригинального ПО (Borland C++Builder, 2009).</p>"
                "<p><b>Зависимости:</b> PyQt6, pyserial, openpyxl (для экспорта в Excel).</p>")

        # =================== A6: печать / PDF текущей вкладки ====================
        # Кнопки «Печать…» и «PDF…» в шапке. Один обработчик на оба пути,
        # отличаются только тем, открываем ли мы QPrintDialog или сразу пишем
        # PDF через QPrinter.setOutputFormat(PdfFormat). Каждая вкладка
        # печатается по своему правилу:
        #   A-scan / B-scan  → widget.grab() → QPainter.drawPixmap (вектор-rect
        #                      из бумаги сохраняется, A-scan чёткий)
        #   Отчёт            → self.report_view.document().print(printer)
        #   Сырой буфер      → plain-text → QTextDocument
        #   Поля / УД2 / FDB → QTableWidget → HTML-таблица → QTextDocument
        def _on_print_current_tab(self):
            try:
                from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
            except Exception as e:
                QMessageBox.critical(self, "Печать",
                    f"Не удалось загрузить QtPrintSupport: {e}\n"
                    f"Установите PyQt6-Qt6 / system libQt6PrintSupport."); return
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            dlg = QPrintDialog(printer, self)
            if dlg.exec() != QPrintDialog.DialogCode.Accepted:
                return
            try:
                self._print_current_tab_to(printer)
            except Exception as e:
                QMessageBox.critical(self, "Печать", f"Ошибка печати: {e}")

        def _on_export_pdf_current_tab(self):
            try:
                from PyQt6.QtPrintSupport import QPrinter
            except Exception as e:
                QMessageBox.critical(self, "PDF",
                    f"Не удалось загрузить QtPrintSupport: {e}\n"
                    f"Установите PyQt6-Qt6 / system libQt6PrintSupport."); return
            fn, _ = QFileDialog.getSaveFileName(self, "Сохранить как PDF",
                str(Path.home() / "peleng_tab.pdf"), "PDF (*.pdf)")
            if not fn: return
            if not fn.lower().endswith(".pdf"):
                fn += ".pdf"
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(fn)
            try:
                self._print_current_tab_to(printer)
            except Exception as e:
                QMessageBox.critical(self, "PDF", f"Ошибка сохранения: {e}"); return
            QMessageBox.information(self, "PDF", f"Сохранено: {fn}")

        @staticmethod
        def _table_to_html(table: "QTableWidget") -> str:
            """QTableWidget → HTML-таблица (учёт скрытых строк и сортировки)."""
            cols = table.columnCount()
            headers = []
            for c in range(cols):
                h = table.horizontalHeaderItem(c)
                headers.append(h.text() if h else "")
            html_rows = ["<tr>" + "".join(
                f"<th style='border:1px solid #444;padding:3px;"
                f"background:#eee'>{h}</th>" for h in headers) + "</tr>"]
            for r in range(table.rowCount()):
                if table.isRowHidden(r):
                    continue
                cells = []
                for c in range(cols):
                    it = table.item(r, c)
                    cells.append(f"<td style='border:1px solid #888;"
                                 f"padding:2px'>{(it.text() if it else '')}</td>")
                html_rows.append("<tr>" + "".join(cells) + "</tr>")
            return ("<table style='border-collapse:collapse;font-size:8pt;"
                    "font-family:sans-serif'>" + "".join(html_rows) + "</table>")

        def _print_current_tab_to(self, printer):
            from PyQt6.QtGui import QPainter, QTextDocument
            from PyQt6.QtPrintSupport import QPrinter
            w = self.tabs.currentWidget()
            # Если вкладка — контейнер с фильтром, ищем внутри QTableWidget.
            target_table = None
            if isinstance(w, QTableWidget):
                target_table = w
            else:
                for ch in w.findChildren(QTableWidget):
                    target_table = ch; break

            if w is self.ascan or w is self.bscan:
                pm = w.grab()
                painter = QPainter(printer)
                try:
                    pr = printer.pageLayout().paintRectPixels(
                        int(printer.resolution()))
                    sz = pm.size()
                    sz.scale(pr.size(), Qt.AspectRatioMode.KeepAspectRatio)
                    painter.drawPixmap(pr.x(), pr.y(),
                                       sz.width(), sz.height(), pm)
                    # подпись внизу
                    title = ("А-развёртка" if w is self.ascan else "B-развёртка")
                    painter.setPen(QColor(40, 40, 40))
                    painter.drawText(pr.x(),
                        pr.y() + sz.height() + 24,
                        f"{title} — Peleng-Clone v{__version__}")
                finally:
                    painter.end()
                return
            if w is self.report_view:
                self.report_view.document().print(printer); return
            if w is self.raw_view:
                doc = QTextDocument()
                doc.setDefaultFont(QFont("monospace", 8))
                doc.setPlainText(self.raw_view.toPlainText())
                doc.print(printer); return
            if target_table is not None:
                doc = QTextDocument()
                doc.setHtml(self._table_to_html(target_table))
                doc.print(printer); return
            # fallback — растром текущей вкладки
            pm = w.grab()
            painter = QPainter(printer)
            try:
                pr = printer.pageLayout().paintRectPixels(
                    int(printer.resolution()))
                sz = pm.size()
                sz.scale(pr.size(), Qt.AspectRatioMode.KeepAspectRatio)
                painter.drawPixmap(pr.x(), pr.y(),
                                   sz.width(), sz.height(), pm)
            finally:
                painter.end()

        def _refresh_table(self, select_number=None):
            rows = self._db.all()
            self.table.setRowCount(len(rows))
            select_row = -1
            for i, row in enumerate(rows):
                hdr = row.header()
                self.table.setItem(i, 0, QTableWidgetItem(str(row.number)))
                self.table.setItem(i, 1, QTableWidgetItem(hdr.date_str()))
                self.table.setItem(i, 2, QTableWidgetItem(hdr.time_str()))
                self.table.setItem(i, 3, QTableWidgetItem(hdr.firmware_version()))
                self.table.setItem(i, 4, QTableWidgetItem(str(row.block_len)))
                if select_number is not None and row.number == select_number:
                    select_row = i
            if select_row >= 0: self.table.selectRow(select_row)
            elif rows:          self.table.selectRow(0)
            else:               self._show_records([])
            self._update_status_widgets()

        def _selected_number(self):
            rows = self.table.selectionModel().selectedRows()
            if not rows: return None
            item = self.table.item(rows[0].row(), 0)
            return int(item.text()) if item else None

        def _decoded_records_for_selected(self):
            n = self._selected_number()
            if n is None: return None
            row = self._db.get(n)
            if row is None: return None
            firmware = row.header().firmware_version()
            return decode_all(list(parse_tlv(row.block)), firmware)

        def _on_selection_changed(self):
            recs = self._decoded_records_for_selected()
            # Снимаем «зону» (SHORTPROT2 подсвечивает её в _on_fdb_selection_changed).
            try: self.ascan.set_zone_label("", "")
            except Exception: pass
            self._show_records(recs or [])

        def _show_records(self, recs, raw_block=None):
            a = next((r for r in recs if isinstance(r, AScanRecord)), None); self.ascan.set_record(a)
            b = next((r for r in recs if isinstance(r, BScanRecord)), None); self.bscan.set_record(b)
            rep = next((r for r in recs if isinstance(r, ReportRecord)), None)
            if rep is None:
                self.report_view.setHtml("<i>(в записи нет блока Report)</i>")
            else:
                # Рисуем отчёт как структурированную HTML-таблицу из .dal-полей.
                rows_html = ""
                for fld in rep.fields:
                    rows_html += (f"<tr><td><b>{fld.label or fld.name}</b></td>"
                                  f"<td><tt>{fld.name}</tt></td>"
                                  f"<td>{fld.display}</td></tr>")
                self.report_view.setHtml(
                    "<h3>Протокол контроля</h3>"
                    f"<p><b>Оператор:</b> {rep.operator}<br>"
                    f"<b>Объект:</b> {rep.object_}</p>"
                    f"<p><b>Протокол:</b><br>"
                    f"<pre style='white-space: pre-wrap'>{rep.notes}</pre></p>"
                    f"<table border=1 cellspacing=0 cellpadding=4 "
                    f"style='border-collapse:collapse'>"
                    f"<tr><th>Подпись</th><th>Поле</th><th>Значение</th></tr>"
                    f"{rows_html}</table>")
            # Таблица «Поля» — объединяем все DecodedField из всех записей.
            all_fields: list[tuple[str, DecodedField]] = []
            for r in recs:
                src = r.record_type.name
                for f_ in r.fields:
                    all_fields.append((src, f_))
            self.fields_view.setRowCount(len(all_fields))
            for i, (src, f_) in enumerate(all_fields):
                self.fields_view.setItem(i, 0, QTableWidgetItem(f_.name))
                self.fields_view.setItem(i, 1, QTableWidgetItem(f_.label))
                self.fields_view.setItem(i, 2, QTableWidgetItem(src))
                self.fields_view.setItem(i, 3, QTableWidgetItem(f_.display))
            self.fields_view.resizeColumnsToContents()
            # Hex — либо из raw_block (режим FDB-клика), либо из BlockZap-выборки.
            if raw_block is not None:
                self.raw_view.setPlainText(_hexdump(bytes(raw_block), max_lines=64))
            else:
                n = self._selected_number()
                if n is None:
                    self.raw_view.setPlainText("")
                else:
                    row = self._db.get(n)
                    self.raw_view.setPlainText("" if row is None else _hexdump(row.block, max_lines=64))

        # ------- A1/A2: визуализация импортированных FDB-блобов -----------------
        # NASTR2/RESULTS2/SHORTPROT2 → колонка NUMKOD → FdbBlock.Number → блок
        # BLOB (header 16 + TLV) → parse_tlv → decode_all → ровно те же
        # AScanRecord/BScanRecord/SettingsRecord/ReportRecord/Generic, что мы
        # уже умеем рисовать. Никакого нового декодера не понадобилось.
        def _on_fdb_selection_changed(self):
            rows = self.fdb_table.selectionModel().selectedRows()
            if not rows:
                return
            head = self.fdb_table.item(rows[0].row(), 0)
            meta = head.data(int(Qt.ItemDataRole.UserRole)) if head else None
            if not isinstance(meta, dict):
                return
            num_kod = meta.get("num_kod")
            if num_kod is None:
                return
            block = self._db.fdb_block_get(int(num_kod))
            if block is None:
                # NUMKOD не нашёлся в FdbBlock — частая ситуация в SHORTPROT2,
                # когда оригинальная программа не сохранила blob для строки.
                self._show_records([], raw_block=None)
                try:
                    self.ascan.set_zone_label("",
                        f"NUMKOD={num_kod} не найден в BLOCKZAP")
                except Exception: pass
                return
            # Версия прошивки нужна для выбора схемы декодирования полей
            # (.dal-таблицы 1.x vs 2.x), берём её из FdbRecord.NumVers.
            firmware = (meta.get("num_vers") or "").strip()
            try:
                recs = decode_all(list(parse_tlv(block)), firmware)
            except Exception:
                recs = []
            self._show_records(recs, raw_block=block)
            # SHORTPROT2 → текстовая подсветка зоны.
            kind = (meta.get("kind") or "").upper()
            num_zap = meta.get("num_zap")
            if kind == "SHORTPROT" and num_zap is not None:
                ext = []
                obj = (meta.get("num_obj") or "").strip()
                cdf = (meta.get("code_def") or "").strip()
                if obj: ext.append(obj)
                if cdf: ext.append(f"код {cdf}")
                try:
                    self.ascan.set_zone_label(
                        f"Зона № {num_zap}", " / ".join(ext))
                except Exception: pass
            else:
                try: self.ascan.set_zone_label("", "")
                except Exception: pass
            # Переключаем фокус на самую релевантную вкладку результата
            # (но «Импорт FDB» оставляем активной — пользователь сам решит,
            # куда смотреть).
            try:
                a = next((r for r in recs if isinstance(r, AScanRecord)), None)
                b = next((r for r in recs if isinstance(r, BScanRecord)), None)
                if a is not None: self.tabs.setCurrentWidget(self.ascan)
                elif b is not None: self.tabs.setCurrentWidget(self.bscan)
                else: self.tabs.setCurrentWidget(self.fields_view)
            except Exception: pass

        # ------- Двойной клик по строке в «Записи (УД2-102)» ------------------
        def _ud2_row_meta(self, table: QTableWidget, row: int) -> dict | None:
            """Достаём словарь Measurements-row, который мы спрятали в
            UserRole первой ячейки в _refresh_measurements."""
            if row < 0 or row >= table.rowCount():
                return None
            it = table.item(row, 0)
            if it is None:
                return None
            meta = it.data(int(Qt.ItemDataRole.UserRole))
            return meta if isinstance(meta, dict) else None

        def _load_blockzap_for_row(self, meta: dict
                                   ) -> tuple[bytes | None, str]:
            """Возвращает (block_bytes, firmware). Для FDB-* — берём NUMKOD;
            для UD2 — поднимаем BlockZap.Block по реальному адресу."""
            source = str(meta.get("source") or "UD2")
            if source.startswith("FDB-"):
                num_kod = meta.get("num_kod")
                if num_kod is None:
                    return (None, "")
                block = self._db.fdb_block_get(int(num_kod))
                firmware = ""
                # NumVers лежит во FdbRecord — у Measurements его прямо нет,
                # но у Settings/RESULTS/SHORTPROT он одинаковый, поэтому
                # достаём первый попавшийся из FdbRecord.
                row = self._db._con.execute(
                    "SELECT NumVers FROM FdbRecord WHERE NumKod=? LIMIT 1",
                    (int(num_kod),)).fetchone()
                if row and row[0]:
                    firmware = str(row[0]).strip()
                return (block, firmware)
            # UD2 (UART-приём): Addr напрямую = record_id в BlockZap.
            addr = int(meta.get("addr") or 0)
            br = self._db.get(addr)
            if br is None:
                return (None, "")
            return (bytes(br.block), "")

        def _on_ud2_protocol_selection_changed(self):
            """Одинарный клик (выделение строки) в «Протоколы (А-развёртка)» →
            подгружаем BLOCKZAP, разбираем TLV, рисуем A-/B-scan в правой
            панели БЕЗ переключения активной вкладки."""
            rows = self.ud2_table_protocols.selectionModel().selectedRows()
            if not rows:
                return
            row = rows[0].row()
            meta = self._ud2_row_meta(self.ud2_table_protocols, row)
            if meta is None:
                return
            block, firmware = self._load_blockzap_for_row(meta)
            if block is None:
                self._show_records([], raw_block=None)
                try:
                    self.ascan.set_zone_label("",
                        f"NUMKOD={meta.get('num_kod')} не найден в BLOCKZAP")
                except Exception:
                    pass
                return
            try:
                recs = decode_all(list(parse_tlv(block)), firmware)
            except Exception:
                recs = []
            self._show_records(recs, raw_block=block)
            # Подпись зоны.
            try:
                obj = (meta.get("num_obj") or "").strip()
                cdf = (meta.get("code_def") or "").strip()
                ext = " / ".join(x for x in (obj, f"код {cdf}" if cdf else "") if x)
                self.ascan.set_zone_label("", ext)
            except Exception:
                pass

        def _on_ud2_open_protocol(self, row: int, *, settings: bool = False):
            """Двойной клик по протоколу/настройке → подгружаем BLOCKZAP,
            разбираем TLV, рисуем A-/B-развёртку (или поля настроек) и
            переходим на соответствующую вкладку."""
            table = (self.ud2_table_settings if settings
                     else self.ud2_table_protocols)
            meta = self._ud2_row_meta(table, row)
            if meta is None:
                return
            block, firmware = self._load_blockzap_for_row(meta)
            if block is None:
                QMessageBox.information(self, "Протокол",
                    "Для этой записи нет сохранённого BLOCKZAP-блока "
                    "(NUMKOD не найден).")
                return
            try:
                recs = decode_all(list(parse_tlv(block)), firmware)
            except Exception:
                recs = []
            self._show_records(recs, raw_block=block)
            # Подпись зоны (для SHORTPROT-строк сюда не попадаем — они
            # уходят в _on_ud2_open_report; для RESULTS-строк зону рисуем,
            # если в PROTOCOL/Defect полях что-то есть).
            try:
                obj = (meta.get("num_obj") or "").strip()
                cdf = (meta.get("code_def") or "").strip()
                ext = " / ".join(x for x in (obj, f"код {cdf}" if cdf else "") if x)
                self.ascan.set_zone_label("", ext)
            except Exception:
                pass
            # Переключение на самую релевантную вкладку:
            try:
                a = next((r for r in recs if isinstance(r, AScanRecord)), None)
                b = next((r for r in recs if isinstance(r, BScanRecord)), None)
                if a is not None:
                    self.tabs.setCurrentWidget(self.ascan)
                elif b is not None:
                    self.tabs.setCurrentWidget(self.bscan)
                else:
                    self.tabs.setCurrentWidget(self.fields_view)
            except Exception:
                pass

        def _on_ud2_open_report(self, row: int):
            """Двойной клик по строке-отчёту (FDB-SHORTPROT или UD2) →
            переключаемся на вкладку «Отчёт» и автозаполняем её HTML-данными
            из Measurements + (если есть) разобранным BLOCKZAP-телом."""
            meta = self._ud2_row_meta(self.ud2_table_reports, row)
            if meta is None:
                return
            block, firmware = self._load_blockzap_for_row(meta)
            recs = []
            if block is not None:
                try:
                    recs = decode_all(list(parse_tlv(block)), firmware)
                except Exception:
                    recs = []
            # Базовая шапка «Отчёт по контролю» — собираем из Measurements,
            # независимо от того, нашёлся ли BLOCKZAP-блок.
            def _esc(s: object) -> str:
                return html.escape("" if s is None else str(s))
            dash = "—"
            make_year = _fdb_make_year(meta.get("make_time") or 0)
            make_year_s = (str(make_year) if make_year else
                           (str(meta.get("make_time"))
                            if meta.get("make_time") else dash))
            head_rows = [
                ("Источник",       meta.get("source") or "UD2"),
                ("Дата контроля",  _strip_unprintable(str(meta.get("date") or ""))
                                   or dash),
                ("Время",          meta.get("time_str") or dash),
                ("Тип",            meta.get("type_name") or dash),
                ("№ объекта",      meta.get("num_obj")   or dash),
                ("Плавка",         meta.get("plavka")    or dash),
                ("Клеймо",         meta.get("stamp")     or dash),
                ("Год",            meta.get("god")       or dash),
                ("Сторона",        meta.get("side")      or dash),
                ("Шейка",          meta.get("sheika")    or dash),
                ("Обод",           meta.get("obod")      or dash),
                ("Обточка",        meta.get("obtochka")  or dash),
                ("Гребень",        meta.get("greben")    or dash),
                ("Наплавка",       meta.get("naplavka")  or dash),
                ("Завод (инд.)",   meta.get("ind_maker") or dash),
                ("Дата изг.",      make_year_s),
                ("Дефект",         meta.get("defekt")    or dash),
                ("Код деф.",       meta.get("code_def")  or dash),
                ("Оператор",       meta.get("operator")  or dash),
            ]
            parts: list[str] = ["<h2>Отчёт по контролю</h2>",
                                "<table border='1' cellpadding='4' "
                                "cellspacing='0'>"]
            for k, v in head_rows:
                parts.append(f"<tr><th align='left'>{_esc(k)}</th>"
                             f"<td>{_esc(v)}</td></tr>")
            parts.append("</table>")
            # Если у нас есть TLV-разбор — добавляем «полную» таблицу полей,
            # тот же формат, что и _show_records даёт во вкладке «Отчёт».
            for r in recs:
                if not r.fields:
                    continue
                parts.append(f"<h3>{_esc(r.record_type.name)}</h3>")
                parts.append("<table border='1' cellpadding='4' cellspacing='0'>"
                             "<tr><th>Подпись</th><th>Поле</th>"
                             "<th>Значение</th></tr>")
                for f_ in r.fields:
                    parts.append(f"<tr><td>{_esc(f_.label)}</td>"
                                 f"<td>{_esc(f_.name)}</td>"
                                 f"<td>{_esc(f_.display)}</td></tr>")
                parts.append("</table>")
            self.report_view.setHtml("".join(parts))
            # Если есть TLV — заодно обновим таблицу «Поля» и hex.
            if recs:
                self._show_records(recs, raw_block=block)
            try:
                self.tabs.setCurrentWidget(self.report_view)
            except Exception:
                pass

        def closeEvent(self, ev):
            self._db.close(); super().closeEvent(ev)

    app = QApplication.instance() or QApplication(argv)
    app.setApplicationName("Peleng-Clone")
    w = MainWindow(); w.show()
    if demo:
        # Сразу вставляем встроенный демо-блок — пользователю не нужен ни
        # прибор, ни сторонние скрипты, чтобы увидеть рабочий GUI.
        try:
            w._on_demo_block()
        except Exception as e:
            QMessageBox.warning(w, "Демо-блок",
                f"Не удалось вставить демо-блок: {e}")
    return app.exec()


# =============================================================================
# 7. CLI / ENTRY POINT
# =============================================================================
def cli_proto_test(args) -> int:
    cfg = PortConfig(name=args.port, baud=args.baud, timeout_ms=args.timeout_ms)
    with PelengPort(cfg) as p:
        hdr = p.test_port()
    print(f"record_id   = 0x{hdr.record_id:04x}")
    print(f"time        = {hdr.time_str()}")
    print(f"date        = {hdr.date_str()}")
    print(f"firmware    = {hdr.firmware_version()}")
    print(f"payload_len = {hdr.payload_len}")
    print(f"raw         = {hdr.raw.hex(' ')}")
    return 0


def cli_proto_block(args) -> int:
    cfg = PortConfig(name=args.port, baud=args.baud, timeout_ms=args.timeout_ms)
    with PelengPort(cfg) as p:
        blob = p.request_block(args.len)
    hdr = ProtocolHeader.parse(blob)
    print("HEADER:", hdr)
    for r in parse_tlv(blob):
        print("  ", r)
    return 0


def cli_fake_device(args) -> int:
    try:
        fake_device_loop(args.port, args.baud)
    except KeyboardInterrupt:
        return 0
    return 0


def cli_gui(args) -> int:
    demo = bool(getattr(args, "demo", False))
    return _launch_gui(sys.argv[:1], demo=demo)


def cli_demo(args) -> int:
    """Sanity-режим: собираем демо-блок и выводим summary без GUI."""
    blob, hdr = make_fake_block(seed=int(getattr(args, "seed", 0) or 0))
    print(f"demo block: {len(blob)} bytes, "
          f"firmware={hdr.firmware_version()}, "
          f"record_id=0x{hdr.record_id:04x}, "
          f"date={hdr.date_str()}, time={hdr.time_str()}")
    records = decode_all(list(parse_tlv(blob)), hdr.firmware_version())
    for r in records:
        print(" •", r.summary())
    out = getattr(args, "out", None)
    if out:
        Path(out).write_bytes(blob)
        print(f"raw block written to {out}")
    return 0


# --- УД2-102 CLI ----------------------------------------------------------

def cli_ud2_handshake(args) -> int:
    """4× 0x55 → распечатать заголовок и заводской № прибора."""
    cfg = UD2_102PortConfig(name=args.port, baud=args.baud)
    with UD2_102Port(cfg) as p:
        hdr = p.handshake()
    print(f"device_no = {hdr.device_no}")
    print(f"raw[16]   = {hdr.raw.hex(' ')}")
    return 0


def cli_ud2_scan(args) -> int:
    """Полный обход памяти прибора + печать summary каждой записи."""
    cfg = UD2_102PortConfig(name=args.port, baud=args.baud)
    db: BlockZapDB | None = None
    if args.db:
        Path(args.db).parent.mkdir(parents=True, exist_ok=True)
        db = BlockZapDB(args.db)
    cnt = 0
    try:
        with UD2_102Port(cfg) as p:
            hdr = p.handshake()
            print(f"device_no = {hdr.device_no}")
            for rec in p.scan_memory():
                cnt += 1
                print(rec.summary())
                if db is not None:
                    db.upsert_measurement(hdr.device_no, rec)
    finally:
        if db is not None:
            db.close()
    print(f"---\n{cnt} записей" + (f" сохранено в {args.db}" if args.db else ""))
    return 0


def cli_ud2_screen_dump(args) -> int:
    cfg = UD2_102PortConfig(name=args.port, baud=args.baud)
    with UD2_102Port(cfg) as p:
        p.handshake()
        data = p.screen_dump()
    out = args.out or f"ud2_screen_{int(time.time())}.bin"
    Path(out).write_bytes(data)
    print(f"saved {len(data)} bytes -> {out}")
    return 0


def cli_ud2_rtc_sync(args) -> int:
    cfg = UD2_102PortConfig(name=args.port, baud=args.baud)
    with UD2_102Port(cfg) as p:
        p.handshake()
        p.rtc_sync()
    print("RTC synced.")
    return 0


def cli_demo_ud2(args) -> int:
    """Демо-обход для УД2-102: парсим набор фейковых записей без прибора."""
    seed = int(getattr(args, "seed", 0) or 0)
    fake_hs = UD2_102Header.parse(make_fake_ud2102_handshake(
        device_no=int(getattr(args, "device_no", 4242) or 4242)))
    print(f"FAKE handshake: device_no = {fake_hs.device_no}")
    db: BlockZapDB | None = None
    if args.db:
        Path(args.db).parent.mkdir(parents=True, exist_ok=True)
        db = BlockZapDB(args.db)
    cnt = 0
    for base in UD2_BASES[:int(args.bases)]:
        for idx in range(int(args.per_base)):
            addr, raw86 = make_fake_ud2102_record(
                seed=seed * 1000 + idx, addr=base + idx)
            rec = parse_ud2102_record(addr, raw86[2:])
            if rec is None:
                continue
            cnt += 1
            print(rec.summary())
            if db is not None:
                db.upsert_measurement(fake_hs.device_no, rec)
    if db is not None:
        db.close()
    print(f"---\n{cnt} записей" + (f" сохранено в {args.db}" if args.db else ""))
    return 0


def cli_rev_info(args) -> int:
    """Печать всех таблиц реверса (LUT тегов, TYPEVAR, TypeField, .dal-схемы).

    Команда без побочных эффектов — нужна, чтобы быстро убедиться, что
    словари TLV-LUT, TYPEVAR и TypeField внутри peleng_clone.py соответствуют
    тому, что выкопано из 102_203dll.dll, zapis2.exe и PelengPCtest.dal.
    """
    print("=== TLV-LUT (102_203dll.dll @ 0x401670) ===")
    print("tcode | case | RecordType")
    print("------+------+--------------------------------")
    for tcode, case in enumerate(TLV_DLL_LUT):
        info = TLV_CASE_MAP.get(case)
        rt_name = info[0].name if info else "—"
        print(f"  {tcode:3d}  |  {case:2d}  | {rt_name}")
    print()

    print("=== TYPEVAR_DICT (из zapis2.exe + PelengPC.fdb) ===")
    print(f"Всего {len(TYPEVAR_DICT)} записей.")
    for (tcode, sub), name in sorted(TYPEVAR_DICT.items(),
                                     key=lambda kv: (kv[0][0], kv[0][1] or 0)):
        sub_str = f"{sub:2d}" if sub is not None else "—"
        print(f"  tcode={tcode:5d}  sub={sub_str}  → {name}")
    print()

    print("=== TypeField bit-flags (.dal → fcn.00402980) ===")
    print("raw   | name           | описание                                | max-len")
    print("------+----------------+----------------------------------------+--------")
    for code, (name, descr) in sorted(TYPEFIELD_DECODER.items()):
        max_len = (code >> 6) & 0x3F
        print(f" 0x{code:03x} | {name:14s} | {descr:38s} | {max_len}")
    print()

    print("=== UD2_BASES (адреса баз памяти прибора) ===")
    print(", ".join(f"{b:5d}" for b in UD2_BASES))
    print(f"индексов в базе: {UD2_INDICES_PER_BASE}")
    print()

    print("=== Эталонные значения ===")
    print(f"АЦП               : {ADC_RATE_MHZ} МГц")
    print(f"V_сталь           : {US_VEL_MM_PER_US} мм/мкс")
    print(f"Полная амплитуда  : 0..255 → 0..{AMPL_FULL_DB} дБ")
    print(f"Inter-byte (UD2)  : {UD2_INTER_BYTE_S * 1000:.1f} мс")
    print(f"Inter-byte (PC1.2): {INTER_BYTE_DELAY_S * 1000:.1f} мс  "
          f"(busy-wait по GetTickCount в fcn.00423a58)")
    print()

    print("=== PelengPC: КОМАНДЫ → ПРИБОР (полная карта) ===")
    print("op  | name        | описание")
    print("----+-------------+------------------------------------------------")
    for op in sorted(PELENGPC_COMMANDS):
        name, desc = PELENGPC_COMMANDS[op]
        print(f" 0x{op:02x} | {name:11s} | {desc}")
    print()

    print("=== PelengPC.exe — ключевые функции (RVA) ===")
    for name, addr in PELENGPC_EXE_FUNCS.items():
        print(f"  0x{addr:08x}  {name}")
    print()

    print("=== 102_203dll.dll — ключевые функции (RVA) ===")
    for name, addr in PELENGPC_DLL_FUNCS.items():
        print(f"  0x{addr:08x}  {name}")
    print()

    print("=== ПРОВЕРКИ ОТСУТСТВИЯ ===")
    print(" - CRC / контрольной суммы НЕТ: ни в EXE, ни в DLL нет импортов")
    print("   из crypt32 / ntdll!RtlCrc*; полиномы 0x8005/0x1021 в .rdata —")
    print("   случайные попадания, а не lookup-таблицы (нет 256-байтных")
    print("   массивов). Проверено: rabin2 -i 102_203dll.dll | grep -i crc.")
    print(" - RTC-sync ПК→прибор НЕТ: SetLocalTime / SetSystemTime / ")
    print("   SetSystemTimeAdjustment не импортированы. GetLocalTime в EXE")
    print("   используется только для штампа отчётов на ПК-стороне.")
    print(" - Calibration/zeroing: специальной команды нет — калибровка")
    print("   живёт внутри Settings-записи (tail-блок Settings, offsets")
    print("   +0x65..+0x87 — поля gain/threshold/gate/velocity/angle).")
    print(" - ARD curves: тоже хранятся в Settings (см. UD2_AcousticSettings),")
    print("   read-out — через тот же 'B' LL HH, а не отдельная команда.")
    print(" - Дополнительных опкодов (0x4D / 0x4C / 0x45 / 0x46) нет: ")
    print("   fcn.00423f0c имеет ровно 5 callers, опкоды только {0x55,0x42,")
    print("   0x74,0x9a}, см. карту PELENGPC_COMMANDS.")
    return 0


def cli_fbembed_install(args) -> int:
    """Скачивает Firebird 2.5.9 embed (Windows) или подсказывает apt-команду
    (Linux/Mac). После — печатает путь к найденному fbembed.dll / libfbembed.so.
    """
    existing = _fdb_find_client_lib()
    if existing and not getattr(args, "force", False):
        print(f"libfbembed уже найден: {existing}")
        print("Используй --force, чтобы перекачать заново.")
        return 0
    if sys.platform != "win32":
        # На Linux/Mac бросаем подсказку — авто-установки нет.
        try:
            fdb_install_fbembed()
        except PelengError as e:
            print(f"[fbembed-install] {e}", file=sys.stderr)
        # …но если в системе всё-таки нашли lib через _fdb_find_client_lib —
        # сообщим путь.
        found = _fdb_find_client_lib()
        if found:
            print(f"libfbembed обнаружен системно: {found}")
            return 0
        return 2

    print(f"Скачиваю Firebird-2.5.9 embed → {PELENG_FB25_DIR} …")
    last_pct = [-1]
    def _cb(done: int, total: int) -> None:
        if total <= 0:
            return
        pct = int(done * 100 / total)
        if pct != last_pct[0]:
            last_pct[0] = pct
            print(f"  {pct:3d}%  ({done/1024/1024:.1f} / "
                  f"{total/1024/1024:.1f} MB)")
    try:
        path = fdb_install_fbembed(progress_cb=_cb,
                                   force=getattr(args, "force", False))
    except PelengError as e:
        print(f"[fbembed-install] {e}", file=sys.stderr)
        return 2
    print(f"OK: {path}")
    return 0


def cli_fdb_info(args) -> int:
    """Печатает структуру и счётчики таблиц PelengPC.fdb (без записи в SQLite).

    Удобно для проверки, что libfbembed подхватился и .fdb читается.
    """
    try:
        counts = fdb_count_rows(args.fdb)
    except PelengError as e:
        print(f"[fdb-info] {e}", file=sys.stderr)
        return 2
    if not counts:
        print(f"[fdb-info] В {args.fdb} нет ни одной NASTR*/RESULTS*/"
              "SHORTPROT*/BLOCKZAP таблицы.", file=sys.stderr)
        return 1
    print(f"FDB: {args.fdb}")
    width = max(len(t) for t in counts)
    for t in sorted(counts):
        print(f"  {t.ljust(width)}  {counts[t]:>10} rows")
    print(f"  --\n  Всего записей в *1/*2/*3-таблицах:  "
          f"{sum(v for t, v in counts.items() if t != FDB_BLOCKZAP_TABLE)}")
    if FDB_BLOCKZAP_TABLE in counts:
        print(f"  BLOB-кадров в {FDB_BLOCKZAP_TABLE}: "
              f"{counts[FDB_BLOCKZAP_TABLE]}")
    if args.sample:
        print("\n[fdb-info] первые записи (по одной из каждой таблицы):")
        seen = set()
        for tname, kind, variant, rec, blob in fdb_iter_records(args.fdb):
            if tname in seen:
                continue
            seen.add(tname)
            print(f"\n  --- {tname}  (kind={kind} variant={variant}) ---")
            for k, v in rec.items():
                disp = _opt_str(v) if v is not None else None
                if disp is not None and len(disp) > 60:
                    disp = disp[:60] + "…"
                print(f"    {k:<20} = {disp!r}")
            if blob is not None:
                print(f"    [BLOCK len={len(blob)}; "
                      f"first 16 = {blob[:16].hex()}]")
    return 0


def cli_fdb_import(args) -> int:
    """Импорт PelengPC.fdb → SQLite (FdbRecord + FdbBlock).

    После импорта (если --decode) проходит по импортированным NUMKOD и
    запускает parse_tlv() на каждом FdbBlock.Block — проверяет, что наш
    TLV-парсер видит ту же TLV-структуру, что и оригинальный PelengPC.
    """
    db_path = args.db or str(Path.home() / ".peleng_clone" / "blockzap.sqlite3")
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    db = BlockZapDB(db_path)
    try:
        last_t = [time.time()]
        def _cb(p: FdbImportProgress) -> None:
            now = time.time()
            if now - last_t[0] >= 0.5 or p.current == p.total:
                last_t[0] = now
                print(f"  [{p.stage}]  {p.current}/{p.total}")
        try:
            stats = fdb_import_to_db(args.fdb, db,
                                     variants=tuple(args.variants),
                                     progress_cb=_cb)
        except PelengError as e:
            print(f"[fdb-import] {e}", file=sys.stderr)
            return 2

        print()
        print(f"Импорт завершён → {db_path}")
        print(f"  записей всего:      {stats.records_total}")
        print(f"  BLOB-кадров:        {stats.blocks_total}")
        print(f"  пропущено (no blk): {stats.skipped_missing}")
        for t, n in sorted(stats.by_table.items()):
            if n:
                print(f"  {t:<14} {n}")

        if args.decode:
            print("\n[fdb-import] парсим TLV на импортированных BLOB-кадрах…")
            n_ok = n_fail = 0
            cur = db._con.execute("SELECT Number, Block FROM FdbBlock")
            for number, blob in cur:
                if not blob or len(blob) < HEADER_LEN:
                    continue
                try:
                    hdr = ProtocolHeader.parse(bytes(blob[:HEADER_LEN]))
                    _ = parse_tlv(bytes(blob[HEADER_LEN:HEADER_LEN
                                                 + hdr.payload_len]))
                    n_ok += 1
                except Exception:
                    n_fail += 1
            print(f"  TLV ok:   {n_ok}")
            print(f"  TLV fail: {n_fail}")
        return 0
    finally:
        db.close()


def _launch_gui(argv: list[str], demo: bool = False) -> int:
    ok, hint = _check_runtime_deps()
    if not ok:
        # Печатаем подсказку в консоль и пытаемся показать диалог через
        # Tk — он есть в стандартной библиотеке Python почти везде, в
        # отличие от PyQt6, которого, собственно, и не хватает.
        print("[peleng-clone] " + hint, file=sys.stderr)
        try:
            import tkinter
            from tkinter import messagebox
            root = tkinter.Tk(); root.withdraw()
            messagebox.showerror("Peleng-Clone — нет зависимостей", hint)
            root.destroy()
        except Exception:
            pass
        return 2
    return run_gui(argv, demo=demo)


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]
    ap = argparse.ArgumentParser(
        prog="peleng_clone",
        description="Однофайловый клон PelengPC v1.2 (PyQt6).",
    )
    ap.add_argument("--demo", action="store_true",
                    help="запустить GUI с уже подготовленным демо-блоком "
                         "(без прибора, без COM-порта, без сторонних скриптов)")
    sub = ap.add_subparsers(dest="cmd")
    p_gui = sub.add_parser("gui", help="запустить GUI (по умолчанию)")
    p_gui.add_argument("--demo", action="store_true",
                       help="вставить демо-блок в БД сразу при старте")
    p_gui.set_defaults(func=cli_gui)
    p_demo = sub.add_parser("demo", help="напечатать summary демо-блока (без GUI)")
    p_demo.add_argument("--seed", type=int, default=0)
    p_demo.add_argument("--out",  type=str, default=None,
                        help="записать сырой блок (заголовок+TLV) в файл")
    p_demo.set_defaults(func=cli_demo)
    p_pt = sub.add_parser("proto-test", help="хэндшейк 'U' и распечатать заголовок")
    p_pt.add_argument("--port", required=True)
    p_pt.add_argument("--baud", type=int, default=19200)
    p_pt.add_argument("--timeout-ms", type=int, default=2000)
    p_pt.set_defaults(func=cli_proto_test)
    p_pb = sub.add_parser("proto-block", help="запросить блок 'B LL HH'")
    p_pb.add_argument("--port", required=True)
    p_pb.add_argument("--baud", type=int, default=19200)
    p_pb.add_argument("--timeout-ms", type=int, default=2000)
    p_pb.add_argument("--len", type=int, default=512)
    p_pb.set_defaults(func=cli_proto_block)
    p_fd = sub.add_parser("fake-device", help="запустить заглушку-«прибор» (опционально, для теста через socat)")
    p_fd.add_argument("--port", required=True)
    p_fd.add_argument("--baud", type=int, default=19200)
    p_fd.set_defaults(func=cli_fake_device)

    # --- УД2-102 CLI ---
    p_ud2_hs = sub.add_parser("ud2102-handshake",
        help="УД2-102: 4× 0x55 → распечатать заголовок и заводской № прибора")
    p_ud2_hs.add_argument("--port", required=True)
    p_ud2_hs.add_argument("--baud", type=int, default=19200)
    p_ud2_hs.set_defaults(func=cli_ud2_handshake)
    p_ud2_sc = sub.add_parser("ud2102-scan",
        help="УД2-102: полный обход памяти с heuristic-обрывом по пустым кадрам")
    p_ud2_sc.add_argument("--port", required=True)
    p_ud2_sc.add_argument("--baud", type=int, default=19200)
    p_ud2_sc.add_argument("--db",   type=str, default=None,
                          help="опц. путь к sqlite, в который писать Measurements")
    p_ud2_sc.set_defaults(func=cli_ud2_scan)
    p_ud2_dump = sub.add_parser("ud2102-screen-dump",
        help="УД2-102: команда 0x49 — снимок экрана прибора")
    p_ud2_dump.add_argument("--port", required=True)
    p_ud2_dump.add_argument("--baud", type=int, default=19200)
    p_ud2_dump.add_argument("--out",  type=str, default=None)
    p_ud2_dump.set_defaults(func=cli_ud2_screen_dump)
    p_ud2_rtc = sub.add_parser("ud2102-rtc-sync",
        help="УД2-102: команда 0x54 — отправить на прибор текущее время ПК")
    p_ud2_rtc.add_argument("--port", required=True)
    p_ud2_rtc.add_argument("--baud", type=int, default=19200)
    p_ud2_rtc.set_defaults(func=cli_ud2_rtc_sync)
    p_demo_ud2 = sub.add_parser("demo-ud2102",
        help="УД2-102: напечатать summary набора фейковых записей (без прибора)")
    p_demo_ud2.add_argument("--seed",      type=int, default=0)
    p_demo_ud2.add_argument("--device-no", type=int, default=4242,
                            dest="device_no")
    p_demo_ud2.add_argument("--bases",     type=int, default=2,
                            help="сколько баз сканировать (1..8)")
    p_demo_ud2.add_argument("--per-base",  type=int, default=8,
                            help="сколько индексов в базе")
    p_demo_ud2.add_argument("--db",        type=str, default=None,
                            help="опц. путь к sqlite, в который писать Measurements")
    p_demo_ud2.set_defaults(func=cli_demo_ud2)

    p_rev = sub.add_parser("rev-info",
        help="распечатать таблицы реверса (LUT тегов, TYPEVAR, TypeField, эталоны)")
    p_rev.set_defaults(func=cli_rev_info)

    # --- нативная установка fbembed (Firebird 2.5.9) ---
    p_fbi = sub.add_parser("fbembed-install",
        help="скачать и распаковать Firebird 2.5.9 embed в ~/.peleng_clone/fb25/")
    p_fbi.add_argument("--force", action="store_true",
        help="перекачать, даже если fbembed уже найден")
    p_fbi.set_defaults(func=cli_fbembed_install)

    # --- импорт PelengPC.fdb (вагонная версия) ---
    p_fi = sub.add_parser("fdb-info",
        help="посмотреть содержимое PelengPC.fdb (Firebird 2.x) без записи в БД")
    p_fi.add_argument("fdb",
        help="путь к PelengPC.fdb (или другому Firebird 2.x .fdb)")
    p_fi.add_argument("--sample", action="store_true",
        help="напечатать по одной первой записи из каждой таблицы")
    p_fi.set_defaults(func=cli_fdb_info)
    p_fim = sub.add_parser("fdb-import",
        help="импортировать NASTR/RESULTS/SHORTPROT + BLOCKZAP в локальную SQLite")
    p_fim.add_argument("fdb", help="путь к PelengPC.fdb")
    p_fim.add_argument("--db", type=str, default=None,
        help="путь к SQLite (по умолчанию ~/.peleng_clone/blockzap.sqlite3)")
    p_fim.add_argument("--variants", nargs="+", default=["1", "2", "3"],
        choices=["1", "2", "3"],
        help="какие варианты *1/*2/*3 импортировать (по умолчанию все)")
    p_fim.add_argument("--decode", action="store_true",
        help="после импорта прогнать parse_tlv() по всем BLOB-кадрам")
    p_fim.set_defaults(func=cli_fdb_import)

    # Без аргументов — сразу GUI. С --help / неизвестной командой —
    # пусть argparse напечатает help и выйдет.
    if not argv:
        return _launch_gui(sys.argv[:1])
    args = ap.parse_args(argv)
    if hasattr(args, "func"):
        return args.func(args)
    return _launch_gui(sys.argv[:1], demo=bool(getattr(args, "demo", False)))


if __name__ == "__main__":
    sys.exit(main())
