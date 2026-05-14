#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
peleng_clone_v2.py -- 1:1 клон PelengPC v1.2 (PyQt6), ультразвуковой дефектоскоп.

Однофайловый Python-скрипт: протокол COM-порта, декодеры TLV, SQLite,
A-scan / B-scan отрисовка, экспорт в Excel, эмулятор прибора.

Использование:
    python peleng_clone_v2.py                   # GUI
    python peleng_clone_v2.py proto-test        # тест порта
    python peleng_clone_v2.py proto-block       # чтение блока
    python peleng_clone_v2.py fake-device       # эмулятор
    python peleng_clone_v2.py demo              # демо-данные

(c) 2024  |  Реверс: PELENG_REVERSE_ROUND5.md  |  www.altek.info
"""

from __future__ import annotations

# ============= 1. IMPORTS AND CONSTANTS =============

import argparse
import datetime as _dt
import math
import os
import sqlite3
import struct
import sys
import time
import threading
import logging
from collections.abc import Callable, Iterator, Sequence
from dataclasses import dataclass, field
from enum import IntEnum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

# Ленивые импорты: PyQt6, serial, openpyxl -- могут отсутствовать
_PYQT_AVAILABLE = False
_SERIAL_AVAILABLE = False
_OPENPYXL_AVAILABLE = False

try:
    import serial
    from serial.tools import list_ports as _list_ports
    _SERIAL_AVAILABLE = True
except ImportError:
    serial = None  # type: ignore
    _list_ports = None  # type: ignore

try:
    from PyQt6.QtWidgets import (
        QApplication, QMainWindow, QTabWidget, QTableWidget, QTableWidgetItem,
        QWidget, QVBoxLayout, QHBoxLayout, QSplitter, QDialog, QDialogButtonBox,
        QLabel, QComboBox, QSpinBox, QLineEdit, QPushButton, QMenuBar, QMenu,
        QToolBar, QStatusBar, QProgressBar, QFileDialog, QMessageBox,
        QHeaderView, QAbstractItemView, QGroupBox, QFormLayout, QTextEdit,
        QProgressDialog
    )
    from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal, QRect, QPoint, QSize, QObject
    from PyQt6.QtGui import (
        QAction, QPixmap, QImage, QPainter, QPen, QBrush, QColor, QFont,
        QPolygonF, QKeySequence, QPainterPath
    )
    from PyQt6.QtCore import QPointF
    _PYQT_AVAILABLE = True
except ImportError:
    pass

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None  # type: ignore


__version__ = "2.0.0"

# --- Протокольные константы (из ROUND5 реверса) ---

# Скорость порта по умолчанию: 19200 (НЕ 9600!)
DEFAULT_BAUD: int = 19200
AUTO_DETECT_BAUDS: Tuple[int, ...] = (19200, 57600)
# Формат: 8 бит данных, EVEN (четная) четность, 1 стоп-бит
PARITY_EVEN: str = "E"
DATA_BITS: int = 8
STOP_BITS: int = 1

# Задержка между байтами при передаче команд
INTER_BYTE_DELAY_PELENG: float = 0.010   # 10 мс (PelengPC)
INTER_BYTE_DELAY_UD2: float = 0.0048     # 4.8 мс (UD2-102)

# Опкоды протокола
OP_HANDSHAKE: int = 0x55       # Хэндшейк, ответ 16 байт
OP_BLOCK_READ: int = 0x42      # Чтение блока: 'B' + LL + HH
OP_FLASH_DUMP: int = 0x74      # Дамп flash-памяти
OP_SCREEN_DUMP: int = 0x9A     # Дамп экрана
OP_KEYBOARD: int = 0x4B        # Клавиатура
OP_SCREEN: int = 0x49          # Экран
OP_RTC_SYNC: int = 0x54        # Синхронизация часов

# Длина ответа на хэндшейк
HANDSHAKE_REPLY_LEN: int = 16
# Максимальный размер ответа при полном приёме (header 16 + каталог до 512KB)
HANDSHAKE_FULL_BUFFER: int = 0x80010  # 524304 байт
# TLV: тег(LE16) + длина(LE16) + тело
TLV_HEADER_SIZE: int = 4
TLV_PADDING_TAG: int = 0xFFFF

# АЦП и физика
ADC_RATE_MHZ: int = 50         # Частота дискретизации 50 МГц
AMPL_FULL_DB: float = 26.0     # Полная шкала амплитуды 0..26 дБ
DEFAULT_VELOCITY_MPS: int = 5900  # Скорость УЗ в стали, м/с

# DLL output format
# [decoder_type(1B), n_fields(1B), offset_table(4B*n: LE16 offset + LE16 size), field_data, raw_record]
# Integer fields: BIG-ENDIAN in output buffer
# String fields: fixed-size, NUL-terminated, zero-padded
# Date fields: 3 bytes (day, month, year%100)

# Категории по sweep_addr / 1000
CATEGORY_DISPATCH: Dict[int, Tuple[str, int]] = {
    0:  ("GENERIC", 1),      # decoder_type=1 (неизвестный тип, но не UNKNOWN)
    1:  ("A_SCAN", 2),       # decoder_type=2
    2:  ("A_SCAN", 2),       # decoder_type=2 (дополнительный A-scan)
    3:  ("SETTINGS", 1),     # decoder_type=1 (настройки прибора)
    4:  ("GENERIC", 1),      # decoder_type=1
    5:  ("CALIBRATION", 1),  # decoder_type=1 (настройки)
    6:  ("GENERIC", 1),      # decoder_type=1
    7:  ("GENERIC", 1),      # decoder_type=1
    8:  ("GENERIC", 1),      # decoder_type=1
    9:  ("GENERIC", 1),      # decoder_type=1
    10: ("B_SCAN", 3),       # decoder_type=3
    11: ("B_SCAN", 3),
    12: ("B_SCAN", 3),
    13: ("B_SCAN", 3),
    14: ("B_SCAN", 3),
    15: ("B_SCAN", 3),
    16: ("B_SCAN", 3),
    17: ("B_SCAN", 3),
    18: ("B_SCAN", 3),
    19: ("B_SCAN", 3),
    20: ("SHORT_PROTO", 4),  # decoder_type=4
    21: ("SHORT_PROTO", 4),
    22: ("SHORT_PROTO", 4),
    23: ("SHORT_PROTO", 4),
    24: ("SHORT_PROTO", 4),
    25: ("SHORT_PROTO", 4),
    26: ("SHORT_PROTO", 4),
    27: ("SHORT_PROTO", 4),
    28: ("SHORT_PROTO", 4),
    29: ("SHORT_PROTO", 4),
}

# PASSPORT_LUT -- 124-символьная таблица подстановки
# Формат: '\x00' + цифры + ' .-/' + кириллица (верх+нижн) + латиница + доп.цифры + padding
PASSPORT_LUT: str = (
    '\x00'
    '1234567890'
    ' .-/'
    '\u0410\u0411\u0412\u0413\u0414\u0415\u0401\u0416\u0417\u0418'
    '\u0419\u041a\u041b\u041c\u041d\u041e\u041f\u0420\u0421\u0422'
    '\u0423\u0424\u0425\u0426\u0427\u0428\u0429\u042a\u042b\u042c'
    '\u042d\u042e\u042f'
    '\u0430\u0431\u0432\u0433\u0434\u0435\u0451\u0436\u0437\u0438'
    '\u0439\u043a\u043b\u043c\u043d\u043e\u043f\u0440\u0441\u0442'
    '\u0443\u0444\u0445\u0446\u0447\u0448\u0449\u044a\u044b\u044c'
    '\u044d\u044e\u044f'
    'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    '0123456789+'
    '\x00' * 6
)

# Смещения в хвосте настроек (Settings tail, от начала body)
SETTINGS_OFFSET_GAIN_DB: int = 0x66         # +0x66: gain_db (u8, dB)
SETTINGS_OFFSET_VELOCITY: int = 0x67        # +0x67: velocity_ms (LE16, m/s)
SETTINGS_OFFSET_ANGLE: int = 0x69           # +0x69: angle_deg (u8, degrees)
SETTINGS_OFFSET_DELAY: int = 0x6A           # +0x6A: delay_us (u8, microseconds)
SETTINGS_OFFSET_GATE_START: int = 0x6E      # +0x6E: gate_start (LE16, ADC counts)
SETTINGS_OFFSET_GATE_WIDTH: int = 0x70      # +0x70: gate_width (LE16, ADC counts)
SETTINGS_OFFSET_THRESHOLD: int = 0x74       # +0x74: threshold_pct (u8, %)
SETTINGS_OFFSET_VRT_TABLE: int = 0x80       # +0x80..+0x87: VRT table (8 bytes)

# RAW Record Body Fields (от начала тела)
BODY_OFFSET_DEVICE_ID: int = 0x00           # +0x00: device_id (LE16)
BODY_OFFSET_VERSION_FLAGS: int = 0x02       # +0x02: version_flags/info_byte
BODY_OFFSET_SWEEP_ID_HI: int = 0x04        # +0x04: sweep_id high (byte*100)
BODY_OFFSET_SWEEP_ID_LO: int = 0x05        # +0x05: sweep_id low (+byte)
BODY_OFFSET_DATE_DAY: int = 0x07           # +0x07: day
BODY_OFFSET_DATE_MONTH: int = 0x08         # +0x08: month
BODY_OFFSET_DATE_YEAR: int = 0x09          # +0x09: year - 2000
BODY_OFFSET_TIME_HOUR: int = 0x0A          # +0x0A: hour
BODY_OFFSET_TIME_MINUTE: int = 0x0B        # +0x0B: minute
BODY_OFFSET_DEFECT_FLAG: int = 0x0C        # +0x0C: defect_flag
# NOTE: sweep_addr occupies +0x10..+0x11 as LE16, and passport_primary starts at +0x11.
# This one-byte overlap is intentional and matches the original DLL behavior: sweep_addr
# is read first (at dispatch time) from offset 0x10 as a 2-byte LE16 for category routing,
# then the same byte range is reinterpreted as part of the passport during field extraction.
# See ROUND5 finding #191 (sweep_addr read) and #327 (passport_primary layout).
BODY_OFFSET_SWEEP_ADDR: int = 0x10         # +0x10..+0x11: block_addr/sweep_addr (LE16)
BODY_OFFSET_PASSPORT_PRIMARY: int = 0x11   # +0x11..+0x1B: passport_primary (11 bytes)
BODY_OFFSET_PASSPORT_SECONDARY: int = 0x21 # +0x21..+0x27: passport_secondary (7 bytes)

# info_byte (handshake response[2]) bit masks
INFO_BIT_EXTENDED_FLASH: int = 0x01    # bit0: extended flash (512KB vs 128KB)
INFO_BIT_WAGON_LCD: int = 0x04         # bit2: wagon LCD mode (5253 vs 4293 bytes)
INFO_BIT_DISTANCE_SCALE: int = 0x08    # bit3: distance scale selection
INFO_BITS_PEP_FREQ: int = 0x30         # bits 4+5: PEP frequency table selection

# Размеры экрана LCD
LCD_SIZE_NORMAL: int = 4293
LCD_SIZE_WAGON: int = 5253

# Максимальный размер payload
MAX_PAYLOAD: int = 0xFFFF

# Logging
LOG = logging.getLogger("peleng_clone_v2")



# ============= 2. PROTOCOL LAYER =============


@dataclass
class PortConfig:
    """Конфигурация COM-порта для связи с прибором Пеленг."""
    port_name: str = "COM3"
    baud_rate: int = DEFAULT_BAUD
    data_bits: int = DATA_BITS
    parity: str = PARITY_EVEN
    stop_bits: int = STOP_BITS
    inter_byte_delay: float = INTER_BYTE_DELAY_PELENG
    timeout_s: float = 2.0
    auto_detect: bool = True


class PelengPort:
    """Низкоуровневый драйвер COM-порта для связи с дефектоскопом Пеленг."""

    def __init__(self, config: PortConfig):
        """Инициализация драйвера порта с заданной конфигурацией."""
        self._config = config
        self._serial: Any = None
        self._lock = threading.Lock()
        self._connected = False

    @property
    def is_open(self) -> bool:
        """Возвращает True если порт открыт."""
        return self._serial is not None and self._serial.is_open

    def open(self) -> bool:
        """Открывает COM-порт с параметрами из конфигурации."""
        if not _SERIAL_AVAILABLE:
            LOG.error("pyserial не установлен")
            return False
        try:
            self._serial = serial.Serial(
                port=self._config.port_name,
                baudrate=self._config.baud_rate,
                bytesize=self._config.data_bits,
                parity=self._config.parity,
                stopbits=self._config.stop_bits,
                timeout=self._config.timeout_s,
            )
            self._connected = True
            LOG.info("Порт %s открыт на скорости %d", self._config.port_name, self._config.baud_rate)
            return True
        except Exception as exc:
            LOG.error("Ошибка открытия порта: %s", exc)
            self._connected = False
            return False

    def close(self) -> None:
        """Закрывает COM-порт."""
        if self._serial and self._serial.is_open:
            self._serial.close()
        self._serial = None
        self._connected = False
        LOG.info("Порт закрыт")

    def _send_byte(self, b: int) -> None:
        """Отправляет один байт с межбайтовой задержкой."""
        if self._serial:
            self._serial.write(bytes([b]))
            time.sleep(self._config.inter_byte_delay)

    def _send_bytes(self, data: bytes) -> None:
        """Отправляет последовательность байт с межбайтовой задержкой."""
        for b in data:
            self._send_byte(b)

    def _read_bytes(self, count: int) -> bytes:
        """Читает заданное количество байт из порта."""
        if self._serial:
            return self._serial.read(count)
        return b""

    def test_port(self) -> Optional[bytes]:
        """Отправляет хэндшейк (0x55) и ожидает 16-байтный ответ."""
        # NOTE: No checksum/CRC validation is performed on received data.
        # The original PelengPC protocol does NOT use CRC or checksums on the wire
        # (per ROUND5 finding #501). Integrity relies on the RS-232 parity bit (8-E-1).
        with self._lock:
            if not self.is_open:
                if not self.open():
                    return None
            self._send_byte(OP_HANDSHAKE)
            reply = self._read_bytes(HANDSHAKE_REPLY_LEN)
            if len(reply) == HANDSHAKE_REPLY_LEN:
                LOG.info("Хэндшейк: получен ответ %d байт", len(reply))
                return reply
            LOG.warning("Хэндшейк: получено %d байт (ожидалось %d)", len(reply), HANDSHAKE_REPLY_LEN)
            return None

    def auto_detect_baud(self) -> Optional[int]:
        """Автоматическое определение скорости порта (19200 и 57600)."""
        for baud in AUTO_DETECT_BAUDS:
            self._config.baud_rate = baud
            if self.is_open:
                self.close()
            if self.open():
                reply = self.test_port()
                if reply is not None:
                    LOG.info("Определена скорость: %d бод", baud)
                    return baud
                self.close()
        return None

    def handshake_full(self) -> Optional[bytes]:
        """Отправляет 0x55 и читает ВСЕ доступные данные (каталог адресов).

        Используется при полном приёме данных с прибора (F5).
        Ответ = 16 байт header + каталог (массив LE16 адресов).
        Вместо блокирующего чтения 524304 байт (ждёт timeout) читаем порциями.
        """
        with self._lock:
            if not self.is_open:
                if not self.open():
                    return None
            # Очищаем буфер перед отправкой
            self._serial.reset_input_buffer()
            self._send_byte(OP_HANDSHAKE)
            # Читаем данные порциями с коротким таймаутом
            # Прибор отвечает быстро, но объём ответа неизвестен заранее
            old_timeout = self._serial.timeout
            self._serial.timeout = 0.5  # 500ms на первый ответ
            chunks = []
            while True:
                chunk = self._serial.read(4096)
                if not chunk:
                    break
                chunks.append(chunk)
                self._serial.timeout = 0.1  # 100ms на последующие порции
            self._serial.timeout = old_timeout
            reply = b''.join(chunks)
            if reply and len(reply) >= HANDSHAKE_REPLY_LEN:
                LOG.info("Полный хэндшейк: получено %d байт", len(reply))
                return reply
            LOG.warning("Полный хэндшейк: получено %d байт", len(reply) if reply else 0)
            return None

    def read_block_by_addr(self, addr: int) -> Optional[bytes]:
        """Запрашивает блок по адресу из каталога: 0x42 + addr_lo + addr_hi.

        Читает 16 байт header + payload (размер из TLV header).
        Возвращает полные сырые данные блока (header + body).
        """
        with self._lock:
            if not self.is_open:
                return None
            addr_lo = addr & 0xFF
            addr_hi = (addr >> 8) & 0xFF
            # Уменьшаем таймаут для быстрого сканирования
            old_timeout = self._serial.timeout
            self._serial.timeout = 0.3  # 300ms достаточно для одного блока
            try:
                self._send_byte(OP_BLOCK_READ)
                self._send_byte(addr_lo)
                self._send_byte(addr_hi)
                # Сначала читаем TLV header (4 байта: tag LE16 + length LE16)
                hdr = self._read_bytes(TLV_HEADER_SIZE)
                if len(hdr) < TLV_HEADER_SIZE:
                    return None
                tag = struct.unpack_from('<H', hdr, 0)[0]
                length = struct.unpack_from('<H', hdr, 2)[0]
                if tag == TLV_PADDING_TAG or length == 0:
                    return None
                body = self._read_bytes(length)
                if len(body) < length:
                    LOG.warning("Блок 0x%04X: получено %d/%d байт", addr, len(body), length)
                return hdr + body
            finally:
                self._serial.timeout = old_timeout

    def request_block(self, block_len: int) -> Optional[bytes]:
        """Запрашивает блок данных: отправляет 0x42 + LL + HH, читает block_len байт."""
        # NOTE: No checksum/CRC validation on received block data - the original PelengPC
        # protocol does not use checksums (per ROUND5 finding #501).
        with self._lock:
            if not self.is_open:
                return None
            ll = block_len & 0xFF
            hh = (block_len >> 8) & 0xFF
            self._send_byte(OP_BLOCK_READ)
            self._send_byte(ll)
            self._send_byte(hh)
            data = self._read_bytes(block_len)
            if len(data) == block_len:
                LOG.info("Блок: получено %d байт", block_len)
                return data
            LOG.warning("Блок: получено %d байт (ожидалось %d)", len(data), block_len)
            return data if data else None

    def request_flash_dump(self) -> Optional[bytes]:
        """Запрашивает дамп flash-памяти (опкод 0x74)."""
        with self._lock:
            if not self.is_open:
                return None
            self._send_byte(OP_FLASH_DUMP)
            # Flash dump может быть 128KB или 512KB
            header = self._read_bytes(4)
            if len(header) < 4:
                return None
            size = struct.unpack_from('<I', header)[0]
            data = self._read_bytes(size)
            return header + data

    def request_screen_dump(self) -> Optional[bytes]:
        """Запрашивает дамп экрана LCD (опкод 0x9A)."""
        with self._lock:
            if not self.is_open:
                return None
            self._send_byte(OP_SCREEN_DUMP)
            # Размер экрана зависит от info_byte
            data = self._read_bytes(LCD_SIZE_WAGON)
            return data if data else None

    def send_keyboard(self, key_code: int) -> None:
        """Отправляет код нажатия клавиши (опкод 0x4B)."""
        with self._lock:
            if not self.is_open:
                return
            self._send_byte(OP_KEYBOARD)
            self._send_byte(key_code)

    def sync_rtc(self, dt: Optional[_dt.datetime] = None) -> None:
        """Синхронизирует часы прибора с текущим временем (опкод 0x54)."""
        with self._lock:
            if not self.is_open:
                return
            if dt is None:
                dt = _dt.datetime.now()
            self._send_byte(OP_RTC_SYNC)
            self._send_byte(dt.second)
            self._send_byte(dt.minute)
            self._send_byte(dt.hour)
            self._send_byte(dt.day)
            self._send_byte(dt.month)
            self._send_byte(dt.year % 100)

    def parse_info_byte(self, reply: bytes) -> Dict[str, Any]:
        """Разбирает info_byte из ответа хэндшейка (байт [2])."""
        if len(reply) < 3:
            return {}
        ib = reply[2]
        return {
            "extended_flash": bool(ib & INFO_BIT_EXTENDED_FLASH),
            "wagon_lcd": bool(ib & INFO_BIT_WAGON_LCD),
            "distance_scale": bool(ib & INFO_BIT_DISTANCE_SCALE),
            "pep_freq_bits": (ib & INFO_BITS_PEP_FREQ) >> 4,
            "flash_size_kb": 512 if (ib & INFO_BIT_EXTENDED_FLASH) else 128,
            "lcd_size": LCD_SIZE_WAGON if (ib & INFO_BIT_WAGON_LCD) else LCD_SIZE_NORMAL,
        }


def parse_catalog(handshake_reply: bytes) -> List[int]:
    """Извлекает список LE16 адресов из ответа на полный хэндшейк.

    Формат ответа: 16 байт header + каталог (массив LE16 адресов).
    catalog_count = (len(reply) - 16) / 2 - 1.
    Отфильтровывает 0xFFFF (padding).
    """
    if len(handshake_reply) <= HANDSHAKE_REPLY_LEN:
        return []
    catalog_bytes = handshake_reply[HANDSHAKE_REPLY_LEN:]
    catalog_count = len(catalog_bytes) // 2 - 1
    if catalog_count <= 0:
        return []
    addresses: List[int] = []
    for i in range(catalog_count):
        addr = struct.unpack_from('<H', catalog_bytes, i * 2)[0]
        if addr != 0xFFFF:
            addresses.append(addr)
    return addresses


class UD2_102Port(PelengPort):
    """Драйвер порта для прибора UD2-102 (модифицированный протокол)."""

    def __init__(self, config: PortConfig):
        """Инициализация драйвера UD2-102 с уменьшенной задержкой."""
        config.inter_byte_delay = INTER_BYTE_DELAY_UD2
        super().__init__(config)
        self._streak_count: int = 0
        self._max_streak: int = 100

    def handshake(self) -> Optional[bytes]:
        """Хэндшейк UD2-102: отправляет 4x 0x55, ожидает ответ."""
        with self._lock:
            if not self.is_open:
                if not self.open():
                    return None
            for _ in range(4):
                self._send_byte(OP_HANDSHAKE)
            reply = self._read_bytes(HANDSHAKE_REPLY_LEN)
            if len(reply) == HANDSHAKE_REPLY_LEN:
                return reply
            return None

    def read_record(self, addr_lo: int, addr_hi: int) -> Optional[bytes]:
        """Чтение одной записи по адресу из flash UD2-102."""
        with self._lock:
            if not self.is_open:
                return None
            self._send_byte(OP_BLOCK_READ)
            self._send_byte(addr_lo)
            self._send_byte(addr_hi)
            # Сначала читаем 4 байта TLV header
            hdr = self._read_bytes(TLV_HEADER_SIZE)
            if len(hdr) < TLV_HEADER_SIZE:
                return None
            tag = struct.unpack_from('<H', hdr, 0)[0]
            length = struct.unpack_from('<H', hdr, 2)[0]
            if tag == TLV_PADDING_TAG or length == 0:
                return None
            body = self._read_bytes(length)
            return hdr + body

    def scan_memory(self, start_addr: int = 0, end_addr: int = 0xFFFF,
                    callback: Optional[Callable[[int, int], None]] = None) -> List[bytes]:
        """Сканирование памяти прибора UD2-102 с контролем непрерывных блоков."""
        records: List[bytes] = []
        self._streak_count = 0
        addr = start_addr
        while addr <= end_addr:
            lo = addr & 0xFF
            hi = (addr >> 8) & 0xFF
            rec = self.read_record(lo, hi)
            if rec is None:
                self._streak_count += 1
                if self._streak_count >= self._max_streak:
                    LOG.info("scan_memory: прервано после %d пустых адресов", self._max_streak)
                    break
            else:
                self._streak_count = 0
                records.append(rec)
            addr += 1
            if callback:
                callback(addr, end_addr)
        return records



# ============= 3. DECODERS AND DATA STRUCTURES =============


class RecordType(IntEnum):
    """Типы записей дефектоскопа."""
    A_SCAN = 1
    B_SCAN = 2
    SETTINGS = 3
    REPORT = 4
    GENERIC = 5
    SHORT_PROTO = 6
    UNKNOWN = 0xFF


@dataclass
class TLVRecord:
    """Структура TLV-записи: тег(LE16) + длина(LE16) + тело."""
    tag: int
    length: int
    body: bytes
    raw: bytes

    @classmethod
    def parse(cls, data: bytes) -> Optional['TLVRecord']:
        """Разбирает TLV-запись из буфера байт."""
        if len(data) < TLV_HEADER_SIZE:
            return None
        tag = struct.unpack_from('<H', data, 0)[0]
        length = struct.unpack_from('<H', data, 2)[0]
        if tag == TLV_PADDING_TAG:
            return None
        if len(data) < TLV_HEADER_SIZE + length:
            return None
        body = data[TLV_HEADER_SIZE:TLV_HEADER_SIZE + length]
        return cls(tag=tag, length=length, body=body, raw=data[:TLV_HEADER_SIZE + length])

    @classmethod
    def parse_all(cls, data: bytes) -> List['TLVRecord']:
        """Разбирает все TLV-записи из блока данных."""
        records: List['TLVRecord'] = []
        offset = 0
        while offset < len(data) - TLV_HEADER_SIZE:
            tag = struct.unpack_from('<H', data, offset)[0]
            if tag == TLV_PADDING_TAG:
                offset += 2
                continue
            length = struct.unpack_from('<H', data, offset + 2)[0]
            if length == 0 or offset + TLV_HEADER_SIZE + length > len(data):
                break
            body = data[offset + TLV_HEADER_SIZE:offset + TLV_HEADER_SIZE + length]
            raw = data[offset:offset + TLV_HEADER_SIZE + length]
            records.append(cls(tag=tag, length=length, body=body, raw=raw))
            offset += TLV_HEADER_SIZE + length
        return records


def get_category(sweep_addr: int) -> Tuple[str, int]:
    """Определяет категорию записи по sweep_addr / 1000."""
    cat_key = sweep_addr // 1000
    return CATEGORY_DISPATCH.get(cat_key, ("UNKNOWN", 0))


def decode_date(body: bytes, offset: int = BODY_OFFSET_DATE_DAY) -> Optional[_dt.date]:
    """Декодирует дату из 3 байт тела записи (день, месяц, год-2000)."""
    if len(body) <= offset + 2:
        return None
    day = body[offset]
    month = body[offset + 1]
    year_raw = body[offset + 2]
    year = 2000 + year_raw
    # Валидация: если дата невалидна, попробовать альтернативный порядок
    if not (1 <= day <= 31 and 1 <= month <= 12 and 2000 < year < 2100):
        # Попытка: может быть порядок (month, day, year) - иногда встречается
        if 1 <= month <= 31 and 1 <= day <= 12:
            day, month = month, day
            year = 2000 + year_raw
        else:
            return None
    try:
        return _dt.date(year, month, day)
    except (ValueError, OverflowError):
        return None


def decode_time(body: bytes, offset: int = BODY_OFFSET_TIME_HOUR) -> Optional[_dt.time]:
    """Декодирует время из 2 байт тела записи (час, минута)."""
    if len(body) <= offset + 1:
        return None
    hour = body[offset]
    minute = body[offset + 1]
    try:
        return _dt.time(hour, minute)
    except (ValueError, OverflowError):
        return None


def decode_passport(raw_bytes: bytes) -> str:
    """Декодирует паспорт через PASSPORT_LUT (124 символа), байты в обратном порядке."""
    result_chars: List[str] = []
    # Итерация в ОБРАТНОМ порядке, пропуск NUL (индекс 0)
    started = False
    for b in reversed(raw_bytes):
        idx = int(b)
        if idx == 0 and not started:
            continue  # пропуск начальных нулей
        started = True
        if 0 <= idx < len(PASSPORT_LUT):
            ch = PASSPORT_LUT[idx]
            if ch == '\x00':
                continue
            result_chars.append(ch)
        else:
            result_chars.append('?')
    return ''.join(result_chars)


def decode_sweep_id(body: bytes) -> int:
    """Декодирует sweep_id из двух байт: byte*100 + byte."""
    if len(body) <= BODY_OFFSET_SWEEP_ID_LO:
        return 0
    hi = body[BODY_OFFSET_SWEEP_ID_HI]
    lo = body[BODY_OFFSET_SWEEP_ID_LO]
    return hi * 100 + lo


def decode_device_id(body: bytes) -> int:
    """Декодирует device_id (LE16) из начала тела записи."""
    if len(body) < 2:
        return 0
    return struct.unpack_from('<H', body, BODY_OFFSET_DEVICE_ID)[0]


def decode_defect_flag(body: bytes) -> int:
    """Декодирует флаг дефекта из тела записи."""
    if len(body) <= BODY_OFFSET_DEFECT_FLAG:
        return 0
    return body[BODY_OFFSET_DEFECT_FLAG]


@dataclass
class UD2_AcousticSettings:
    """Акустические настройки прибора (из хвоста Settings)."""
    gain_db: int = 0
    velocity_mps: int = DEFAULT_VELOCITY_MPS
    angle_deg: int = 0
    delay_us: int = 0
    gate_start: int = 0
    gate_width: int = 0
    threshold_pct: int = 0
    vrt_table: bytes = b'\x00' * 8

    @classmethod
    def from_body(cls, body: bytes) -> 'UD2_AcousticSettings':
        """Извлекает настройки из тела записи по известным смещениям."""
        settings = cls()
        if len(body) > SETTINGS_OFFSET_GAIN_DB:
            settings.gain_db = body[SETTINGS_OFFSET_GAIN_DB]
        if len(body) > SETTINGS_OFFSET_VELOCITY + 1:
            settings.velocity_mps = struct.unpack_from('<H', body, SETTINGS_OFFSET_VELOCITY)[0]
        if len(body) > SETTINGS_OFFSET_ANGLE:
            settings.angle_deg = body[SETTINGS_OFFSET_ANGLE]
        if len(body) > SETTINGS_OFFSET_DELAY:
            settings.delay_us = body[SETTINGS_OFFSET_DELAY]
        if len(body) > SETTINGS_OFFSET_GATE_START + 1:
            settings.gate_start = struct.unpack_from('<H', body, SETTINGS_OFFSET_GATE_START)[0]
        if len(body) > SETTINGS_OFFSET_GATE_WIDTH + 1:
            settings.gate_width = struct.unpack_from('<H', body, SETTINGS_OFFSET_GATE_WIDTH)[0]
        if len(body) > SETTINGS_OFFSET_THRESHOLD:
            settings.threshold_pct = body[SETTINGS_OFFSET_THRESHOLD]
        if len(body) > SETTINGS_OFFSET_VRT_TABLE + 7:
            settings.vrt_table = body[SETTINGS_OFFSET_VRT_TABLE:SETTINGS_OFFSET_VRT_TABLE + 8]
        return settings

    def gate_start_us(self) -> float:
        """Начало строба в микросекундах."""
        return self.gate_start / ADC_RATE_MHZ

    def gate_width_us(self) -> float:
        """Ширина строба в микросекундах."""
        return self.gate_width / ADC_RATE_MHZ

    def gate_start_mm(self) -> float:
        """Начало строба в миллиметрах."""
        return self.gate_start_us() * self.velocity_mps / 2000.0

    def gate_end_mm(self) -> float:
        """Конец строба в миллиметрах."""
        return (self.gate_start_us() + self.gate_width_us()) * self.velocity_mps / 2000.0

    def threshold_db(self) -> float:
        """Порог в дБ."""
        return self.threshold_pct * AMPL_FULL_DB / 100.0


@dataclass
class DecodedRecord:
    """Базовый декодированный объект записи."""
    record_type: RecordType = RecordType.UNKNOWN
    category_name: str = ""
    decoder_type: int = 0
    device_id: int = 0
    sweep_id: int = 0
    sweep_addr: int = 0
    date: Optional[_dt.date] = None
    time: Optional[_dt.time] = None
    defect_flag: int = 0
    passport_primary: str = ""
    passport_secondary: str = ""
    raw_body: bytes = b""
    fields: Dict[str, Any] = field(default_factory=dict)


@dataclass
class AScanRecord(DecodedRecord):
    """Декодированная запись A-скана."""
    samples: bytes = b""
    n_samples: int = 0
    settings: Optional[UD2_AcousticSettings] = None

    @classmethod
    def decode(cls, body: bytes, sweep_addr: int) -> 'AScanRecord':
        """Декодирует A-скан из тела TLV-записи."""
        rec = cls()
        rec.record_type = RecordType.A_SCAN
        rec.category_name = "A_SCAN"
        rec.decoder_type = 2
        rec.sweep_addr = sweep_addr
        rec.device_id = decode_device_id(body)
        rec.sweep_id = decode_sweep_id(body)
        rec.date = decode_date(body)
        rec.time = decode_time(body)
        rec.defect_flag = decode_defect_flag(body)
        # Паспорт
        if len(body) > BODY_OFFSET_PASSPORT_PRIMARY + 10:
            rec.passport_primary = decode_passport(
                body[BODY_OFFSET_PASSPORT_PRIMARY:BODY_OFFSET_PASSPORT_PRIMARY + 11]
            )
        if len(body) > BODY_OFFSET_PASSPORT_SECONDARY + 6:
            rec.passport_secondary = decode_passport(
                body[BODY_OFFSET_PASSPORT_SECONDARY:BODY_OFFSET_PASSPORT_SECONDARY + 7]
            )
        # Настройки
        rec.settings = UD2_AcousticSettings.from_body(body)
        # Отсчеты A-скана идут после основных полей
        sample_offset = 0x88  # после VRT таблицы
        if len(body) > sample_offset:
            rec.samples = body[sample_offset:]
            rec.n_samples = len(rec.samples)
        rec.raw_body = body
        rec.fields = {
            "device_id": rec.device_id,
            "sweep_id": rec.sweep_id,
            "date": str(rec.date) if rec.date else "",
            "time": str(rec.time) if rec.time else "",
            "passport_primary": rec.passport_primary,
            "passport_secondary": rec.passport_secondary,
            "defect_flag": rec.defect_flag,
            "n_samples": rec.n_samples,
            "gain_db": rec.settings.gain_db if rec.settings else 0,
            "velocity_mps": rec.settings.velocity_mps if rec.settings else 0,
            "angle_deg": rec.settings.angle_deg if rec.settings else 0,
            "gate_start": rec.settings.gate_start if rec.settings else 0,
            "gate_width": rec.settings.gate_width if rec.settings else 0,
            "threshold_pct": rec.settings.threshold_pct if rec.settings else 0,
        }
        return rec


@dataclass
class BScanRecord(DecodedRecord):
    """Декодированная запись B-скана."""
    width: int = 0
    height: int = 0
    pixels: bytes = b""
    settings: Optional[UD2_AcousticSettings] = None

    @classmethod
    def decode(cls, body: bytes, sweep_addr: int) -> 'BScanRecord':
        """Декодирует B-скан из тела TLV-записи."""
        rec = cls()
        rec.record_type = RecordType.B_SCAN
        rec.category_name = "B_SCAN"
        rec.decoder_type = 3
        rec.sweep_addr = sweep_addr
        rec.device_id = decode_device_id(body)
        rec.sweep_id = decode_sweep_id(body)
        rec.date = decode_date(body)
        rec.time = decode_time(body)
        rec.defect_flag = decode_defect_flag(body)
        # Паспорт
        if len(body) > BODY_OFFSET_PASSPORT_PRIMARY + 10:
            rec.passport_primary = decode_passport(
                body[BODY_OFFSET_PASSPORT_PRIMARY:BODY_OFFSET_PASSPORT_PRIMARY + 11]
            )
        if len(body) > BODY_OFFSET_PASSPORT_SECONDARY + 6:
            rec.passport_secondary = decode_passport(
                body[BODY_OFFSET_PASSPORT_SECONDARY:BODY_OFFSET_PASSPORT_SECONDARY + 7]
            )
        # Настройки
        rec.settings = UD2_AcousticSettings.from_body(body)
        # B-scan данные: после полей схемы -> LE16 width + LE16 height + pixels
        bscan_offset = 0x88
        if len(body) > bscan_offset + 4:
            rec.width = struct.unpack_from('<H', body, bscan_offset)[0]
            rec.height = struct.unpack_from('<H', body, bscan_offset + 2)[0]
            pixel_start = bscan_offset + 4
            expected = rec.width * rec.height
            if len(body) >= pixel_start + expected:
                rec.pixels = body[pixel_start:pixel_start + expected]
            else:
                rec.pixels = body[pixel_start:]
        rec.raw_body = body
        rec.fields = {
            "device_id": rec.device_id,
            "sweep_id": rec.sweep_id,
            "date": str(rec.date) if rec.date else "",
            "time": str(rec.time) if rec.time else "",
            "passport_primary": rec.passport_primary,
            "passport_secondary": rec.passport_secondary,
            "width": rec.width,
            "height": rec.height,
            "pixel_count": len(rec.pixels),
            "gain_db": rec.settings.gain_db if rec.settings else 0,
            "velocity_mps": rec.settings.velocity_mps if rec.settings else 0,
        }
        return rec


@dataclass
class SettingsRecord(DecodedRecord):
    """Декодированная запись настроек прибора."""
    settings: Optional[UD2_AcousticSettings] = None

    @classmethod
    def decode(cls, body: bytes, sweep_addr: int) -> 'SettingsRecord':
        """Декодирует запись настроек из тела TLV-записи."""
        rec = cls()
        rec.record_type = RecordType.SETTINGS
        rec.category_name = "CALIBRATION"
        rec.decoder_type = 1
        rec.sweep_addr = sweep_addr
        rec.device_id = decode_device_id(body)
        rec.sweep_id = decode_sweep_id(body)
        rec.date = decode_date(body)
        rec.time = decode_time(body)
        rec.settings = UD2_AcousticSettings.from_body(body)
        rec.raw_body = body
        if rec.settings:
            rec.fields = {
                "device_id": rec.device_id,
                "date": str(rec.date) if rec.date else "",
                "time": str(rec.time) if rec.time else "",
                "gain_db": rec.settings.gain_db,
                "velocity_mps": rec.settings.velocity_mps,
                "angle_deg": rec.settings.angle_deg,
                "delay_us": rec.settings.delay_us,
                "gate_start": rec.settings.gate_start,
                "gate_width": rec.settings.gate_width,
                "threshold_pct": rec.settings.threshold_pct,
                "vrt_table": rec.settings.vrt_table.hex(),
            }
        return rec


@dataclass
class ReportRecord(DecodedRecord):
    """Декодированная запись отчета (Generic/ShortProto)."""
    thickness_mm: float = 0.0
    amplitude_db: float = 0.0
    verdict: str = ""

    @classmethod
    def decode(cls, body: bytes, sweep_addr: int) -> 'ReportRecord':
        """Декодирует запись отчета из тела TLV-записи."""
        rec = cls()
        cat_name, dec_type = get_category(sweep_addr)
        rec.record_type = RecordType.REPORT if dec_type == 1 else RecordType.GENERIC
        rec.category_name = cat_name
        rec.decoder_type = dec_type
        rec.sweep_addr = sweep_addr
        rec.device_id = decode_device_id(body)
        rec.sweep_id = decode_sweep_id(body)
        rec.date = decode_date(body)
        rec.time = decode_time(body)
        rec.defect_flag = decode_defect_flag(body)
        # Паспорт
        if len(body) > BODY_OFFSET_PASSPORT_PRIMARY + 10:
            rec.passport_primary = decode_passport(
                body[BODY_OFFSET_PASSPORT_PRIMARY:BODY_OFFSET_PASSPORT_PRIMARY + 11]
            )
        if len(body) > BODY_OFFSET_PASSPORT_SECONDARY + 6:
            rec.passport_secondary = decode_passport(
                body[BODY_OFFSET_PASSPORT_SECONDARY:BODY_OFFSET_PASSPORT_SECONDARY + 7]
            )
        # Толщина и амплитуда из фиксированных смещений (если есть)
        if len(body) > 0x30:
            rec.thickness_mm = struct.unpack_from('<H', body, 0x2E)[0] / 100.0
        if len(body) > 0x32:
            rec.amplitude_db = body[0x32] * AMPL_FULL_DB / 255.0
        # Вердикт
        rec.verdict = compute_verdict_simple(rec.defect_flag)
        rec.raw_body = body
        rec.fields = {
            "device_id": rec.device_id,
            "sweep_id": rec.sweep_id,
            "date": str(rec.date) if rec.date else "",
            "time": str(rec.time) if rec.time else "",
            "passport_primary": rec.passport_primary,
            "passport_secondary": rec.passport_secondary,
            "defect_flag": rec.defect_flag,
            "thickness_mm": rec.thickness_mm,
            "amplitude_db": round(rec.amplitude_db, 2),
            "verdict": rec.verdict,
        }
        return rec


def compute_verdict_simple(defect_flag: int) -> str:
    """Простой вердикт по флагу дефекта."""
    if defect_flag == 0:
        return "ГОДЕН"
    elif defect_flag == 1:
        return "КОНТРОЛЬ"
    else:
        return "БРАК"


def dispatch_decode(tlv: TLVRecord) -> DecodedRecord:
    """Диспетчер декодирования: определяет тип записи и вызывает соответствующий декодер."""
    body = tlv.body
    if len(body) < 0x12:
        # Тело слишком короткое для полного декодирования, но ставим категорию UNKNOWN
        # чтобы запись не потерялась при фильтрации (пустая строка не попадает ни в один фильтр)
        rec = DecodedRecord()
        rec.raw_body = body
        rec.category_name = "UNKNOWN"
        rec.decoder_type = 0
        return rec
    sweep_addr = struct.unpack_from('<H', body, BODY_OFFSET_SWEEP_ADDR)[0]
    cat_name, dec_type = get_category(sweep_addr)
    if dec_type == 2:
        return AScanRecord.decode(body, sweep_addr)
    elif dec_type == 3:
        return BScanRecord.decode(body, sweep_addr)
    elif dec_type == 1 and cat_name == "CALIBRATION":
        return SettingsRecord.decode(body, sweep_addr)
    elif dec_type in (1, 4):
        return ReportRecord.decode(body, sweep_addr)
    else:
        rec = DecodedRecord()
        rec.raw_body = body
        rec.sweep_addr = sweep_addr
        rec.category_name = cat_name
        rec.decoder_type = dec_type
        return rec


# Таблица известных TypeVar значений (из DLL)
TYPEVAR_TABLE: Dict[int, str] = {
    0x0001: "SINGLE_VALUE",
    0x0002: "SETTINGS_BLOCK",
    0x0003: "B_SCAN_DATA",
    0x0004: "A_SCAN_DATA",
    0x0010: "MEASUREMENT_RESULT",
    0x0020: "CALIBRATION_DATA",
    0x0030: "PROTOCOL_HEADER",
    0x0040: "REPORT_FIELDS",
    0x0050: "PASSPORT_DATA",
    0x0060: "OPERATOR_INFO",
    0x0070: "DEVICE_CONFIG",
    0x0080: "FIRMWARE_VER",
}


# Словарь полей тела записи (для отображения в GUI)
PELENG_BODY_FIELDS: Dict[str, Tuple[int, int, str]] = {
    # name: (offset, size_bytes, description)
    "device_id": (0x00, 2, "ID прибора (LE16)"),
    "version_flags": (0x02, 1, "Флаги версии"),
    "info_byte": (0x03, 1, "Информационный байт"),
    "sweep_id_hi": (0x04, 1, "Номер прохода (старш.)"),
    "sweep_id_lo": (0x05, 1, "Номер прохода (младш.)"),
    "reserved_06": (0x06, 1, "Резерв"),
    "date_day": (0x07, 1, "День"),
    "date_month": (0x08, 1, "Месяц"),
    "date_year": (0x09, 1, "Год (- 2000)"),
    "time_hour": (0x0A, 1, "Час"),
    "time_minute": (0x0B, 1, "Минута"),
    "defect_flag": (0x0C, 1, "Флаг дефекта"),
    "reserved_0D": (0x0D, 3, "Резерв"),
    "sweep_addr": (0x10, 2, "Адрес прохода (LE16)"),
    "passport_primary": (0x11, 11, "Паспорт основной (11 байт)"),
    "reserved_1C": (0x1C, 5, "Резерв"),
    "passport_secondary": (0x21, 7, "Паспорт вторичный (7 байт)"),
}



# ============= 4. DATABASE (SQLite) =============


def _strip_unprintable(s: str) -> str:
    """Remove control characters from FDB strings."""
    if not s:
        return s
    out = []
    for ch in s:
        o = ord(ch)
        if o < 0x20 and o not in (0x09, 0x0A, 0x0D):
            continue
        if o == 0x7F:
            continue
        if 0x80 <= o <= 0x9F:
            continue
        out.append(ch)
    return "".join(out).strip()


def _opt_int(v) -> Optional[int]:
    """FDB value to int or None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None


def _opt_str(v) -> Optional[str]:
    """FDB value to string with cp1251 handling."""
    if v is None:
        return None
    if isinstance(v, bytes):
        for enc in ("cp1251", "utf-8"):
            try:
                s = v.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            s = v.hex()
    else:
        s = str(v)
    s = _strip_unprintable(s.rstrip(" \x00"))
    return s if s else None


def _date_iso(v) -> Optional[str]:
    """Convert datetime/date/string to 'YYYY-MM-DD'. Handles DD.MM.YYYY strings."""
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    s = _opt_str(v)
    if not s:
        return None
    # Try DD.MM.YYYY format
    parts = s.strip().split(".")
    if len(parts) == 3 and len(parts[2]) == 4:
        try:
            return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        except Exception:
            pass
    return s


def _fmt_date_display(iso_str: str) -> str:
    """Конвертирует YYYY-MM-DD в DD.MM.YYYY для отображения в таблицах."""
    if not iso_str or len(iso_str) < 10:
        return iso_str
    parts = iso_str[:10].split("-")
    if len(parts) == 3 and len(parts[0]) == 4:
        return f"{parts[2]}.{parts[1]}.{parts[0]}"
    return iso_str


def _fdb_make_year(maketime) -> int:
    """Extract year from MAKETIME integer (YYYYMM or YYYY format)."""
    if not maketime:
        return 0
    v = int(maketime)
    if v >= 190000:  # YYYYMM
        return v // 100
    if 1900 <= v <= 2100:  # plain YYYY
        return v
    return 0


class BlockZapDB:
    """SQLite-хранилище записей дефектоскопа (аналог Firebird BlockZap)."""

    CREATE_SQL = """
    CREATE TABLE IF NOT EXISTS BlockZap (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        block_addr INTEGER,
        sweep_addr INTEGER,
        category TEXT,
        decoder_type INTEGER,
        device_id INTEGER,
        sweep_id INTEGER,
        date_str TEXT,
        time_str TEXT,
        defect_flag INTEGER DEFAULT 0,
        passport_primary TEXT DEFAULT '',
        passport_secondary TEXT DEFAULT '',
        verdict TEXT DEFAULT '',
        thickness_mm REAL DEFAULT 0.0,
        amplitude_db REAL DEFAULT 0.0,
        gain_db INTEGER DEFAULT 0,
        velocity_mps INTEGER DEFAULT 5900,
        angle_deg INTEGER DEFAULT 0,
        gate_start INTEGER DEFAULT 0,
        gate_width INTEGER DEFAULT 0,
        threshold_pct INTEGER DEFAULT 0,
        source TEXT DEFAULT '',
        block_blob BLOB,
        UNIQUE(id)
    );
    CREATE TABLE IF NOT EXISTS Measurements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        addr INTEGER,
        date_str TEXT,
        time_str TEXT,
        type_name TEXT,
        num_obj TEXT DEFAULT '',
        plavka TEXT DEFAULT '',
        stamp TEXT DEFAULT '',
        god TEXT DEFAULT '',
        side TEXT DEFAULT '',
        sheika TEXT DEFAULT '',
        obod TEXT DEFAULT '',
        obtochka TEXT DEFAULT '',
        greben TEXT DEFAULT '',
        naplavka TEXT DEFAULT '',
        source TEXT DEFAULT '',
        defekt TEXT DEFAULT '',
        code_def TEXT DEFAULT '',
        operator TEXT DEFAULT '',
        make_time TEXT DEFAULT '',
        ind_maker TEXT DEFAULT '',
        block_blob BLOB,
        UNIQUE(addr, date_str, time_str, source)
    );
    """

    def __init__(self, db_path: str = ":memory:"):
        """Открывает или создает базу данных SQLite."""
        self._path = db_path
        self._conn = sqlite3.connect(db_path)
        self._conn.row_factory = sqlite3.Row
        self._conn.executescript(self.CREATE_SQL)
        self._conn.commit()
        LOG.info("БД открыта: %s", db_path)

    def close(self) -> None:
        """Закрывает соединение с базой данных."""
        if self._conn:
            self._conn.close()
            self._conn = None  # type: ignore
            LOG.info("БД закрыта")

    def upsert(self, record: DecodedRecord, blob: bytes) -> int:
        """Вставляет или обновляет запись в таблице BlockZap."""
        cursor = self._conn.execute(
            """INSERT OR REPLACE INTO BlockZap
               (block_addr, sweep_addr, category, decoder_type, device_id, sweep_id,
                date_str, time_str, defect_flag, passport_primary, passport_secondary,
                verdict, thickness_mm, amplitude_db, gain_db, velocity_mps,
                angle_deg, gate_start, gate_width, threshold_pct, source, block_blob)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                record.sweep_addr, record.sweep_addr, record.category_name,
                record.decoder_type, record.device_id, record.sweep_id,
                str(record.date) if record.date else "",
                str(record.time) if record.time else "",
                record.defect_flag, record.passport_primary, record.passport_secondary,
                record.fields.get("verdict", ""),
                record.fields.get("thickness_mm", 0.0),
                record.fields.get("amplitude_db", 0.0),
                record.fields.get("gain_db", 0),
                record.fields.get("velocity_mps", DEFAULT_VELOCITY_MPS),
                record.fields.get("angle_deg", 0),
                record.fields.get("gate_start", 0),
                record.fields.get("gate_width", 0),
                record.fields.get("threshold_pct", 0),
                "", blob,
            )
        )
        self._conn.commit()
        return cursor.lastrowid or 0

    def upsert_measurement(self, addr: int, fields: Dict[str, Any], blob: bytes) -> int:
        """Вставляет запись в таблицу Measurements."""
        cursor = self._conn.execute(
            """INSERT OR REPLACE INTO Measurements
               (addr, date_str, time_str, type_name, num_obj, plavka, stamp,
                god, side, sheika, obod, obtochka, greben, naplavka, source,
                defekt, code_def, operator, make_time, ind_maker, block_blob)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                addr,
                fields.get("date_str", ""),
                fields.get("time_str", ""),
                fields.get("type_name", ""),
                fields.get("num_obj", ""),
                fields.get("plavka", ""),
                fields.get("stamp", ""),
                fields.get("god", ""),
                fields.get("side", ""),
                fields.get("sheika", ""),
                fields.get("obod", ""),
                fields.get("obtochka", ""),
                fields.get("greben", ""),
                fields.get("naplavka", ""),
                fields.get("source", ""),
                fields.get("defekt", ""),
                fields.get("code_def", ""),
                fields.get("operator", ""),
                fields.get("make_time", ""),
                fields.get("ind_maker", ""),
                blob,
            )
        )
        self._conn.commit()
        return cursor.lastrowid or 0

    def all_records(self) -> List[Dict[str, Any]]:
        """Возвращает все записи из BlockZap как список словарей."""
        cursor = self._conn.execute("SELECT * FROM BlockZap ORDER BY id")
        rows = cursor.fetchall()
        return [dict(row) for row in rows]

    def all_measurements(self) -> List[Dict[str, Any]]:
        """Возвращает все записи из Measurements как список словарей."""
        cursor = self._conn.execute("SELECT * FROM Measurements ORDER BY id")
        rows = cursor.fetchall()
        return [dict(row) for row in rows]

    def count(self) -> int:
        """Возвращает количество записей в BlockZap."""
        cursor = self._conn.execute("SELECT COUNT(*) FROM BlockZap")
        return cursor.fetchone()[0]

    def has(self, record_id: int) -> bool:
        """Проверяет существование записи по id."""
        cursor = self._conn.execute("SELECT 1 FROM BlockZap WHERE id=?", (record_id,))
        return cursor.fetchone() is not None

    def get_block(self, record_id: int) -> Optional[bytes]:
        """Возвращает blob записи по id."""
        cursor = self._conn.execute("SELECT block_blob FROM BlockZap WHERE id=?", (record_id,))
        row = cursor.fetchone()
        if row:
            return row[0]
        return None

    def get_record_dict(self, record_id: int) -> Optional[Dict[str, Any]]:
        """Возвращает запись как словарь по id."""
        cursor = self._conn.execute("SELECT * FROM BlockZap WHERE id=?", (record_id,))
        row = cursor.fetchone()
        if row:
            return dict(row)
        return None

    def delete(self, record_id: int) -> bool:
        """Удаляет запись по id."""
        self._conn.execute("DELETE FROM BlockZap WHERE id=?", (record_id,))
        self._conn.commit()
        return True

    def filter_by_category(self, categories: List[str]) -> List[Dict[str, Any]]:
        """Фильтрует записи по списку категорий."""
        placeholders = ','.join('?' * len(categories))
        cursor = self._conn.execute(
            f"SELECT * FROM BlockZap WHERE category IN ({placeholders}) ORDER BY id",
            categories
        )
        return [dict(row) for row in cursor.fetchall()]

    def filter_protocols(self) -> List[Dict[str, Any]]:
        """Возвращает записи протоколов (A-scan, B-scan)."""
        return self.filter_by_category(["A_SCAN", "B_SCAN"])

    def filter_reports(self) -> List[Dict[str, Any]]:
        """Возвращает записи отчетов (Generic, ShortProto, а также UNKNOWN и нераспознанные)."""
        return self.filter_by_category(["GENERIC", "SHORT_PROTO", "UNKNOWN", ""])

    def filter_settings(self) -> List[Dict[str, Any]]:
        """Возвращает записи настроек (Calibration, Settings)."""
        return self.filter_by_category(["CALIBRATION", "SETTINGS"])

    def import_fdb(self, fdb_path: str) -> int:
        """Импорт записей из файла FDB (Firebird 2.x база данных PelengPC).

        Стратегия (БЕЗ обязательного fbembed):
        1. Попытка через пакет fdb с автоопределением библиотеки (без load_api)
        2. Бинарное сканирование .fdb файла на TLV-записи напрямую
        3. Парсинг как flash dump (16-byte header + TLV поток)
        """
        if not os.path.isfile(fdb_path):
            LOG.error("Файл не найден: %s", fdb_path)
            return 0

        # Попытка 1: Firebird database через пакет fdb (без обязательного fbembed)
        try:
            import fdb as _fdb
            # Пробуем подключиться БЕЗ явного load_api -- fdb сам найдет
            # fbclient.dll/libfbclient.so если они в PATH/LD_LIBRARY_PATH
            try:
                conn = _fdb.connect(
                    dsn=fdb_path,
                    user='SYSDBA',
                    password='masterkey',
                    charset='WIN1251'
                )
            except Exception:
                # Если не сработало -- попробуем указать путь к embedded
                fb_lib = None
                for candidate in [
                    os.environ.get('PELENG_FBCLIENT', ''),
                    str(Path.home() / '.peleng_clone' / 'fb25' / 'fbembed.dll'),
                    'fbembed.dll', 'fbclient.dll',
                    '/usr/lib/libfbembed.so.2.5',
                    '/usr/lib/x86_64-linux-gnu/libfbclient.so',
                ]:
                    if candidate and os.path.isfile(candidate):
                        fb_lib = candidate
                        break
                if not fb_lib:
                    raise ImportError("fbclient не найден")
                if getattr(_fdb, 'api', None) is None:
                    _fdb.load_api(fb_lib)
                conn = _fdb.connect(
                    dsn=fdb_path,
                    user='SYSDBA',
                    password='masterkey',
                    charset='WIN1251'
                )

            count = 0
            cursor = conn.cursor()
            # Phase A: Read text-field tables (NASTR/RESULTS/SHORTPROT)
            for suffix in ('2', '1', '3'):
                for base_table in ('NASTR', 'RESULTS', 'SHORTPROT'):
                    table_name = f'{base_table}{suffix}'
                    try:
                        cursor.execute(f'SELECT * FROM {table_name}')
                        col_names = [desc[0].upper() for desc in cursor.description]
                        for row in cursor:
                            row_dict = dict(zip(col_names, row))
                            # Handle BlobReader objects
                            for k, val in list(row_dict.items()):
                                if hasattr(val, 'read'):
                                    try:
                                        row_dict[k] = val.read()
                                    except Exception:
                                        row_dict[k] = None

                            number = _opt_int(row_dict.get('NUMBER')) or 0
                            if number <= 0:
                                continue

                            # Extract fields
                            date_iso = _date_iso(row_dict.get('DATEFORM')) or ""
                            time_str = _opt_str(row_dict.get('TIMEFORM')) or ""
                            type_var = _opt_int(row_dict.get('TYPEVAR'))
                            type_zap = _opt_str(row_dict.get('TYPEZAP')) or ""
                            type_name = type_zap or (f"TypeVar={type_var}" if type_var else "")

                            num_obj = _opt_str(row_dict.get('NUMOBJ')) or ""
                            smelting = _opt_str(row_dict.get('SMELTING')) or ""
                            ind_maker = _opt_str(row_dict.get('INDMAKER')) or ""
                            make_time = _opt_int(row_dict.get('MAKETIME'))
                            defekt = _opt_str(row_dict.get('DEFEKT')) or ""
                            code_def = _opt_str(row_dict.get('CODEDEF')) or ""
                            name_opera = _opt_str(row_dict.get('NAMEOPERA')) or ""

                            # Date as "YYYY-MM-DD HH:MM"
                            if date_iso and time_str:
                                date_full = f"{date_iso} {time_str[:5]}"
                            else:
                                date_full = date_iso or time_str or ""

                            # Parse PROTOCOL/BCDDATA blob for side/sheika/etc
                            side = sheika = obod = obtochka = greben = naplavka = ""
                            stamp = 0
                            god = 0
                            bcd_blob = None
                            for blob_key in ('PROTOCOL', 'BCDDATA', 'DATA'):
                                bcd_blob = row_dict.get(blob_key)
                                if isinstance(bcd_blob, (bytes, bytearray)) and len(bcd_blob) >= 70:
                                    break
                                bcd_blob = None
                            if bcd_blob is not None:
                                try:
                                    # BCD passport fields at known offsets in UD2 protocol blob
                                    # Offsets: stamp(2b @40), god(2b @42), side(1b @44),
                                    # sheika(4b @45), obod(4b @49), obtochka(4b @53),
                                    # greben(4b @57), naplavka(4b @61)
                                    stamp_raw = int.from_bytes(bcd_blob[40:42], 'little')
                                    if stamp_raw:
                                        stamp = stamp_raw
                                    god_raw = int.from_bytes(bcd_blob[42:44], 'little')
                                    if god_raw:
                                        god = god_raw
                                    side_b = bcd_blob[44]
                                    if side_b:
                                        side = str(side_b)
                                    sheika_raw = int.from_bytes(bcd_blob[45:49], 'little')
                                    if sheika_raw:
                                        sheika = f"{sheika_raw / 100:.2f}"
                                    obod_raw = int.from_bytes(bcd_blob[49:53], 'little')
                                    if obod_raw:
                                        obod = f"{obod_raw / 100:.2f}"
                                    obtochka_raw = int.from_bytes(bcd_blob[53:57], 'little')
                                    if obtochka_raw:
                                        obtochka = f"{obtochka_raw / 100:.2f}"
                                    greben_raw = int.from_bytes(bcd_blob[57:61], 'little')
                                    if greben_raw:
                                        greben = f"{greben_raw / 100:.2f}"
                                    naplavka_raw = int.from_bytes(bcd_blob[61:65], 'little')
                                    if naplavka_raw:
                                        naplavka = f"{naplavka_raw / 100:.2f}"
                                except (IndexError, ValueError, struct.error):
                                    pass  # blob too short or malformed, keep defaults

                            # Encode table kind in high nibble to avoid address collisions
                            _table_addr_prefix = {
                                'NASTR': 0xF1000000,
                                'RESULTS': 0xF2000000,
                                'SHORTPROT': 0xF3000000,
                            }
                            addr = _table_addr_prefix.get(base_table, 0xF0000000) | number
                            source = f"FDB-{base_table}"

                            fields = {
                                "date_str": date_full,
                                "time_str": time_str,
                                "type_name": type_name,
                                "num_obj": num_obj,
                                "plavka": smelting,
                                "stamp": str(stamp) if stamp else "",
                                "god": str(god) if god else (str(_fdb_make_year(make_time)) if make_time and _fdb_make_year(make_time) else ""),
                                "side": side,
                                "sheika": sheika,
                                "obod": obod,
                                "obtochka": obtochka,
                                "greben": greben,
                                "naplavka": naplavka,
                                "source": source,
                                "defekt": defekt,
                                "code_def": code_def,
                                "operator": name_opera,
                                "make_time": str(make_time) if make_time else "",
                                "ind_maker": ind_maker,
                            }
                            self.upsert_measurement(addr, fields, b"")
                            count += 1
                    except Exception as e:
                        LOG.debug("Table %s: %s", table_name, e)
                        continue

            # Phase B: Read BLOCKZAP blobs for raw TLV decoding
            for suffix in ('', '1', '2', '3'):
                table_name = f'BLOCKZAP{suffix}' if suffix else 'BLOCKZAP'
                try:
                    try:
                        cursor.execute(f'SELECT NUMBER, BLOCKLEN, BLOCK FROM {table_name}')
                    except Exception:
                        # Fallback: old schema may use BLOCK_BLOB instead of BLOCK
                        cursor.execute(f'SELECT * FROM {table_name}')
                    col_names_b = [desc[0].upper() for desc in cursor.description]
                    for row in cursor:
                        row_b = dict(zip(col_names_b, row))
                        num = row_b.get('NUMBER', 0)
                        blob_data = row_b.get('BLOCK') or row_b.get('BLOCK_BLOB')
                        if hasattr(blob_data, 'read'):
                            blob_data = blob_data.read()
                        if blob_data and len(blob_data) > TLV_HEADER_SIZE:
                            if isinstance(blob_data, str):
                                blob_data = blob_data.encode('cp1251')
                            blob_data = bytes(blob_data)
                            tlv = TLVRecord.parse(blob_data)
                            if tlv:
                                decoded = dispatch_decode(tlv)
                                self.upsert(decoded, tlv.raw)
                                count += 1
                except Exception as e:
                    LOG.debug("Table %s: %s", table_name, e)
                    continue
            conn.close()
            LOG.info("FDB импорт (Firebird): %d записей из %s", count, fdb_path)
            return count

        except ImportError:
            LOG.info("Пакет fdb не установлен, пробую бинарное сканирование")
        except Exception as e:
            LOG.info("FDB подключение не удалось (%s), пробую бинарное сканирование", e)

        # Попытка 2: бинарное сканирование .fdb файла на TLV-записи
        # Firebird хранит BLOB данные в страницах "как есть", поэтому TLV-записи
        # можно найти прямым сканированием файла без клиентской библиотеки
        with open(fdb_path, 'rb') as f:
            data = f.read()
        if len(data) < 17:
            LOG.error("Файл слишком мал: %d байт", len(data))
            return 0

        count = self._extract_tlv_from_binary(data)
        if count > 0:
            LOG.info("Бинарное сканирование FDB: %d записей из %s", count, fdb_path)
            return count

        # Попытка 3: парсинг как flash dump (16-байт header + raw TLV поток)
        for skip in (0, 16):
            flash_data = data[skip:]
            tlv_records = TLVRecord.parse_all(flash_data)
            if tlv_records:
                for tlv in tlv_records:
                    decoded = dispatch_decode(tlv)
                    self.upsert(decoded, tlv.raw)
                    count += 1
                break

        if count == 0:
            LOG.warning("Не удалось распознать формат файла %s", fdb_path)
        else:
            LOG.info("Flash dump импорт: %d записей из %s", count, fdb_path)
        return count

    def _extract_tlv_from_binary(self, data: bytes) -> int:
        """Извлечение TLV-записей из бинарного файла FDB прямым сканированием.

        Сканирует файл на наличие валидных TLV-записей без Firebird клиента.
        Работает благодаря тому что BLOB данные хранятся в страницах FDB как есть.
        Проверяет: tag в диапазоне sweep_addr (1000-29999), разумная длина,
        sweep_addr в body совпадает с tag по семейству.
        """
        count = 0
        offset = 0
        file_len = len(data)

        while offset < file_len - TLV_HEADER_SIZE:
            # Читаем потенциальный TLV заголовок
            tag = struct.unpack_from('<H', data, offset)[0]

            # Пропуск padding и нулей
            if tag == TLV_PADDING_TAG or tag == 0:
                offset += 2
                continue

            # tag должен быть валидным sweep_addr (диапазон 1000-29999)
            if not (1000 <= tag <= 29999):
                offset += 1
                continue

            length = struct.unpack_from('<H', data, offset + 2)[0]

            # Длина должна быть разумной: минимум 0x12 (чтобы содержать sweep_addr),
            # максимум 8192 байт (самая большая запись -- B-scan ~4K)
            if length < 0x12 or length > 8192:
                offset += 1
                continue

            # Проверяем что хватает данных
            if offset + TLV_HEADER_SIZE + length > file_len:
                offset += 1
                continue

            body = data[offset + TLV_HEADER_SIZE: offset + TLV_HEADER_SIZE + length]

            # Дополнительная валидация: sweep_addr внутри body (смещение 0x10)
            # должен совпадать с tag или быть в том же семействе (sweep_addr//1000 == tag//1000)
            body_sweep = struct.unpack_from('<H', body, BODY_OFFSET_SWEEP_ADDR)[0]
            if body_sweep != tag and (body_sweep // 1000 != tag // 1000):
                offset += 1
                continue

            # Дополнительная проверка: дата должна быть правдоподобной
            # body[7]=день (1-31), body[8]=месяц (1-12), body[9]=год (0-99)
            day = body[BODY_OFFSET_DATE_DAY] if len(body) > BODY_OFFSET_DATE_DAY else 0
            month = body[BODY_OFFSET_DATE_DAY + 1] if len(body) > BODY_OFFSET_DATE_DAY + 1 else 0
            if not (1 <= day <= 31 and 1 <= month <= 12):
                offset += 1
                continue

            # Все проверки пройдены -- это валидная TLV запись
            raw = data[offset: offset + TLV_HEADER_SIZE + length]
            tlv = TLVRecord(tag=tag, length=length, body=body, raw=raw)
            decoded = dispatch_decode(tlv)
            self.upsert(decoded, tlv.raw)
            count += 1
            offset += TLV_HEADER_SIZE + length

        return count

    def export_to_path(self, out_path: str) -> bool:
        """Экспортирует базу данных в файл SQLite."""
        import shutil
        if self._path == ":memory:":
            # Для in-memory нужен backup
            dst = sqlite3.connect(out_path)
            self._conn.backup(dst)
            dst.close()
        else:
            shutil.copy2(self._path, out_path)
        return True



# ============= 5. A-SCAN FORMULAS =============


def ascan_sample_to_time_us(sample_index: int) -> float:
    """Преобразует индекс отсчета АЦП во время в микросекундах (t_us = idx / 50)."""
    return sample_index / ADC_RATE_MHZ


def ascan_sample_to_depth_mm(sample_index: int, velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Преобразует индекс отсчета в глубину в миллиметрах (d_mm = t_us * v / 2000)."""
    t_us = ascan_sample_to_time_us(sample_index)
    return t_us * velocity_mps / 2000.0


def ascan_byte_to_db(byte_value: int) -> float:
    """Преобразует байт амплитуды в децибелы (db = byte * 26 / 255)."""
    return byte_value * AMPL_FULL_DB / 255.0


def ascan_byte_to_percent(byte_value: int) -> float:
    """Преобразует байт амплитуды в проценты экрана (0..100%)."""
    return byte_value * 100.0 / 255.0


def ascan_gate_to_us(gate_counts: int) -> float:
    """Преобразует строб из отсчетов АЦП в микросекунды."""
    return gate_counts / ADC_RATE_MHZ


def ascan_gate_to_mm(gate_counts: int, velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Преобразует строб из отсчетов АЦП в миллиметры."""
    return ascan_gate_to_us(gate_counts) * velocity_mps / 2000.0


def ascan_threshold_db(threshold_pct: int) -> float:
    """Вычисляет порог в дБ из процентов (thr_db = threshold_pct * 26 / 100)."""
    return threshold_pct * AMPL_FULL_DB / 100.0


def ascan_threshold_byte(threshold_pct: int) -> int:
    """Вычисляет пороговый уровень в единицах АЦП (0..255)."""
    return int(threshold_pct * 255.0 / 100.0)


def ascan_peak_detect(samples: bytes, start: int = 0, end: int = -1) -> Tuple[int, int]:
    """Находит пиковое значение и его индекс в массиве отсчетов (с 3-точечным сглаживанием)."""
    if not samples:
        return (0, 0)
    if end < 0:
        end = len(samples)
    max_val = 0
    max_idx = start
    for i in range(start + 1, min(end - 1, len(samples) - 1)):
        # 3-точечное сглаживание
        smoothed = (samples[i - 1] + samples[i] + samples[i + 1]) / 3.0
        if smoothed > max_val:
            max_val = smoothed
            max_idx = i
    return (int(max_val), max_idx)


def ascan_peak_in_gate(samples: bytes, gate_start: int, gate_width: int) -> Tuple[int, int]:
    """Находит пик в зоне строба."""
    start_sample = gate_start
    end_sample = gate_start + gate_width
    return ascan_peak_detect(samples, start_sample, end_sample)


def ascan_vrt_correction(samples: bytes, vrt_table: bytes) -> List[float]:
    """Применяет таблицу ВРЧ (временной регулировки чувствительности) к отсчетам."""
    if not vrt_table or len(vrt_table) < 8:
        return [float(s) for s in samples]
    n = len(samples)
    corrected: List[float] = []
    # Интерполяция ВРЧ по 8 точкам
    step = n / 7.0 if n > 7 else 1.0
    for i in range(n):
        # Определяем между какими точками ВРЧ находимся
        pos = i / step
        idx_lo = int(pos)
        idx_hi = min(idx_lo + 1, 7)
        frac = pos - idx_lo
        vrt_val = vrt_table[idx_lo] * (1.0 - frac) + vrt_table[idx_hi] * frac
        # ВРЧ добавляет усиление (в dB-подобных единицах)
        correction_factor = 10.0 ** (vrt_val / 20.0 / 10.0) if vrt_val > 0 else 1.0
        corrected.append(float(samples[i]) * correction_factor)
    return corrected


def ascan_compute_duration_us(n_samples: int) -> float:
    """Вычисляет длительность A-скана в микросекундах."""
    return n_samples / ADC_RATE_MHZ


def ascan_compute_max_depth_mm(n_samples: int, velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Вычисляет максимальную глубину A-скана в миллиметрах."""
    duration_us = ascan_compute_duration_us(n_samples)
    return duration_us * velocity_mps / 2000.0


def ascan_vtract_correction(depth_mm: float, radius_mm: float, angle_deg: float) -> float:
    """Коррекция V-тракта для цилиндрической геометрии (закон косинусов)."""
    if radius_mm <= 0 or angle_deg == 0:
        return depth_mm
    angle_rad = math.radians(angle_deg)
    # Закон косинусов: d_corr = sqrt(R^2 + d^2 - 2*R*d*cos(alpha))
    r2 = radius_mm ** 2
    d2 = depth_mm ** 2
    cos_a = math.cos(angle_rad)
    val = r2 + d2 - 2.0 * radius_mm * depth_mm * cos_a
    if val < 0:
        return depth_mm
    return math.sqrt(val)



# ============= 6. B-SCAN FORMULAS =============


def bscan_pixel_to_depth_mm(pixel_y: int, height: int, velocity_mps: int = DEFAULT_VELOCITY_MPS,
                             duration_us: float = 100.0) -> float:
    """Преобразует координату пикселя Y в глубину (мм)."""
    if height <= 0:
        return 0.0
    t_us = (pixel_y / height) * duration_us
    return t_us * velocity_mps / 2000.0


def bscan_pixel_to_position_mm(pixel_x: int, width: int, scan_length_mm: float = 100.0) -> float:
    """Преобразует координату пикселя X в позицию сканирования (мм)."""
    if width <= 0:
        return 0.0
    return (pixel_x / width) * scan_length_mm


def bscan_amplitude_to_db(pixel_byte: int) -> float:
    """Преобразует байт амплитуды B-скана в дБ (аналогично A-скану)."""
    return pixel_byte * AMPL_FULL_DB / 255.0


def bscan_amplitude_to_color(pixel_byte: int) -> Tuple[int, int, int]:
    """Преобразует байт амплитуды в RGB цвет (палитра blue->green->yellow->red)."""
    v = pixel_byte
    if v < 64:
        # Синий -> Голубой
        r = 0
        g = v * 4
        b = 255
    elif v < 128:
        # Голубой -> Зеленый
        r = 0
        g = 255
        b = 255 - (v - 64) * 4
    elif v < 192:
        # Зеленый -> Желтый
        r = (v - 128) * 4
        g = 255
        b = 0
    else:
        # Желтый -> Красный
        r = 255
        g = 255 - (v - 192) * 4
        b = 0
    return (max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b)))


def bscan_build_palette() -> List[Tuple[int, int, int]]:
    """Строит полную палитру из 256 цветов для B-скана."""
    return [bscan_amplitude_to_color(i) for i in range(256)]


def bscan_extract_dimensions(body: bytes, offset: int = 0x88) -> Tuple[int, int]:
    """Извлекает размеры B-скана (LE16 width + LE16 height) из тела записи."""
    if len(body) < offset + 4:
        return (0, 0)
    width = struct.unpack_from('<H', body, offset)[0]
    height = struct.unpack_from('<H', body, offset + 2)[0]
    return (width, height)


def bscan_extract_pixels(body: bytes, offset: int = 0x88) -> Tuple[int, int, bytes]:
    """Извлекает пиксели B-скана из тела записи."""
    width, height = bscan_extract_dimensions(body, offset)
    if width == 0 or height == 0:
        return (0, 0, b"")
    pixel_start = offset + 4
    expected = width * height
    pixels = body[pixel_start:pixel_start + expected]
    return (width, height, pixels)


def lcd_bitmap_to_pixels(lcd_data: bytes, width: int = 240, stride: int = 30) -> List[List[int]]:
    """Преобразует LCD-битмап (LSB-first, stride=30 байт, 240px) в матрицу пикселей."""
    rows: List[List[int]] = []
    offset = 0
    while offset + stride <= len(lcd_data):
        row: List[int] = []
        for byte_idx in range(stride):
            b = lcd_data[offset + byte_idx]
            for bit in range(8):  # LSB-first
                if len(row) < width:
                    row.append(1 if (b >> bit) & 1 else 0)
        rows.append(row)
        offset += stride
    return rows


# Предвычисленная палитра
BSCAN_PALETTE: List[Tuple[int, int, int]] = bscan_build_palette()



# ============= 7. GUI MAIN WINDOW =============


# ARD Verdict formula (from ROUND5)
def compute_ard_verdict(a_peak: float, a_gate: float, s_fact: float, s_req: float) -> Tuple[str, float]:
    """Вычисляет вердикт по формуле АРД: delta_db = 20*log10(a_peak/a_gate) + (s_fact - s_req).
    
    Возвращает (вердикт, delta_db):
      delta >= 0: БРАК (красный)
      -6 <= delta < 0: КОНТРОЛЬ (желтый)
      delta < -6: ГОДЕН/ПОИСК (зеленый)
    """
    if a_gate <= 0 or a_peak <= 0:
        return ("НЕТ ДАННЫХ", 0.0)
    delta_db = 20.0 * math.log10(a_peak / a_gate) + (s_fact - s_req)
    if delta_db >= 0:
        return ("БРАК", delta_db)
    elif delta_db >= -6.0:
        return ("КОНТРОЛЬ", delta_db)
    else:
        return ("ГОДЕН", delta_db)


def _create_gui_classes():
    """Создает классы GUI только при наличии PyQt6 (ленивая инициализация)."""
    if not _PYQT_AVAILABLE:
        return None

    class PelengMainWindow(QMainWindow):
        """Главное окно приложения PelengPC Clone v2."""

        def __init__(self):
            """Инициализация главного окна с вкладками и меню."""
            super().__init__()
            self.setWindowTitle("PelengPC Clone v2.0 -- Ультразвуковой дефектоскоп")
            self.setMinimumSize(1024, 700)
            self._db: Optional[BlockZapDB] = None
            self._port: Optional[PelengPort] = None
            self._port_config = PortConfig()

            self._init_ui()
            self._init_menu()
            self._init_toolbar()
            self._init_statusbar()

            # Открываем БД в памяти по умолчанию
            self._db = BlockZapDB()

        def _init_ui(self) -> None:
            """Инициализация основного интерфейса с тремя вкладками."""
            central = QWidget()
            self.setCentralWidget(central)
            layout = QVBoxLayout(central)

            # Три вкладки как в оригинальном PelengPC
            self._tabs = QTabWidget()
            layout.addWidget(self._tabs)

            # Вкладка 1: Протоколы (A-scan, B-scan)
            self._tab_protocols = QWidget()
            self._init_tab_protocols()
            self._tabs.addTab(self._tab_protocols, "\u041f\u0440\u043e\u0442\u043e\u043a\u043e\u043b\u044b")

            # Вкладка 2: Отчеты (Generic, ShortProto)
            self._tab_reports = QWidget()
            self._init_tab_reports()
            self._tabs.addTab(self._tab_reports, "\u041e\u0442\u0447\u0451\u0442\u044b")

            # Вкладка 3: Настройки (Calibration)
            self._tab_settings = QWidget()
            self._init_tab_settings()
            self._tabs.addTab(self._tab_settings, "\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438")

        def _init_tab_protocols(self) -> None:
            """Инициализация вкладки Протоколы с таблицей записей A-scan и B-scan."""
            layout = QVBoxLayout(self._tab_protocols)
            self._tbl_protocols = QTableWidget()
            _vagon_cols = [
                "\u0410\u0434\u0440\u0435\u0441", "\u0414\u0430\u0442\u0430",
                "\u0412\u0440\u0435\u043c\u044f", "\u0422\u0438\u043f",
                "\u2116 \u043e\u0431\u044a\u0435\u043a\u0442\u0430",
                "\u041f\u043b\u0430\u0432\u043a\u0430",
                "\u041a\u043b\u0435\u0439\u043c\u043e",
                "\u0413\u043e\u0434",
                "\u0421\u0442\u043e\u0440\u043e\u043d\u0430",
                "\u0428\u0435\u0439\u043a\u0430",
                "\u041e\u0431\u043e\u0434",
                "\u041e\u0431\u0442\u043e\u0447\u043a\u0430",
                "\u0413\u0440\u0435\u0431\u0435\u043d\u044c",
                "\u041d\u0430\u043f\u043b\u0430\u0432\u043a\u0430",
                "\u0417\u0430\u0432\u043e\u0434 (\u0438\u043d\u0434.)",
                "\u0414\u0430\u0442\u0430 \u0438\u0437\u0433.",
                "\u0414\u0435\u0444\u0435\u043a\u0442",
                "\u041a\u043e\u0434 \u0434\u0435\u0444.",
                "\u041e\u043f\u0435\u0440\u0430\u0442\u043e\u0440",
                "\u0418\u0441\u0442\u043e\u0447\u043d\u0438\u043a",
            ]
            self._tbl_protocols.setColumnCount(len(_vagon_cols))
            self._tbl_protocols.setHorizontalHeaderLabels(_vagon_cols)
            self._tbl_protocols.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            self._tbl_protocols.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            self._tbl_protocols.horizontalHeader().setStretchLastSection(True)
            self._tbl_protocols.doubleClicked.connect(self._on_protocol_double_click)
            layout.addWidget(self._tbl_protocols)

        def _init_tab_reports(self) -> None:
            """Инициализация вкладки Отчеты с таблицей отчетных записей."""
            layout = QVBoxLayout(self._tab_reports)
            self._tbl_reports = QTableWidget()
            _vagon_cols = [
                "\u0410\u0434\u0440\u0435\u0441", "\u0414\u0430\u0442\u0430",
                "\u0412\u0440\u0435\u043c\u044f", "\u0422\u0438\u043f",
                "\u2116 \u043e\u0431\u044a\u0435\u043a\u0442\u0430",
                "\u041f\u043b\u0430\u0432\u043a\u0430",
                "\u041a\u043b\u0435\u0439\u043c\u043e",
                "\u0413\u043e\u0434",
                "\u0421\u0442\u043e\u0440\u043e\u043d\u0430",
                "\u0428\u0435\u0439\u043a\u0430",
                "\u041e\u0431\u043e\u0434",
                "\u041e\u0431\u0442\u043e\u0447\u043a\u0430",
                "\u0413\u0440\u0435\u0431\u0435\u043d\u044c",
                "\u041d\u0430\u043f\u043b\u0430\u0432\u043a\u0430",
                "\u0417\u0430\u0432\u043e\u0434 (\u0438\u043d\u0434.)",
                "\u0414\u0430\u0442\u0430 \u0438\u0437\u0433.",
                "\u0414\u0435\u0444\u0435\u043a\u0442",
                "\u041a\u043e\u0434 \u0434\u0435\u0444.",
                "\u041e\u043f\u0435\u0440\u0430\u0442\u043e\u0440",
                "\u0418\u0441\u0442\u043e\u0447\u043d\u0438\u043a",
            ]
            self._tbl_reports.setColumnCount(len(_vagon_cols))
            self._tbl_reports.setHorizontalHeaderLabels(_vagon_cols)
            self._tbl_reports.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            self._tbl_reports.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            self._tbl_reports.horizontalHeader().setStretchLastSection(True)
            self._tbl_reports.doubleClicked.connect(self._on_report_double_click)
            layout.addWidget(self._tbl_reports)

        def _init_tab_settings(self) -> None:
            """Инициализация вкладки Настройки с таблицей параметров прибора."""
            layout = QVBoxLayout(self._tab_settings)
            self._tbl_settings = QTableWidget()
            _vagon_cols = [
                "\u0410\u0434\u0440\u0435\u0441", "\u0414\u0430\u0442\u0430",
                "\u0412\u0440\u0435\u043c\u044f", "\u0422\u0438\u043f",
                "\u2116 \u043e\u0431\u044a\u0435\u043a\u0442\u0430",
                "\u041f\u043b\u0430\u0432\u043a\u0430",
                "\u041a\u043b\u0435\u0439\u043c\u043e",
                "\u0413\u043e\u0434",
                "\u0421\u0442\u043e\u0440\u043e\u043d\u0430",
                "\u0428\u0435\u0439\u043a\u0430",
                "\u041e\u0431\u043e\u0434",
                "\u041e\u0431\u0442\u043e\u0447\u043a\u0430",
                "\u0413\u0440\u0435\u0431\u0435\u043d\u044c",
                "\u041d\u0430\u043f\u043b\u0430\u0432\u043a\u0430",
                "\u0417\u0430\u0432\u043e\u0434 (\u0438\u043d\u0434.)",
                "\u0414\u0430\u0442\u0430 \u0438\u0437\u0433.",
                "\u0414\u0435\u0444\u0435\u043a\u0442",
                "\u041a\u043e\u0434 \u0434\u0435\u0444.",
                "\u041e\u043f\u0435\u0440\u0430\u0442\u043e\u0440",
                "\u0418\u0441\u0442\u043e\u0447\u043d\u0438\u043a",
            ]
            self._tbl_settings.setColumnCount(len(_vagon_cols))
            self._tbl_settings.setHorizontalHeaderLabels(_vagon_cols)
            self._tbl_settings.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            self._tbl_settings.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            self._tbl_settings.horizontalHeader().setStretchLastSection(True)
            self._tbl_settings.doubleClicked.connect(self._on_settings_double_click)
            layout.addWidget(self._tbl_settings)

        def _refresh_tables(self) -> None:
            """Обновляет все таблицы из базы данных."""
            if not self._db:
                return
            self._refresh_protocols_table()
            self._refresh_reports_table()
            self._refresh_settings_table()

        def _refresh_protocols_table(self) -> None:
            """Обновляет таблицу протоколов из БД (все записи BlockZap + FDB-RESULTS)."""
            if not self._db:
                return
            DASH = "\u2014"
            rows = []
            # ВСЕ записи из BlockZap (от прибора через UART или TLV-парсинг)
            for rec in self._db.all_records():
                rows.append({
                    "addr": rec.get("block_addr", rec.get("sweep_addr", rec.get("id", 0))),
                    "date_str": rec.get("date_str", ""),
                    "time_str": rec.get("time_str", ""),
                    "type_name": rec.get("category", ""),
                    "num_obj": rec.get("passport_primary", ""),
                    "plavka": "",
                    "stamp": "",
                    "god": "",
                    "side": "",
                    "sheika": "",
                    "obod": "",
                    "obtochka": "",
                    "greben": "",
                    "naplavka": "",
                    "ind_maker": "",
                    "make_time": "",
                    "defekt": rec.get("verdict", ""),
                    "code_def": "",
                    "operator": rec.get("passport_secondary", ""),
                    "source": "UART",
                })
            # Записи FDB-RESULTS из Measurements
            for m in self._db.all_measurements():
                src = m.get("source", "")
                if src == "FDB-RESULTS":
                    rows.append(m)
            self._tbl_protocols.setSortingEnabled(False)
            self._tbl_protocols.setRowCount(len(rows))
            for i, r in enumerate(rows):
                addr = r.get("addr", 0)
                source = r.get("source", "")
                if source.startswith("FDB-") and isinstance(addr, int) and addr >= 0xF0000000:
                    num = addr & 0x0FFFFFFF
                    addr_str = f"Results#{num}"
                else:
                    addr_str = f"0x{addr:04x}" if isinstance(addr, int) else str(addr)
                cells = [
                    addr_str,
                    _fmt_date_display(r.get("date_str", "") or "") or DASH,
                    r.get("time_str", "") or DASH,
                    r.get("type_name", "") or DASH,
                    r.get("num_obj", "") or DASH,
                    r.get("plavka", "") or DASH,
                    r.get("stamp", "") or DASH,
                    r.get("god", "") if r.get("god", "") not in ("", "0") else DASH,
                    r.get("side", "") or DASH,
                    r.get("sheika", "") or DASH,
                    r.get("obod", "") or DASH,
                    r.get("obtochka", "") or DASH,
                    r.get("greben", "") or DASH,
                    r.get("naplavka", "") or DASH,
                    r.get("ind_maker", "") or DASH,
                    _fmt_date_display(r.get("make_time", "") or "") or DASH,
                    r.get("defekt", "") or DASH,
                    r.get("code_def", "") or DASH,
                    r.get("operator", "") or DASH,
                    r.get("source", "") or DASH,
                ]
                for j, cell in enumerate(cells):
                    item = QTableWidgetItem(str(cell))
                    if j == 0:
                        item.setData(Qt.ItemDataRole.UserRole, r)
                    self._tbl_protocols.setItem(i, j, item)
            self._tbl_protocols.setSortingEnabled(True)
            self._tbl_protocols.resizeColumnsToContents()

        def _refresh_reports_table(self) -> None:
            """Обновляет таблицу отчетов из БД (FDB-SHORTPROT/UD2 + BlockZap SHORT_PROTO/GENERIC/UNKNOWN)."""
            if not self._db:
                return
            DASH = "\u2014"
            rows = []
            # Записи FDB-SHORTPROT и UD2 из Measurements
            for m in self._db.all_measurements():
                src = m.get("source", "")
                if src in ("FDB-SHORTPROT", "UD2"):
                    rows.append(m)
            # Записи SHORT_PROTO/GENERIC/UNKNOWN из BlockZap
            for rec in self._db.filter_reports():
                rows.append({
                    "addr": rec.get("block_addr", rec.get("id", 0)),
                    "date_str": rec.get("date_str", ""),
                    "time_str": rec.get("time_str", ""),
                    "type_name": rec.get("category", ""),
                    "num_obj": rec.get("passport_primary", ""),
                    "plavka": "",
                    "stamp": "",
                    "god": "",
                    "side": "",
                    "sheika": "",
                    "obod": "",
                    "obtochka": "",
                    "greben": "",
                    "naplavka": "",
                    "ind_maker": "",
                    "make_time": "",
                    "defekt": rec.get("verdict", ""),
                    "code_def": "",
                    "operator": rec.get("passport_secondary", ""),
                    "source": "UART",
                })
            self._tbl_reports.setSortingEnabled(False)
            self._tbl_reports.setRowCount(len(rows))
            for i, r in enumerate(rows):
                addr = r.get("addr", 0)
                source = r.get("source", "")
                if source.startswith("FDB-") and isinstance(addr, int) and addr >= 0xF0000000:
                    num = addr & 0x0FFFFFFF
                    kind_short = source[4:].title()
                    addr_str = f"{kind_short}#{num}"
                else:
                    addr_str = f"0x{addr:04x}" if isinstance(addr, int) else str(addr)
                cells = [
                    addr_str,
                    _fmt_date_display(r.get("date_str", "") or "") or DASH,
                    r.get("time_str", "") or DASH,
                    r.get("type_name", "") or DASH,
                    r.get("num_obj", "") or DASH,
                    r.get("plavka", "") or DASH,
                    r.get("stamp", "") or DASH,
                    r.get("god", "") if r.get("god", "") not in ("", "0") else DASH,
                    r.get("side", "") or DASH,
                    r.get("sheika", "") or DASH,
                    r.get("obod", "") or DASH,
                    r.get("obtochka", "") or DASH,
                    r.get("greben", "") or DASH,
                    r.get("naplavka", "") or DASH,
                    r.get("ind_maker", "") or DASH,
                    _fmt_date_display(r.get("make_time", "") or "") or DASH,
                    r.get("defekt", "") or DASH,
                    r.get("code_def", "") or DASH,
                    r.get("operator", "") or DASH,
                    r.get("source", "") or DASH,
                ]
                for j, cell in enumerate(cells):
                    item = QTableWidgetItem(str(cell))
                    if j == 0:
                        item.setData(Qt.ItemDataRole.UserRole, r)
                    self._tbl_reports.setItem(i, j, item)
            self._tbl_reports.setSortingEnabled(True)
            self._tbl_reports.resizeColumnsToContents()

        def _refresh_settings_table(self) -> None:
            """Обновляет таблицу настроек из БД (FDB-NASTR + BlockZap CALIBRATION/SETTINGS)."""
            if not self._db:
                return
            DASH = "\u2014"
            rows = []
            # Записи FDB-NASTR из Measurements
            for m in self._db.all_measurements():
                if m.get("source") == "FDB-NASTR":
                    rows.append(m)
            # Записи CALIBRATION/SETTINGS из BlockZap
            for rec in self._db.filter_settings():
                rows.append({
                    "addr": rec.get("block_addr", rec.get("id", 0)),
                    "date_str": rec.get("date_str", ""),
                    "time_str": rec.get("time_str", ""),
                    "type_name": rec.get("category", ""),
                    "num_obj": rec.get("passport_primary", ""),
                    "plavka": "",
                    "stamp": "",
                    "god": "",
                    "side": "",
                    "sheika": "",
                    "obod": "",
                    "obtochka": "",
                    "greben": "",
                    "naplavka": "",
                    "ind_maker": "",
                    "make_time": "",
                    "defekt": rec.get("verdict", ""),
                    "code_def": "",
                    "operator": rec.get("passport_secondary", ""),
                    "source": "UART",
                })
            self._tbl_settings.setSortingEnabled(False)
            self._tbl_settings.setRowCount(len(rows))
            for i, r in enumerate(rows):
                addr = r.get("addr", 0)
                source = r.get("source", "")
                if source.startswith("FDB-") and isinstance(addr, int) and addr >= 0xF0000000:
                    num = addr & 0x0FFFFFFF
                    addr_str = f"Nastr#{num}"
                else:
                    addr_str = f"0x{addr:04x}" if isinstance(addr, int) else str(addr)
                cells = [
                    addr_str,
                    _fmt_date_display(r.get("date_str", "") or "") or DASH,
                    r.get("time_str", "") or DASH,
                    r.get("type_name", "") or DASH,
                    r.get("num_obj", "") or DASH,
                    r.get("plavka", "") or DASH,
                    r.get("stamp", "") or DASH,
                    r.get("god", "") if r.get("god", "") not in ("", "0") else DASH,
                    r.get("side", "") or DASH,
                    r.get("sheika", "") or DASH,
                    r.get("obod", "") or DASH,
                    r.get("obtochka", "") or DASH,
                    r.get("greben", "") or DASH,
                    r.get("naplavka", "") or DASH,
                    r.get("ind_maker", "") or DASH,
                    _fmt_date_display(r.get("make_time", "") or "") or DASH,
                    r.get("defekt", "") or DASH,
                    r.get("code_def", "") or DASH,
                    r.get("operator", "") or DASH,
                    r.get("source", "") or DASH,
                ]
                for j, cell in enumerate(cells):
                    item = QTableWidgetItem(str(cell))
                    if j == 0:
                        item.setData(Qt.ItemDataRole.UserRole, r)
                    self._tbl_settings.setItem(i, j, item)
            self._tbl_settings.setSortingEnabled(True)
            self._tbl_settings.resizeColumnsToContents()

        # Placeholder methods referenced by menu/toolbar -- implemented in block 8 and 11
        def _on_protocol_double_click(self, index) -> None:
            """Обработчик двойного клика по строке таблицы протоколов."""
            row = index.row()
            _handle_protocol_double_click(self, row)

        def _on_report_double_click(self, index) -> None:
            """Обработчик двойного клика по строке таблицы отчетов."""
            row = index.row()
            _handle_report_double_click(self, row)

        def _on_settings_double_click(self, index) -> None:
            """Обработчик двойного клика по строке таблицы настроек."""
            row = index.row()
            _handle_settings_double_click(self, row)

        def _init_menu(self) -> None:
            """Инициализация главного меню приложения."""
            _setup_menu(self)

        def _init_toolbar(self) -> None:
            """Инициализация панели инструментов."""
            _setup_toolbar(self)

        def _init_statusbar(self) -> None:
            """Инициализация строки состояния."""
            self._statusbar = QStatusBar()
            self.setStatusBar(self._statusbar)
            self._statusbar.showMessage("\u0413\u043e\u0442\u043e\u0432")

        def _action_receive(self) -> None:
            """Действие: Принять данные от прибора (F5) -- полный приём по каталогу."""
            # Открываем порт при необходимости
            if not self._port:
                self._port = PelengPort(self._port_config)
            if not self._port.is_open:
                if not self._port.open():
                    self._statusbar.showMessage("Ошибка открытия порта")
                    return
            # Отправляем 0x55 и читаем полный ответ (header + каталог)
            reply = self._port.handshake_full()
            if reply is None or len(reply) < 17:
                QMessageBox.warning(self, "Приём", "Нет ответа от прибора")
                self._statusbar.showMessage("Нет ответа от прибора")
                return
            # Парсим info_byte
            info_byte = reply[2] if len(reply) > 2 else 0
            LOG.info("Приём: info_byte=0x%02X, всего %d байт", info_byte, len(reply))
            # Извлекаем каталог адресов
            catalog_addrs = parse_catalog(reply)
            if not catalog_addrs:
                QMessageBox.information(self, "Приём",
                                        "Каталог пуст -- нет записей в приборе")
                self._statusbar.showMessage("Каталог пуст")
                return
            self._statusbar.showMessage(
                f"Чтение каталога... {len(catalog_addrs)} адресов"
            )
            # Открываем диалог прогресса приёма
            self._recv_dialog = ReceiveProgressDialog(self)
            self._recv_dialog.set_status(
                f"Чтение каталога... {len(catalog_addrs)} адресов"
            )
            self._recv_dialog.show()
            # Создаем QThread и воркер для фонового чтения блоков
            self._recv_thread = QThread()
            self._recv_worker = ReceiveWorker(self._port, catalog_addrs)
            self._recv_worker.moveToThread(self._recv_thread)
            # Соединяем сигналы воркера с обработчиками
            self._recv_worker.progress.connect(self._recv_dialog.update_progress)
            self._recv_worker.status.connect(self._recv_dialog.set_status)
            self._recv_worker.record_found.connect(self._recv_dialog.add_record)
            self._recv_worker.record_found.connect(self._on_receive_record_found)
            self._recv_worker.finished.connect(self._on_receive_finished)
            self._recv_worker.error.connect(self._on_receive_error)
            # Кнопка "Остановить" устанавливает флаг отмены в воркере
            self._recv_dialog.btn_stop.clicked.connect(
                self._recv_worker.abort, Qt.ConnectionType.DirectConnection)
            # Запуск фонового потока
            self._recv_thread.started.connect(self._recv_worker.run)
            self._recv_thread.start()

        def _on_receive_record_found(self, raw: bytes, decoded_info: str) -> None:
            """Обработка найденной записи при приёме: вставка в БД."""
            if self._db:
                tlv = TLVRecord.parse(raw)
                if tlv:
                    decoded = dispatch_decode(tlv)
                    self._db.upsert(decoded, tlv.raw)

        def _on_receive_finished(self, count: int) -> None:
            """Завершение приёма: закрытие диалога и обновление таблиц."""
            if hasattr(self, '_recv_dialog') and self._recv_dialog:
                self._recv_dialog.close()
                self._recv_dialog = None
            if hasattr(self, '_recv_thread') and self._recv_thread:
                self._recv_thread.quit()
                self._recv_thread.wait()
                self._recv_thread = None
            self._recv_worker = None
            self._refresh_tables()
            self._statusbar.showMessage(f"Приём завершён: получено {count} записей")

        def _on_receive_error(self, msg: str) -> None:
            """Обработка ошибки приёма."""
            if hasattr(self, '_recv_dialog') and self._recv_dialog:
                self._recv_dialog.close()
                self._recv_dialog = None
            if hasattr(self, '_recv_thread') and self._recv_thread:
                self._recv_thread.quit()
                self._recv_thread.wait()
                self._recv_thread = None
            self._recv_worker = None
            QMessageBox.critical(self, "Приём", f"Ошибка приёма: {msg}")

        def _action_test_port(self) -> None:
            """Действие: Тест порта (F6)."""
            if not self._port:
                self._port = PelengPort(self._port_config)
            result = self._port.test_port()
            if result:
                QMessageBox.information(self, "\u0422\u0435\u0441\u0442 \u043f\u043e\u0440\u0442\u0430",
                                        f"\u041f\u043e\u0440\u0442 OK. \u041e\u0442\u0432\u0435\u0442: {result.hex()}")
            else:
                QMessageBox.warning(self, "\u0422\u0435\u0441\u0442 \u043f\u043e\u0440\u0442\u0430",
                                    "\u041d\u0435\u0442 \u043e\u0442\u0432\u0435\u0442\u0430 \u043e\u0442 \u043f\u0440\u0438\u0431\u043e\u0440\u0430")

        def _action_port_settings(self) -> None:
            """Действие: Настройка COM-порта."""
            # Используем реальный диалог PortDialogImpl вместо заглушки
            _, _, PortDialogReal = _create_dialog_classes()
            if PortDialogReal is None:
                return
            dlg = PortDialogReal(self, self._port_config)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                self._port_config = dlg.get_config()
                # Закрываем старый порт, чтобы при следующем использовании
                # он был пересоздан с новыми настройками
                if self._port is not None:
                    try:
                        self._port.close()
                    except Exception:
                        pass
                    self._port = None
                self._statusbar.showMessage(
                    f"\u041f\u043e\u0440\u0442: {self._port_config.port_name} @ {self._port_config.baud_rate}")

        def _action_demo_block(self) -> None:
            """Действие: Загрузка демо-блока (F8)."""
            demo_data = generate_demo_data()
            if self._db:
                for tlv in TLVRecord.parse_all(demo_data):
                    decoded = dispatch_decode(tlv)
                    self._db.upsert(decoded, tlv.raw)
                self._refresh_tables()
                self._statusbar.showMessage("\u0414\u0435\u043c\u043e-\u0434\u0430\u043d\u043d\u044b\u0435 \u0437\u0430\u0433\u0440\u0443\u0436\u0435\u043d\u044b")

        def _action_open_db(self) -> None:
            """Действие: Открыть файл БД."""
            path, _ = QFileDialog.getOpenFileName(self, "\u041e\u0442\u043a\u0440\u044b\u0442\u044c \u0411\u0414",
                                                   "", "SQLite (*.db *.sqlite);;All (*)")
            if path:
                if self._db:
                    self._db.close()
                self._db = BlockZapDB(path)
                self._refresh_tables()
                self._statusbar.showMessage(f"\u0411\u0414 \u043e\u0442\u043a\u0440\u044b\u0442\u0430: {path}")

        def _action_save_as(self) -> None:
            """Действие: Сохранить БД как..."""
            path, _ = QFileDialog.getSaveFileName(self, "\u0421\u043e\u0445\u0440\u0430\u043d\u0438\u0442\u044c \u0411\u0414",
                                                   "", "SQLite (*.db);;All (*)")
            if path and self._db:
                self._db.export_to_path(path)
                self._statusbar.showMessage(f"\u0411\u0414 \u0441\u043e\u0445\u0440\u0430\u043d\u0435\u043d\u0430: {path}")

        def _action_delete_record(self) -> None:
            """Действие: Удалить выбранную запись."""
            current_tab = self._tabs.currentIndex()
            if current_tab == 0:
                tbl = self._tbl_protocols
            elif current_tab == 1:
                tbl = self._tbl_reports
            else:
                tbl = self._tbl_settings
            row = tbl.currentRow()
            if row < 0:
                return
            id_item = tbl.item(row, 0)
            if id_item and self._db:
                rec_id = int(id_item.text())
                self._db.delete(rec_id)
                self._refresh_tables()

        def _action_import_fdb(self) -> None:
            """Действие: Импорт из FDB файла (Firebird БД или бинарный дамп)."""
            path, _ = QFileDialog.getOpenFileName(self, "\u0418\u043c\u043f\u043e\u0440\u0442 FDB",
                                                   str(Path.home()),
                                                   "Firebird DB (*.fdb *.FDB);;Flash dump (*.dat *.bin);;All (*)")
            if path and self._db:
                self._statusbar.showMessage("\u0418\u043c\u043f\u043e\u0440\u0442...")
                QApplication.processEvents()
                count = self._db.import_fdb(path)
                self._refresh_tables()
                if count > 0:
                    self._statusbar.showMessage(f"\u0418\u043c\u043f\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u0430\u043d\u043e {count} \u0437\u0430\u043f\u0438\u0441\u0435\u0439 \u0438\u0437 {Path(path).name}")
                    QMessageBox.information(self, "\u0418\u043c\u043f\u043e\u0440\u0442 FDB",
                        f"\u0423\u0441\u043f\u0435\u0448\u043d\u043e \u0438\u043c\u043f\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u0430\u043d\u043e {count} \u0437\u0430\u043f\u0438\u0441\u0435\u0439")
                else:
                    self._statusbar.showMessage("\u0418\u043c\u043f\u043e\u0440\u0442: \u0437\u0430\u043f\u0438\u0441\u0438 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u044b")
                    QMessageBox.warning(self, "\u0418\u043c\u043f\u043e\u0440\u0442 FDB",
                        "\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u0438\u043c\u043f\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u0430\u0442\u044c \u0437\u0430\u043f\u0438\u0441\u0438.\n\n"
                        "\u0412\u043e\u0437\u043c\u043e\u0436\u043d\u044b\u0435 \u043f\u0440\u0438\u0447\u0438\u043d\u044b:\n"
                        "\u2022 \u0424\u0430\u0439\u043b \u043d\u0435 \u0441\u043e\u0434\u0435\u0440\u0436\u0438\u0442 \u0437\u0430\u043f\u0438\u0441\u0435\u0439 \u0411\u041b\u041e\u041a\u0417\u0410\u041f\n"
                        "\u2022 \u0424\u0430\u0439\u043b \u043f\u0443\u0441\u0442 \u0438\u043b\u0438 \u043f\u043e\u0432\u0440\u0435\u0436\u0434\u0451\u043d\n"
                        "\u2022 \u0424\u043e\u0440\u043c\u0430\u0442 \u043d\u0435 \u0440\u0430\u0441\u043f\u043e\u0437\u043d\u0430\u043d (Firebird 2.x / flash dump)")

        def _action_export_excel(self) -> None:
            """Действие: Экспорт в Excel."""
            if not _OPENPYXL_AVAILABLE:
                QMessageBox.warning(self, "\u042d\u043a\u0441\u043f\u043e\u0440\u0442",
                                    "\u041d\u0435\u043e\u0431\u0445\u043e\u0434\u0438\u043c\u043e \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u0438\u0442\u044c openpyxl:\npip install openpyxl")
                return
            path, _ = QFileDialog.getSaveFileName(self, "\u042d\u043a\u0441\u043f\u043e\u0440\u0442 Excel",
                                                   "", "Excel (*.xlsx);;All (*)")
            if path and self._db:
                _export_to_excel(self._db, path)
                self._statusbar.showMessage(f"\u042d\u043a\u0441\u043f\u043e\u0440\u0442: {path}")

        def _action_about(self) -> None:
            """Действие: О программе."""
            QMessageBox.about(self, "\u041e \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u0435",
                              "PelengPC Clone v2.0\n\n"
                              "\u041a\u043b\u043e\u043d \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b PelengPC v1.2\n"
                              "\u0434\u043b\u044f \u0443\u043b\u044c\u0442\u0440\u0430\u0437\u0432\u0443\u043a\u043e\u0432\u043e\u0433\u043e \u0434\u0435\u0444\u0435\u043a\u0442\u043e\u0441\u043a\u043e\u043f\u0430.\n\n"
                              "www.altek.info\n\n"
                              "\u041f\u0440\u043e\u0442\u043e\u043a\u043e\u043b: 19200/57600, 8-E-1\n"
                              "\u0410\u0426\u041f: 50 \u041c\u0413\u0446, \u0448\u043a\u0430\u043b\u0430 0..26 \u0434\u0411")

        def _action_ud2_handshake(self) -> None:
            """Действие: Хэндшейк UD2-102."""
            ud2 = UD2_102Port(self._port_config)
            reply = ud2.handshake()
            if reply:
                QMessageBox.information(self, "UD2-102",
                                        f"\u041e\u0442\u0432\u0435\u0442: {reply.hex()}")
            else:
                QMessageBox.warning(self, "UD2-102", "\u041d\u0435\u0442 \u043e\u0442\u0432\u0435\u0442\u0430")

        def _action_ud2_scan(self) -> None:
            """Действие: Сканирование памяти UD2-102 (асинхронно через QThread)."""
            ud2 = UD2_102Port(self._port_config)
            if not ud2.is_open:
                if not ud2.open():
                    QMessageBox.warning(self, "UD2-102", "Не удалось открыть порт")
                    return

            # Создаем диалог прогресса сканирования
            self._scan_dialog = ScanProgressDialog(self)
            self._scan_dialog.show()

            # Создаем QThread и воркер для фонового сканирования
            self._scan_thread = QThread()
            self._scan_worker = ScanWorker(ud2, start_addr=0, end_addr=0xFFFF)
            self._scan_worker.moveToThread(self._scan_thread)

            # Соединяем сигналы воркера с обработчиками
            self._scan_worker.progress.connect(self._scan_dialog.update_progress)
            self._scan_worker.record_found.connect(self._scan_dialog.add_record)
            self._scan_worker.record_found.connect(self._on_scan_record_found)
            self._scan_worker.finished.connect(self._on_scan_finished)
            self._scan_worker.error.connect(self._on_scan_error)

            # Кнопка "Остановить" устанавливает флаг отмены в воркере
            self._scan_dialog.btn_stop.clicked.connect(
                self._scan_worker.abort, Qt.ConnectionType.DirectConnection)

            # Запуск фонового потока
            self._scan_thread.started.connect(self._scan_worker.run)
            self._scan_thread.start()

        def _on_scan_record_found(self, raw: bytes, decoded_info: str) -> None:
            """Обработка найденной записи: вставка в БД."""
            if self._db:
                tlv = TLVRecord.parse(raw)
                if tlv:
                    decoded = dispatch_decode(tlv)
                    self._db.upsert(decoded, tlv.raw)

        def _on_scan_finished(self, count: int) -> None:
            """Завершение сканирования: закрытие диалога и обновление таблиц."""
            if hasattr(self, '_scan_dialog') and self._scan_dialog:
                self._scan_dialog.close()
                self._scan_dialog = None
            if hasattr(self, '_scan_thread') and self._scan_thread:
                self._scan_thread.quit()
                self._scan_thread.wait()
                self._scan_thread = None
            self._scan_worker = None
            self._refresh_tables()
            self._statusbar.showMessage(f"UD2-102: получено {count} записей")

        def _on_scan_error(self, msg: str) -> None:
            """Обработка ошибки сканирования."""
            if hasattr(self, '_scan_dialog') and self._scan_dialog:
                self._scan_dialog.close()
                self._scan_dialog = None
            if hasattr(self, '_scan_thread') and self._scan_thread:
                self._scan_thread.quit()
                self._scan_thread.wait()
                self._scan_thread = None
            self._scan_worker = None
            QMessageBox.critical(self, "UD2-102", f"Ошибка сканирования: {msg}")

        def _action_edit_settings(self) -> None:
            """Действие: Редактирование настроек."""
            QMessageBox.information(self, "\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438",
                                    "\u0420\u0435\u0434\u0430\u043a\u0442\u043e\u0440 \u043d\u0430\u0441\u0442\u0440\u043e\u0435\u043a \u0432 \u0440\u0430\u0437\u0440\u0430\u0431\u043e\u0442\u043a\u0435")


    # --------- ScanWorker: фоновый поток сканирования памяти ---------

    class ScanWorker(QObject):
        """Воркер для сканирования памяти UD2-102 в фоновом потоке."""

        # Сигналы для обратной связи с GUI
        progress = pyqtSignal(int, int)        # (текущий_адрес, конечный_адрес)
        record_found = pyqtSignal(bytes, str)  # (сырые_данные, расшифрованная_информация)
        finished = pyqtSignal(int)             # количество найденных записей
        error = pyqtSignal(str)                # сообщение об ошибке

        def __init__(self, ud2: UD2_102Port, start_addr: int = 0,
                     end_addr: int = 0xFFFF):
            """Инициализация воркера сканирования."""
            super().__init__()
            self._ud2 = ud2
            self._start_addr = start_addr
            self._end_addr = end_addr
            self._abort_event = threading.Event()
            self._max_streak = 100

        def abort(self) -> None:
            """Установка флага прерывания сканирования (потокобезопасно)."""
            self._abort_event.set()

        def run(self) -> None:
            """Основной цикл сканирования (выполняется в фоновом потоке)."""
            try:
                count = 0
                streak = 0
                addr = self._start_addr
                while addr <= self._end_addr:
                    # Проверяем флаг отмены
                    if self._abort_event.is_set():
                        break

                    lo = addr & 0xFF
                    hi = (addr >> 8) & 0xFF
                    rec = self._ud2.read_record(lo, hi)

                    if rec is None:
                        streak += 1
                        if streak >= self._max_streak:
                            break
                    else:
                        streak = 0
                        count += 1
                        # Декодируем для отображения в таблице
                        decoded_info = self._make_decoded_info(rec)
                        self.record_found.emit(rec, decoded_info)

                    addr += 1
                    # Отправляем прогресс каждые 16 адресов для снижения нагрузки
                    if addr % 16 == 0:
                        self.progress.emit(addr, self._end_addr)

                self.finished.emit(count)
            except Exception as e:
                self.error.emit(str(e))

        def _make_decoded_info(self, raw: bytes) -> str:
            """Формирование строки с информацией о записи для отображения."""
            try:
                tlv = TLVRecord.parse(raw)
                if tlv is None:
                    return ""
                decoded = dispatch_decode(tlv)
                parts = []
                if decoded.category_name:
                    parts.append(decoded.category_name)
                if decoded.date:
                    parts.append(decoded.date.strftime("%d.%m.%Y"))
                if decoded.passport_primary:
                    parts.append(decoded.passport_primary)
                return "|".join(parts)
            except Exception:
                return ""

    # --------- ScanProgressDialog: диалог прогресса сканирования ---------

    class ScanProgressDialog(QDialog):
        """Диалог отображения прогресса сканирования памяти UD2-102."""

        def __init__(self, parent=None):
            """Инициализация диалога прогресса с таблицей найденных записей."""
            super().__init__(parent)
            self.setWindowTitle("Сканирование памяти UD2-102")
            self.setMinimumSize(700, 400)
            self.setModal(True)
            self._record_count = 0
            self._init_ui()

        def _init_ui(self) -> None:
            """Создание элементов интерфейса диалога."""
            layout = QVBoxLayout(self)

            # Метка текущего адреса
            self._lbl_status = QLabel("Адрес: 0x0000 / 0xFFFF")
            layout.addWidget(self._lbl_status)

            # Прогресс-бар
            self._progress_bar = QProgressBar()
            self._progress_bar.setMinimum(0)
            self._progress_bar.setMaximum(0xFFFF)
            self._progress_bar.setValue(0)
            layout.addWidget(self._progress_bar)

            # Метка количества найденных записей
            self._lbl_count = QLabel("Найдено записей: 0")
            layout.addWidget(self._lbl_count)

            # Таблица найденных записей
            self._table = QTableWidget()
            self._table.setColumnCount(5)
            self._table.setHorizontalHeaderLabels([
                "№", "Адрес", "Тип", "Дата", "Паспорт"
            ])
            self._table.horizontalHeader().setStretchLastSection(True)
            layout.addWidget(self._table)

            # Кнопка "Остановить"
            self.btn_stop = QPushButton("Остановить")
            layout.addWidget(self.btn_stop)

        def update_progress(self, current_addr: int, end_addr: int) -> None:
            """Обновление прогресс-бара и метки адреса."""
            self._progress_bar.setMaximum(end_addr)
            self._progress_bar.setValue(current_addr)
            self._lbl_status.setText(
                f"Адрес: 0x{current_addr:04X} / 0x{end_addr:04X}"
            )

        def add_record(self, raw: bytes, decoded_info: str) -> None:
            """Добавление найденной записи в таблицу."""
            self._record_count += 1
            self._lbl_count.setText(f"Найдено записей: {self._record_count}")

            row = self._table.rowCount()
            self._table.insertRow(row)
            self._table.setItem(row, 0, QTableWidgetItem(str(self._record_count)))

            # Адрес из сырых данных (первые 4 байта TLV: tag + length)
            addr_str = f"0x{len(raw):04X}" if len(raw) < 8 else f"{len(raw)} B"
            self._table.setItem(row, 1, QTableWidgetItem(addr_str))

            # Разбираем decoded_info (формат: "тип|дата|паспорт")
            parts = decoded_info.split("|") if decoded_info else []
            rec_type = parts[0] if len(parts) > 0 else ""
            rec_date = parts[1] if len(parts) > 1 else ""
            rec_passport = parts[2] if len(parts) > 2 else ""

            self._table.setItem(row, 2, QTableWidgetItem(rec_type))
            self._table.setItem(row, 3, QTableWidgetItem(rec_date))
            self._table.setItem(row, 4, QTableWidgetItem(rec_passport))

            # Прокрутка к последней строке
            self._table.scrollToBottom()

    # --------- ReceiveWorker: фоновый поток приёма данных (F5) ---------

    class ReceiveWorker(QObject):
        """Воркер для полного приёма данных с прибора в фоновом потоке."""

        # Сигналы для обратной связи с GUI
        progress = pyqtSignal(int, int)        # (текущий_индекс, общее_количество)
        status = pyqtSignal(str)               # текстовое сообщение статуса
        record_found = pyqtSignal(bytes, str)  # (сырые_данные, расшифрованная_информация)
        finished = pyqtSignal(int)             # количество прочитанных записей
        error = pyqtSignal(str)                # сообщение об ошибке

        def __init__(self, port: PelengPort, catalog_addrs: List[int]):
            """Инициализация воркера приёма данных по каталогу."""
            super().__init__()
            self._port = port
            self._catalog_addrs = catalog_addrs
            self._abort_event = threading.Event()

        def abort(self) -> None:
            """Установка флага прерывания приёма (потокобезопасно)."""
            self._abort_event.set()

        def run(self) -> None:
            """Основной цикл чтения блоков по адресам каталога."""
            try:
                total = len(self._catalog_addrs)
                count = 0
                for idx, addr in enumerate(self._catalog_addrs):
                    if self._abort_event.is_set():
                        break
                    self.status.emit(
                        f"Чтение блока {idx + 1}/{total}: 0x{addr:04X}"
                    )
                    self.progress.emit(idx + 1, total)
                    # Читаем блок по адресу
                    raw = self._port.read_block_by_addr(addr)
                    if raw is not None and len(raw) > TLV_HEADER_SIZE:
                        count += 1
                        decoded_info = self._make_decoded_info(raw)
                        self.record_found.emit(raw, decoded_info)

                self.finished.emit(count)
            except Exception as e:
                self.error.emit(str(e))

        def _make_decoded_info(self, raw: bytes) -> str:
            """Формирование строки с информацией о записи для отображения."""
            try:
                tlv = TLVRecord.parse(raw)
                if tlv is None:
                    return ""
                decoded = dispatch_decode(tlv)
                parts = []
                if decoded.category_name:
                    parts.append(decoded.category_name)
                if decoded.date:
                    parts.append(decoded.date.strftime("%d.%m.%Y"))
                if decoded.passport_primary:
                    parts.append(decoded.passport_primary)
                return "|".join(parts)
            except Exception:
                return ""

    # --------- ReceiveProgressDialog: диалог прогресса приёма данных ---------

    class ReceiveProgressDialog(QDialog):
        """Диалог отображения прогресса приёма данных от прибора."""

        def __init__(self, parent=None):
            """Инициализация диалога прогресса приёма."""
            super().__init__(parent)
            self.setWindowTitle("Приём данных от прибора")
            self.setMinimumSize(700, 400)
            self.setModal(True)
            self._record_count = 0
            self._init_ui()

        def _init_ui(self) -> None:
            """Создание элементов интерфейса диалога приёма."""
            layout = QVBoxLayout(self)

            # Метка статуса
            self._lbl_status = QLabel("Чтение каталога...")
            layout.addWidget(self._lbl_status)

            # Прогресс-бар
            self._progress_bar = QProgressBar()
            self._progress_bar.setMinimum(0)
            self._progress_bar.setMaximum(100)
            self._progress_bar.setValue(0)
            layout.addWidget(self._progress_bar)

            # Метка количества найденных записей
            self._lbl_count = QLabel("Найдено записей: 0")
            layout.addWidget(self._lbl_count)

            # Таблица найденных записей
            self._table = QTableWidget()
            self._table.setColumnCount(5)
            self._table.setHorizontalHeaderLabels([
                "\u2116", "Адрес", "Тип", "Дата", "Паспорт"
            ])
            self._table.horizontalHeader().setStretchLastSection(True)
            layout.addWidget(self._table)

            # Кнопка "Остановить"
            self.btn_stop = QPushButton("Остановить")
            layout.addWidget(self.btn_stop)

        def set_status(self, text: str) -> None:
            """Обновление текста статуса."""
            self._lbl_status.setText(text)

        def update_progress(self, current: int, total: int) -> None:
            """Обновление прогресс-бара."""
            self._progress_bar.setMaximum(total)
            self._progress_bar.setValue(current)
            self._lbl_status.setText(
                f"Чтение блока {current}/{total}"
            )

        def add_record(self, raw: bytes, decoded_info: str) -> None:
            """Добавление найденной записи в таблицу."""
            self._record_count += 1
            self._lbl_count.setText(f"Найдено записей: {self._record_count}")

            row = self._table.rowCount()
            self._table.insertRow(row)
            self._table.setItem(row, 0, QTableWidgetItem(str(self._record_count)))

            # Адрес из размера данных
            addr_str = f"{len(raw)} B"
            self._table.setItem(row, 1, QTableWidgetItem(addr_str))

            # Разбираем decoded_info (формат: "тип|дата|паспорт")
            parts = decoded_info.split("|") if decoded_info else []
            rec_type = parts[0] if len(parts) > 0 else ""
            rec_date = parts[1] if len(parts) > 1 else ""
            rec_passport = parts[2] if len(parts) > 2 else ""

            self._table.setItem(row, 2, QTableWidgetItem(rec_type))
            self._table.setItem(row, 3, QTableWidgetItem(rec_date))
            self._table.setItem(row, 4, QTableWidgetItem(rec_passport))

            # Прокрутка к последней строке
            self._table.scrollToBottom()


    return PelengMainWindow



# ============= 8. DOUBLE-CLICK HANDLERS =============


def _handle_protocol_double_click(window, row: int) -> None:
    """Обработчик двойного клика по строке протоколов: открывает граф + расшифровку."""
    if not _PYQT_AVAILABLE:
        return
    if not window._db:
        return
    id_item = window._tbl_protocols.item(row, 0)
    if not id_item:
        return
    rec_id = int(id_item.text())
    blob = window._db.get_block(rec_id)
    if not blob:
        return
    # Разбираем TLV
    tlv = TLVRecord.parse(blob)
    if not tlv:
        return
    decoded = dispatch_decode(tlv)
    # Открываем комбинированное окно граф+расшифровка
    dlg = GraphDecryptionWindow(window, decoded)
    dlg.exec()


def _handle_report_double_click(window, row: int) -> None:
    """Обработчик двойного клика по строке отчетов: открывает расшифровку полей."""
    if not _PYQT_AVAILABLE:
        return
    if not window._db:
        return
    id_item = window._tbl_reports.item(row, 0)
    if not id_item:
        return
    rec_id = int(id_item.text())
    blob = window._db.get_block(rec_id)
    if not blob:
        return
    tlv = TLVRecord.parse(blob)
    if not tlv:
        return
    decoded = dispatch_decode(tlv)
    # Открываем диалог расшифровки
    dlg = DecryptionDialog(window, decoded)
    dlg.exec()


def _handle_settings_double_click(window, row: int) -> None:
    """Обработчик двойного клика по строке настроек: детальный просмотр."""
    if not _PYQT_AVAILABLE:
        return
    if not window._db:
        return
    id_item = window._tbl_settings.item(row, 0)
    if not id_item:
        return
    rec_id = int(id_item.text())
    blob = window._db.get_block(rec_id)
    if not blob:
        return
    tlv = TLVRecord.parse(blob)
    if not tlv:
        return
    decoded = dispatch_decode(tlv)
    # Открываем детальный диалог настроек
    dlg = SettingsDetailDialog(window, decoded)
    dlg.exec()



# ============= 9. A-SCAN WIDGET =============


class AScanWidget:
    """Виджет отрисовки A-скана (создается только при наличии PyQt6)."""
    pass


def _create_ascan_widget_class():
    """Создает класс виджета A-скана при наличии PyQt6."""
    if not _PYQT_AVAILABLE:
        return None

    class AScanWidgetImpl(QWidget):
        """Виджет отрисовки A-скана: амплитуда vs время/глубина."""

        def __init__(self, parent=None, record: Optional[AScanRecord] = None):
            """Инициализация виджета A-скана с данными записи."""
            super().__init__(parent)
            self._record = record
            self._samples: bytes = record.samples if record else b""
            self._settings = record.settings if record else None
            self._velocity = self._settings.velocity_mps if self._settings else DEFAULT_VELOCITY_MPS
            self._n_samples = len(self._samples)
            self._margin_left = 60
            self._margin_right = 20
            self._margin_top = 30
            self._margin_bottom = 40
            self.setMinimumSize(400, 300)

        def paintEvent(self, event) -> None:
            """Отрисовка A-скана: сетка, трасса, строб, порог, метки."""
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            w = self.width()
            h = self.height()

            # Фон
            painter.fillRect(0, 0, w, h, QColor(20, 20, 40))

            # Область графика
            plot_x = self._margin_left
            plot_y = self._margin_top
            plot_w = w - self._margin_left - self._margin_right
            plot_h = h - self._margin_top - self._margin_bottom

            if plot_w <= 0 or plot_h <= 0:
                painter.end()
                return

            # Рамка графика
            painter.setPen(QPen(QColor(100, 100, 100), 1))
            painter.drawRect(plot_x, plot_y, plot_w, plot_h)

            # Сетка
            painter.setPen(QPen(QColor(50, 50, 70), 1, Qt.PenStyle.DotLine))
            for i in range(1, 10):
                gx = plot_x + int(plot_w * i / 10)
                painter.drawLine(gx, plot_y, gx, plot_y + plot_h)
            for i in range(1, 5):
                gy = plot_y + int(plot_h * i / 5)
                painter.drawLine(plot_x, gy, plot_x + plot_w, gy)

            # Строб (если есть настройки)
            if self._settings and self._settings.gate_width > 0:
                gate_x1 = plot_x + int(plot_w * self._settings.gate_start / max(self._n_samples, 1))
                gate_x2 = plot_x + int(plot_w * (self._settings.gate_start + self._settings.gate_width) / max(self._n_samples, 1))
                gate_x2 = min(gate_x2, plot_x + plot_w)
                painter.fillRect(gate_x1, plot_y, gate_x2 - gate_x1, plot_h,
                                 QColor(0, 100, 0, 40))
                painter.setPen(QPen(QColor(0, 200, 0), 1, Qt.PenStyle.DashLine))
                painter.drawLine(gate_x1, plot_y, gate_x1, plot_y + plot_h)
                painter.drawLine(gate_x2, plot_y, gate_x2, plot_y + plot_h)

            # Порог
            if self._settings and self._settings.threshold_pct > 0:
                thr_y = plot_y + plot_h - int(plot_h * self._settings.threshold_pct / 100)
                painter.setPen(QPen(QColor(255, 255, 0), 1, Qt.PenStyle.DashDotLine))
                painter.drawLine(plot_x, thr_y, plot_x + plot_w, thr_y)

            # Трасса A-скана
            if self._samples and self._n_samples > 1:
                painter.setPen(QPen(QColor(0, 255, 0), 1))
                points: List[QPointF] = []
                step = max(1, self._n_samples // plot_w)
                for i in range(0, self._n_samples, step):
                    x = plot_x + int(plot_w * i / self._n_samples)
                    amp_pct = self._samples[i] / 255.0
                    y = plot_y + plot_h - int(plot_h * amp_pct)
                    points.append(QPointF(x, y))

                # Заполнение под кривой (как в оригинале)
                if points:
                    fill_path = QPainterPath()
                    fill_path.moveTo(QPointF(plot_x, plot_y + plot_h))
                    for pt in points:
                        fill_path.lineTo(pt)
                    fill_path.lineTo(QPointF(points[-1].x(), plot_y + plot_h))
                    fill_path.closeSubpath()
                    painter.setBrush(QBrush(QColor(0, 200, 0, 60)))
                    painter.setPen(Qt.PenStyle.NoPen)
                    painter.drawPath(fill_path)

                    # Линия трассы
                    painter.setBrush(Qt.BrushStyle.NoBrush)
                    painter.setPen(QPen(QColor(0, 255, 0), 1))
                    for i in range(len(points) - 1):
                        painter.drawLine(points[i].toPoint(), points[i + 1].toPoint())

            # Оси: X - время (мкс) / глубина (мм)
            painter.setPen(QPen(QColor(200, 200, 200)))
            font = QFont("Monospace", 8)
            painter.setFont(font)
            duration_us = ascan_compute_duration_us(self._n_samples)
            max_depth = ascan_compute_max_depth_mm(self._n_samples, self._velocity)
            for i in range(0, 11, 2):
                x = plot_x + int(plot_w * i / 10)
                t = duration_us * i / 10
                d = max_depth * i / 10
                painter.drawText(x - 15, plot_y + plot_h + 12, f"{t:.0f}")
                painter.drawText(x - 15, plot_y + plot_h + 25, f"{d:.1f}")

            # Ось Y - амплитуда (%) и дБ
            for i in range(0, 6):
                y = plot_y + plot_h - int(plot_h * i / 5)
                pct = i * 20
                db = pct * AMPL_FULL_DB / 100
                painter.drawText(plot_x - 55, y + 4, f"{pct}% {db:.0f}\u0434\u0411")

            # Заголовок
            painter.setPen(QPen(QColor(255, 255, 255)))
            title_font = QFont("Monospace", 10, QFont.Weight.Bold)
            painter.setFont(title_font)
            title = "A-\u0441\u043a\u0430\u043d"
            if self._record:
                title += f"  |  {self._record.passport_primary}  |  {self._record.date}"
            painter.drawText(plot_x, plot_y - 10, title)

            # Подписи осей
            painter.setPen(QPen(QColor(150, 150, 150)))
            painter.setFont(font)
            painter.drawText(plot_x + plot_w // 2 - 30, plot_y + plot_h + 35,
                             "\u0412\u0440\u0435\u043c\u044f (\u043c\u043a\u0441) / \u0413\u043b\u0443\u0431\u0438\u043d\u0430 (\u043c\u043c)")

            painter.end()

    return AScanWidgetImpl



# ============= 10. B-SCAN WIDGET =============


def _create_bscan_widget_class():
    """Создает класс виджета B-скана при наличии PyQt6."""
    if not _PYQT_AVAILABLE:
        return None

    class BScanWidgetImpl(QWidget):
        """Виджет отрисовки B-скана: 2D тепловая карта амплитуд."""

        def __init__(self, parent=None, record: Optional[BScanRecord] = None):
            """Инициализация виджета B-скана с данными записи."""
            super().__init__(parent)
            self._record = record
            self._width = record.width if record else 0
            self._height = record.height if record else 0
            self._pixels = record.pixels if record else b""
            self._settings = record.settings if record else None
            self._velocity = self._settings.velocity_mps if self._settings else DEFAULT_VELOCITY_MPS
            self._margin_left = 60
            self._margin_right = 20
            self._margin_top = 30
            self._margin_bottom = 40
            self._cursor_x = -1
            self._cursor_y = -1
            self._image: Optional[QImage] = None
            self.setMinimumSize(400, 300)
            self.setMouseTracking(True)
            self._build_image()

        def _build_image(self) -> None:
            """Строит QImage из пиксельных данных B-скана."""
            if self._width <= 0 or self._height <= 0 or not self._pixels:
                return
            self._image = QImage(self._width, self._height, QImage.Format.Format_RGB32)
            for y in range(self._height):
                for x in range(self._width):
                    idx = y * self._width + x
                    if idx < len(self._pixels):
                        amp = self._pixels[idx]
                        r, g, b = BSCAN_PALETTE[amp]
                        self._image.setPixelColor(x, y, QColor(r, g, b))
                    else:
                        self._image.setPixelColor(x, y, QColor(0, 0, 0))

        def mouseMoveEvent(self, event) -> None:
            """Обработка движения мыши для отображения координат курсора."""
            self._cursor_x = event.pos().x()
            self._cursor_y = event.pos().y()
            self.update()

        def paintEvent(self, event) -> None:
            """Отрисовка B-скана: тепловая карта, оси, курсор."""
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            w = self.width()
            h = self.height()

            # Фон
            painter.fillRect(0, 0, w, h, QColor(20, 20, 40))

            plot_x = self._margin_left
            plot_y = self._margin_top
            plot_w = w - self._margin_left - self._margin_right
            plot_h = h - self._margin_top - self._margin_bottom

            if plot_w <= 0 or plot_h <= 0:
                painter.end()
                return

            # Отрисовка изображения
            if self._image:
                target_rect = QRect(plot_x, plot_y, plot_w, plot_h)
                painter.drawImage(target_rect, self._image)
            else:
                painter.fillRect(plot_x, plot_y, plot_w, plot_h, QColor(30, 30, 50))

            # Рамка
            painter.setPen(QPen(QColor(150, 150, 150), 1))
            painter.drawRect(plot_x, plot_y, plot_w, plot_h)

            # Оси
            font = QFont("Monospace", 8)
            painter.setFont(font)
            painter.setPen(QPen(QColor(200, 200, 200)))

            # X-ось: позиция (мм)
            scan_length_mm = 100.0  # условная длина сканирования
            for i in range(0, 11, 2):
                x = plot_x + int(plot_w * i / 10)
                pos = scan_length_mm * i / 10
                painter.drawText(x - 10, plot_y + plot_h + 15, f"{pos:.0f}")

            # Y-ось: глубина (мм)
            duration_us = 100.0  # условная длительность
            max_depth = duration_us * self._velocity / 2000.0
            for i in range(0, 6):
                y = plot_y + int(plot_h * i / 5)
                d = max_depth * i / 5
                painter.drawText(plot_x - 55, y + 4, f"{d:.1f} \u043c\u043c")

            # Курсор-перекрестие
            if (plot_x <= self._cursor_x <= plot_x + plot_w and
                plot_y <= self._cursor_y <= plot_y + plot_h):
                painter.setPen(QPen(QColor(255, 255, 0), 1, Qt.PenStyle.DotLine))
                painter.drawLine(self._cursor_x, plot_y, self._cursor_x, plot_y + plot_h)
                painter.drawLine(plot_x, self._cursor_y, plot_x + plot_w, self._cursor_y)
                # Координаты
                rel_x = (self._cursor_x - plot_x) / plot_w
                rel_y = (self._cursor_y - plot_y) / plot_h
                pos_mm = rel_x * scan_length_mm
                depth_mm = rel_y * max_depth
                painter.setPen(QPen(QColor(255, 255, 0)))
                painter.drawText(self._cursor_x + 5, self._cursor_y - 5,
                                 f"X={pos_mm:.1f}\u043c\u043c Y={depth_mm:.1f}\u043c\u043c")

            # Заголовок
            painter.setPen(QPen(QColor(255, 255, 255)))
            title_font = QFont("Monospace", 10, QFont.Weight.Bold)
            painter.setFont(title_font)
            title = f"B-\u0441\u043a\u0430\u043d [{self._width}x{self._height}]"
            if self._record:
                title += f"  |  {self._record.passport_primary}  |  {self._record.date}"
            painter.drawText(plot_x, plot_y - 10, title)

            # Подпись X-оси
            painter.setPen(QPen(QColor(150, 150, 150)))
            painter.setFont(font)
            painter.drawText(plot_x + plot_w // 2 - 40, plot_y + plot_h + 30,
                             "\u041f\u043e\u0437\u0438\u0446\u0438\u044f \u0441\u043a\u0430\u043d\u0438\u0440\u043e\u0432\u0430\u043d\u0438\u044f (\u043c\u043c)")

            # Палитра (цветовая шкала справа)
            palette_x = plot_x + plot_w + 5
            palette_w = 15
            if palette_x + palette_w < w:
                for py in range(plot_h):
                    amp = int(255 * (1.0 - py / plot_h))
                    r, g, b = BSCAN_PALETTE[amp]
                    painter.setPen(QPen(QColor(r, g, b)))
                    painter.drawLine(palette_x, plot_y + py, palette_x + palette_w, plot_y + py)

            painter.end()

    return BScanWidgetImpl



# ============= 11. MENU AND TOOLBAR =============


def _setup_menu(window) -> None:
    """Настройка главного меню приложения (как в оригинальном PelengPC)."""
    if not _PYQT_AVAILABLE:
        return

    menubar = window.menuBar()

    # Меню "Прибор"
    menu_device = menubar.addMenu("\u041f\u0440\u0438\u0431\u043e\u0440")

    act_receive = QAction("\u041f\u0440\u0438\u043d\u044f\u0442\u044c (F5)", window)
    act_receive.setShortcut(QKeySequence("F5"))
    act_receive.triggered.connect(window._action_receive)
    menu_device.addAction(act_receive)

    act_test = QAction("\u0422\u0435\u0441\u0442 \u043f\u043e\u0440\u0442\u0430 (F6)", window)
    act_test.setShortcut(QKeySequence("F6"))
    act_test.triggered.connect(window._action_test_port)
    menu_device.addAction(act_test)

    act_port = QAction("\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0430 \u043f\u043e\u0440\u0442\u0430...", window)
    act_port.triggered.connect(window._action_port_settings)
    menu_device.addAction(act_port)

    menu_device.addSeparator()

    # Подменю UD2-102
    menu_ud2 = menu_device.addMenu("UD2-102")
    act_ud2_hs = QAction("\u0425\u044d\u043d\u0434\u0448\u0435\u0439\u043a", window)
    act_ud2_hs.triggered.connect(window._action_ud2_handshake)
    menu_ud2.addAction(act_ud2_hs)
    act_ud2_scan = QAction("\u0421\u043a\u0430\u043d\u0438\u0440\u043e\u0432\u0430\u043d\u0438\u0435 \u043f\u0430\u043c\u044f\u0442\u0438", window)
    act_ud2_scan.triggered.connect(window._action_ud2_scan)
    menu_ud2.addAction(act_ud2_scan)

    menu_device.addSeparator()

    act_demo = QAction("\u0414\u0435\u043c\u043e-\u0431\u043b\u043e\u043a (F8)", window)
    act_demo.setShortcut(QKeySequence("F8"))
    act_demo.triggered.connect(window._action_demo_block)
    menu_device.addAction(act_demo)

    menu_device.addSeparator()

    act_exit = QAction("\u0412\u044b\u0445\u043e\u0434", window)
    act_exit.setShortcut(QKeySequence("Ctrl+Q"))
    act_exit.triggered.connect(window.close)
    menu_device.addAction(act_exit)

    # Меню "База"
    menu_db = menubar.addMenu("\u0411\u0430\u0437\u0430")

    act_open = QAction("\u041e\u0442\u043a\u0440\u044b\u0442\u044c \u0411\u0414...", window)
    act_open.setShortcut(QKeySequence("Ctrl+O"))
    act_open.triggered.connect(window._action_open_db)
    menu_db.addAction(act_open)

    act_save = QAction("\u0421\u043e\u0445\u0440\u0430\u043d\u0438\u0442\u044c \u043a\u0430\u043a...", window)
    act_save.setShortcut(QKeySequence("Ctrl+S"))
    act_save.triggered.connect(window._action_save_as)
    menu_db.addAction(act_save)

    act_del = QAction("\u0423\u0434\u0430\u043b\u0438\u0442\u044c \u0437\u0430\u043f\u0438\u0441\u044c", window)
    act_del.setShortcut(QKeySequence("Delete"))
    act_del.triggered.connect(window._action_delete_record)
    menu_db.addAction(act_del)

    menu_db.addSeparator()

    act_import = QAction("\u0418\u043c\u043f\u043e\u0440\u0442 FDB...", window)
    act_import.triggered.connect(window._action_import_fdb)
    menu_db.addAction(act_import)

    # Меню "Отчет"
    menu_report = menubar.addMenu("\u041e\u0442\u0447\u0451\u0442")

    act_excel = QAction("\u042d\u043a\u0441\u043f\u043e\u0440\u0442 Excel...", window)
    act_excel.triggered.connect(window._action_export_excel)
    menu_report.addAction(act_excel)

    # Меню "Настройки"
    menu_settings = menubar.addMenu("\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438")

    act_edit = QAction("\u0420\u0435\u0434\u0430\u043a\u0442\u0438\u0440\u043e\u0432\u0430\u0442\u044c...", window)
    act_edit.triggered.connect(window._action_edit_settings)
    menu_settings.addAction(act_edit)

    # Меню "Помощь"
    menu_help = menubar.addMenu("\u041f\u043e\u043c\u043e\u0449\u044c")

    act_about = QAction("\u041e \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u0435...", window)
    act_about.triggered.connect(window._action_about)
    menu_help.addAction(act_about)


def _setup_toolbar(window) -> None:
    """Настройка панели инструментов (основные действия)."""
    if not _PYQT_AVAILABLE:
        return

    toolbar = QToolBar("\u0418\u043d\u0441\u0442\u0440\u0443\u043c\u0435\u043d\u0442\u044b")
    window.addToolBar(toolbar)

    btn_receive = QPushButton("\u041f\u0440\u0438\u043d\u044f\u0442\u044c")
    btn_receive.clicked.connect(window._action_receive)
    toolbar.addWidget(btn_receive)

    btn_test = QPushButton("\u0422\u0435\u0441\u0442")
    btn_test.clicked.connect(window._action_test_port)
    toolbar.addWidget(btn_test)

    btn_demo = QPushButton("\u0414\u0435\u043c\u043e")
    btn_demo.clicked.connect(window._action_demo_block)
    toolbar.addWidget(btn_demo)

    btn_open = QPushButton("\u041e\u0442\u043a\u0440\u044b\u0442\u044c \u0411\u0414")
    btn_open.clicked.connect(window._action_open_db)
    toolbar.addWidget(btn_open)

    btn_excel = QPushButton("\u042d\u043a\u0441\u043f\u043e\u0440\u0442 XLS")
    btn_excel.clicked.connect(window._action_export_excel)
    toolbar.addWidget(btn_excel)



# ============= 12. DECRYPTION DIALOG =============


class DecryptionDialog:
    """Заглушка для диалога расшифровки (создается при PyQt6)."""
    pass


class SettingsDetailDialog:
    """Заглушка для диалога детальных настроек (создается при PyQt6)."""
    pass


class PortDialog:
    """Заглушка для диалога настройки порта (создается при PyQt6)."""
    pass


def _create_dialog_classes():
    """Создает классы диалогов при наличии PyQt6."""
    if not _PYQT_AVAILABLE:
        return None, None, None

    class DecryptionDialogImpl(QDialog):
        """Диалог расшифровки записи: отображает все декодированные поля."""

        def __init__(self, parent=None, record: Optional[DecodedRecord] = None):
            """Инициализация диалога расшифровки с декодированной записью."""
            super().__init__(parent)
            self.setWindowTitle("\u0420\u0430\u0441\u0448\u0438\u0444\u0440\u043e\u0432\u043a\u0430 \u0437\u0430\u043f\u0438\u0441\u0438")
            self.setMinimumSize(700, 500)
            self._record = record
            self._init_ui()

        def _init_ui(self) -> None:
            """Инициализация интерфейса диалога расшифровки."""
            layout = QVBoxLayout(self)

            # Заголовок
            header = QLabel(f"\u0422\u0438\u043f: {self._record.category_name if self._record else ''} | "
                            f"Decoder: {self._record.decoder_type if self._record else ''}")
            header.setStyleSheet("font-weight: bold; font-size: 12px;")
            layout.addWidget(header)

            # Таблица полей
            self._table = QTableWidget()
            self._table.setColumnCount(5)
            self._table.setHorizontalHeaderLabels([
                "\u041f\u043e\u043b\u0435", "Raw Hex",
                "\u0417\u043d\u0430\u0447\u0435\u043d\u0438\u0435",
                "\u0415\u0434.\u0438\u0437\u043c.", "\u041a\u043e\u043c\u043c\u0435\u043d\u0442\u0430\u0440\u0438\u0439"
            ])
            self._table.horizontalHeader().setStretchLastSection(True)
            self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            layout.addWidget(self._table)

            # Кнопки
            buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            buttons.accepted.connect(self.accept)
            layout.addWidget(buttons)

            self._populate_table()

        def _populate_table(self) -> None:
            """Заполняет таблицу декодированными полями записи."""
            if not self._record:
                return

            rows_data: List[Tuple[str, str, str, str, str]] = []

            # Основные поля из raw body
            body = self._record.raw_body
            for name, (offset, size, desc) in PELENG_BODY_FIELDS.items():
                if len(body) > offset:
                    raw_slice = body[offset:offset + size]
                    hex_str = raw_slice.hex()
                    # Декодированное значение
                    val = self._record.fields.get(name, "")
                    rows_data.append((name, hex_str, str(val), "", desc))

            # Дополнительные поля из fields
            for key, val in self._record.fields.items():
                if key not in PELENG_BODY_FIELDS:
                    rows_data.append((key, "", str(val), "", ""))

            # Параметры для A-скана
            if isinstance(self._record, AScanRecord) and self._record.settings:
                s = self._record.settings
                rows_data.append(("\u0434\u043b\u0438\u0442\u0435\u043b\u044c\u043d\u043e\u0441\u0442\u044c_\u043c\u043a\u0441",
                                  "", f"{ascan_compute_duration_us(self._record.n_samples):.1f}", "\u043c\u043a\u0441", ""))
                rows_data.append(("\u043c\u0430\u043a\u0441_\u0433\u043b\u0443\u0431\u0438\u043d\u0430_\u043c\u043c",
                                  "", f"{ascan_compute_max_depth_mm(self._record.n_samples, s.velocity_mps):.1f}", "\u043c\u043c", ""))
                rows_data.append(("\u0441\u0442\u0440\u043e\u0431_\u043d\u0430\u0447_\u043c\u043a\u0441",
                                  "", f"{s.gate_start_us():.2f}", "\u043c\u043a\u0441", ""))
                rows_data.append(("\u0441\u0442\u0440\u043e\u0431_\u0448\u0438\u0440_\u043c\u043a\u0441",
                                  "", f"{s.gate_width_us():.2f}", "\u043c\u043a\u0441", ""))
                rows_data.append(("\u043f\u043e\u0440\u043e\u0433_\u0434\u0411",
                                  "", f"{s.threshold_db():.1f}", "\u0434\u0411", ""))

            # Параметры для B-скана
            if isinstance(self._record, BScanRecord):
                rows_data.append(("\u0440\u0430\u0437\u043c\u0435\u0440_\u0438\u0437\u043e\u0431\u0440\u0430\u0436\u0435\u043d\u0438\u044f",
                                  "", f"{self._record.width}x{self._record.height}", "px", ""))

            # Настройки
            if isinstance(self._record, SettingsRecord) and self._record.settings:
                s = self._record.settings
                rows_data.append(("VRT_\u0442\u0430\u0431\u043b\u0438\u0446\u0430",
                                  s.vrt_table.hex(), "", "", "8 \u0431\u0430\u0439\u0442 \u0412\u0420\u0427"))

            # Для отчета
            if isinstance(self._record, ReportRecord):
                rows_data.append(("\u0442\u043e\u043b\u0449\u0438\u043d\u0430",
                                  "", f"{self._record.thickness_mm:.2f}", "\u043c\u043c", ""))
                rows_data.append(("\u0430\u043c\u043f\u043b\u0438\u0442\u0443\u0434\u0430",
                                  "", f"{self._record.amplitude_db:.1f}", "\u0434\u0411", ""))
                rows_data.append(("\u0432\u0435\u0440\u0434\u0438\u043a\u0442",
                                  "", self._record.verdict, "", ""))

            self._table.setRowCount(len(rows_data))
            for i, (name, hex_val, value, units, comment) in enumerate(rows_data):
                self._table.setItem(i, 0, QTableWidgetItem(name))
                self._table.setItem(i, 1, QTableWidgetItem(hex_val))
                self._table.setItem(i, 2, QTableWidgetItem(value))
                self._table.setItem(i, 3, QTableWidgetItem(units))
                self._table.setItem(i, 4, QTableWidgetItem(comment))

    class SettingsDetailDialogImpl(QDialog):
        """Диалог детального просмотра настроек прибора."""

        def __init__(self, parent=None, record: Optional[DecodedRecord] = None):
            """Инициализация диалога детальных настроек."""
            super().__init__(parent)
            self.setWindowTitle("\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438 \u043f\u0440\u0438\u0431\u043e\u0440\u0430")
            self.setMinimumSize(500, 400)
            self._record = record
            self._init_ui()

        def _init_ui(self) -> None:
            """Инициализация интерфейса диалога настроек."""
            layout = QVBoxLayout(self)

            # Группа основных параметров
            group = QGroupBox("\u0410\u043a\u0443\u0441\u0442\u0438\u0447\u0435\u0441\u043a\u0438\u0435 \u043f\u0430\u0440\u0430\u043c\u0435\u0442\u0440\u044b")
            form = QFormLayout(group)

            settings = None
            if isinstance(self._record, SettingsRecord):
                settings = self._record.settings

            form.addRow("\u0423\u0441\u0438\u043b\u0435\u043d\u0438\u0435:",
                        QLabel(f"{settings.gain_db if settings else 0} \u0434\u0411"))
            form.addRow("\u0421\u043a\u043e\u0440\u043e\u0441\u0442\u044c:",
                        QLabel(f"{settings.velocity_mps if settings else 0} \u043c/\u0441"))
            form.addRow("\u0423\u0433\u043e\u043b \u0432\u0432\u043e\u0434\u0430:",
                        QLabel(f"{settings.angle_deg if settings else 0}\u00b0"))
            form.addRow("\u0417\u0430\u0434\u0435\u0440\u0436\u043a\u0430:",
                        QLabel(f"{settings.delay_us if settings else 0} \u043c\u043a\u0441"))
            form.addRow("\u0421\u0442\u0440\u043e\u0431 \u043d\u0430\u0447\u0430\u043b\u043e:",
                        QLabel(f"{settings.gate_start if settings else 0} \u043e\u0442\u0441\u0447. ({settings.gate_start_us() if settings else 0:.2f} \u043c\u043a\u0441)"))
            form.addRow("\u0421\u0442\u0440\u043e\u0431 \u0448\u0438\u0440\u0438\u043d\u0430:",
                        QLabel(f"{settings.gate_width if settings else 0} \u043e\u0442\u0441\u0447. ({settings.gate_width_us() if settings else 0:.2f} \u043c\u043a\u0441)"))
            form.addRow("\u041f\u043e\u0440\u043e\u0433:",
                        QLabel(f"{settings.threshold_pct if settings else 0} % ({settings.threshold_db() if settings else 0:.1f} \u0434\u0411)"))
            form.addRow("\u0412\u0420\u0427 \u0442\u0430\u0431\u043b\u0438\u0446\u0430:",
                        QLabel(f"{settings.vrt_table.hex() if settings else ''}"))

            layout.addWidget(group)

            # Группа вычисленных параметров
            group2 = QGroupBox("\u0412\u044b\u0447\u0438\u0441\u043b\u0435\u043d\u043d\u044b\u0435 \u043f\u0430\u0440\u0430\u043c\u0435\u0442\u0440\u044b")
            form2 = QFormLayout(group2)
            if settings:
                form2.addRow("\u0421\u0442\u0440\u043e\u0431 \u043d\u0430\u0447. (\u043c\u043c):",
                             QLabel(f"{settings.gate_start_mm():.2f}"))
                form2.addRow("\u0421\u0442\u0440\u043e\u0431 \u043a\u043e\u043d. (\u043c\u043c):",
                             QLabel(f"{settings.gate_end_mm():.2f}"))
            layout.addWidget(group2)

            # Кнопки
            buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            buttons.accepted.connect(self.accept)
            layout.addWidget(buttons)

    class PortDialogImpl(QDialog):
        """Диалог настройки COM-порта."""

        def __init__(self, parent=None, config: Optional[PortConfig] = None):
            """Инициализация диалога настроек порта."""
            super().__init__(parent)
            self.setWindowTitle("\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0430 \u043f\u043e\u0440\u0442\u0430")
            self.setFixedSize(350, 250)
            self._config = config or PortConfig()
            self._init_ui()

        def _init_ui(self) -> None:
            """Инициализация интерфейса диалога порта."""
            layout = QFormLayout(self)

            # Выпадающий список COM-портов (с возможностью ручного ввода)
            self._port_combo = QComboBox()
            self._port_combo.setEditable(True)
            # Перечисление доступных COM-портов через pyserial
            if _list_ports is not None:
                try:
                    ports = _list_ports.comports()
                    for p in sorted(ports, key=lambda x: x.device):
                        self._port_combo.addItem(f"{p.device} - {p.description}", p.device)
                except Exception:
                    pass
            # Если портов не найдено, добавляем типовые имена
            if self._port_combo.count() == 0:
                for name in ["COM1", "COM2", "COM3", "COM4",
                             "/dev/ttyUSB0", "/dev/ttyACM0"]:
                    self._port_combo.addItem(name, name)
            # Устанавливаем текущее значение из конфигурации
            self._port_combo.setCurrentText(self._config.port_name)
            layout.addRow("\u041f\u043e\u0440\u0442:", self._port_combo)

            self._baud_combo = QComboBox()
            self._baud_combo.addItems(["19200", "57600", "9600", "115200"])
            self._baud_combo.setCurrentText(str(self._config.baud_rate))
            layout.addRow("\u0421\u043a\u043e\u0440\u043e\u0441\u0442\u044c:", self._baud_combo)

            self._parity_combo = QComboBox()
            self._parity_combo.addItems(["E (Even)", "N (None)", "O (Odd)"])
            layout.addRow("\u0427\u0435\u0442\u043d\u043e\u0441\u0442\u044c:", self._parity_combo)

            self._delay_spin = QSpinBox()
            self._delay_spin.setRange(1, 50)
            self._delay_spin.setValue(int(self._config.inter_byte_delay * 1000))
            self._delay_spin.setSuffix(" \u043c\u0441")
            layout.addRow("\u0417\u0430\u0434\u0435\u0440\u0436\u043a\u0430:", self._delay_spin)

            buttons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
            )
            buttons.accepted.connect(self.accept)
            buttons.rejected.connect(self.reject)
            layout.addRow(buttons)

        def get_config(self) -> PortConfig:
            """Возвращает конфигурацию порта из диалога."""
            config = PortConfig()
            # Если выбран элемент из списка, берем device из userData, иначе текст
            idx = self._port_combo.currentIndex()
            user_data = self._port_combo.itemData(idx) if idx >= 0 else None
            if user_data:
                config.port_name = user_data
            else:
                config.port_name = self._port_combo.currentText()
            config.baud_rate = int(self._baud_combo.currentText())
            parity_text = self._parity_combo.currentText()
            if "E" in parity_text:
                config.parity = "E"
            elif "O" in parity_text:
                config.parity = "O"
            else:
                config.parity = "N"
            config.inter_byte_delay = self._delay_spin.value() / 1000.0
            return config

    return DecryptionDialogImpl, SettingsDetailDialogImpl, PortDialogImpl



# ============= 13. GRAPH+DECRYPTION WINDOW =============


class GraphDecryptionWindow:
    """Заглушка для комбинированного окна граф+расшифровка."""
    pass


def _create_graph_decryption_window():
    """Создает класс комбинированного окна граф+расшифровка."""
    if not _PYQT_AVAILABLE:
        return None

    AScanWidgetCls = _create_ascan_widget_class()
    BScanWidgetCls = _create_bscan_widget_class()
    DecryptionDialogCls, _, _ = _create_dialog_classes()

    class GraphDecryptionWindowImpl(QDialog):
        """Комбинированное окно: график (A-scan или B-scan) + таблица расшифровки полей."""

        def __init__(self, parent=None, record: Optional[DecodedRecord] = None):
            """Инициализация комбинированного окна граф+расшифровка."""
            super().__init__(parent)
            self.setWindowTitle("\u041f\u0440\u043e\u0442\u043e\u043a\u043e\u043b: \u0413\u0440\u0430\u0444\u0438\u043a + \u0420\u0430\u0441\u0448\u0438\u0444\u0440\u043e\u0432\u043a\u0430")
            self.setMinimumSize(1100, 600)
            self._record = record
            self._init_ui()

        def _init_ui(self) -> None:
            """Инициализация интерфейса: QSplitter с графиком слева и полями справа."""
            layout = QVBoxLayout(self)

            splitter = QSplitter(Qt.Orientation.Horizontal)
            layout.addWidget(splitter)

            # Левая панель: график
            if isinstance(self._record, AScanRecord) and AScanWidgetCls:
                graph_widget = AScanWidgetCls(record=self._record)
            elif isinstance(self._record, BScanRecord) and BScanWidgetCls:
                graph_widget = BScanWidgetCls(record=self._record)
            else:
                graph_widget = QLabel("\u041d\u0435\u0442 \u0433\u0440\u0430\u0444\u0438\u0447\u0435\u0441\u043a\u0438\u0445 \u0434\u0430\u043d\u043d\u044b\u0445")
            splitter.addWidget(graph_widget)

            # Правая панель: таблица полей
            right_panel = QWidget()
            right_layout = QVBoxLayout(right_panel)
            right_layout.setContentsMargins(0, 0, 0, 0)

            lbl = QLabel("\u0420\u0430\u0441\u0448\u0438\u0444\u0440\u043e\u0432\u043a\u0430 \u043f\u043e\u043b\u0435\u0439")
            lbl.setStyleSheet("font-weight: bold;")
            right_layout.addWidget(lbl)

            self._fields_table = QTableWidget()
            self._fields_table.setColumnCount(4)
            self._fields_table.setHorizontalHeaderLabels([
                "\u041f\u043e\u043b\u0435", "Hex",
                "\u0417\u043d\u0430\u0447\u0435\u043d\u0438\u0435", "\u0415\u0434.\u0438\u0437\u043c."
            ])
            self._fields_table.horizontalHeader().setStretchLastSection(True)
            self._fields_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            right_layout.addWidget(self._fields_table)

            splitter.addWidget(right_panel)
            splitter.setSizes([600, 500])

            self._populate_fields()

            # Кнопки
            buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            buttons.accepted.connect(self.accept)
            layout.addWidget(buttons)

        def _populate_fields(self) -> None:
            """Заполняет таблицу расшифрованными полями записи."""
            if not self._record:
                return

            rows_data: List[Tuple[str, str, str, str]] = []

            # Информация о записи
            rows_data.append(("\u041a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u044f", "",
                              self._record.category_name, ""))
            rows_data.append(("Device ID", "",
                              str(self._record.device_id), ""))
            rows_data.append(("Sweep ID", "",
                              str(self._record.sweep_id), ""))
            rows_data.append(("\u0414\u0430\u0442\u0430", "",
                              str(self._record.date) if self._record.date else "", ""))
            rows_data.append(("\u0412\u0440\u0435\u043c\u044f", "",
                              str(self._record.time) if self._record.time else "", ""))
            rows_data.append(("\u041f\u0430\u0441\u043f\u043e\u0440\u0442 \u043e\u0441\u043d.", "",
                              self._record.passport_primary, ""))
            rows_data.append(("\u041f\u0430\u0441\u043f\u043e\u0440\u0442 \u0434\u043e\u043f.", "",
                              self._record.passport_secondary, ""))
            rows_data.append(("\u0414\u0435\u0444\u0435\u043a\u0442", "",
                              str(self._record.defect_flag), ""))

            # Поля из fields dict
            for key, val in self._record.fields.items():
                if key not in ("device_id", "sweep_id", "date", "time",
                               "passport_primary", "passport_secondary", "defect_flag"):
                    # Попытка получить hex из raw body
                    hex_str = ""
                    if key in PELENG_BODY_FIELDS:
                        offset, size, _ = PELENG_BODY_FIELDS[key]
                        body = self._record.raw_body
                        if len(body) > offset:
                            hex_str = body[offset:offset + size].hex()
                    units = ""
                    if "db" in key.lower() or "gain" in key.lower():
                        units = "\u0434\u0411"
                    elif "mps" in key.lower() or "velocity" in key.lower():
                        units = "\u043c/\u0441"
                    elif "mm" in key.lower() or "depth" in key.lower():
                        units = "\u043c\u043c"
                    elif "us" in key.lower() or "delay" in key.lower():
                        units = "\u043c\u043a\u0441"
                    elif "deg" in key.lower() or "angle" in key.lower():
                        units = "\u00b0"
                    elif "pct" in key.lower() or "threshold" in key.lower():
                        units = "%"
                    rows_data.append((key, hex_str, str(val), units))

            # Специфичные для A-скана
            if isinstance(self._record, AScanRecord):
                n = self._record.n_samples
                v = self._record.settings.velocity_mps if self._record.settings else DEFAULT_VELOCITY_MPS
                rows_data.append(("\u0414\u043b\u0438\u0442\u0435\u043b\u044c\u043d\u043e\u0441\u0442\u044c", "",
                                  f"{ascan_compute_duration_us(n):.1f}", "\u043c\u043a\u0441"))
                rows_data.append(("\u041c\u0430\u043a\u0441. \u0433\u043b\u0443\u0431\u0438\u043d\u0430", "",
                                  f"{ascan_compute_max_depth_mm(n, v):.1f}", "\u043c\u043c"))
                # Пик
                peak_val, peak_idx = ascan_peak_detect(self._record.samples)
                rows_data.append(("\u041f\u0438\u043a \u0430\u043c\u043f\u043b\u0438\u0442\u0443\u0434\u044b", "",
                                  f"{ascan_byte_to_db(peak_val):.1f}", "\u0434\u0411"))
                rows_data.append(("\u041f\u0438\u043a \u043f\u043e\u0437\u0438\u0446\u0438\u044f", "",
                                  f"{ascan_sample_to_depth_mm(peak_idx, v):.2f}", "\u043c\u043c"))

            # Специфичные для B-скана
            if isinstance(self._record, BScanRecord):
                rows_data.append(("\u0420\u0430\u0437\u043c\u0435\u0440", "",
                                  f"{self._record.width}x{self._record.height}", "px"))
                rows_data.append(("\u041f\u0438\u043a\u0441\u0435\u043b\u0435\u0439", "",
                                  str(len(self._record.pixels)), ""))

            self._fields_table.setRowCount(len(rows_data))
            for i, (name, hex_val, value, units) in enumerate(rows_data):
                self._fields_table.setItem(i, 0, QTableWidgetItem(name))
                self._fields_table.setItem(i, 1, QTableWidgetItem(hex_val))
                self._fields_table.setItem(i, 2, QTableWidgetItem(value))
                self._fields_table.setItem(i, 3, QTableWidgetItem(units))

    return GraphDecryptionWindowImpl



# ============= 14. ENTRY POINT AND CLI =============


def _export_to_excel(db: BlockZapDB, path: str) -> None:
    """Экспорт данных из БД в файл Excel (xlsx)."""
    if not _OPENPYXL_AVAILABLE:
        LOG.error("openpyxl не установлен")
        return
    wb = openpyxl.Workbook()

    # Лист протоколов
    ws_proto = wb.active
    ws_proto.title = "\u041f\u0440\u043e\u0442\u043e\u043a\u043e\u043b\u044b"
    ws_proto.append(["\u2116", "\u0414\u0430\u0442\u0430", "\u0412\u0440\u0435\u043c\u044f", "\u0422\u0438\u043f",
                     "\u041e\u0431\u044a\u0435\u043a\u0442", "\u041e\u043f\u0435\u0440\u0430\u0442\u043e\u0440", "\u0412\u0435\u0440\u0434\u0438\u043a\u0442"])
    for rec in db.filter_protocols():
        ws_proto.append([rec["id"], rec["date_str"], rec["time_str"],
                         rec["category"], rec["passport_primary"],
                         rec["passport_secondary"], rec["verdict"]])

    # Лист отчетов
    ws_rep = wb.create_sheet("\u041e\u0442\u0447\u0451\u0442\u044b")
    ws_rep.append(["\u2116", "\u0414\u0430\u0442\u0430", "\u0412\u0440\u0435\u043c\u044f", "\u0422\u0438\u043f",
                   "\u041e\u0431\u044a\u0435\u043a\u0442", "\u0422\u043e\u043b\u0449\u0438\u043d\u0430", "\u0410\u043c\u043f\u043b\u0438\u0442\u0443\u0434\u0430", "\u0412\u0435\u0440\u0434\u0438\u043a\u0442"])
    for rec in db.filter_reports():
        ws_rep.append([rec["id"], rec["date_str"], rec["time_str"],
                       rec["category"], rec["passport_primary"],
                       rec["thickness_mm"], rec["amplitude_db"], rec["verdict"]])

    # Лист настроек
    ws_set = wb.create_sheet("\u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438")
    ws_set.append(["\u2116", "\u0414\u0430\u0442\u0430", "\u0423\u0441\u0438\u043b\u0435\u043d\u0438\u0435 \u0434\u0411", "\u0421\u043a\u043e\u0440\u043e\u0441\u0442\u044c",
                   "\u0423\u0433\u043e\u043b", "\u0421\u0442\u0440\u043e\u0431 \u043d\u0430\u0447.", "\u0421\u0442\u0440\u043e\u0431 \u0448\u0438\u0440.", "\u041f\u043e\u0440\u043e\u0433 %"])
    for rec in db.filter_settings():
        ws_set.append([rec["id"], rec["date_str"], rec["gain_db"],
                       rec["velocity_mps"], rec["angle_deg"],
                       rec["gate_start"], rec["gate_width"], rec["threshold_pct"]])

    wb.save(path)
    LOG.info("\u042d\u043a\u0441\u043f\u043e\u0440\u0442 \u0432 %s \u0437\u0430\u0432\u0435\u0440\u0448\u0435\u043d", path)


def generate_demo_data() -> bytes:
    """Генерирует демонстрационные данные для тестирования без прибора."""
    import random
    random.seed(42)
    result = bytearray()

    # Генерация нескольких TLV-записей разных типов
    for i in range(15):
        # Выбираем тип записи
        if i < 5:
            sweep_addr = 1000 + i  # A-scan (category 1)
        elif i < 8:
            sweep_addr = 10000 + i  # B-scan (category 10)
        elif i < 11:
            sweep_addr = 5000 + i  # Calibration (category 5)
        elif i < 13:
            sweep_addr = 4000 + i  # Generic (category 4)
        else:
            sweep_addr = 20000 + i  # ShortProto (category 20)

        # Формируем тело записи
        body = bytearray(256)
        # device_id
        struct.pack_into('<H', body, 0x00, 1234)
        # sweep_id
        body[0x04] = (i // 100) & 0xFF
        body[0x05] = (i % 100) & 0xFF
        # date
        body[0x07] = 15  # day
        body[0x08] = 6   # month
        body[0x09] = 24  # year (2024)
        # time
        body[0x0A] = 10 + i  # hour
        body[0x0B] = 30      # minute
        # defect_flag
        body[0x0C] = random.randint(0, 2)
        # sweep_addr
        struct.pack_into('<H', body, 0x10, sweep_addr)
        # passport (простые данные)
        for j in range(11):
            body[0x11 + j] = random.randint(1, 40)
        for j in range(7):
            body[0x21 + j] = random.randint(1, 30)

        # Settings tail
        body[0x66] = random.randint(20, 60)  # gain_db
        struct.pack_into('<H', body, 0x67, random.choice([5900, 5920, 3260]))  # velocity
        body[0x69] = random.choice([0, 45, 50, 60, 65, 70])  # angle
        body[0x6A] = random.randint(0, 20)  # delay
        struct.pack_into('<H', body, 0x6E, random.randint(100, 500))  # gate_start
        struct.pack_into('<H', body, 0x70, random.randint(200, 1000))  # gate_width
        body[0x74] = random.randint(20, 80)  # threshold
        # VRT table
        for j in range(8):
            body[0x80 + j] = random.randint(0, 20)

        # A-scan: добавляем отсчеты
        if sweep_addr // 1000 == 1:
            samples = bytearray(512)
            for s in range(512):
                # Синусоподобный сигнал с шумом
                val = int(127 + 100 * math.sin(s * 0.05) * math.exp(-s / 200))
                val = max(0, min(255, val + random.randint(-10, 10)))
                samples[s] = val
            body = body + samples

        # B-scan: добавляем пиксели
        if sweep_addr // 1000 in range(10, 20):
            bw, bh = 64, 128
            struct.pack_into('<H', body, 0x88, bw)
            struct.pack_into('<H', body, 0x8A, bh)
            pixels = bytearray(bw * bh)
            for py in range(bh):
                for px in range(bw):
                    # Градиент + шум
                    val = int((py / bh) * 200 + random.randint(0, 50))
                    pixels[py * bw + px] = max(0, min(255, val))
            body = body[:0x8C] + pixels

        # Упаковка TLV
        tag = sweep_addr & 0xFFFF
        length = len(body)
        tlv_header = struct.pack('<HH', tag, length)
        result += tlv_header + bytes(body)

    return bytes(result)


def _run_fake_device(port_name: str) -> None:
    """Запускает эмулятор прибора на указанном порту."""
    if not _SERIAL_AVAILABLE:
        print("\u041e\u0448\u0438\u0431\u043a\u0430: pyserial \u043d\u0435 \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d")
        return
    print(f"\u042d\u043c\u0443\u043b\u044f\u0442\u043e\u0440 \u043f\u0440\u0438\u0431\u043e\u0440\u0430 \u043d\u0430 {port_name}")
    print("\u041e\u0436\u0438\u0434\u0430\u043d\u0438\u0435 \u043a\u043e\u043c\u0430\u043d\u0434...")
    try:
        ser = serial.Serial(port_name, DEFAULT_BAUD, parity='E', timeout=1)
    except Exception as exc:
        print(f"\u041e\u0448\u0438\u0431\u043a\u0430 \u043e\u0442\u043a\u0440\u044b\u0442\u0438\u044f \u043f\u043e\u0440\u0442\u0430: {exc}")
        return

    # Генерируем демо-блок
    demo_block = generate_demo_data()

    while True:
        try:
            cmd = ser.read(1)
            if not cmd:
                continue
            opcode = cmd[0]
            if opcode == OP_HANDSHAKE:
                # Ответ на хэндшейк: 16 байт
                reply = bytearray(16)
                reply[0] = 0x55  # echo
                reply[1] = 0x01  # version
                reply[2] = 0x05  # info_byte (extended flash + wagon)
                reply[3] = 0x02  # firmware sub
                struct.pack_into('<I', reply, 4, len(demo_block))
                ser.write(reply)
                print(f"  -> \u0425\u044d\u043d\u0434\u0448\u0435\u0439\u043a: \u043e\u0442\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u043e {len(reply)} \u0431\u0430\u0439\u0442")
            elif opcode == OP_BLOCK_READ:
                # Читаем LL HH
                params = ser.read(2)
                if len(params) == 2:
                    block_len = params[0] | (params[1] << 8)
                    # Отправляем данные (или столько, сколько есть)
                    to_send = demo_block[:block_len]
                    ser.write(to_send)
                    print(f"  -> \u0411\u043b\u043e\u043a: \u043e\u0442\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u043e {len(to_send)} \u0431\u0430\u0439\u0442")
            elif opcode == OP_SCREEN_DUMP:
                # Экран: отправляем пустой
                screen = bytearray(LCD_SIZE_WAGON)
                ser.write(screen)
                print(f"  -> \u042d\u043a\u0440\u0430\u043d: \u043e\u0442\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u043e {len(screen)} \u0431\u0430\u0439\u0442")
            else:
                print(f"  ?? \u041d\u0435\u0438\u0437\u0432\u0435\u0441\u0442\u043d\u044b\u0439 \u043e\u043f\u043a\u043e\u0434: 0x{opcode:02X}")
        except KeyboardInterrupt:
            break
        except Exception as exc:
            print(f"\u041e\u0448\u0438\u0431\u043a\u0430: {exc}")
            break

    ser.close()
    print("\u042d\u043c\u0443\u043b\u044f\u0442\u043e\u0440 \u043e\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d")


def _run_proto_test(port_name: str, baud: int) -> None:
    """CLI: тест соединения с прибором (хэндшейк)."""
    config = PortConfig(port_name=port_name, baud_rate=baud)
    port = PelengPort(config)
    if not port.open():
        print(f"\u041e\u0448\u0438\u0431\u043a\u0430 \u043e\u0442\u043a\u0440\u044b\u0442\u0438\u044f \u043f\u043e\u0440\u0442\u0430 {port_name}")
        return
    reply = port.test_port()
    if reply:
        print(f"\u0425\u044d\u043d\u0434\u0448\u0435\u0439\u043a OK: {reply.hex()}")
        info = port.parse_info_byte(reply)
        print(f"\u0418\u043d\u0444\u043e: {info}")
    else:
        print("\u041d\u0435\u0442 \u043e\u0442\u0432\u0435\u0442\u0430 \u043e\u0442 \u043f\u0440\u0438\u0431\u043e\u0440\u0430")
    port.close()


def _run_proto_block(port_name: str, baud: int, block_len: int) -> None:
    """CLI: запрос блока данных от прибора."""
    config = PortConfig(port_name=port_name, baud_rate=baud)
    port = PelengPort(config)
    if not port.open():
        print(f"\u041e\u0448\u0438\u0431\u043a\u0430 \u043e\u0442\u043a\u0440\u044b\u0442\u0438\u044f \u043f\u043e\u0440\u0442\u0430 {port_name}")
        return
    data = port.request_block(block_len)
    if data:
        print(f"\u041f\u043e\u043b\u0443\u0447\u0435\u043d\u043e {len(data)} \u0431\u0430\u0439\u0442")
        # Парсим TLV
        records = TLVRecord.parse_all(data)
        print(f"TLV \u0437\u0430\u043f\u0438\u0441\u0435\u0439: {len(records)}")
        for i, tlv in enumerate(records):
            decoded = dispatch_decode(tlv)
            print(f"  [{i}] {decoded.category_name}: sweep={decoded.sweep_addr}, "
                  f"date={decoded.date}, passport={decoded.passport_primary}")
    else:
        print("\u041d\u0435\u0442 \u0434\u0430\u043d\u043d\u044b\u0445")
    port.close()


def _run_demo() -> None:
    """CLI: генерация и анализ демо-данных."""
    print("\u0413\u0435\u043d\u0435\u0440\u0430\u0446\u0438\u044f \u0434\u0435\u043c\u043e-\u0434\u0430\u043d\u043d\u044b\u0445...")
    data = generate_demo_data()
    print(f"\u0421\u0433\u0435\u043d\u0435\u0440\u0438\u0440\u043e\u0432\u0430\u043d\u043e {len(data)} \u0431\u0430\u0439\u0442")

    records = TLVRecord.parse_all(data)
    print(f"\nTLV \u0437\u0430\u043f\u0438\u0441\u0435\u0439: {len(records)}")

    db = BlockZapDB()
    for tlv in records:
        decoded = dispatch_decode(tlv)
        db.upsert(decoded, tlv.raw)
        print(f"  {decoded.category_name:15s} | dev={decoded.device_id:5d} | "
              f"date={decoded.date} | passport={decoded.passport_primary}")

    print(f"\n\u0411\u0414: {db.count()} \u0437\u0430\u043f\u0438\u0441\u0435\u0439")
    print(f"  \u041f\u0440\u043e\u0442\u043e\u043a\u043e\u043b\u044b: {len(db.filter_protocols())}")
    print(f"  \u041e\u0442\u0447\u0435\u0442\u044b:    {len(db.filter_reports())}")
    print(f"  \u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0438: {len(db.filter_settings())}")
    db.close()


def _run_gui() -> None:
    """Запуск графического интерфейса PyQt6."""
    if not _PYQT_AVAILABLE:
        print("\u041e\u0448\u0438\u0431\u043a\u0430: PyQt6 \u043d\u0435 \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d. \u0423\u0441\u0442\u0430\u043d\u043e\u0432\u0438\u0442\u0435: pip install PyQt6")
        sys.exit(1)

    # Инициализация классов GUI
    global DecryptionDialog, SettingsDetailDialog, PortDialog, GraphDecryptionWindow

    PelengMainWindow = _create_gui_classes()
    dialogs = _create_dialog_classes()
    if dialogs:
        DecryptionDialog, SettingsDetailDialog, PortDialog = dialogs
    gd_win = _create_graph_decryption_window()
    if gd_win:
        GraphDecryptionWindow = gd_win

    app = QApplication(sys.argv)
    app.setApplicationName("PelengPC Clone v2")
    app.setOrganizationName("Altek")

    window = PelengMainWindow()
    window.show()

    sys.exit(app.exec())


def main() -> None:
    """Точка входа: разбор аргументов командной строки и запуск."""
    parser = argparse.ArgumentParser(
        description="PelengPC Clone v2.0 -- \u043a\u043b\u043e\u043d \u043f\u0440\u043e\u0433\u0440\u0430\u043c\u043c\u044b PelengPC v1.2"
    )
    subparsers = parser.add_subparsers(dest="command")

    # proto-test
    sp_test = subparsers.add_parser("proto-test",
                                     help="\u0422\u0435\u0441\u0442 \u0441\u043e\u0435\u0434\u0438\u043d\u0435\u043d\u0438\u044f (handshake)")
    sp_test.add_argument("--port", default="/dev/ttyUSB0",
                         help="\u0418\u043c\u044f COM-\u043f\u043e\u0440\u0442\u0430")
    sp_test.add_argument("--baud", type=int, default=DEFAULT_BAUD,
                         help="\u0421\u043a\u043e\u0440\u043e\u0441\u0442\u044c \u043f\u043e\u0440\u0442\u0430")

    # proto-block
    sp_block = subparsers.add_parser("proto-block",
                                      help="\u0427\u0442\u0435\u043d\u0438\u0435 \u0431\u043b\u043e\u043a\u0430 \u0434\u0430\u043d\u043d\u044b\u0445")
    sp_block.add_argument("--port", default="/dev/ttyUSB0")
    sp_block.add_argument("--baud", type=int, default=DEFAULT_BAUD)
    sp_block.add_argument("--len", type=int, default=8873,
                          help="\u0420\u0430\u0437\u043c\u0435\u0440 \u0431\u043b\u043e\u043a\u0430")

    # fake-device
    sp_fake = subparsers.add_parser("fake-device",
                                     help="\u042d\u043c\u0443\u043b\u044f\u0442\u043e\u0440 \u043f\u0440\u0438\u0431\u043e\u0440\u0430")
    sp_fake.add_argument("--port", default="/tmp/peleng_dev_b")

    # demo
    subparsers.add_parser("demo", help="\u0413\u0435\u043d\u0435\u0440\u0430\u0446\u0438\u044f \u0434\u0435\u043c\u043e-\u0434\u0430\u043d\u043d\u044b\u0445")

    args = parser.parse_args()

    if args.command == "proto-test":
        _run_proto_test(args.port, args.baud)
    elif args.command == "proto-block":
        _run_proto_block(args.port, args.baud, getattr(args, 'len', 8873))
    elif args.command == "fake-device":
        _run_fake_device(args.port)
    elif args.command == "demo":
        _run_demo()
    else:
        _run_gui()


if __name__ == "__main__":
    main()


# ============= ADDITIONAL UTILITIES AND HELPERS =============


def format_hex_dump(data: bytes, bytes_per_line: int = 16) -> str:
    """Форматирует байтовый массив как hex-дамп с ASCII представлением."""
    lines: List[str] = []
    for offset in range(0, len(data), bytes_per_line):
        chunk = data[offset:offset + bytes_per_line]
        hex_part = ' '.join(f'{b:02X}' for b in chunk)
        ascii_part = ''.join(chr(b) if 32 <= b < 127 else '.' for b in chunk)
        lines.append(f"{offset:04X}  {hex_part:<{bytes_per_line * 3}}  {ascii_part}")
    return '\n'.join(lines)


def compute_checksum_xor(data: bytes) -> int:
    """Вычисляет контрольную сумму XOR для массива байт."""
    result = 0
    for b in data:
        result ^= b
    return result


def compute_checksum_sum8(data: bytes) -> int:
    """Вычисляет контрольную сумму (сумма по модулю 256)."""
    return sum(data) & 0xFF


def compute_crc16(data: bytes, poly: int = 0xA001) -> int:
    """Вычисляет CRC16 (Modbus) для массива байт."""
    crc = 0xFFFF
    for b in data:
        crc ^= b
        for _ in range(8):
            if crc & 1:
                crc = (crc >> 1) ^ poly
            else:
                crc >>= 1
    return crc


def decode_bcd(byte_val: int) -> int:
    """Декодирует BCD-значение (двоично-десятичный код) в целое число."""
    return (byte_val >> 4) * 10 + (byte_val & 0x0F)


def encode_bcd(value: int) -> int:
    """Кодирует целое число (0-99) в BCD."""
    return ((value // 10) << 4) | (value % 10)


def bytes_to_int_be(data: bytes) -> int:
    """Преобразует массив байт (big-endian) в целое число."""
    result = 0
    for b in data:
        result = (result << 8) | b
    return result


def bytes_to_int_le(data: bytes) -> int:
    """Преобразует массив байт (little-endian) в целое число."""
    result = 0
    for i, b in enumerate(data):
        result |= (b << (8 * i))
    return result


def int_to_bytes_be(value: int, length: int) -> bytes:
    """Преобразует целое число в массив байт (big-endian)."""
    result = bytearray(length)
    for i in range(length - 1, -1, -1):
        result[i] = value & 0xFF
        value >>= 8
    return bytes(result)


def int_to_bytes_le(value: int, length: int) -> bytes:
    """Преобразует целое число в массив байт (little-endian)."""
    result = bytearray(length)
    for i in range(length):
        result[i] = value & 0xFF
        value >>= 8
    return bytes(result)


# --- Дополнительные формулы и утилиты для протокола ---


def compute_near_zone(frequency_mhz: float, diameter_mm: float,
                      velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Вычисляет длину ближней зоны преобразователя (мм).
    N = D^2 * f / (4 * v)
    """
    wavelength_mm = velocity_mps / (frequency_mhz * 1000.0)
    return (diameter_mm ** 2) / (4.0 * wavelength_mm)


def compute_beam_divergence(frequency_mhz: float, diameter_mm: float,
                             velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Вычисляет угол расхождения луча (градусы).
    sin(alpha) = 1.22 * lambda / D
    """
    wavelength_mm = velocity_mps / (frequency_mhz * 1000.0)
    sin_alpha = 1.22 * wavelength_mm / diameter_mm
    if abs(sin_alpha) > 1.0:
        return 90.0
    return math.degrees(math.asin(sin_alpha))


def compute_wavelength(frequency_mhz: float, velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Вычисляет длину волны (мм) для заданной частоты и скорости."""
    return velocity_mps / (frequency_mhz * 1000.0)


def compute_acoustic_impedance(density_kg_m3: float, velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Вычисляет акустический импеданс (кг/(м^2*с))."""
    return density_kg_m3 * velocity_mps


def compute_reflection_coeff(z1: float, z2: float) -> float:
    """Вычисляет коэффициент отражения на границе двух сред."""
    if (z1 + z2) == 0:
        return 0.0
    return (z2 - z1) / (z2 + z1)


def compute_attenuation_db(distance_mm: float, alpha_db_per_mm: float = 0.01) -> float:
    """Вычисляет затухание сигнала (дБ) на заданном расстоянии."""
    return 2.0 * distance_mm * alpha_db_per_mm  # туда-обратно


def compute_snell_angle(angle_deg: float, v1_mps: int, v2_mps: int) -> float:
    """Вычисляет угол преломления по закону Снеллиуса (градусы)."""
    sin_alpha = math.sin(math.radians(angle_deg))
    sin_beta = sin_alpha * v2_mps / v1_mps
    if abs(sin_beta) > 1.0:
        return 90.0  # полное внутреннее отражение
    return math.degrees(math.asin(sin_beta))


def compute_dgs_distance(s_eq_mm2: float, frequency_mhz: float,
                          diameter_mm: float, gain_db: float,
                          velocity_mps: int = DEFAULT_VELOCITY_MPS) -> float:
    """Оценка расстояния по DGS-диаграмме (упрощенная формула)."""
    near_zone = compute_near_zone(frequency_mhz, diameter_mm, velocity_mps)
    wavelength = compute_wavelength(frequency_mhz, velocity_mps)
    if s_eq_mm2 <= 0:
        return 0.0
    # Упрощенная формула: r = sqrt(S_eq / (pi * lambda)) * 10^(gain/40)
    r = math.sqrt(s_eq_mm2 / (math.pi * wavelength))
    r *= 10.0 ** (gain_db / 40.0)
    return r


# --- Дополнительные структуры данных для TypeVar ---


@dataclass
class PEPFrequencyEntry:
    """Запись таблицы частот ПЭП (пьезоэлектрического преобразователя)."""
    index: int = 0
    frequency_mhz: float = 2.5
    diameter_mm: float = 12.0
    angle_deg: int = 0
    name: str = ""


# Таблицы частот ПЭП (4 набора по info_byte bits 4+5)
PEP_FREQUENCY_TABLES: Dict[int, List[PEPFrequencyEntry]] = {
    0: [
        PEPFrequencyEntry(0, 1.25, 20.0, 0, "1.25 MHz / 20mm / 0 deg"),
        PEPFrequencyEntry(1, 1.8, 18.0, 0, "1.8 MHz / 18mm / 0 deg"),
        PEPFrequencyEntry(2, 2.5, 12.0, 0, "2.5 MHz / 12mm / 0 deg"),
        PEPFrequencyEntry(3, 2.5, 12.0, 45, "2.5 MHz / 12mm / 45 deg"),
        PEPFrequencyEntry(4, 2.5, 12.0, 50, "2.5 MHz / 12mm / 50 deg"),
        PEPFrequencyEntry(5, 2.5, 12.0, 65, "2.5 MHz / 12mm / 65 deg"),
        PEPFrequencyEntry(6, 5.0, 6.0, 0, "5.0 MHz / 6mm / 0 deg"),
        PEPFrequencyEntry(7, 5.0, 6.0, 70, "5.0 MHz / 6mm / 70 deg"),
    ],
    1: [
        PEPFrequencyEntry(0, 1.25, 24.0, 0, "1.25 MHz / 24mm / 0 deg"),
        PEPFrequencyEntry(1, 2.5, 14.0, 0, "2.5 MHz / 14mm / 0 deg"),
        PEPFrequencyEntry(2, 2.5, 14.0, 40, "2.5 MHz / 14mm / 40 deg"),
        PEPFrequencyEntry(3, 2.5, 14.0, 50, "2.5 MHz / 14mm / 50 deg"),
        PEPFrequencyEntry(4, 2.5, 14.0, 65, "2.5 MHz / 14mm / 65 deg"),
        PEPFrequencyEntry(5, 5.0, 8.0, 0, "5.0 MHz / 8mm / 0 deg"),
        PEPFrequencyEntry(6, 5.0, 8.0, 50, "5.0 MHz / 8mm / 50 deg"),
        PEPFrequencyEntry(7, 5.0, 8.0, 70, "5.0 MHz / 8mm / 70 deg"),
    ],
    2: [
        PEPFrequencyEntry(0, 1.8, 20.0, 0, "1.8 MHz / 20mm / 0 deg (rail)"),
        PEPFrequencyEntry(1, 2.5, 12.0, 0, "2.5 MHz / 12mm / 0 deg (rail)"),
        PEPFrequencyEntry(2, 2.5, 12.0, 42, "2.5 MHz / 12mm / 42 deg (rail)"),
        PEPFrequencyEntry(3, 2.5, 12.0, 55, "2.5 MHz / 12mm / 55 deg (rail)"),
        PEPFrequencyEntry(4, 2.5, 12.0, 70, "2.5 MHz / 12mm / 70 deg (rail)"),
        PEPFrequencyEntry(5, 5.0, 6.0, 0, "5.0 MHz / 6mm / 0 deg (rail)"),
        PEPFrequencyEntry(6, 5.0, 6.0, 55, "5.0 MHz / 6mm / 55 deg (rail)"),
        PEPFrequencyEntry(7, 5.0, 6.0, 70, "5.0 MHz / 6mm / 70 deg (rail)"),
    ],
    3: [
        PEPFrequencyEntry(0, 2.5, 10.0, 0, "2.5 MHz / 10mm / 0 deg (pipe)"),
        PEPFrequencyEntry(1, 2.5, 10.0, 45, "2.5 MHz / 10mm / 45 deg (pipe)"),
        PEPFrequencyEntry(2, 2.5, 10.0, 60, "2.5 MHz / 10mm / 60 deg (pipe)"),
        PEPFrequencyEntry(3, 4.0, 8.0, 0, "4.0 MHz / 8mm / 0 deg (pipe)"),
        PEPFrequencyEntry(4, 4.0, 8.0, 45, "4.0 MHz / 8mm / 45 deg (pipe)"),
        PEPFrequencyEntry(5, 5.0, 6.0, 0, "5.0 MHz / 6mm / 0 deg (pipe)"),
        PEPFrequencyEntry(6, 5.0, 6.0, 45, "5.0 MHz / 6mm / 45 deg (pipe)"),
        PEPFrequencyEntry(7, 10.0, 4.0, 0, "10.0 MHz / 4mm / 0 deg (pipe)"),
    ],
}


def get_pep_frequency_table(info_byte: int) -> List[PEPFrequencyEntry]:
    """Возвращает таблицу частот ПЭП по битам 4+5 info_byte."""
    table_idx = (info_byte & INFO_BITS_PEP_FREQ) >> 4
    return PEP_FREQUENCY_TABLES.get(table_idx, PEP_FREQUENCY_TABLES[0])


# --- Константы материалов ---

MATERIAL_VELOCITIES: Dict[str, Tuple[int, int]] = {
    # Материал: (скорость продольной волны м/с, скорость поперечной волны м/с)
    "steel_carbon": (5920, 3230),
    "steel_stainless": (5740, 3100),
    "aluminum": (6320, 3130),
    "copper": (4700, 2260),
    "cast_iron": (4600, 2200),
    "titanium": (6100, 3120),
    "glass": (5660, 3280),
    "water": (1480, 0),
    "plexiglass": (2730, 1430),
    "polyethylene": (2340, 0),
}


def get_material_velocity(material: str, wave_type: str = "longitudinal") -> int:
    """Возвращает скорость звука в материале (м/с)."""
    if material not in MATERIAL_VELOCITIES:
        return DEFAULT_VELOCITY_MPS
    vl, vt = MATERIAL_VELOCITIES[material]
    if wave_type == "transverse" or wave_type == "shear":
        return vt
    return vl


# --- Расширенный парсер TLV с DLL output format ---


@dataclass
class DLLOutputRecord:
    """Формат вывода DLL: [decoder_type(1B), n_fields(1B), offset_table, field_data, raw]."""
    decoder_type: int = 0
    n_fields: int = 0
    field_offsets: List[Tuple[int, int]] = field(default_factory=list)  # (offset, size)
    field_data: bytes = b""
    raw_record: bytes = b""

    @classmethod
    def from_bytes(cls, data: bytes) -> Optional['DLLOutputRecord']:
        """Разбирает DLL output record из массива байт."""
        if len(data) < 2:
            return None
        rec = cls()
        rec.decoder_type = data[0]
        rec.n_fields = data[1]
        # Таблица смещений: 4 байта на поле (LE16 offset + LE16 size)
        table_size = rec.n_fields * 4
        if len(data) < 2 + table_size:
            return None
        rec.field_offsets = []
        for i in range(rec.n_fields):
            off = struct.unpack_from('<H', data, 2 + i * 4)[0]
            sz = struct.unpack_from('<H', data, 2 + i * 4 + 2)[0]
            rec.field_offsets.append((off, sz))
        # Данные полей начинаются после таблицы
        field_start = 2 + table_size
        # Определяем конец field_data по максимальному offset+size
        if rec.field_offsets:
            max_end = max(off + sz for off, sz in rec.field_offsets)
            rec.field_data = data[field_start:field_start + max_end]
            raw_start = field_start + max_end
        else:
            rec.field_data = b""
            raw_start = field_start
        rec.raw_record = data[raw_start:]
        return rec

    def get_field_int(self, field_idx: int) -> int:
        """Извлекает целочисленное поле (BIG-ENDIAN) по индексу."""
        if field_idx >= len(self.field_offsets):
            return 0
        offset, size = self.field_offsets[field_idx]
        if offset + size > len(self.field_data):
            return 0
        raw = self.field_data[offset:offset + size]
        return bytes_to_int_be(raw)

    def get_field_string(self, field_idx: int) -> str:
        """Извлекает строковое поле (NUL-terminated, zero-padded) по индексу."""
        if field_idx >= len(self.field_offsets):
            return ""
        offset, size = self.field_offsets[field_idx]
        if offset + size > len(self.field_data):
            return ""
        raw = self.field_data[offset:offset + size]
        # NUL-terminated
        nul_pos = raw.find(b'\x00')
        if nul_pos >= 0:
            raw = raw[:nul_pos]
        try:
            return raw.decode('cp1251')
        except UnicodeDecodeError:
            return raw.hex()

    def get_field_date(self, field_idx: int) -> Optional[_dt.date]:
        """Извлекает поле даты (3 байта: day, month, year%100) по индексу."""
        if field_idx >= len(self.field_offsets):
            return None
        offset, size = self.field_offsets[field_idx]
        if size < 3 or offset + 3 > len(self.field_data):
            return None
        day = self.field_data[offset]
        month = self.field_data[offset + 1]
        year = 2000 + self.field_data[offset + 2]
        try:
            return _dt.date(year, month, day)
        except (ValueError, OverflowError):
            return None


# --- Flash memory layout ---


@dataclass
class FlashLayout:
    """Разметка flash-памяти прибора."""
    total_size: int = 128 * 1024  # 128KB по умолчанию
    block_size: int = 256
    header_offset: int = 0
    data_offset: int = 256
    record_count: int = 0

    @classmethod
    def from_info_byte(cls, info_byte: int) -> 'FlashLayout':
        """Создает разметку flash на основе info_byte."""
        layout = cls()
        if info_byte & INFO_BIT_EXTENDED_FLASH:
            layout.total_size = 512 * 1024  # 512KB
        layout.record_count = (layout.total_size - layout.data_offset) // layout.block_size
        return layout


def parse_flash_dump(data: bytes, layout: Optional[FlashLayout] = None) -> List[TLVRecord]:
    """Разбирает дамп flash-памяти на TLV-записи."""
    if layout is None:
        layout = FlashLayout()
    records: List[TLVRecord] = []
    offset = layout.data_offset
    while offset < min(len(data), layout.total_size):
        if offset + TLV_HEADER_SIZE > len(data):
            break
        tag = struct.unpack_from('<H', data, offset)[0]
        if tag == TLV_PADDING_TAG or tag == 0x0000:
            offset += layout.block_size
            continue
        length = struct.unpack_from('<H', data, offset + 2)[0]
        if length == 0 or length > MAX_PAYLOAD:
            offset += layout.block_size
            continue
        end = offset + TLV_HEADER_SIZE + length
        if end > len(data):
            break
        body = data[offset + TLV_HEADER_SIZE:end]
        raw = data[offset:end]
        records.append(TLVRecord(tag=tag, length=length, body=body, raw=raw))
        # Выравнивание на размер блока
        offset = ((end + layout.block_size - 1) // layout.block_size) * layout.block_size
    return records


# --- Утилиты для работы с паспортами и строками ---


def passport_encode(text: str) -> bytes:
    """Кодирует текстовую строку в формат паспорта (обратный PASSPORT_LUT)."""
    # Обратная таблица
    reverse_lut: Dict[str, int] = {}
    for i, ch in enumerate(PASSPORT_LUT):
        if ch != '\x00' and ch not in reverse_lut:
            reverse_lut[ch] = i
    encoded: List[int] = []
    for ch in reversed(text):
        idx = reverse_lut.get(ch, 0)
        encoded.append(idx)
    # Дополняем нулями
    encoded.reverse()
    return bytes(encoded)


def format_record_summary(record: DecodedRecord) -> str:
    """Форматирует краткую сводку по записи для отображения в списке."""
    parts: List[str] = []
    parts.append(f"[{record.category_name}]")
    if record.date:
        parts.append(str(record.date))
    if record.time:
        parts.append(str(record.time))
    if record.passport_primary:
        parts.append(record.passport_primary)
    if record.defect_flag:
        parts.append(f"DEF={record.defect_flag}")
    return ' | '.join(parts)


def format_settings_summary(settings: UD2_AcousticSettings) -> str:
    """Форматирует краткую сводку настроек прибора."""
    return (f"G={settings.gain_db}dB V={settings.velocity_mps}m/s "
            f"A={settings.angle_deg}deg Gate={settings.gate_start}-"
            f"{settings.gate_start + settings.gate_width} "
            f"Thr={settings.threshold_pct}%")


# --- Расширенный ARD расчет ---


@dataclass
class ARDResult:
    """Результат вычисления ARD (Amplitude-Range-Distance) вердикта."""
    delta_db: float = 0.0
    verdict: str = ""
    a_peak: float = 0.0
    a_gate: float = 0.0
    s_fact: float = 0.0
    s_req: float = 0.0
    color: Tuple[int, int, int] = (0, 0, 0)


def compute_ard_full(samples: bytes, settings: UD2_AcousticSettings,
                     s_req: float = 0.0, frequency_mhz: float = 2.5,
                     diameter_mm: float = 12.0) -> ARDResult:
    """Полный расчет ARD вердикта с учетом всех параметров.
    
    delta_db = 20*log10(a_peak/a_gate) + (s_fact - s_req)
    delta >= 0: БРАК (красный)
    -6 <= delta < 0: КОНТРОЛЬ (желтый)
    delta < -6: ГОДЕН/ПОИСК (зеленый)
    """
    result = ARDResult()
    result.s_req = s_req

    # Находим пик в зоне строба
    if not samples or settings.gate_width == 0:
        result.verdict = "НЕТ ДАННЫХ"
        return result

    peak_val, peak_idx = ascan_peak_in_gate(
        samples, settings.gate_start, settings.gate_width
    )
    result.a_peak = peak_val / 255.0 if peak_val > 0 else 0.0

    # Уровень строба (порог)
    threshold_level = settings.threshold_pct * 255 / 100
    result.a_gate = threshold_level / 255.0 if threshold_level > 0 else 0.001

    # Вычисляем эквивалентную площадь дефекта (упрощенно)
    depth_mm = ascan_sample_to_depth_mm(peak_idx, settings.velocity_mps)
    near_zone = compute_near_zone(frequency_mhz, diameter_mm, settings.velocity_mps)
    if near_zone > 0 and depth_mm > 0:
        # s_fact = pi * D^2 / 4 * (depth / near_zone) * (a_peak / a_gate)
        s_area = math.pi * (diameter_mm ** 2) / 4.0
        ratio = depth_mm / near_zone
        amp_ratio = result.a_peak / result.a_gate if result.a_gate > 0 else 0
        result.s_fact = s_area * ratio * amp_ratio
    else:
        result.s_fact = 0.0

    # Формула ARD
    if result.a_gate > 0 and result.a_peak > 0:
        result.delta_db = 20.0 * math.log10(result.a_peak / result.a_gate) + (result.s_fact - s_req)
    else:
        result.delta_db = -100.0

    # Вердикт
    if result.delta_db >= 0:
        result.verdict = "БРАК"
        result.color = (255, 0, 0)
    elif result.delta_db >= -6.0:
        result.verdict = "КОНТРОЛЬ"
        result.color = (255, 255, 0)
    else:
        result.verdict = "ГОДЕН"
        result.color = (0, 255, 0)

    return result


# --- Вспомогательные функции для экрана LCD ---


def parse_lcd_screen(data: bytes, info_byte: int = 0) -> Dict[str, Any]:
    """Разбирает данные экрана LCD прибора."""
    is_wagon = bool(info_byte & INFO_BIT_WAGON_LCD)
    expected_size = LCD_SIZE_WAGON if is_wagon else LCD_SIZE_NORMAL
    result: Dict[str, Any] = {
        "is_wagon": is_wagon,
        "expected_size": expected_size,
        "actual_size": len(data),
        "valid": len(data) >= expected_size,
        "width": 240,
        "height": expected_size // 30,  # stride = 30 bytes
    }
    if result["valid"]:
        result["pixels"] = lcd_bitmap_to_pixels(data[:expected_size])
    return result


# --- Дополнительные утилиты отчетов ---


def format_verdict_html(verdict: str) -> str:
    """Форматирует вердикт как HTML с цветовым кодированием."""
    if verdict in ("БРАК",):
        return f'<span style="color: red; font-weight: bold;">{verdict}</span>'
    elif verdict in ("КОНТРОЛЬ",):
        return f'<span style="color: orange; font-weight: bold;">{verdict}</span>'
    elif verdict in ("ГОДЕН", "ПОИСК"):
        return f'<span style="color: green; font-weight: bold;">{verdict}</span>'
    return f'<span>{verdict}</span>'


def compute_statistics(records: List[DecodedRecord]) -> Dict[str, Any]:
    """Вычисляет статистику по набору записей."""
    stats: Dict[str, Any] = {
        "total": len(records),
        "by_category": {},
        "by_verdict": {"БРАК": 0, "КОНТРОЛЬ": 0, "ГОДЕН": 0, "НЕТ ДАННЫХ": 0},
        "date_range": (None, None),
    }
    dates: List[_dt.date] = []
    for rec in records:
        cat = rec.category_name
        stats["by_category"][cat] = stats["by_category"].get(cat, 0) + 1
        verdict = rec.fields.get("verdict", "НЕТ ДАННЫХ")
        if verdict in stats["by_verdict"]:
            stats["by_verdict"][verdict] += 1
        if rec.date:
            dates.append(rec.date)
    if dates:
        stats["date_range"] = (min(dates), max(dates))
    return stats
